#!/usr/bin/env python3
"""Test suite for parlay full-leg-set backfill (Testing strategy #9) — money-safety.

Tests:
 - Two terminal prop legs (WIN+LOSS) persisted; MANUAL REVIEW parlay over both
   → parlay resolves to LOSS from the FULL persisted leg set (not WIN/PENDING)
 - A parlay with one still-missing/non-terminal leg → ABSTAINS (stays at prior result)

Run from scripts/: python3 test_parlay_leg_backfill.py
"""
from __future__ import annotations

import importlib
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

_SCRIPTS = Path(__file__).parent
if str(_SCRIPTS) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS))

runner = importlib.import_module("sports_system_runner")

grade_game_in_workbook = runner.grade_game_in_workbook
result_headers = runner.result_headers
ensure_ws_columns = runner.ensure_ws_columns
RESULT_HEADERS = runner.RESULT_HEADERS
PROPS_HEADERS = runner.PROPS_HEADERS
PARLAY_HEADERS = runner.PARLAY_HEADERS
save_workbook_atomic = runner.save_workbook_atomic


# ---------------------------------------------------------------------------
# Player stats: one player with Hits=3 (WIN for Over 2.5) and another with Hits=1 (LOSS for Over 2.5)
# ---------------------------------------------------------------------------

_PLAYER_STATS = {
    "carlos correa": {
        "batting": {
            "hits": 3.0, "runs": 1.0, "rbis": 2.0, "homeruns": 0.0,
            "walks": 1.0, "strikeouts": 1.0,
            "_hit_counts": {"single": 2, "double": 1, "triple": 0, "home-run": 0},
        },
        "pitching": {},
    },
    "jose abreu": {
        "batting": {
            "hits": 1.0, "runs": 0.0, "rbis": 0.0, "homeruns": 0.0,
            "walks": 0.0, "strikeouts": 2.0,
            "_hit_counts": {"single": 1, "double": 0, "triple": 0, "home-run": 0},
        },
        "pitching": {},
    },
}

_GAME = {
    "event_id": "test-event-parlay-001",
    "id": "test-event-parlay-001",
    "status": "final",
    "status_name": "STATUS_FINAL",
    "completed": True,
    "home_team": "Twins",
    "away_team": "Tigers",
    "home_score": 5,
    "away_score": 3,
}

_GAME_INCOMPLETE = {
    "event_id": "test-event-parlay-002",
    "id": "test-event-parlay-002",
    "status": "final",
    "status_name": "STATUS_FINAL",
    "completed": True,
    "home_team": "Cubs",
    "away_team": "Cardinals",
    "home_score": 4,
    "away_score": 2,
}


def _add_result_row(ws, date: str, sport: str, ref: str, result: str) -> None:
    """Append a Results row for pre-seeding terminal leg results."""
    cols = result_headers(ws)
    row_data = {h: None for h in RESULT_HEADERS}
    row_data["Date"] = date
    row_data["Sport"] = sport
    row_data["Pick Ref"] = ref
    row_data["Result"] = result
    row_data["Units"] = 1.0
    row_data["PnL"] = 0.0
    row_data["Pick Type"] = "PROP"
    ws.append([row_data.get(h) for h in RESULT_HEADERS])


def _count_rows(wb, sheet: str, date: str, sport: str, ref: str) -> int:
    ws = wb[sheet]
    cols = result_headers(ws)
    count = 0
    for r in range(2, ws.max_row + 1):
        if (str(ws.cell(r, cols["Date"]).value or "")[:10] == date
                and str(ws.cell(r, cols["Sport"]).value or "") == sport
                and str(ws.cell(r, cols["Pick Ref"]).value or "") == ref):
            count += 1
    return count


def _get_result(wb, sheet: str, date: str, sport: str, ref: str) -> str | None:
    ws = wb[sheet]
    cols = result_headers(ws)
    for r in range(2, ws.max_row + 1):
        if (str(ws.cell(r, cols["Date"]).value or "")[:10] == date
                and str(ws.cell(r, cols["Sport"]).value or "") == sport
                and str(ws.cell(r, cols["Pick Ref"]).value or "") == ref):
            return str(ws.cell(r, cols["Result"]).value or "")
    return None


class TestParlayFullLegSetBackfill(unittest.TestCase):
    """Parlay verdict must use FULL persisted leg set, not only this-run graded legs."""

    def setUp(self) -> None:
        self._tmpdir = tempfile.TemporaryDirectory()
        tmp = Path(self._tmpdir.name)
        self._pnl_dir = tmp / "pnl"
        self._pnl_dir.mkdir(parents=True)
        self._data_dir = tmp / "data"
        self._mlb_dir = self._data_dir / "mlb"
        self._mlb_dir.mkdir(parents=True)
        self._locks_dir = tmp / "locks"
        self._locks_dir.mkdir()
        self._backups_dir = self._data_dir / "backups" / "workbooks"
        self._backups_dir.mkdir(parents=True)
        self._log_dir = self._pnl_dir / "logs"
        self._log_dir.mkdir()

        self._patches = [
            patch.object(runner, "PNL_DIR", self._pnl_dir),
            patch.object(runner, "NBA_DIR", tmp / "data" / "nba"),
            patch.object(runner, "MLB_DIR", self._mlb_dir),
            patch.object(runner, "DATA", self._data_dir),
            patch.object(runner, "LOG_DIR", self._log_dir),
            patch.object(runner, "WORKBOOK_LOCK_DIR", self._locks_dir),
            patch.object(runner, "WORKBOOK_BACKUP_DIR", self._backups_dir),
            patch.object(runner, "BANKROLL", self._pnl_dir / "bankroll.json"),
            patch.object(runner, "GAME_STATUS_CACHE_FILE", self._pnl_dir / "game_status_cache.json"),
        ]
        for p in self._patches:
            p.start()
        (tmp / "data" / "nba").mkdir(parents=True, exist_ok=True)

        self._fn_patches = [
            patch.object(runner, "espn_player_stats_by_event", return_value=_PLAYER_STATS),
            patch.object(runner, "send_telegram", return_value=None),
            patch.object(runner, "obsidian_update_results_section", return_value=None),
            patch.object(runner, "obsidian_update_bankroll_files", return_value=None),
        ]
        for p in self._fn_patches:
            p.start()

    def tearDown(self) -> None:
        for p in self._fn_patches:
            p.stop()
        for p in self._patches:
            p.stop()
        self._tmpdir.cleanup()

    def test_parlay_resolves_to_loss_from_persisted_legs(self) -> None:
        """MANUAL REVIEW parlay over a persisted WIN + LOSS must resolve to LOSS (full leg set)."""
        date = "2026-06-08"
        # Create workbook
        path = runner.ensure_workbook("mlb", date)
        from openpyxl import load_workbook as lw
        wb = lw(path)

        # Seed Results sheet with two TERMINAL prop legs (WIN + LOSS)
        results_ws = wb["Results"]
        ref_win = "PROP:Carlos Correa Hits 2.5"
        ref_loss = "PROP:Jose Abreu Hits 2.5"
        _add_result_row(results_ws, date, "MLB", ref_win, "WIN")
        _add_result_row(results_ws, date, "MLB", ref_loss, "LOSS")

        # Seed MANUAL REVIEW parlay row in Results sheet (the parlay we want to re-grade)
        parlay_ref = "PARLAY:SGP-Twins-001"
        _add_result_row(results_ws, date, "MLB", parlay_ref, "MANUAL REVIEW")

        # Seed Correlated Parlays sheet with a parlay entry referencing those legs
        parlays_ws = wb["Correlated Parlays"]
        parlay_cols = ensure_ws_columns(parlays_ws, PARLAY_HEADERS)
        parlays_ws.append([None] * len(PARLAY_HEADERS))
        r = parlays_ws.max_row
        parlays_ws.cell(r, parlay_cols["Date"]).value = date
        parlays_ws.cell(r, parlay_cols["Sport"]).value = "MLB"
        parlays_ws.cell(r, parlay_cols["Parlay Name"]).value = "SGP-Twins-001"
        parlays_ws.cell(r, parlay_cols["Legs"]).value = f"{ref_win}|{ref_loss}"
        parlays_ws.cell(r, parlay_cols["Units"]).value = 0.5

        # Also seed Props so the game_matches_row can link props to the game
        # But since legs are TERMINAL, they'll be skipped in props loop
        # The parlay must be graded from persisted Results
        props_ws = wb["Props"]
        props_cols = ensure_ws_columns(props_ws, PROPS_HEADERS)
        # Add Carlos Correa prop (will be SKIPPED because WIN is terminal)
        props_ws.append([None] * len(PROPS_HEADERS))
        r2 = props_ws.max_row
        props_ws.cell(r2, props_cols["Date"]).value = date
        props_ws.cell(r2, props_cols["Sport"]).value = "MLB"
        props_ws.cell(r2, props_cols["Player Name"]).value = "Carlos Correa"
        props_ws.cell(r2, props_cols["Stat"]).value = "Hits"
        props_ws.cell(r2, props_cols["Line"]).value = 2.5
        props_ws.cell(r2, props_cols["Opponent/Description"]).value = "Over"
        # Add Jose Abreu prop (will be SKIPPED because LOSS is terminal)
        props_ws.append([None] * len(PROPS_HEADERS))
        r3 = props_ws.max_row
        props_ws.cell(r3, props_cols["Date"]).value = date
        props_ws.cell(r3, props_cols["Sport"]).value = "MLB"
        props_ws.cell(r3, props_cols["Player Name"]).value = "Jose Abreu"
        props_ws.cell(r3, props_cols["Stat"]).value = "Hits"
        props_ws.cell(r3, props_cols["Line"]).value = 2.5
        props_ws.cell(r3, props_cols["Opponent/Description"]).value = "Over"

        save_workbook_atomic(wb, path)

        # Re-grade — both prop legs are terminal (skipped), parlay should use persisted legs
        grade_game_in_workbook("mlb", _GAME, date, dry_run=False)

        wb2 = lw(path)
        parlay_result = _get_result(wb2, "Results", date, "MLB", parlay_ref)
        self.assertIn(
            parlay_result, {"LOSS", "WIN", "PUSH"},
            f"Parlay should have been graded to a terminal result, got {parlay_result!r}"
        )
        # Because one leg is LOSS, the parlay MUST be LOSS
        self.assertEqual(
            parlay_result, "LOSS",
            f"Parlay with WIN+LOSS legs must resolve to LOSS, got {parlay_result!r}"
        )

    def test_parlay_abstains_when_leg_missing(self) -> None:
        """A parlay with a still-missing/non-terminal leg must ABSTAIN (stay prior result)."""
        date = "2026-06-08"
        path = runner.ensure_workbook("mlb", date)
        from openpyxl import load_workbook as lw
        wb = lw(path)

        # Seed only ONE terminal prop leg (WIN); the other leg is absent from Results
        results_ws = wb["Results"]
        ref_win = "PROP:Carlos Correa Hits 2.5"
        ref_missing = "PROP:Jose Abreu Hits 2.5"  # NOT in Results
        _add_result_row(results_ws, date, "MLB", ref_win, "WIN")

        # Parlay over both legs — second leg is missing → must ABSTAIN
        parlay_ref = "PARLAY:SGP-Twins-002"
        # Pre-seed the parlay as MANUAL REVIEW so we can detect if it stays or changes
        _add_result_row(results_ws, date, "MLB", parlay_ref, "MANUAL REVIEW")

        parlays_ws = wb["Correlated Parlays"]
        parlay_cols = ensure_ws_columns(parlays_ws, PARLAY_HEADERS)
        parlays_ws.append([None] * len(PARLAY_HEADERS))
        r = parlays_ws.max_row
        parlays_ws.cell(r, parlay_cols["Date"]).value = date
        parlays_ws.cell(r, parlay_cols["Sport"]).value = "MLB"
        parlays_ws.cell(r, parlay_cols["Parlay Name"]).value = "SGP-Twins-002"
        parlays_ws.cell(r, parlay_cols["Legs"]).value = f"{ref_win}|{ref_missing}"
        parlays_ws.cell(r, parlay_cols["Units"]).value = 0.5

        # Seed Props with only one prop (Carlos Correa — TERMINAL → skipped)
        # Jose Abreu prop NOT added → truly missing
        props_ws = wb["Props"]
        props_cols = ensure_ws_columns(props_ws, PROPS_HEADERS)
        props_ws.append([None] * len(PROPS_HEADERS))
        r2 = props_ws.max_row
        props_ws.cell(r2, props_cols["Date"]).value = date
        props_ws.cell(r2, props_cols["Sport"]).value = "MLB"
        props_ws.cell(r2, props_cols["Player Name"]).value = "Carlos Correa"
        props_ws.cell(r2, props_cols["Stat"]).value = "Hits"
        props_ws.cell(r2, props_cols["Line"]).value = 2.5
        props_ws.cell(r2, props_cols["Opponent/Description"]).value = "Over"

        save_workbook_atomic(wb, path)

        # Re-grade: one leg terminal (WIN), one leg absent → parlay must ABSTAIN
        grade_game_in_workbook("mlb", _GAME, date, dry_run=False)

        wb2 = lw(path)
        parlay_result = _get_result(wb2, "Results", date, "MLB", parlay_ref)
        # Parlay should remain MANUAL REVIEW (abstain = stay at prior result)
        # It must NOT be graded to WIN or any terminal value
        self.assertNotIn(
            parlay_result, {"WIN", "LOSS", "PUSH"},
            f"Parlay with missing leg must ABSTAIN, got {parlay_result!r}"
        )


if __name__ == "__main__":
    unittest.main()
