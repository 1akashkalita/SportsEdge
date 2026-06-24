#!/usr/bin/env python3
"""test_dashboard_views.py — VIEW-01/02/03 tests for Phase 2 view accessors and routes.

Covers:
- TestTodayBoard: unit tests for get_today_board() using synthetic in-memory workbooks
- TestSlipsAccessor: unit tests for get_all_slips() using in-memory master_pnl-shaped workbook
- TestHistoryAccessor: unit tests for get_history_data() using in-memory workbook
- TestRoutes: route smoke tests for /, /slips, /history via Flask test_client()
"""
from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path
from typing import Any
from unittest.mock import patch

from openpyxl import Workbook

SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

import dashboard_data  # noqa: E402
import dashboard  # noqa: E402


# ---------------------------------------------------------------------------
# Fixture builders — synthetic in-memory workbooks using exact runner headers
# ---------------------------------------------------------------------------

def _make_today_wb(
    today: str,
    picks_rows: list[dict[str, Any]] | None = None,
    skipped_rows: list[dict[str, Any]] | None = None,
) -> Workbook:
    """Build a synthetic sport workbook with Picks and Skipped Picks sheets.

    Uses the exact column headers from sports_system_runner.py PICKS_HEADERS
    (lines 277-283) and SKIPPED_PICK_HEADERS (line 295).
    """
    wb = Workbook()

    # ---- Picks sheet (PICKS_HEADERS from runner lines 277-283) ----
    ws_picks = wb.active
    ws_picks.title = "Picks"
    picks_headers = [
        "Date", "Sport", "Pick Type", "Selection", "Line", "Confidence", "Units",
        "Status", "Platform", "Player/Team", "Model Projection", "Edge",
        "Model Over Probability", "EV", "Injury Flag",
    ]
    ws_picks.append(picks_headers)
    for row in (picks_rows or []):
        ws_picks.append([row.get(h) for h in picks_headers])

    # ---- Skipped Picks sheet (SKIPPED_PICK_HEADERS from runner line 295) ----
    ws_skipped = wb.create_sheet("Skipped Picks")
    skipped_headers = [
        "Date", "Sport", "Pick", "Gate Failed", "Reason",
        "What Edge Would Have Been", "Pick Type", "Player/Team", "Line",
        "Probability", "EV", "Platform",
    ]
    ws_skipped.append(skipped_headers)
    for row in (skipped_rows or []):
        ws_skipped.append([row.get(h) for h in skipped_headers])

    return wb


def _make_master_pnl_wb(
    slip_rows: list[dict[str, Any]] | None = None,
    pick_history_rows: list[dict[str, Any]] | None = None,
    chart_rows: list[dict[str, Any]] | None = None,
) -> Workbook:
    """Build a synthetic master_pnl workbook with Slip History, Pick History,
    and Bankroll Chart Data sheets.

    Uses exact column headers from slip_payouts.py SLIP_HISTORY_HEADERS (lines 18-24).
    """
    wb = Workbook()

    # ---- Slip History sheet (SLIP_HISTORY_HEADERS from slip_payouts.py lines 18-24) ----
    ws_slips = wb.active
    ws_slips.title = "Slip History"
    slip_headers = [
        "Date", "Slip ID", "Platform", "Slip Type", "Number of Legs", "Legs",
        "Slip Result", "Standard Payout Multiplier", "Estimated Payout Multiplier",
        "Net PnL", "Gross Return", "Winning Legs", "Losing Legs",
        "Contains Demon", "Contains Goblin", "Notes",
    ]
    ws_slips.append(slip_headers)
    for row in (slip_rows or []):
        ws_slips.append([row.get(h) for h in slip_headers])

    # ---- Pick History sheet ----
    ws_history = wb.create_sheet("Pick History")
    history_headers = [
        "Date", "Sport", "Pick Ref", "Result", "Units", "PnL", "Graded At",
        "Notes", "Game", "Actual", "Platform", "Player/Team", "Pick",
        "Pick Type", "Line", "Odds", "Confidence Tier", "Model Projection",
        "Edge", "Model Over Probability", "EV",
    ]
    ws_history.append(history_headers)
    for row in (pick_history_rows or []):
        ws_history.append([row.get(h) for h in history_headers])

    # ---- Bankroll Chart Data sheet ----
    ws_chart = wb.create_sheet("Bankroll Chart Data")
    chart_headers = ["Date", "Bankroll", "ROI", "Updated At"]
    ws_chart.append(chart_headers)
    for row in (chart_rows or []):
        ws_chart.append([row.get(h) for h in chart_headers])

    return wb


# ---------------------------------------------------------------------------
# TestTodayBoard — unit tests for get_today_board() (VIEW-01)
# ---------------------------------------------------------------------------

class TestTodayBoard(unittest.TestCase):
    """VIEW-01 tests for get_today_board() — approved picks, skipped picks,
    lock tolerance, and EV coercion."""

    def test_approved_picks(self) -> None:
        """get_today_board() returns approved Picks rows date-filtered to today_str(),
        each with status_label == '✓ Approved'."""
        today = dashboard_data.today_str()
        picks_rows = [
            {
                "Date": today, "Sport": "NBA", "Pick Type": "PROP",
                "Selection": "LeBron James Over 25.5 Points", "Line": 25.5,
                "Confidence": "A", "Units": 3, "Status": "APPROVED",
                "Platform": "PrizePicks", "Player/Team": "LeBron James",
                "Model Projection": 27.1, "Edge": 1.5,
                "Model Over Probability": 0.63, "EV": 0.12, "Injury Flag": None,
            },
            # A different date row that should be filtered out
            {
                "Date": "2026-01-01", "Sport": "NBA", "Pick Type": "PROP",
                "Selection": "Old Pick Over 10.0 Points", "Line": 10.0,
                "Confidence": "B", "Units": 2, "Status": "APPROVED",
                "Platform": "PrizePicks", "Player/Team": "Someone",
                "Model Projection": 11.0, "Edge": 1.0,
                "Model Over Probability": 0.55, "EV": 0.05, "Injury Flag": None,
            },
        ]

        wb = _make_today_wb(today, picks_rows=picks_rows)
        with tempfile.TemporaryDirectory() as tmpdir:
            xlsx_path = Path(tmpdir) / f"nba_{today}.xlsx"
            wb.save(xlsx_path)

            orig_nba = dashboard_data.NBA_DIR
            orig_mlb = dashboard_data.MLB_DIR
            try:
                dashboard_data.NBA_DIR = Path(tmpdir)
                dashboard_data.MLB_DIR = Path(tmpdir)  # empty — no MLB workbook
                result = dashboard_data.get_today_board(date=today)
            finally:
                dashboard_data.NBA_DIR = orig_nba
                dashboard_data.MLB_DIR = orig_mlb

        self.assertEqual(result["date"], today)
        self.assertFalse(result["locked"])
        # Only the today-dated approved row should appear
        approved = result["approved"]
        self.assertEqual(len(approved), 1, f"Expected 1 approved pick for today, got {len(approved)}")
        self.assertEqual(approved[0]["status_label"], "✓ Approved")
        self.assertEqual(approved[0]["Selection"], "LeBron James Over 25.5 Points")

    def test_skipped_picks_gate_label(self) -> None:
        """A Skipped Picks row with Gate Failed 'GATE 1 — MINIMUM EDGE' yields
        status_label == 'Skip: MINIMUM EDGE'."""
        today = dashboard_data.today_str()
        skipped_rows = [
            {
                "Date": today, "Sport": "NBA",
                "Pick": "LeBron James Under 8.5 Assists",
                "Gate Failed": "GATE 1 — MINIMUM EDGE",
                "Reason": "Edge below threshold",
                "What Edge Would Have Been": 0.3, "Pick Type": "PROP",
                "Player/Team": "LeBron James", "Line": 8.5,
                "Probability": 0.48, "EV": 0.02, "Platform": "PrizePicks",
            },
        ]

        wb = _make_today_wb(today, skipped_rows=skipped_rows)
        with tempfile.TemporaryDirectory() as tmpdir:
            xlsx_path = Path(tmpdir) / f"nba_{today}.xlsx"
            wb.save(xlsx_path)

            orig_nba = dashboard_data.NBA_DIR
            orig_mlb = dashboard_data.MLB_DIR
            try:
                dashboard_data.NBA_DIR = Path(tmpdir)
                dashboard_data.MLB_DIR = Path(tmpdir)
                result = dashboard_data.get_today_board(date=today)
            finally:
                dashboard_data.NBA_DIR = orig_nba
                dashboard_data.MLB_DIR = orig_mlb

        skipped = result["skipped"]
        self.assertEqual(len(skipped), 1, f"Expected 1 skipped pick, got {len(skipped)}")
        self.assertEqual(skipped[0]["status_label"], "Skip: MINIMUM EDGE")

    def test_locked_state(self) -> None:
        """When read_sheet_rows returns None (locked workbook), get_today_board()
        returns locked=True and does not raise.

        Creates a real workbook file first so the accessor sees a present-but-locked
        file (file-existence check passes) then patches read_sheet_rows to return None
        to simulate a mid-write lock (WorkbookAccessError path, D-01).
        """
        today = dashboard_data.today_str()

        with tempfile.TemporaryDirectory() as tmpdir:
            # Create empty (but real) workbook file to pass the existence check
            nba_path = Path(tmpdir) / f"nba_{today}.xlsx"
            Workbook().save(nba_path)

            orig_nba = dashboard_data.NBA_DIR
            orig_mlb = dashboard_data.MLB_DIR
            try:
                dashboard_data.NBA_DIR = Path(tmpdir)
                dashboard_data.MLB_DIR = Path(tmpdir)
                # Patch read_sheet_rows to return None, simulating mid-write lock
                with patch.object(dashboard_data, "read_sheet_rows", return_value=None):
                    result = dashboard_data.get_today_board(date=today)
            finally:
                dashboard_data.NBA_DIR = orig_nba
                dashboard_data.MLB_DIR = orig_mlb

        self.assertTrue(result["locked"], "Expected locked=True when read_sheet_rows returns None")
        self.assertEqual(result["approved"], [], "Expected empty approved list when locked")
        self.assertEqual(result["skipped"], [], "Expected empty skipped list when locked")
        # Must not raise — return value is the only assertion needed

    def test_ev_coercion(self) -> None:
        """A skipped row with EV == 'unavailable' yields ev_float == None (no exception)."""
        today = dashboard_data.today_str()
        skipped_rows = [
            {
                "Date": today, "Sport": "NBA",
                "Pick": "James Harden Under 30.5 Points",
                "Gate Failed": "GATE 2 — MINIMUM PROBABILITY",
                "Reason": "Probability too low", "What Edge Would Have Been": 0.5,
                "Pick Type": "PROP", "Player/Team": "James Harden", "Line": 30.5,
                "Probability": "unavailable", "EV": "unavailable",
                "Platform": "PrizePicks",
            },
        ]

        wb = _make_today_wb(today, skipped_rows=skipped_rows)
        with tempfile.TemporaryDirectory() as tmpdir:
            xlsx_path = Path(tmpdir) / f"nba_{today}.xlsx"
            wb.save(xlsx_path)

            orig_nba = dashboard_data.NBA_DIR
            orig_mlb = dashboard_data.MLB_DIR
            try:
                dashboard_data.NBA_DIR = Path(tmpdir)
                dashboard_data.MLB_DIR = Path(tmpdir)
                result = dashboard_data.get_today_board(date=today)
            finally:
                dashboard_data.NBA_DIR = orig_nba
                dashboard_data.MLB_DIR = orig_mlb

        skipped = result["skipped"]
        self.assertEqual(len(skipped), 1)
        self.assertIsNone(
            skipped[0]["ev_float"],
            "ev_float must be None for EV='unavailable', not a crash",
        )


# ---------------------------------------------------------------------------
# TestSlipsAccessor — unit tests for get_all_slips() (VIEW-02)
# ---------------------------------------------------------------------------

class TestSlipsAccessor(unittest.TestCase):
    """VIEW-02 tests for get_all_slips() — sorting, leg parsing, why_paired."""

    def test_slips_sorted(self) -> None:
        """get_all_slips() returns slips sorted Date descending."""
        slip_rows = [
            {
                "Date": "2026-06-10", "Slip ID": "2026-06-10:highest_ev:aaaa1111",
                "Platform": "PrizePicks", "Slip Type": "power", "Number of Legs": 2,
                "Legs": "A; B", "Slip Result": "GRADED",
                "Standard Payout Multiplier": 3.0, "Estimated Payout Multiplier": None,
                "Net PnL": 2.0, "Gross Return": 3.0, "Winning Legs": 2,
                "Losing Legs": 0, "Contains Demon": False, "Contains Goblin": False,
                "Notes": None,
            },
            {
                "Date": "2026-06-15", "Slip ID": "2026-06-15:diversified:bbbb2222",
                "Platform": "PrizePicks", "Slip Type": "flex", "Number of Legs": 3,
                "Legs": "X; Y; Z", "Slip Result": "GRADED",
                "Standard Payout Multiplier": 5.0, "Estimated Payout Multiplier": None,
                "Net PnL": 4.0, "Gross Return": 5.0, "Winning Legs": 3,
                "Losing Legs": 0, "Contains Demon": False, "Contains Goblin": False,
                "Notes": None,
            },
            {
                "Date": "2026-06-08", "Slip ID": "2026-06-08:safest_2_leg:cccc3333",
                "Platform": "PrizePicks", "Slip Type": "power", "Number of Legs": 2,
                "Legs": "M; N", "Slip Result": "GRADED",
                "Standard Payout Multiplier": 3.0, "Estimated Payout Multiplier": None,
                "Net PnL": -1.0, "Gross Return": 0.0, "Winning Legs": 1,
                "Losing Legs": 1, "Contains Demon": False, "Contains Goblin": False,
                "Notes": None,
            },
        ]

        wb = _make_master_pnl_wb(slip_rows=slip_rows)
        with tempfile.TemporaryDirectory() as tmpdir:
            master_path = Path(tmpdir) / "master_pnl.xlsx"
            wb.save(master_path)

            orig_pnl = dashboard_data.PNL_DIR
            try:
                dashboard_data.PNL_DIR = Path(tmpdir)
                result = dashboard_data.get_all_slips()
            finally:
                dashboard_data.PNL_DIR = orig_pnl

        self.assertFalse(result["locked"])
        slips = result["slips"]
        self.assertEqual(len(slips), 3)
        dates = [s["Date"] for s in slips]
        self.assertEqual(dates, sorted(dates, reverse=True), f"Slips must be date-descending, got {dates}")

    def test_legs_parsed(self) -> None:
        """A slip with Legs 'A; B; C' yields legs_list == ['A', 'B', 'C']."""
        slip_rows = [
            {
                "Date": "2026-06-15", "Slip ID": "2026-06-15:diversified:dddd4444",
                "Platform": "PrizePicks", "Slip Type": "power", "Number of Legs": 3,
                "Legs": "A; B; C", "Slip Result": "GRADED",
                "Standard Payout Multiplier": 5.0, "Estimated Payout Multiplier": None,
                "Net PnL": 4.0, "Gross Return": 5.0, "Winning Legs": 3,
                "Losing Legs": 0, "Contains Demon": False, "Contains Goblin": False,
                "Notes": None,
            },
        ]

        wb = _make_master_pnl_wb(slip_rows=slip_rows)
        with tempfile.TemporaryDirectory() as tmpdir:
            master_path = Path(tmpdir) / "master_pnl.xlsx"
            wb.save(master_path)

            orig_pnl = dashboard_data.PNL_DIR
            try:
                dashboard_data.PNL_DIR = Path(tmpdir)
                result = dashboard_data.get_all_slips()
            finally:
                dashboard_data.PNL_DIR = orig_pnl

        slips = result["slips"]
        self.assertEqual(len(slips), 1)
        self.assertEqual(
            slips[0]["legs_list"],
            ["A", "B", "C"],
            f"Expected legs_list=['A','B','C'], got {slips[0].get('legs_list')}",
        )

    def test_why_paired_derived(self) -> None:
        """Slip ID '...:correlated_upside:...' yields why_paired starting 'Correlated upside';
        an unknown category yields the independent-legs fallback string."""
        slip_rows = [
            {
                "Date": "2026-06-15", "Slip ID": "2026-06-15:correlated_upside:eeee5555",
                "Platform": "PrizePicks", "Slip Type": "power", "Number of Legs": 2,
                "Legs": "Leg1; Leg2", "Slip Result": "GRADED",
                "Standard Payout Multiplier": 3.0, "Estimated Payout Multiplier": None,
                "Net PnL": 2.0, "Gross Return": 3.0, "Winning Legs": 2,
                "Losing Legs": 0, "Contains Demon": False, "Contains Goblin": False,
                "Notes": None,
            },
            {
                "Date": "2026-06-14", "Slip ID": "2026-06-14:unknown_category:ffff6666",
                "Platform": "PrizePicks", "Slip Type": "power", "Number of Legs": 2,
                "Legs": "Leg3; Leg4", "Slip Result": "GRADED",
                "Standard Payout Multiplier": 3.0, "Estimated Payout Multiplier": None,
                "Net PnL": -1.0, "Gross Return": 0.0, "Winning Legs": 1,
                "Losing Legs": 1, "Contains Demon": False, "Contains Goblin": False,
                "Notes": None,
            },
        ]

        wb = _make_master_pnl_wb(slip_rows=slip_rows)
        with tempfile.TemporaryDirectory() as tmpdir:
            master_path = Path(tmpdir) / "master_pnl.xlsx"
            wb.save(master_path)

            orig_pnl = dashboard_data.PNL_DIR
            try:
                dashboard_data.PNL_DIR = Path(tmpdir)
                result = dashboard_data.get_all_slips()
            finally:
                dashboard_data.PNL_DIR = orig_pnl

        slips_by_id = {s["Slip ID"]: s for s in result["slips"]}

        correlated = slips_by_id["2026-06-15:correlated_upside:eeee5555"]
        self.assertTrue(
            correlated["why_paired"].startswith("Correlated upside"),
            f"Expected why_paired to start with 'Correlated upside', got {correlated['why_paired']!r}",
        )

        unknown = slips_by_id["2026-06-14:unknown_category:ffff6666"]
        self.assertIn(
            "Independent legs",
            unknown["why_paired"],
            f"Expected fallback 'Independent legs' string in why_paired, got {unknown['why_paired']!r}",
        )

    def test_correlated_parlays_read_once_per_date(self) -> None:
        """REGRESSION (CR-01): the Tier-1 Correlated Parlays lookup must read each
        per-sport workbook at most once per DISTINCT slip date, never once per slip.

        The original implementation called _lookup_correlated_parlays() inside the
        per-slip loop, re-opening BOTH per-sport workbooks for every slip. Each open
        does a 1s wait_for_stable_file sleep, so an 88-slip workbook took ~184s.
        This test pins the structural fix: with N slips across D distinct dates, the
        number of 'Correlated Parlays' reads must stay <= 2 (sports) * D, independent
        of N. It does not measure time — it counts workbook opens, which is the exact
        property that made the route unusable at scale.
        """
        # 12 slips spread across only 2 distinct dates.
        slip_rows: list[dict[str, Any]] = []
        for i in range(6):
            slip_rows.append({
                "Date": "2026-06-10",
                "Slip ID": f"2026-06-10:highest_ev:a{i:03d}aaaa",
                "Legs": "A; B", "Slip Result": "GRADED",
            })
        for i in range(6):
            slip_rows.append({
                "Date": "2026-06-11",
                "Slip ID": f"2026-06-11:highest_ev:b{i:03d}bbbb",
                "Legs": "C; D", "Slip Result": "GRADED",
            })

        calls = {"Correlated Parlays": 0}

        def fake_read_sheet_rows(xlsx: Any, sheet: str, delay: float = 1.0) -> Any:
            if sheet == "Slip History":
                return slip_rows
            if sheet == "Correlated Parlays":
                calls["Correlated Parlays"] += 1
                return []  # no Tier-1 match -> falls through to Tier-2
            return []

        with patch.object(dashboard_data, "read_sheet_rows", side_effect=fake_read_sheet_rows):
            result = dashboard_data.get_all_slips()

        self.assertEqual(len(result["slips"]), 12)
        # 2 distinct dates * 2 sports = 4 max. The O(N) bug would produce 12 * 2 = 24.
        self.assertLessEqual(
            calls["Correlated Parlays"],
            4,
            f"CR-01 regression: 'Correlated Parlays' read {calls['Correlated Parlays']} times "
            f"for 12 slips across 2 dates; expected <= 4 (once per sport per distinct date). "
            f"A per-slip lookup has been reintroduced.",
        )


# ---------------------------------------------------------------------------
# TestHistoryAccessor — unit tests for get_history_data() (VIEW-03)
# ---------------------------------------------------------------------------

class TestHistoryAccessor(unittest.TestCase):
    """VIEW-03 tests for get_history_data() — tier breakdown, None-tier handling,
    daily chart, weekly chart aggregation."""

    def test_tier_breakdown(self) -> None:
        """per-tier dict has W, L, hit_pct, roi_pct, n keys for A/B/C/UNKNOWN."""
        pick_history_rows = [
            {"Date": "2026-06-10", "Sport": "NBA", "Result": "WIN",
             "Units": 3.0, "PnL": 2.0, "Confidence Tier": "A"},
            {"Date": "2026-06-11", "Sport": "NBA", "Result": "LOSS",
             "Units": 3.0, "PnL": -3.0, "Confidence Tier": "A"},
            {"Date": "2026-06-12", "Sport": "MLB", "Result": "WIN",
             "Units": 2.0, "PnL": 1.5, "Confidence Tier": "B"},
            {"Date": "2026-06-13", "Sport": "MLB", "Result": "WIN",
             "Units": 1.0, "PnL": 0.8, "Confidence Tier": "C"},
            {"Date": "2026-06-14", "Sport": "NBA", "Result": "LOSS",
             "Units": 1.0, "PnL": -1.0, "Confidence Tier": None},
        ]

        wb = _make_master_pnl_wb(pick_history_rows=pick_history_rows)
        with tempfile.TemporaryDirectory() as tmpdir:
            master_path = Path(tmpdir) / "master_pnl.xlsx"
            wb.save(master_path)

            orig_pnl = dashboard_data.PNL_DIR
            try:
                dashboard_data.PNL_DIR = Path(tmpdir)
                result = dashboard_data.get_history_data()
            finally:
                dashboard_data.PNL_DIR = orig_pnl

        by_tier = result["by_tier"]
        for tier in ("A", "B", "C", "UNKNOWN"):
            self.assertIn(tier, by_tier, f"Expected tier '{tier}' in by_tier dict")
            t = by_tier[tier]
            for key in ("W", "L", "hit_pct", "roi_pct", "n"):
                self.assertIn(key, t, f"Expected key '{key}' in tier '{tier}' dict")

        # Verify A-tier counts: 1 WIN, 1 LOSS
        self.assertEqual(by_tier["A"]["W"], 1)
        self.assertEqual(by_tier["A"]["L"], 1)
        self.assertEqual(by_tier["A"]["n"], 2)

    def test_none_tier_as_unknown(self) -> None:
        """A Pick History row with Confidence Tier None is counted under the UNKNOWN
        tier, not dropped."""
        pick_history_rows = [
            {"Date": "2026-06-14", "Sport": "NBA", "Result": "WIN",
             "Units": 2.0, "PnL": 1.5, "Confidence Tier": None},
            {"Date": "2026-06-15", "Sport": "NBA", "Result": "LOSS",
             "Units": 2.0, "PnL": -2.0, "Confidence Tier": None},
        ]

        wb = _make_master_pnl_wb(pick_history_rows=pick_history_rows)
        with tempfile.TemporaryDirectory() as tmpdir:
            master_path = Path(tmpdir) / "master_pnl.xlsx"
            wb.save(master_path)

            orig_pnl = dashboard_data.PNL_DIR
            try:
                dashboard_data.PNL_DIR = Path(tmpdir)
                result = dashboard_data.get_history_data()
            finally:
                dashboard_data.PNL_DIR = orig_pnl

        by_tier = result["by_tier"]
        unknown = by_tier["UNKNOWN"]
        self.assertEqual(unknown["n"], 2, "Both None-tier rows must be counted under UNKNOWN")
        self.assertEqual(unknown["W"], 1)
        self.assertEqual(unknown["L"], 1)
        # Other tiers should be empty (n==0)
        for tier in ("A", "B", "C"):
            self.assertEqual(by_tier[tier]["n"], 0, f"Tier '{tier}' should be empty (n==0)")

    def test_chart_daily(self) -> None:
        """chart_daily has labels and bankroll lists of equal length from Bankroll Chart Data."""
        chart_rows = [
            {"Date": "2026-06-08", "Bankroll": 100.0, "ROI": 0.0, "Updated At": "2026-06-08T10:00:00"},
            {"Date": "2026-06-10", "Bankroll": 105.5, "ROI": 5.5, "Updated At": "2026-06-10T10:00:00"},
            {"Date": "2026-06-15", "Bankroll": 110.0, "ROI": 10.0, "Updated At": "2026-06-15T10:00:00"},
        ]

        wb = _make_master_pnl_wb(chart_rows=chart_rows)
        with tempfile.TemporaryDirectory() as tmpdir:
            master_path = Path(tmpdir) / "master_pnl.xlsx"
            wb.save(master_path)

            orig_pnl = dashboard_data.PNL_DIR
            try:
                dashboard_data.PNL_DIR = Path(tmpdir)
                result = dashboard_data.get_history_data()
            finally:
                dashboard_data.PNL_DIR = orig_pnl

        chart_daily = result["chart_daily"]
        self.assertIn("labels", chart_daily)
        self.assertIn("bankroll", chart_daily)
        labels = chart_daily["labels"]
        bankroll = chart_daily["bankroll"]
        self.assertEqual(len(labels), len(bankroll), "labels and bankroll must have equal length")
        self.assertEqual(len(labels), 3, f"Expected 3 chart points, got {len(labels)}")
        self.assertEqual(labels[0], "2026-06-08")

    def test_chart_weekly(self) -> None:
        """Two Bankroll Chart Data rows in the same ISO week collapse to one weekly label
        (last point wins)."""
        # 2026-06-08 (Mon) and 2026-06-11 (Thu) are in the same ISO week (W24)
        # 2026-06-15 (Mon) is in a different ISO week (W25)
        chart_rows = [
            {"Date": "2026-06-08", "Bankroll": 100.0, "ROI": 0.0, "Updated At": "2026-06-08T10:00:00"},
            {"Date": "2026-06-11", "Bankroll": 103.0, "ROI": 3.0, "Updated At": "2026-06-11T10:00:00"},
            {"Date": "2026-06-15", "Bankroll": 110.0, "ROI": 10.0, "Updated At": "2026-06-15T10:00:00"},
        ]

        wb = _make_master_pnl_wb(chart_rows=chart_rows)
        with tempfile.TemporaryDirectory() as tmpdir:
            master_path = Path(tmpdir) / "master_pnl.xlsx"
            wb.save(master_path)

            orig_pnl = dashboard_data.PNL_DIR
            try:
                dashboard_data.PNL_DIR = Path(tmpdir)
                result = dashboard_data.get_history_data()
            finally:
                dashboard_data.PNL_DIR = orig_pnl

        chart_weekly = result["chart_weekly"]
        self.assertIn("labels", chart_weekly)
        self.assertIn("bankroll", chart_weekly)
        labels = chart_weekly["labels"]
        bankroll = chart_weekly["bankroll"]
        self.assertEqual(len(labels), len(bankroll), "labels and bankroll must have equal length")
        # Two ISO weeks: W24 (Jun 8+11 → last point Jun 11) and W25 (Jun 15)
        self.assertEqual(len(labels), 2, f"Expected 2 weekly labels (2 ISO weeks), got {len(labels)}: {labels}")
        # The W24 entry should have bankroll 103.0 (last point in the week)
        w24_idx = labels.index("2026-W24") if "2026-W24" in labels else None
        self.assertIsNotNone(w24_idx, f"Expected '2026-W24' in weekly labels, got {labels}")
        self.assertEqual(
            bankroll[w24_idx], 103.0,
            f"W24 bankroll must be 103.0 (last point wins), got {bankroll[w24_idx]}",
        )


# ---------------------------------------------------------------------------
# TestRoutes — route smoke tests for /, /slips, /history (Plan 02)
# ---------------------------------------------------------------------------

class TestRoutes(unittest.TestCase):
    """Smoke tests for the three dashboard GET routes added in Plan 02.

    Uses Flask's test_client() — no network socket needed.
    test_history_200 checks for 'chart.js' in the response body; this will be
    RED until Plan 03 ships history.html (an acceptable Wave-2 state per the plan).
    """

    def setUp(self) -> None:
        self.client = dashboard.app.test_client()

    def test_index_200(self) -> None:
        """GET / returns 200 and the response body contains 'EV' (the EV column header)."""
        resp = self.client.get("/")
        self.assertEqual(resp.status_code, 200)
        self.assertIn(b"EV", resp.data)

    def test_slips_200(self) -> None:
        """GET /slips returns 200."""
        resp = self.client.get("/slips")
        self.assertEqual(resp.status_code, 200)

    def test_history_200(self) -> None:
        """GET /history returns 200 and the lowercased body contains 'chart.js'.

        This will be RED until Plan 03 ships history.html (the route is defined here
        but the template ships later — acceptable Wave-2 state until Plan 03 merges).
        """
        resp = self.client.get("/history")
        self.assertEqual(resp.status_code, 200)
        self.assertIn(b"chart.js", resp.data.lower())


if __name__ == "__main__":
    unittest.main()
