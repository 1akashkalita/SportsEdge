"""NBA Prop Monitor - Monitors line movements from The Odds API against baseline"""

import json
import requests
import os
from datetime import datetime, timedelta
from collections import defaultdict
from openpyxl import load_workbook
from pathlib import Path

API_KEY = "[REDACTED_OLD_THE_ODDS_API_KEY]"
API_KEY = "[REDACTED_OLD_THE_ODDS_API_KEY]"
BASE_URL = "https://api.the-odds-api.com/v4/sports/basketball_nba/odds"

PROPS_DIR = "/Users/akashkalita/sports_picks/data/nba"

# Load fresh PrizePicks data
PRIZEPICKS_PATH = os.path.join(PROPS_DIR, "prizepicks_nba_latest.json")
BASELINE_XLSX = os.path.join(PROPS_DIR, "nba_YYYY-MM-DD.xlsx")

print("="*70)
print("NBA PROP LINE MONITOR")
print("="*70)
print(f"Looking for baseline file: {BASELINE_XLSX}")
print()

if not os.path.exists(BASELINE_XLSX):
    print("ERROR: No xlsx files found in sports_picks/data/nba directory!")
    print("Attempting to find any xlsx file...")
    
    # Search for recent xlsx files
    import glob
    xlsx_files = glob.glob(os.path.join(PROPS_DIR, "**/*.xlsx"), recursive=True)
    print(f"\nFound {len(xlsx_files)} files:")
    for f in sorted(xlsx_files):
        print(f"  - {os.path.basename(f)}")
    
    # Use the found file or exit
    if xlsx_files:
        BASELINE_XLSX = sorted(xlsx_files)[-1]
        print(f"\nUsing latest file: {BASELINE_XLSX}")
    else:
        print("\nNo files found - cannot monitor props without baseline!")
        exit(1)

print("\n" + "="*70)
print("STEP 1: Loading baseline from Excel...")
print("="*70)

try:
    wb = load_workbook(BASELINE_XLSX, read_only=True, data_only=True)
    
    # Check available sheets
    print(f"Available sheets: {list(wb.sheetnames)}")
    
    # Load Player Props sheet if exists
    if "Player Props" in wb.sheetnames:
        props_sheet = wb["Player Props"]
    elif "Player Props Sheet" in wb.sheetnames:
        props_sheet = wb["Player Props Sheet"]
    elif "Props" in wb.sheetnames:
        props_sheet = wb["Props"]
    else:
        wb.close()
        print("ERROR: No Player Props sheet found!")
        exit(1)
    
    print(f"\nLoaded {props_sheet.max_row} rows from Player Props sheet")
    print("\nColumns:", props_sheet[1])
    
except Exception as e:
    print(f"Error loading workbook: {e}")
    import traceback
    traceback.print_exc()
    exit(1)

print("\n" + "="*70)
print("STEP 2: Pulling current lines from The Odds API...")
print("="*70)
print(f"Fetching player prop data for NBA...\n")

# Market types we need to pull
markets_to_fetch = ["player_points", "player_rebounds", "player_assists", "player_threes", "player_blocks"]

all_current_lines = {}  # {book: {player_xxx_prop_type: line_value}}

for market in markets_to_fetch:
    print(f"\nFetching market: {market}")
    
    params = {
        "apiKey": API_KEY,
        "regions": "us",
        "markets": market
    }
    
    try:
        resp = requests.get(BASE_URL, params=params, timeout=45)
        
        if resp.status_code == 200:
            market_data = resp.json()
            print(f"  Retrieved {len(market_data)} markets")
            
            # Process outcomes from this market
            for mkt in market_data:
                book_name = mkt.get('bookName', 'Unknown')
                
                for outcome in mkt.get('outcomes', []):
                    if outcome.get('player_name') and outcome.get('line') is not None:
                        player = outcome['player_name']
                        team = outcome.get('team', '')
                        line = outcome['line']
                        prop_type = market
                        
                        key = f"{player}_{prop_type}" if team else f"{key}_{prop_type}"
                        
                        all_current_lines.setdefault(book_name, {})
                        all_current_lines[book_name][f"{player}_{prop_type}"] = line
            
            print(f"  {sum(1 for outcomes in market_data.values() for o in outcomes.get('outcomes', []) if o.get('line'))} prop lines loaded")
        else:
            print(f"  Error: Status code {resp.status_code}")
    
    except Exception as e:
        print(f"  Error fetching market '{market}': {e}")

print("\n" + "="*70)
print("STEP 3: Comparing current lines to baseline...")
print("="*70)

movements_found = []
skipped_listed = []
injury_watch_alerts = []
arb_opportunities = []

for row in props_sheet.iter_rows(values_only=True):
    if len(row) < 4:
       continue  
    
    player = row[0] if row[0] else 'N/A'
    team = row[1] if len(row) > 1 and row[1] else 'N/A'
    prop_type = row[2] if len(row) > 2 else 'unknown'
    morning_line = row[3] if len(row) > 3 else None
    
    status = row[8] if len(row) > 8 else 'pending'
    
    if not player or not prop_type:
        continue
    
    # Skip deleted/removed props
    if 'deleted' in str(status).lower() or 'removed' in str(status).lower():
        continue
    
    print(f"\nAnalyzing: {player} - {team} - {prop_type}")
    print(f"  Morning Line: {morning_line}, Current Status: {status}")
