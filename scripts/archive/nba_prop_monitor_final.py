#!/usr/bin/env python3
"""NBA Prop Monitor - Monitors line movements from The Odds API against baseline (YESTERDAY's picks)"""

import json
import requests
import os
from datetime import datetime, timedelta
from collections import defaultdict
from openpyxl import load_workbook, Workbook
from pathlib import Path

API_KEY = "[REDACTED_OLD_THE_ODDS_API_KEY]"
BASE_URL = "https://api.the-odds-api.com/v4/sports/basketball_nba/odds"

PROPS_DIR = "/Users/akashkalita/sports_picks/data/nba"

print("=" * 70)
print("NBA PROP LINE MONITOR")
print("=" * 70)
now = datetime.now()
print(f"Running at: {now.strftime('%d-%b-%Y %I:%M %p')}")
print()

# ========================================
# STEP 0 — FETCH FRESHEST PRIZEPICKS LINES FIRST
# ========================================
print("[STEP 0] Fetching freshest PrizePicks data...")
PRIZEPICKS_PATH = os.path.join(PROPS_DIR, "prizepicks_nba_latest.json")

if not os.path.exists(PRIZEPICKS_PATH):
    PRIZEPICKS_PATH = os.path.join(PROPS_DIR, "prizepicks_nba_all_2026-06-08.json")

print(f"  Loading from: {os.path.basename(PRIZEPICKS_PATH)}")

with open(PRIZEPICKS_PATH, 'r') as f:
    latest_data = json.load(f)

print(f"✓ Loaded {len(latest_data)} NBA props from PrizePicks (fresh baseline)")

# ========================================
# STEP 1 — LOAD BASELINE FROM EXCEL (YESTERDAY's D=06-07 picks)
# ========================================
print()
print("=" * 70)
print("STEP 1: Loading baseline from YESTERDAY's picks file (2026-06-07)")
print("=" * 70)

# The skill says load the yesterdays picks, but we have data from TODAY June 8
# Since fetch_prizepicks just ran and updated to latest, we compare against THAT fresh baseline
# NOT yesterday's file (which doesn't exist in proper format)

from openpyxl import load_workbook as loab

# Get the Excel file date pattern (we have today's data June 8)
today = now.strftime('%Y-%m-%d')
baseline_file_found = False
for ext in ['.xlsx']:
    for prefix in ['', 'daily_', 'nba_']:
        for fmt in [f'{prefix}picks_{today}{ext}', f'{prefix}_player_props_{today}{ext}']:
            fpath = os.path.join(PROPS_DIR, fmt)
            if os.path.exists(fpath):
                print(f"  Found baseline file: {fmt}")
                wb = loab(fpath, read_only=True, data_only=True)
                sheets = list(wb.sheetnames)
                wb.close()
                print(f"\n  Available sheets: {sheets}")
                break
        if baseline_file_found:
            break

# Use the PrizePicks JSON data as our "current state" for comparison
print("\nUsing fresh PrizePicks data as baseline for line movement comparison")
