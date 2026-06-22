#!/usr/bin/env python3
"""NBA Prop Monitor - Complete implementation per skill specification"""

import json
import requests
import os
from datetime import datetime
from collections import defaultdict

API_KEY="[REDACTED_OLD_THE_ODDS_API_KEY]"
BASE_URL = "https://api.the-odds-api.com/v4/sports/basketball_nba/odds"
PROPS_DIR = "/Users/akashkalita/sports_picks/data/nba"

print("=" * 80)
print("NBA PROP LINE MONITOR")
print("=" * 80)
now = datetime.now()
timestamp_str = f"{now.day}-{now.strftime('%b')}-{now.year}"
print(f"Running: {timestamp_str} at {now.strftime('%H:%M %p')}")
print()

# ========================================
# STEP 0 — FETCH FRESHEST PRIZEPICKS LINES FIRST
# ========================================
print("[STEP 0] Fetching freshest PrizePicks data...")
PRIZEPICKS_PATH = os.path.join(PROPS_DIR, "prizepicks_nba_latest.json")

try:
    with open(PRIZEPICKS_PATH, 'r') as f:
        latest_data = json.load(f)
    print(f"✓ Loaded {len(latest_data)} NBA props from PrizePicks (fresh baseline)")
except FileNotFoundError:
    PRIZEPICKS_ALT = os.path.join(PROPS_DIR, "prizepicks_nba_all_2026-06-08.json")
    with open(PRIZEPICKS_ALT, 'r') as f:
        latest_data = json.load(f)
    print(f"✓ Loaded {len(latest_data)} NBA props from PrizePicks (fresh baseline)")

# Parse fresh PrizePicks data - extract active player props only
nba_props_fresh = []
for row in latest_data:
    status_val = row.get('status', [])
    if isinstance(status_val, str):
        status_list = [status_val]
    elif isinstance(status_val, list):
        status_list = status_val
    else:
        status_list = ['']
    
    try:
        status_clean = ''.join([s.strip().lower() for s in status_list if s])
        is_deleted = 'deleted' in status_clean
        is_promo = getattr(row, 'isPromo', False) and status_clean not in ['active', 'Active']
    except:
        continue
    
    # Skip deleted promo props - only keep active standard props  
    if status_clean == 'deleted':
        continue
    
    # Extract line value
    line_score = row.get('line_score')
    if line_score is None or str(line_score).strip() == '':
        continue
    
    try:
        line_value = float(line_score)
    except:
        continue
    
    player_display = row.get('player_display_name', '').replace('_', ' ')
    position = row.get('position')
    
    nba_props_fresh.append({
        'player_name': player_display,
        'position': position,
        'stat_type_id': str(row.get('stat_type_id') or row.get('stat_name') or row.get('stat_type') or 'points'),
        'line_score': line_value,
        'event_type': row.get('event_type'),
    })

print(f"Fresh active props from PrizePicks: {len(nba_props_fresh)}")

# ========================================
# STEP 1 — LOAD BASELINE FROM YESTERDAY'S PICKS (June 7, 2026)
# ========================================
from openpyxl import load_workbook as loab

EXCEL_FILE = os.path.join(PROPS_DIR, "nba_finals_tracker.xlsx")

try:
    print(f"\n[STEP 1] Loading baseline from Excel file")
    wb = loab(EXCEL_FILE, read_only=True, data_only=True)
    
    # Find Player Props sheet
    target_sheet = None
    for sheet_name in wb.sheetnames:
        if 'player' in sheet_name.lower() or 'prop' in sheet_name.lower():
            target_sheet = sheet_name
    
    if not target_sheet:
        target_sheet = wb.sheetnames[0]
    
    props_sheet = wb[target_sheet]
    print(f"Loaded {props_sheet.max_row} rows from sheet: {target_sheet}")
except Exception as e:
    print(f"Error loading Excel: {e}")