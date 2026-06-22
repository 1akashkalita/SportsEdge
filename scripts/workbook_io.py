#!/usr/bin/env python3
"""Safe XLSX load/save helpers for SportsEdge workbooks.

This module intentionally only handles workbook I/O safety: stable-file checks,
retry-on-transient read errors, cooperative lock files, validated temp saves,
and timestamped backups. It does not contain betting/gate/projection logic.
"""
from __future__ import annotations

import json
import os
import shutil
import time
import zipfile
from contextlib import contextmanager
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

HOME = Path.home()
ROOT = HOME / "sports_picks"
DATA = ROOT / "data"
LOCK_DIR = ROOT / "locks"
BACKUP_DIR = DATA / "backups" / "workbooks"
LOG_FILE = DATA / "pnl" / "logs" / "run_log.txt"


class WorkbookAccessError(RuntimeError):
    """Raised when a workbook remains unreadable or locked after retries."""


def now_iso() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def today_str() -> str:
    return datetime.now().strftime("%Y-%m-%d")


def log(message: str) -> None:
    try:
        LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
        with LOG_FILE.open("a") as fh:
            fh.write(f"{now_iso()} {message}\n")
    except Exception:
        pass


def _ensure_dirs() -> None:
    LOCK_DIR.mkdir(parents=True, exist_ok=True)
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)


def workbook_lock_path(path: Path) -> Path:
    return LOCK_DIR / f"{Path(path).name}.lock"


def workbook_is_valid(path: Path) -> bool:
    path = Path(path)
    if not path.exists() or not zipfile.is_zipfile(path):
        return False
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        wb.close()
        return True
    except Exception:
        return False


def wait_for_stable_file(path: Path, delay: float = 1.0) -> tuple[int, int]:
    first = Path(path).stat().st_size
    time.sleep(delay)
    second = Path(path).stat().st_size
    return first, second


@contextmanager
def workbook_file_lock(path: Path, wait_seconds: int = 120, stale_seconds: int = 600):
    _ensure_dirs()
    path = Path(path)
    lock_path = workbook_lock_path(path)
    start = time.time()
    acquired = False
    while not acquired:
        try:
            fd = os.open(lock_path, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            with os.fdopen(fd, "w") as fh:
                fh.write(json.dumps({"pid": os.getpid(), "path": str(path), "acquired_at": now_iso()}) + "\n")
            acquired = True
            log(f"Acquired workbook lock {lock_path}")
        except FileExistsError:
            try:
                age = time.time() - lock_path.stat().st_mtime
            except FileNotFoundError:
                continue
            if age > stale_seconds:
                log(f"WARNING stale workbook lock removed: {lock_path} age_seconds={age:.1f}")
                try:
                    lock_path.unlink()
                except FileNotFoundError:
                    pass
                continue
            if time.time() - start >= wait_seconds:
                raise WorkbookAccessError(f"Timed out waiting for workbook lock {lock_path}")
            log(f"Workbook lock exists; waiting path={path} lock={lock_path} age_seconds={age:.1f}")
            time.sleep(2.0)
    try:
        yield lock_path
    finally:
        if acquired:
            try:
                lock_path.unlink()
                log(f"Released workbook lock {lock_path}")
            except FileNotFoundError:
                pass


def safe_load_workbook(path: Path, retries: int = 5, delay: float = 1.0, **kwargs: Any):
    path = Path(path)
    last_error: Exception | None = None
    for attempt in range(1, retries + 1):
        try:
            if not path.exists():
                raise FileNotFoundError(str(path))
            size_a, size_b = wait_for_stable_file(path, delay=delay)
            if size_a != size_b:
                raise WorkbookAccessError(f"Workbook size not stable ({size_a} -> {size_b})")
            if not zipfile.is_zipfile(path):
                raise zipfile.BadZipFile(f"Not a valid xlsx/zip: {path}")
            wb = load_workbook(path, **kwargs)
            if attempt > 1:
                log(f"safe_load_workbook succeeded after retry path={path} attempts={attempt}")
            return wb
        except (EOFError, zipfile.BadZipFile, PermissionError, WorkbookAccessError, OSError) as exc:
            last_error = exc
            log(f"WARNING safe_load_workbook retry path={path} attempt={attempt}/{retries} error={type(exc).__name__}: {exc}")
            if attempt < retries:
                time.sleep(delay)
    raise WorkbookAccessError(
        f"Workbook unreadable after {retries} attempts: {path}; "
        f"last_error={type(last_error).__name__}: {last_error}"
    )


def safe_save_workbook(wb: Any, path: Path) -> Path | None:
    _ensure_dirs()
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_name(f"{path.name}.tmp.{os.getpid()}.xlsx")
    backup_path: Path | None = None
    try:
        wb.save(tmp)
        if not zipfile.is_zipfile(tmp):
            raise zipfile.BadZipFile(f"Temp workbook is not a valid zip: {tmp}")
        test_wb = load_workbook(tmp, read_only=True, data_only=True)
        test_wb.close()
        if path.exists() and zipfile.is_zipfile(path):
            day_dir = BACKUP_DIR / today_str()
            day_dir.mkdir(parents=True, exist_ok=True)
            stamp = datetime.now().strftime("%H%M%S")
            backup_path = day_dir / f"{path.name}.{stamp}.xlsx"
            shutil.copy2(path, backup_path)
        os.replace(tmp, path)
        log(f"Atomic workbook save complete path={path} backup={backup_path}")
        return backup_path
    finally:
        try:
            if tmp.exists():
                tmp.unlink()
        except Exception as exc:
            log(f"WARNING failed to remove temp workbook {tmp}: {exc}")
