#!/usr/bin/env python3
"""Tests for metrics_report.py — slip ROI aggregation, prop hit-rate, WoW arrows.

Covers:
- TestSlipRoiAggregation: staked-only ROI = Σ Net PnL / Σ Stake, sport bucketing
- TestZeroStakeSeparation: zero-stake slips counted separately, not blended
- TestPropHitRateAggregation: reads Prop Accuracy sheet → {(week,sport): rate}
- TestWowArrow: increasing→↑, flat→→, decreasing→↓, None→→
"""
from __future__ import annotations

import tempfile
import unittest
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook


# ---------------------------------------------------------------------------
# Helpers to build synthetic per-sport workbooks and master_pnl workbooks
# ---------------------------------------------------------------------------

def _make_slip_history_wb(rows: list[dict[str, Any]]) -> Workbook:
    """Build an in-memory workbook with a Slip History sheet populated from rows.

    Each row dict should have keys matching SLIP_HISTORY_HEADERS column names.
    """
    from slip_payouts import SLIP_HISTORY_HEADERS
    wb = Workbook()
    ws = wb.active
    ws.title = "Slip History"
    ws.append(SLIP_HISTORY_HEADERS)
    for row in rows:
        ws.append([row.get(h) for h in SLIP_HISTORY_HEADERS])
    return wb


def _make_prop_accuracy_wb(rows: list[dict[str, Any]]) -> Workbook:
    """Build an in-memory master_pnl workbook with a Prop Accuracy sheet."""
    PROP_ACCURACY_HEADERS = ["Week", "Sport", "Total Props", "Wins", "Losses", "Pushes", "Hit Rate", "Updated At"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Prop Accuracy"
    ws.append(PROP_ACCURACY_HEADERS)
    for row in rows:
        ws.append([row.get(h) for h in PROP_ACCURACY_HEADERS])
    return wb


# ---------------------------------------------------------------------------
# TestSlipRoiAggregation
# ---------------------------------------------------------------------------

class TestSlipRoiAggregation(unittest.TestCase):
    """Tests for aggregate_slip_roi_by_week_sport: staked slips, ROI formula, sport bucketing."""

    def _aggregate(self, sport: str, rows: list[dict[str, Any]]) -> dict:
        """Helper: write a per-sport workbook to a tmpdir, call the aggregation function, return result."""
        import metrics_report
        from slip_payouts import SLIP_HISTORY_HEADERS

        with tempfile.TemporaryDirectory() as tmpdir:
            sport_dir = Path(tmpdir) / sport
            sport_dir.mkdir()
            # Build one dated workbook
            wb = _make_slip_history_wb(rows)
            wb_path = sport_dir / f"{sport}_2026-06-15.xlsx"
            wb.save(wb_path)

            # Temporarily override module dirs
            orig_nba = metrics_report.NBA_DIR
            orig_mlb = metrics_report.MLB_DIR
            orig_inception = metrics_report.INCEPTION_DATE
            try:
                if sport == "nba":
                    metrics_report.NBA_DIR = sport_dir
                    metrics_report.MLB_DIR = Path(tmpdir) / "mlb_empty"
                    metrics_report.MLB_DIR.mkdir()
                else:
                    metrics_report.MLB_DIR = sport_dir
                    metrics_report.NBA_DIR = Path(tmpdir) / "nba_empty"
                    metrics_report.NBA_DIR.mkdir()
                metrics_report.INCEPTION_DATE = "2026-06-01"  # include test dates
                result = metrics_report.aggregate_slip_roi_by_week_sport()
            finally:
                metrics_report.NBA_DIR = orig_nba
                metrics_report.MLB_DIR = orig_mlb
                metrics_report.INCEPTION_DATE = orig_inception

        return result

    def test_roi_formula_staked_slips(self):
        """ROI = Σ Net PnL / Σ Stake over staked slips; acceptance_criteria example."""
        # Two staked MLB slips in the same ISO-week: stake=1.0 net=+2.0; stake=1.0 net=-1.0
        # Expected ROI = (2.0 - 1.0) / (1.0 + 1.0) = 0.5
        rows = [
            {"Date": "2026-06-15", "Stake Units": 1.0, "Net PnL": 2.0,
             "Slip Result": "WIN", "Needs Payout Reconciliation": None},
            {"Date": "2026-06-15", "Stake Units": 1.0, "Net PnL": -1.0,
             "Slip Result": "LOSS", "Needs Payout Reconciliation": None},
        ]
        result = self._aggregate("mlb", rows)
        # 2026-06-15 is ISO week 2026-W25
        from datetime import date as d
        iso_w = f"{d(2026, 6, 15).isocalendar().year}-W{d(2026, 6, 15).isocalendar().week:02d}"
        key = (iso_w, "MLB")
        self.assertIn(key, result, f"Expected key {key} in {list(result.keys())}")
        rec = result[key]
        self.assertAlmostEqual(rec["roi"], 0.5, places=6)
        self.assertEqual(rec["total_stake"], 2.0)
        self.assertAlmostEqual(rec["total_pnl"], 1.0, places=6)
        self.assertEqual(rec["staked"], 2)
        self.assertEqual(rec["zero_stake"], 0)

    def test_wins_and_losses_counted_for_staked_slips(self):
        """WIN and LOSS Slip Results are counted in wins/losses for staked slips."""
        rows = [
            {"Date": "2026-06-15", "Stake Units": 1.0, "Net PnL": 2.0,
             "Slip Result": "WIN", "Needs Payout Reconciliation": None},
            {"Date": "2026-06-15", "Stake Units": 1.0, "Net PnL": -1.0,
             "Slip Result": "LOSS", "Needs Payout Reconciliation": None},
        ]
        result = self._aggregate("mlb", rows)
        from datetime import date as d
        iso_w = f"{d(2026, 6, 15).isocalendar().year}-W{d(2026, 6, 15).isocalendar().week:02d}"
        key = (iso_w, "MLB")
        rec = result[key]
        self.assertEqual(rec["wins"], 1)
        self.assertEqual(rec["losses"], 1)

    def test_recon_flagged_rows_excluded_entirely(self):
        """Rows with Needs Payout Reconciliation truthy are excluded from all counts."""
        rows = [
            {"Date": "2026-06-15", "Stake Units": 1.0, "Net PnL": 5.0,
             "Slip Result": "WIN", "Needs Payout Reconciliation": True},  # should be excluded
            {"Date": "2026-06-15", "Stake Units": 1.0, "Net PnL": -1.0,
             "Slip Result": "LOSS", "Needs Payout Reconciliation": None},
        ]
        result = self._aggregate("mlb", rows)
        from datetime import date as d
        iso_w = f"{d(2026, 6, 15).isocalendar().year}-W{d(2026, 6, 15).isocalendar().week:02d}"
        key = (iso_w, "MLB")
        rec = result[key]
        # Only the LOSS row should be counted (recon row excluded)
        self.assertEqual(rec["staked"], 1)
        self.assertAlmostEqual(rec["total_pnl"], -1.0, places=6)
        self.assertEqual(rec["total_stake"], 1.0)

    def test_graded_row_included_in_roi_not_win_loss(self):
        """GRADED Slip Result: Net PnL in ROI but not counted in wins/losses (Pitfall 3)."""
        rows = [
            {"Date": "2026-06-15", "Stake Units": 1.0, "Net PnL": 3.0,
             "Slip Result": "GRADED", "Needs Payout Reconciliation": None},
        ]
        result = self._aggregate("mlb", rows)
        from datetime import date as d
        iso_w = f"{d(2026, 6, 15).isocalendar().year}-W{d(2026, 6, 15).isocalendar().week:02d}"
        key = (iso_w, "MLB")
        rec = result[key]
        self.assertEqual(rec["staked"], 1)
        self.assertAlmostEqual(rec["total_pnl"], 3.0, places=6)
        # GRADED should NOT appear in wins or losses
        self.assertEqual(rec["wins"], 0)
        self.assertEqual(rec["losses"], 0)

    def test_no_sports_system_runner_import(self):
        """metrics_report must not import sports_system_runner (circular import guard)."""
        import ast
        src = (Path(__file__).parent / "metrics_report.py").read_text(encoding="utf-8")
        tree = ast.parse(src)
        for node in ast.walk(tree):
            if isinstance(node, (ast.Import, ast.ImportFrom)):
                names = [n.name for n in getattr(node, "names", [])]
                module = getattr(node, "module", "") or ""
                self.assertNotIn(
                    "sports_system_runner", str(names + [module]),
                    "metrics_report.py must not import sports_system_runner"
                )


# ---------------------------------------------------------------------------
# TestZeroStakeSeparation
# ---------------------------------------------------------------------------

class TestZeroStakeSeparation(unittest.TestCase):
    """Zero-stake slips are counted separately, never blended into money metrics."""

    def _aggregate(self, sport: str, rows: list[dict[str, Any]]) -> dict:
        import metrics_report
        with tempfile.TemporaryDirectory() as tmpdir:
            sport_dir = Path(tmpdir) / sport
            sport_dir.mkdir()
            wb = _make_slip_history_wb(rows)
            wb_path = sport_dir / f"{sport}_2026-06-15.xlsx"
            wb.save(wb_path)
            orig_nba = metrics_report.NBA_DIR
            orig_mlb = metrics_report.MLB_DIR
            orig_inception = metrics_report.INCEPTION_DATE
            try:
                if sport == "nba":
                    metrics_report.NBA_DIR = sport_dir
                    metrics_report.MLB_DIR = Path(tmpdir) / "mlb_empty"
                    metrics_report.MLB_DIR.mkdir()
                else:
                    metrics_report.MLB_DIR = sport_dir
                    metrics_report.NBA_DIR = Path(tmpdir) / "nba_empty"
                    metrics_report.NBA_DIR.mkdir()
                metrics_report.INCEPTION_DATE = "2026-06-01"
                result = metrics_report.aggregate_slip_roi_by_week_sport()
            finally:
                metrics_report.NBA_DIR = orig_nba
                metrics_report.MLB_DIR = orig_mlb
                metrics_report.INCEPTION_DATE = orig_inception
        return result

    def test_zero_stake_separation(self):
        """Acceptance criteria: 2 staked + 1 zero-stake → ROI=0.5, zero_stake=1, total_stake=2.0."""
        rows = [
            {"Date": "2026-06-15", "Stake Units": 1.0, "Net PnL": 2.0,
             "Slip Result": "WIN", "Needs Payout Reconciliation": None},
            {"Date": "2026-06-15", "Stake Units": 1.0, "Net PnL": -1.0,
             "Slip Result": "LOSS", "Needs Payout Reconciliation": None},
            {"Date": "2026-06-15", "Stake Units": 0.0, "Net PnL": 0.0,
             "Slip Result": "WIN", "Needs Payout Reconciliation": None},  # zero-stake
        ]
        result = self._aggregate("mlb", rows)
        from datetime import date as d
        iso_w = f"{d(2026, 6, 15).isocalendar().year}-W{d(2026, 6, 15).isocalendar().week:02d}"
        key = (iso_w, "MLB")
        rec = result[key]
        self.assertEqual(rec["zero_stake"], 1, "zero_stake count should be 1")
        self.assertEqual(rec["staked"], 2, "staked count should be 2")
        self.assertEqual(rec["total_stake"], 2.0, "total_stake must exclude zero-stake slip")
        self.assertAlmostEqual(rec["roi"], 0.5, places=6)

    def test_none_stake_treated_as_zero(self):
        """Stake Units of None is treated as zero-stake (defensive handling)."""
        rows = [
            {"Date": "2026-06-15", "Stake Units": None, "Net PnL": 2.0,
             "Slip Result": "WIN", "Needs Payout Reconciliation": None},
        ]
        result = self._aggregate("mlb", rows)
        from datetime import date as d
        iso_w = f"{d(2026, 6, 15).isocalendar().year}-W{d(2026, 6, 15).isocalendar().week:02d}"
        key = (iso_w, "MLB")
        if key in result:
            rec = result[key]
            self.assertEqual(rec["zero_stake"], 1)
            self.assertEqual(rec["staked"], 0)

    def test_zero_stake_not_in_total_stake(self):
        """Zero-stake slips do not contribute to total_stake or total_pnl."""
        rows = [
            {"Date": "2026-06-15", "Stake Units": 0.0, "Net PnL": 5.0,
             "Slip Result": "WIN", "Needs Payout Reconciliation": None},
        ]
        result = self._aggregate("mlb", rows)
        from datetime import date as d
        iso_w = f"{d(2026, 6, 15).isocalendar().year}-W{d(2026, 6, 15).isocalendar().week:02d}"
        key = (iso_w, "MLB")
        if key in result:
            rec = result[key]
            self.assertEqual(rec["total_stake"], 0.0)
            self.assertIsNone(rec["roi"], "ROI should be None when total_stake=0")


# ---------------------------------------------------------------------------
# TestPropHitRateAggregation
# ---------------------------------------------------------------------------

class TestPropHitRateAggregation(unittest.TestCase):
    """read_prop_hit_rate_by_week_sport reads Prop Accuracy sheet correctly."""

    def test_reads_two_row_fixture(self):
        """Acceptance criteria: 2-row fixture (W25 MLB 0.55, W26 MLB 0.60) → correct dict."""
        import metrics_report
        wb = _make_prop_accuracy_wb([
            {"Week": "2026-W25", "Sport": "MLB", "Total Props": 10, "Wins": 6, "Losses": 4,
             "Pushes": 0, "Hit Rate": 0.55, "Updated At": "2026-06-15T00:00:00Z"},
            {"Week": "2026-W26", "Sport": "MLB", "Total Props": 5, "Wins": 3, "Losses": 2,
             "Pushes": 0, "Hit Rate": 0.60, "Updated At": "2026-06-22T00:00:00Z"},
        ])
        result = metrics_report.read_prop_hit_rate_by_week_sport(_wb_override=wb)
        self.assertEqual(result, {("2026-W25", "MLB"): 0.55, ("2026-W26", "MLB"): 0.60})

    def test_multiple_sports(self):
        """Reads NBA and MLB rows independently."""
        import metrics_report
        wb = _make_prop_accuracy_wb([
            {"Week": "2026-W25", "Sport": "NBA", "Hit Rate": 0.62},
            {"Week": "2026-W25", "Sport": "MLB", "Hit Rate": 0.58},
        ])
        result = metrics_report.read_prop_hit_rate_by_week_sport(_wb_override=wb)
        self.assertAlmostEqual(result[("2026-W25", "NBA")], 0.62, places=4)
        self.assertAlmostEqual(result[("2026-W25", "MLB")], 0.58, places=4)

    def test_missing_sheet_returns_empty(self):
        """Returns empty dict when Prop Accuracy sheet is absent (SKIP behavior)."""
        import metrics_report
        wb = Workbook()
        # No Prop Accuracy sheet
        result = metrics_report.read_prop_hit_rate_by_week_sport(_wb_override=wb)
        self.assertEqual(result, {})

    def test_case_normalized(self):
        """Sport values are uppercased in result keys."""
        import metrics_report
        wb = _make_prop_accuracy_wb([
            {"Week": "2026-W26", "Sport": "mlb", "Hit Rate": 0.55},
        ])
        result = metrics_report.read_prop_hit_rate_by_week_sport(_wb_override=wb)
        # Should have uppercased key
        keys_sports = {k[1] for k in result.keys()}
        self.assertIn("MLB", keys_sports)


# ---------------------------------------------------------------------------
# TestWowArrow
# ---------------------------------------------------------------------------

class TestWowArrow(unittest.TestCase):
    """wow_arrow returns ↑/→/↓ based on delta vs threshold."""

    def setUp(self):
        import metrics_report
        self.wow_arrow = metrics_report.wow_arrow

    def test_increasing_returns_up(self):
        """wow_arrow(0.52, 0.47) == '↑'"""
        self.assertEqual(self.wow_arrow(0.52, 0.47), "↑")

    def test_decreasing_returns_down(self):
        """wow_arrow(0.47, 0.52) == '↓'"""
        self.assertEqual(self.wow_arrow(0.47, 0.52), "↓")

    def test_flat_returns_right(self):
        """wow_arrow(0.50, 0.50) == '→'"""
        self.assertEqual(self.wow_arrow(0.50, 0.50), "→")

    def test_none_current_returns_right(self):
        """wow_arrow(None, 0.5) == '→'"""
        self.assertEqual(self.wow_arrow(None, 0.5), "→")

    def test_none_prev_returns_right(self):
        """wow_arrow(0.5, None) == '→'"""
        self.assertEqual(self.wow_arrow(0.5, None), "→")

    def test_both_none_returns_right(self):
        """wow_arrow(None, None) == '→'"""
        self.assertEqual(self.wow_arrow(None, None), "→")

    def test_threshold_boundary_just_above(self):
        """Just above threshold → ↑"""
        self.assertEqual(self.wow_arrow(0.506, 0.500), "↑")  # delta=0.006 > 0.005

    def test_threshold_boundary_just_below(self):
        """Within threshold → → (flat)"""
        self.assertEqual(self.wow_arrow(0.5040, 0.500), "→")  # delta=0.004 <= 0.005

    def test_large_decrease(self):
        """Large decrease → ↓"""
        self.assertEqual(self.wow_arrow(-0.20, 0.30), "↓")

    def test_large_increase(self):
        """Large increase → ↑"""
        self.assertEqual(self.wow_arrow(0.80, 0.20), "↑")


# ---------------------------------------------------------------------------
# TestFormatTelegramDigest + TestFillObsidianMarkdown (acceptance criteria checks)
# ---------------------------------------------------------------------------

class TestFormatTelegramDigest(unittest.TestCase):
    """format_telegram_digest renders the D-03 Telegram digest correctly."""

    def _make_report(self):
        """Build a minimal report dict for testing."""
        import metrics_report
        roi_agg = {
            ("2026-W26", "NBA"): {
                "roi": 0.15, "total_stake": 2.0, "total_pnl": 0.30,
                "staked": 2, "zero_stake": 1, "wins": 1, "losses": 1,
            },
            ("2026-W26", "MLB"): {
                "roi": -0.05, "total_stake": 4.0, "total_pnl": -0.20,
                "staked": 4, "zero_stake": 0, "wins": 2, "losses": 2,
            },
        }
        prop_rates = {
            ("2026-W26", "NBA"): 0.60,
            ("2026-W26", "MLB"): 0.55,
        }
        return metrics_report.build_weekly_report(roi_agg=roi_agg, prop_rates=prop_rates)

    def test_contains_iso_week(self):
        import metrics_report
        report = self._make_report()
        digest = metrics_report.format_telegram_digest(report)
        self.assertIn("2026-W26", digest)

    def test_contains_roi_and_hits(self):
        import metrics_report
        report = self._make_report()
        digest = metrics_report.format_telegram_digest(report)
        # Should mention ROI and Hits (or Hit) for each sport
        lower = digest.lower()
        self.assertTrue("roi" in lower or "%" in lower, "Digest should contain ROI value")
        self.assertTrue("hit" in lower or "%" in lower, "Digest should contain hit-rate")

    def test_contains_arrow_characters(self):
        import metrics_report
        report = self._make_report()
        digest = metrics_report.format_telegram_digest(report)
        # At least one arrow character should be present
        self.assertTrue(
            any(c in digest for c in ("↑", "↓", "→")),
            "Digest should contain WoW arrow characters"
        )

    def test_no_verdict_text(self):
        """Digest must not contain 'improving' or 'stagnant' verdict words (D-03)."""
        import metrics_report
        report = self._make_report()
        digest = metrics_report.format_telegram_digest(report)
        self.assertNotIn("improving", digest.lower())
        self.assertNotIn("stagnant", digest.lower())

    def test_contains_zero_stake_count(self):
        """Digest includes a zero-stake informational count line."""
        import metrics_report
        report = self._make_report()
        digest = metrics_report.format_telegram_digest(report)
        # Should contain zero-stake count reference (non-bet / recorded / 0-stake)
        lower = digest.lower()
        self.assertTrue(
            "zero" in lower or "non-bet" in lower or "recorded" in lower or "not staked" in lower,
            "Digest should reference zero-stake count"
        )


class TestFillObsidianMarkdown(unittest.TestCase):
    """fill_obsidian_recap_markdown renders correct Obsidian markdown body."""

    def _make_report(self):
        import metrics_report
        roi_agg = {
            ("2026-W26", "NBA"): {
                "roi": 0.10, "total_stake": 2.0, "total_pnl": 0.20,
                "staked": 2, "zero_stake": 0, "wins": 1, "losses": 1,
            },
            ("2026-W26", "MLB"): {
                "roi": 0.05, "total_stake": 4.0, "total_pnl": 0.20,
                "staked": 4, "zero_stake": 0, "wins": 2, "losses": 2,
            },
        }
        prop_rates = {
            ("2026-W26", "NBA"): 0.55,
            ("2026-W26", "MLB"): 0.60,
        }
        return metrics_report.build_weekly_report(roi_agg=roi_agg, prop_rates=prop_rates)

    def test_contains_by_sport_section(self):
        import metrics_report
        report = self._make_report()
        md = metrics_report.fill_obsidian_recap_markdown(report)
        self.assertIn("By Sport", md)

    def test_contains_nba_and_mlb_rows(self):
        import metrics_report
        report = self._make_report()
        md = metrics_report.fill_obsidian_recap_markdown(report)
        self.assertIn("NBA", md)
        self.assertIn("MLB", md)

    def test_contains_adjustments_heading(self):
        import metrics_report
        report = self._make_report()
        md = metrics_report.fill_obsidian_recap_markdown(report)
        self.assertIn("Adjustments", md)

    def test_calibration_note_included(self):
        import metrics_report
        report = self._make_report()
        note = "Model is slightly overconfident. Factor adjusted to 1.03."
        md = metrics_report.fill_obsidian_recap_markdown(report, calibration_note=note)
        self.assertIn(note, md)

    def test_no_verdict_text(self):
        """Markdown must not contain 'improving' or 'stagnant' verdict words (D-03)."""
        import metrics_report
        report = self._make_report()
        md = metrics_report.fill_obsidian_recap_markdown(report)
        self.assertNotIn("improving", md.lower())
        self.assertNotIn("stagnant", md.lower())


if __name__ == "__main__":
    unittest.main()
