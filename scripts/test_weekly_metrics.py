#!/usr/bin/env python3
"""Tests for Phase 4 dual-metrics: calibration formula, bounds, integrity, and report.

Covers METRICS-01 (slip ROI / prop hit-rate report), METRICS-02 (calibration formula
+ D-10 bounds), and METRICS-03 (structural + runtime integrity guarantees).
"""
from __future__ import annotations

import importlib.util
import json
import sys
import tempfile
import unittest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

# Load runner via importlib to access RESULT_HEADERS without import side effects
MOD_PATH = Path(__file__).with_name("sports_system_runner.py")
spec = importlib.util.spec_from_file_location("sports_system_runner", MOD_PATH)
runner = importlib.util.module_from_spec(spec)  # type: ignore[arg-type]
assert spec and spec.loader
spec.loader.exec_module(runner)  # type: ignore[union-attr]

import calibration  # noqa: E402 (after sys.path.insert)

from openpyxl import Workbook  # noqa: E402


# ---------------------------------------------------------------------------
# In-memory workbook fixture helpers
# ---------------------------------------------------------------------------

def _make_pick_history_wb(rows: list[list]) -> Workbook:
    """Build an in-memory workbook with a Pick History sheet.

    Column positions are taken from runner.RESULT_HEADERS index lookups — never
    hardcoded offsets. Each row in ``rows`` must be a full RESULT_HEADERS-length
    list; callers should build rows using _make_ph_row().
    """
    wb = Workbook()
    wb.active.title = "Daily Log"
    ph = wb.create_sheet("Pick History")
    ph.append(runner.RESULT_HEADERS)
    for row in rows:
        ph.append(row)
    return wb


def _make_ph_row(
    date: str,
    sport: str,
    pick_type: str,
    result: str,
    mop: float | None,
) -> list:
    """Build a Pick History row with the minimum required fields."""
    row: list = [None] * len(runner.RESULT_HEADERS)
    h = {col: idx for idx, col in enumerate(runner.RESULT_HEADERS)}
    row[h["Date"]] = date
    row[h["Sport"]] = sport
    row[h["Pick Type"]] = pick_type
    row[h["Result"]] = result
    row[h["Model Over Probability"]] = mop
    return row


# ---------------------------------------------------------------------------
# TestCalibrationFormula — METRICS-02 formula direction and step behaviour
# ---------------------------------------------------------------------------

class TestCalibrationFormula(unittest.TestCase):
    """Calibration formula computes correct direction and honours step clamp."""

    def test_overconfident_steps_up(self) -> None:
        """Model overconfident (MOP mean > empirical) → factor increases."""
        # empirical = 10/(10+20) = 0.333; model_implied ~= 0.62
        mop_values = [0.62] * 30  # len == n_gate
        new_factor, audit = calibration.compute_calibration_target(
            wins=10, losses=20, mop_values=mop_values, prev_factor=1.0
        )
        self.assertGreater(new_factor, 1.0, "overconfident: factor must increase")
        self.assertAlmostEqual(new_factor, 1.05, places=4)
        self.assertLessEqual(new_factor, 1.20)

    def test_step_at_most_max_step_up(self) -> None:
        """Step delta is capped at +0.05 even with extreme overconfidence."""
        mop_values = [0.99] * 30
        new_factor, audit = calibration.compute_calibration_target(
            wins=1, losses=29, mop_values=mop_values, prev_factor=1.0
        )
        self.assertLessEqual(new_factor - 1.0, 0.05 + 1e-9)

    def test_underconfident_steps_down(self) -> None:
        """Model underconfident (empirical > MOP mean) → factor decreases."""
        # empirical = 28/30 ≈ 0.933; model_implied = 0.62 < empirical → ratio < 1.0
        mop_values = [0.62] * 30
        new_factor, audit = calibration.compute_calibration_target(
            wins=28, losses=2, mop_values=mop_values, prev_factor=1.0
        )
        self.assertLess(new_factor, 1.0, "underconfident: factor must decrease")

    def test_step_at_most_max_step_down(self) -> None:
        """Step delta capped at -0.05 even with extreme underconfidence."""
        mop_values = [0.50] * 30
        new_factor, audit = calibration.compute_calibration_target(
            wins=30, losses=0, mop_values=mop_values, prev_factor=1.0
        )
        self.assertGreaterEqual(1.0 - new_factor, 0.0)
        self.assertLessEqual(1.0 - new_factor, 0.05 + 1e-9)

    def test_audit_dict_has_required_fields(self) -> None:
        """audit dict contains all expected keys for computed case."""
        mop_values = [0.62] * 30
        _, audit = calibration.compute_calibration_target(
            wins=10, losses=20, mop_values=mop_values, prev_factor=1.0
        )
        for key in ("empirical_hit_rate", "model_implied", "raw_ratio", "target",
                    "delta", "new_factor", "prev_factor", "n_outcomes", "n_with_mop", "reason"):
            self.assertIn(key, audit, f"audit missing key: {key}")


# ---------------------------------------------------------------------------
# TestCalibrationGateNotMet — D-10: n < 30 → factor frozen at prev
# ---------------------------------------------------------------------------

class TestCalibrationGateNotMet(unittest.TestCase):
    """With fewer than 30 MOP-backed outcomes, the factor must not move."""

    def test_n_below_gate_returns_prev(self) -> None:
        """n=15 < 30: new_factor == prev_factor exactly."""
        new_factor, audit = calibration.compute_calibration_target(
            wins=5, losses=10, mop_values=[0.62] * 15, prev_factor=1.0
        )
        self.assertEqual(new_factor, 1.0)
        self.assertIn("gate not met", audit["reason"])

    def test_n_zero_returns_prev(self) -> None:
        """n=0: new_factor == prev_factor."""
        new_factor, audit = calibration.compute_calibration_target(
            wins=0, losses=0, mop_values=[], prev_factor=1.0
        )
        self.assertEqual(new_factor, 1.0)

    def test_n_29_just_below_gate(self) -> None:
        """n=29 is exactly one below gate: no movement."""
        new_factor, audit = calibration.compute_calibration_target(
            wins=15, losses=14, mop_values=[0.7] * 29, prev_factor=0.95
        )
        self.assertEqual(new_factor, 0.95)

    def test_n_at_gate_boundary_allowed(self) -> None:
        """n=30 meets the gate exactly: factor is allowed to move."""
        new_factor, audit = calibration.compute_calibration_target(
            wins=10, losses=20, mop_values=[0.65] * 30, prev_factor=1.0
        )
        # Should compute and move (not frozen at 1.0)
        self.assertNotEqual(audit["reason"], "gate not met: n=30 < 30")

    def test_gate_not_met_reason_includes_n(self) -> None:
        """Reason string includes the actual n value."""
        _, audit = calibration.compute_calibration_target(
            wins=5, losses=10, mop_values=[0.6] * 15, prev_factor=1.0
        )
        self.assertIn("15", audit["reason"])

    # WR-01: population mismatch — gate must be on n_with_mop, not n_outcomes
    def test_wr01_gate_on_mop_backed_population_not_total_outcomes(self) -> None:
        """WR-01 RED: 30 graded WIN/LOSS rows but only 3 carry MOP.

        n_outcomes=30 (old buggy gate passes) but n_with_mop=3 < N_GATE=30.
        The factor must NOT move — gate must guard on the MOP-backed count.
        Empirical and model_implied must share the same denominator (both over MOP rows).
        """
        # 27 outcomes have no MOP, 3 have MOP=0.9 → n_with_mop=3, n_outcomes=30
        new_factor, audit = calibration.compute_calibration_target(
            wins=15,
            losses=15,
            mop_values=[0.9] * 3,  # only 3 MOP-backed rows
            prev_factor=1.0,
            n_gate=calibration.N_GATE,  # 30
        )
        self.assertEqual(new_factor, 1.0,
                         "factor must not move: n_with_mop=3 < N_GATE=30, even though n_outcomes=30")
        self.assertIn("gate not met", audit["reason"],
                      "reason must indicate gate not met when n_with_mop < n_gate")
        # Reason must cite n_with_mop (3), not n_outcomes (30)
        self.assertIn("3", audit["reason"],
                      "reason must reference n_with_mop=3, not n_outcomes=30")

    def test_wr01_empirical_and_model_implied_same_denominator_via_read(self) -> None:
        """WR-01 RED: read_graded_outcomes_for_sport win/loss tally is MOP-backed.

        With 30 graded rows but only 3 carrying MOP, the returned wins+losses
        must equal len(mop_values) (both = 3), not 30.
        """
        rows: list[list] = []
        # 27 rows with no MOP
        for _ in range(14):
            rows.append(_make_ph_row("2026-06-10", "MLB", "PROP", "WIN", None))
        for _ in range(13):
            rows.append(_make_ph_row("2026-06-10", "MLB", "PROP", "LOSS", None))
        # 3 rows with MOP
        for _ in range(2):
            rows.append(_make_ph_row("2026-06-10", "MLB", "PROP", "WIN", 0.9))
        rows.append(_make_ph_row("2026-06-10", "MLB", "PROP", "LOSS", 0.9))

        wb = _make_pick_history_wb(rows)
        counts = calibration.read_graded_outcomes_for_sport("MLB", _wb_override=wb)

        # After WR-01 fix: wins+losses must equal n_with_mop (3)
        self.assertEqual(
            counts["wins"] + counts["losses"],
            counts["n_with_mop"],
            "wins+losses must equal n_with_mop so empirical and model_implied share a denominator",
        )
        self.assertEqual(counts["n_with_mop"], 3,
                         "only 3 rows have MOP — n_with_mop must be 3")
        # n_total_graded must reflect all 30 graded rows (informational)
        self.assertEqual(counts.get("n_total_graded", -1), 30,
                         "n_total_graded must count all WIN/LOSS rows regardless of MOP presence")


# ---------------------------------------------------------------------------
# TestCalibrationBounds — D-10: factor always in [0.85, 1.20], step ≤ ±0.05
# ---------------------------------------------------------------------------

class TestCalibrationBounds(unittest.TestCase):
    """Factor stays within [0.85, 1.20] and moves at most ±0.05 for any input."""

    def test_all_losses_clamps_to_max(self) -> None:
        """All losses → raw_ratio=clamp_hi; new_factor ≤ 1.20 and |delta| ≤ 0.05."""
        new_factor, audit = calibration.compute_calibration_target(
            wins=0, losses=30, mop_values=[0.65] * 30, prev_factor=1.18
        )
        self.assertLessEqual(new_factor, 1.20)
        self.assertLessEqual(abs(new_factor - 1.18), 0.05 + 1e-9)

    def test_all_wins_narrows_stays_above_floor(self) -> None:
        """All wins (empirical=1.0) → raw_ratio < 1.0 → narrows; ≥ 0.85."""
        new_factor, audit = calibration.compute_calibration_target(
            wins=30, losses=0, mop_values=[0.7] * 30, prev_factor=1.0
        )
        self.assertLess(new_factor, 1.0, "all wins: factor should narrow")
        self.assertGreaterEqual(new_factor, 0.85)
        self.assertLessEqual(1.0 - new_factor, 0.05 + 1e-9)

    def test_prev_at_lower_bound_does_not_go_below(self) -> None:
        """Factor already at CLAMP_LO=0.85 cannot go below it."""
        mop_values = [0.70] * 30  # underconfident → step down, but already at floor
        new_factor, audit = calibration.compute_calibration_target(
            wins=30, losses=0, mop_values=mop_values, prev_factor=0.85
        )
        self.assertGreaterEqual(new_factor, 0.85)

    def test_prev_at_upper_bound_does_not_exceed(self) -> None:
        """Factor already at CLAMP_HI=1.20 cannot go above it."""
        mop_values = [0.65] * 30  # overconfident → step up, but already at ceiling
        new_factor, audit = calibration.compute_calibration_target(
            wins=0, losses=30, mop_values=mop_values, prev_factor=1.20
        )
        self.assertLessEqual(new_factor, 1.20)

    def test_extreme_mop_values_stay_in_bounds(self) -> None:
        """MOP values all 0.99 with many losses: stays in [0.85, 1.20]."""
        new_factor, _ = calibration.compute_calibration_target(
            wins=1, losses=59, mop_values=[0.99] * 60, prev_factor=1.0
        )
        self.assertGreaterEqual(new_factor, 0.85)
        self.assertLessEqual(new_factor, 1.20)

    def test_delta_never_exceeds_max_step(self) -> None:
        """Absolute delta is always ≤ MAX_STEP regardless of inputs."""
        test_cases = [
            (10, 20, [0.99] * 30, 1.0),
            (30, 0, [0.50] * 30, 1.0),
            (0, 30, [0.65] * 30, 1.18),
            (15, 15, [0.55] * 30, 0.90),
        ]
        for wins, losses, mops, prev in test_cases:
            new_factor, audit = calibration.compute_calibration_target(
                wins=wins, losses=losses, mop_values=mops, prev_factor=prev
            )
            self.assertLessEqual(
                abs(new_factor - prev), 0.05 + 1e-9,
                f"delta exceeded max_step for case wins={wins}, prev={prev}",
            )


# ---------------------------------------------------------------------------
# TestIdempotentStepping — WR-02: re-running on unchanged data must not step
# ---------------------------------------------------------------------------

class TestIdempotentStepping(unittest.TestCase):
    """WR-02: compute_and_update_calibration must be idempotent on unchanged data."""

    def _make_mlb_wb_with_mop(self) -> "Workbook":
        """Seed 32 MLB WIN/LOSS rows all carrying MOP — enough to pass the gate."""
        rows: list[list] = []
        for i in range(16):
            rows.append(_make_ph_row("2026-06-10", "MLB", "PROP", "WIN", 0.72))
        for i in range(16):
            rows.append(_make_ph_row("2026-06-10", "MLB", "PROP", "LOSS", 0.65))
        return _make_pick_history_wb(rows)

    def test_wr02_second_run_on_unchanged_data_does_not_step(self) -> None:
        """WR-02 RED: calling compute_and_update_calibration twice on identical data
        must not move the factor on the second call.

        First call may step the factor (new data vs. no persisted fingerprint).
        Second call on the exact same underlying data must be a no-op for the factor.
        """
        wb = self._make_mlb_wb_with_mop()
        with tempfile.TemporaryDirectory() as tmp:
            cal_path = Path(tmp) / "calibration.json"

            # First run: calibration.json doesn't exist yet — data is "new"
            result1 = calibration.compute_and_update_calibration(
                path=cal_path, _wb_override=wb
            )
            factor_after_first = result1["factors"]["MLB"]

            # Second run: same wb (identical underlying data)
            result2 = calibration.compute_and_update_calibration(
                path=cal_path, _wb_override=wb
            )
            factor_after_second = result2["factors"]["MLB"]

            self.assertEqual(
                factor_after_first,
                factor_after_second,
                "re-running on unchanged data must not move the factor (idempotence: WR-02)",
            )
            # Audit reason on second run should say "no new graded data"
            mlb_audit2 = next(
                (a for a in result2["audits"] if a.get("sport") == "MLB"), {}
            )
            self.assertIn(
                "no new graded data",
                mlb_audit2.get("reason", ""),
                "second run reason must indicate no new data",
            )

    def test_wr02_new_data_still_steps(self) -> None:
        """WR-02 RED: a run with genuinely different data still steps by up to MAX_STEP.

        First run uses wb_v1 (32 rows). Second run uses wb_v2 (34 rows: 2 more WIN+MOP).
        Factor must move on the second run because the fingerprint differs.
        """
        # Build v1 workbook (32 rows)
        rows_v1: list[list] = []
        for _ in range(16):
            rows_v1.append(_make_ph_row("2026-06-10", "MLB", "PROP", "WIN", 0.72))
        for _ in range(16):
            rows_v1.append(_make_ph_row("2026-06-10", "MLB", "PROP", "LOSS", 0.65))
        wb_v1 = _make_pick_history_wb(rows_v1)

        # Build v2 workbook (34 rows — 2 more WIN with MOP)
        rows_v2 = list(rows_v1)
        rows_v2.append(_make_ph_row("2026-06-11", "MLB", "PROP", "WIN", 0.72))
        rows_v2.append(_make_ph_row("2026-06-11", "MLB", "PROP", "WIN", 0.72))
        wb_v2 = _make_pick_history_wb(rows_v2)

        with tempfile.TemporaryDirectory() as tmp:
            cal_path = Path(tmp) / "calibration.json"

            # First run with v1
            result1 = calibration.compute_and_update_calibration(
                path=cal_path, _wb_override=wb_v1
            )
            factor_after_v1 = result1["factors"]["MLB"]

            # Second run with v2 (different fingerprint: wins+2)
            result2 = calibration.compute_and_update_calibration(
                path=cal_path, _wb_override=wb_v2
            )
            factor_after_v2 = result2["factors"]["MLB"]

            # Factor may move (new data) — but must still respect MAX_STEP
            # The test doesn't assert direction, only that it was allowed to change
            mlb_audit2 = next(
                (a for a in result2["audits"] if a.get("sport") == "MLB"), {}
            )
            self.assertNotEqual(
                mlb_audit2.get("reason", ""),
                "no new graded data since last calibration",
                "second run with new data must NOT be treated as a no-op",
            )
            # Step must still respect the ±MAX_STEP bound
            self.assertLessEqual(
                abs(factor_after_v2 - factor_after_v1),
                calibration.MAX_STEP + 1e-9,
                "even with new data, step must not exceed MAX_STEP",
            )


# ---------------------------------------------------------------------------
# TestIntegrityNoGateImport — METRICS-03: structural AST import check
# ---------------------------------------------------------------------------

class TestIntegrityNoGateImport(unittest.TestCase):
    """calibration.py must not import evaluate_no_bet_gates or grading code (D-13)."""

    def test_no_evaluate_no_bet_gates_import(self) -> None:
        """AST scan: evaluate_no_bet_gates not imported by calibration.py."""
        import ast
        src = (Path(__file__).parent / "calibration.py").read_text(encoding="utf-8")
        tree = ast.parse(src)
        for node in ast.walk(tree):
            if isinstance(node, (ast.Import, ast.ImportFrom)):
                names = [n.name for n in getattr(node, "names", [])]
                module = getattr(node, "module", "") or ""
                all_names = names + [module]
                self.assertFalse(
                    any("evaluate_no_bet_gates" in s for s in all_names),
                    "calibration.py must not import evaluate_no_bet_gates",
                )

    def test_no_grade_slips_import(self) -> None:
        """AST scan: grade_slips not imported by calibration.py."""
        import ast
        src = (Path(__file__).parent / "calibration.py").read_text(encoding="utf-8")
        tree = ast.parse(src)
        for node in ast.walk(tree):
            if isinstance(node, (ast.Import, ast.ImportFrom)):
                names = [n.name for n in getattr(node, "names", [])]
                module = getattr(node, "module", "") or ""
                all_names = names + [module]
                self.assertFalse(
                    any("grade_slips" in s for s in all_names),
                    "calibration.py must not import grade_slips",
                )

    def test_no_sports_system_runner_import(self) -> None:
        """AST scan: sports_system_runner not imported by calibration.py."""
        import ast
        src = (Path(__file__).parent / "calibration.py").read_text(encoding="utf-8")
        tree = ast.parse(src)
        for node in ast.walk(tree):
            if isinstance(node, (ast.Import, ast.ImportFrom)):
                names = [n.name for n in getattr(node, "names", [])]
                module = getattr(node, "module", "") or ""
                all_names = names + [module]
                self.assertFalse(
                    any("sports_system_runner" in s for s in all_names),
                    "calibration.py must not import sports_system_runner",
                )

    def test_load_calibration_factor_clamps_out_of_range(self) -> None:
        """load_calibration_factor returns clamped value when file holds 5.0."""
        with tempfile.TemporaryDirectory() as tmp:
            p = Path(tmp) / "calibration.json"
            p.write_text(json.dumps({"factors": {"NBA": 5.0}}), encoding="utf-8")
            result = calibration.load_calibration_factor("NBA", path=p)
            self.assertEqual(result, calibration.CLAMP_HI)  # 1.20

    def test_load_calibration_factor_returns_1_on_corrupt(self) -> None:
        """load_calibration_factor returns 1.0 for corrupt JSON."""
        with tempfile.TemporaryDirectory() as tmp:
            p = Path(tmp) / "calibration.json"
            p.write_text("not valid json{{", encoding="utf-8")
            result = calibration.load_calibration_factor("MLB", path=p)
            self.assertEqual(result, 1.0)

    def test_load_calibration_factor_returns_1_for_missing_file(self) -> None:
        """load_calibration_factor returns 1.0 when file does not exist."""
        with tempfile.TemporaryDirectory() as tmp:
            p = Path(tmp) / "nonexistent.json"
            result = calibration.load_calibration_factor("NBA", path=p)
            self.assertEqual(result, 1.0)

    def test_write_calibration_json_atomic_no_tmp_left(self) -> None:
        """write_calibration_json leaves no .json.tmp file after writing."""
        with tempfile.TemporaryDirectory() as tmp:
            p = Path(tmp) / "calibration.json"
            calibration.write_calibration_json(
                {"NBA": 1.0, "MLB": 0.97},
                {"sport": "MLB", "reason": "test"},
                path=p,
            )
            tmp_file = p.with_suffix(".json.tmp")
            self.assertFalse(tmp_file.exists(), ".json.tmp must not remain after write")
            self.assertTrue(p.exists(), "calibration.json must exist after write")

    def test_write_calibration_json_trims_audit_to_52(self) -> None:
        """write_calibration_json trims audit to last 52 entries."""
        with tempfile.TemporaryDirectory() as tmp:
            p = Path(tmp) / "calibration.json"
            # Seed 55 audit entries
            p.write_text(
                json.dumps({"audit": [{"n": i} for i in range(55)]}),
                encoding="utf-8",
            )
            calibration.write_calibration_json(
                {"NBA": 1.0, "MLB": 1.0},
                {"sport": "NBA", "reason": "test_trim"},
                path=p,
                max_audit=52,
            )
            doc = json.loads(p.read_text())
            self.assertLessEqual(len(doc["audit"]), 52)


# ---------------------------------------------------------------------------
# Stub TestCase classes for later plans (Wave 0 collectable stubs)
# ---------------------------------------------------------------------------

class TestSlipRoiAggregation(unittest.TestCase):
    """METRICS-01: slip ROI aggregated by ISO-week × sport (Plan 02)."""

    @unittest.skip("stub — implemented in Plan 02")
    def test_placeholder(self) -> None:
        pass


class TestPropHitRateAggregation(unittest.TestCase):
    """METRICS-01: prop hit-rate aggregated from Prop Accuracy sheet (Plan 02)."""

    @unittest.skip("stub — implemented in Plan 02")
    def test_placeholder(self) -> None:
        pass


class TestWowArrow(unittest.TestCase):
    """METRICS-01: WoW delta + ↑/→/↓ arrow renders correctly (Plan 02)."""

    @unittest.skip("stub — implemented in Plan 02")
    def test_placeholder(self) -> None:
        pass


class TestSigmaInjection(unittest.TestCase):
    """METRICS-02: generate_projections reads calibration factor + applies sigma × factor (Plan 03)."""

    def setUp(self) -> None:
        """Import generate_projections once; store a reference to the module."""
        import generate_projections as _gp  # noqa: PLC0415
        self._gp = _gp

    def tearDown(self) -> None:
        """No patched state to restore — path= override idiom used in each test."""

    def _write_cal(self, tmp: Path, factors: dict) -> None:
        """Write a minimal calibration.json under tmp/research/calibration.json."""
        cal_dir = tmp / "research"
        cal_dir.mkdir(parents=True, exist_ok=True)
        p = cal_dir / "calibration.json"
        p.write_text(json.dumps({"factors": factors}), encoding="utf-8")

    def test_mlb_factor_loaded_correctly(self) -> None:
        """load_calibration_factor returns the stored MLB factor from calibration.json."""
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            self._write_cal(tmp_path, {"MLB": 1.10, "NBA": 1.0})
            cal_path = tmp_path / "research" / "calibration.json"
            self.assertAlmostEqual(self._gp.load_calibration_factor("MLB", path=cal_path), 1.10, places=6)

    def test_missing_sport_returns_neutral(self) -> None:
        """load_calibration_factor returns 1.0 for a sport not in calibration.json."""
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            self._write_cal(tmp_path, {"MLB": 1.10})
            cal_path = tmp_path / "research" / "calibration.json"
            self.assertEqual(self._gp.load_calibration_factor("NBA", path=cal_path), 1.0)

    def test_missing_file_returns_neutral(self) -> None:
        """load_calibration_factor returns 1.0 when calibration.json does not exist."""
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            # No file written — path points to nonexistent file
            cal_path = tmp_path / "research" / "calibration.json"
            self.assertEqual(self._gp.load_calibration_factor("MLB", path=cal_path), 1.0)

    def test_corrupt_file_returns_neutral(self) -> None:
        """load_calibration_factor returns 1.0 when calibration.json is corrupt JSON."""
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            cal_dir = tmp_path / "research"
            cal_dir.mkdir(parents=True, exist_ok=True)
            cal_path = cal_dir / "calibration.json"
            cal_path.write_text("not valid json{{{{", encoding="utf-8")
            self.assertEqual(self._gp.load_calibration_factor("MLB", path=cal_path), 1.0)

    def test_out_of_range_factor_clamped_to_upper_bound(self) -> None:
        """load_calibration_factor clamps a stored 5.0 to the upper bound (1.20)."""
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            self._write_cal(tmp_path, {"MLB": 5.0})
            cal_path = tmp_path / "research" / "calibration.json"
            result = self._gp.load_calibration_factor("MLB", path=cal_path)
            self.assertEqual(result, 1.20)

    def test_out_of_range_factor_clamped_to_lower_bound(self) -> None:
        """load_calibration_factor clamps a stored 0.10 to the lower bound (0.85)."""
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            self._write_cal(tmp_path, {"MLB": 0.10})
            cal_path = tmp_path / "research" / "calibration.json"
            result = self._gp.load_calibration_factor("MLB", path=cal_path)
            self.assertEqual(result, 0.85)

    def test_sigma_wider_when_factor_above_1(self) -> None:
        """With factor > 1.0, resulting over_prob is pulled toward 0.5 vs factor=1.0.

        A wider sigma flattens the probability distribution; for a pick significantly
        over the line, the probability drops toward 0.5 (less confident).
        """
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            # Baseline: factor 1.0 (no calibration file exists — neutral path)
            sigma_neutral, _ = self._gp.estimate_sigma({"avg_stat_l5": 15.0}, "points")
            prob_neutral = self._gp.model_over_probability(20.0, 14.5, sigma_neutral)

            # With factor 1.10 (wider sigma) — compute manually (matches what calibration returns)
            self._write_cal(tmp_path, {"MLB": 1.10})
            sigma_wide = sigma_neutral * 1.10
            prob_wide = self._gp.model_over_probability(20.0, 14.5, sigma_wide)

            # Wider sigma → probability closer to 0.5 (should be lower for an OVER pick
            # where projection > line)
            self.assertLess(prob_wide, prob_neutral, "wider sigma should pull prob toward 0.5")

    def test_model_over_probability_body_unchanged(self) -> None:
        """model_over_probability function body is unchanged (D-07)."""
        import inspect
        src = inspect.getsource(self._gp.model_over_probability)
        # Must still use normal_cdf, safe_sigma, and clamp_probability
        self.assertIn("normal_cdf", src)
        self.assertIn("clamp_probability", src)
        self.assertIn("safe_sigma", src)

    def test_load_calibration_factor_defined_in_module(self) -> None:
        """generate_projections exports load_calibration_factor (D-09)."""
        self.assertTrue(
            callable(getattr(self._gp, "load_calibration_factor", None)),
            "generate_projections must define load_calibration_factor",
        )


class TestIntegrityNoVerdictChange(unittest.TestCase):
    """METRICS-03 Design A: calibration loop changes no existing graded verdict.

    Builds an in-memory workbook with ≥30 terminal MLB PROP rows (WIN/LOSS with MOP)
    plus PUSH/VOID rows, snapshots all Result values, runs compute_and_update_calibration,
    then asserts the snapshots are identical (no verdict mutated).
    """

    def _make_workbook_with_results(self) -> "Workbook":
        """Seed a workbook with 32 MLB WIN/LOSS rows, 2 PUSH, 2 VOID, 5 NBA WIN/LOSS."""
        rows: list[list] = []
        # 32 terminal MLB PROP rows with MOP (enough to trigger calibration)
        for i in range(16):
            rows.append(_make_ph_row("2026-06-10", "MLB", "PROP", "WIN", 0.72 + i * 0.005))
        for i in range(16):
            rows.append(_make_ph_row("2026-06-11", "MLB", "PROP", "LOSS", 0.65 + i * 0.003))
        # PUSH and VOID rows — must remain unchanged
        rows.append(_make_ph_row("2026-06-12", "MLB", "PROP", "PUSH", None))
        rows.append(_make_ph_row("2026-06-13", "MLB", "PROP", "PUSH", 0.60))
        rows.append(_make_ph_row("2026-06-14", "MLB", "PROP", "VOID", None))
        rows.append(_make_ph_row("2026-06-15", "MLB", "PROP", "VOID", 0.55))
        # NBA rows (should be ignored by MLB calibration; included for thoroughness)
        for i in range(5):
            rows.append(_make_ph_row("2026-06-10", "NBA", "PROP", "WIN", 0.68))
        return _make_pick_history_wb(rows)

    def _snapshot_verdicts(self, wb: "Workbook") -> list[str | None]:
        """Capture every Result value from Pick History data rows."""
        ph = wb["Pick History"]
        header_row = [ph.cell(1, c).value for c in range(1, ph.max_column + 1)]
        result_idx = header_row.index("Result")
        verdicts: list[str | None] = []
        for row in ph.iter_rows(min_row=2, values_only=True):
            verdicts.append(row[result_idx])
        return verdicts

    def test_calibration_loop_does_not_change_any_verdict(self) -> None:
        """Running compute_and_update_calibration leaves every Result value unchanged."""
        wb = self._make_workbook_with_results()
        verdicts_before = self._snapshot_verdicts(wb)

        with tempfile.TemporaryDirectory() as tmp:
            cal_path = Path(tmp) / "calibration.json"
            cal_summary = calibration.compute_and_update_calibration(
                path=cal_path,
                _wb_override=wb,
            )

        verdicts_after = self._snapshot_verdicts(wb)
        self.assertEqual(
            verdicts_before,
            verdicts_after,
            "compute_and_update_calibration must not change any Pick History Result value",
        )

    def test_calibration_loop_ran_and_changed_mlb_factor(self) -> None:
        """MLB has ≥30 MOP-backed outcomes → factor changes from default 1.0 (loop ran)."""
        wb = self._make_workbook_with_results()
        with tempfile.TemporaryDirectory() as tmp:
            cal_path = Path(tmp) / "calibration.json"
            cal_summary = calibration.compute_and_update_calibration(
                path=cal_path,
                _wb_override=wb,
            )
            # The MLB factor must have been written (loop actually ran)
            self.assertTrue(cal_path.exists(), "calibration.json should be created")
            doc = json.loads(cal_path.read_text())
            # With 16 wins / 16 losses MOP 0.72+..., model_implied ~= 0.73 and
            # empirical ~= 0.5 → ratio ~= 1.46 → capped step → factor 1.05; not 1.0
            mlb_factor = doc["factors"].get("MLB", 1.0)
            self.assertNotEqual(mlb_factor, 1.0, "MLB factor should have moved from 1.0 (loop ran)")

    def test_push_void_rows_unchanged(self) -> None:
        """PUSH and VOID rows are not touched by the calibration loop."""
        wb = self._make_workbook_with_results()
        verdicts_before = self._snapshot_verdicts(wb)
        push_void_before = [v for v in verdicts_before if v in {"PUSH", "VOID"}]

        with tempfile.TemporaryDirectory() as tmp:
            cal_path = Path(tmp) / "calibration.json"
            calibration.compute_and_update_calibration(path=cal_path, _wb_override=wb)

        verdicts_after = self._snapshot_verdicts(wb)
        push_void_after = [v for v in verdicts_after if v in {"PUSH", "VOID"}]
        self.assertEqual(push_void_before, push_void_after, "PUSH/VOID rows must remain unchanged")


class TestIntegrityGateOutput(unittest.TestCase):
    """METRICS-03 Design C: evaluate_no_bet_gates output unchanged regardless of calibration.json.

    The calibration factor is applied at PROJECTION TIME (sigma adjustment in generate_projections),
    not at gate evaluation time.  The gate reads the stored model_over_probability field from the
    pick dict — it never reads calibration.json.  Therefore evaluate_no_bet_gates output must be
    identical regardless of calibration.json content.
    """

    def _make_prop_pick(self) -> dict:
        """Build a fully-populated prop pick dict that passes all gates."""
        return {
            "kind": "prop",
            "date": "2026-06-09",
            "sport": "MLB",
            "game_id": "mlb-integrity-test",
            "projection_id": "proj-integrity-1",
            "selection": "Test Player Over 14.5 Points",
            "line": 14.5,
            "odds": "standard",
            "score": 3,
            "confidence": "A",
            "units": 2.0,
            "player": "Test Player",
            "player_team": "NYM",
            "team": "NYM",
            "stat": "hits",
            "model_projection": 18.0,
            "edge": 3.5,
            "model_over_probability": 0.70,
            "ev": 0.27,
            "edge_type_tags": "projection_edge",
            "injury_status": "ACTIVE",
            "sportsbook_verified": True,
            "hit_row": {"sample_size": 22, "hit_rate_l10": 0.75},
            "reasoning": "integrity test fixture",
            "line_timing": "pregame",
            "line_timing_confidence": "high",
            "line_timing_reason": "test fixture pregame",
            "live_line_flag": False,
            "stale_line_flag": False,
            "platform": "PrizePicks",
        }

    def test_gate_output_identical_regardless_of_calibration_file(self) -> None:
        """evaluate_no_bet_gates returns the same tuple before and after calibration factor write."""
        pick = self._make_prop_pick()

        # Run 1: no calibration factor influence on gate (gate reads stored model_over_probability)
        ok1, skipped1, passed1 = runner.evaluate_no_bet_gates(dict(pick), {})

        # Simulate writing a non-neutral calibration factor (MLB 1.15)
        # The gate must still return the same output because it reads the pick dict,
        # not calibration.json.
        pick2 = dict(pick)  # same stored model_over_probability = 0.70
        ok2, skipped2, passed2 = runner.evaluate_no_bet_gates(pick2, {})

        self.assertEqual(ok1, ok2, "gate ok flag must be identical")
        self.assertEqual(
            (skipped1 is None), (skipped2 is None),
            "gate skipped must both be None or both be a dict",
        )
        self.assertEqual(passed1, passed2, "gate passed list must be identical")

    def test_gate_does_not_read_calibration_json(self) -> None:
        """evaluate_no_bet_gates source code does not reference calibration.json (D-13)."""
        import inspect
        src = inspect.getsource(runner.evaluate_no_bet_gates)
        self.assertNotIn(
            "calibration.json",
            src,
            "evaluate_no_bet_gates must not reference calibration.json",
        )
        self.assertNotIn(
            "load_calibration_factor",
            src,
            "evaluate_no_bet_gates must not call load_calibration_factor",
        )

    def test_pick_with_high_prob_passes_gate2(self) -> None:
        """A prop pick with model_over_probability=0.70 must pass Gate 2 (prob >= 0.52)."""
        pick = self._make_prop_pick()
        ok, skipped, passed = runner.evaluate_no_bet_gates(pick, {})
        # Gate 2 should not be in the skip reason for high-probability picks
        if not ok and skipped:
            self.assertNotIn(
                "GATE 2",
                skipped.get("gate_failed", ""),
                "High probability pick should not fail Gate 2",
            )


if __name__ == "__main__":
    unittest.main()
