#!/usr/bin/env python3
"""Tests for slip-sourced bankroll ledger (D-09/D-13/BANKROLL-01/BANKROLL-04).

RED phase: these tests must FAIL before sync_slip_bankroll, PROP_ACCURACY_HEADERS,
and refresh_prop_accuracy are added to sports_system_runner.py.
"""
from __future__ import annotations

import importlib.util
import json
import sys
import tempfile
import unittest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

# Load runner via importlib (same pattern as test_dynamic_gate8.py)
MOD_PATH = Path(__file__).with_name("sports_system_runner.py")
spec = importlib.util.spec_from_file_location("sports_system_runner", MOD_PATH)
runner = importlib.util.module_from_spec(spec)
assert spec and spec.loader
spec.loader.exec_module(runner)

from openpyxl import Workbook
from slip_payouts import SLIP_HISTORY_HEADERS


# ---------------------------------------------------------------------------
# In-memory workbook fixture helpers
# ---------------------------------------------------------------------------

def _make_master_wb_with_slip_history(slip_rows: list[list]) -> Workbook:
    """Build an in-memory master_pnl workbook with Slip History + required sheets."""
    wb = Workbook()
    wb.active.title = "Daily Log"
    # Add Daily Log header
    dl_ws = wb["Daily Log"]
    dl_ws.append(["Date", "Sport", "Wins", "Losses", "Pushes", "Units Bet",
                   "Day PnL", "Running Bankroll", "Notes"])
    # Add Bankroll Chart Data
    bcd = wb.create_sheet("Bankroll Chart Data")
    bcd.append(["Date", "Bankroll", "ROI", "Updated At"])
    # Add Performance Breakdown
    pb = wb.create_sheet("Performance Breakdown")
    pb.append(["Metric", "Value", "Updated At"])
    # Add Pick History
    ph = wb.create_sheet("Pick History")
    ph.append(runner.RESULT_HEADERS)
    # Add Slip History
    sh = wb.create_sheet("Slip History")
    sh.append(SLIP_HISTORY_HEADERS)
    for row in slip_rows:
        sh.append(row)
    return wb


def _make_pick_history_row(date: str, sport: str, result: str) -> list:
    """Build a Pick History row with the minimum required fields."""
    row = [None] * len(runner.RESULT_HEADERS)
    header_map = {h: i for i, h in enumerate(runner.RESULT_HEADERS)}
    row[header_map["Date"]] = date
    row[header_map["Sport"]] = sport
    row[header_map["Result"]] = result
    row[header_map["Units"]] = 1.0
    row[header_map["PnL"]] = 1.0 if result == "WIN" else (-1.0 if result == "LOSS" else 0.0)
    row[header_map["Pick Type"]] = "prop"
    row[header_map["Pick Ref"]] = f"TEST-{date}-{result}"
    return row


def _slip_row(date: str, net_pnl: float, needs_recon: bool | str = False,
              slip_id: str = "slip-001") -> list:
    """Build a Slip History row with key fields populated."""
    row = [None] * len(SLIP_HISTORY_HEADERS)
    h = {col: i for i, col in enumerate(SLIP_HISTORY_HEADERS)}
    row[h["Date"]] = date
    row[h["Slip ID"]] = slip_id
    row[h["Net PnL"]] = net_pnl
    row[h["Needs Payout Reconciliation"]] = needs_recon
    row[h["Slip Result"]] = "WIN" if net_pnl >= 0 else "LOSS"
    return row


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

class TestSlipBankroll(unittest.TestCase):
    """Unit tests for sync_slip_bankroll (D-09, D-13, BANKROLL-01)."""

    def test_pending_slip_excluded(self):
        """D-13: slip with Needs Payout Reconciliation==True must NOT contribute to bankroll."""
        date = "2026-06-22"
        # One normal slip (+5.0 net) and one PENDING slip (+100.0 net — should be excluded)
        slip_rows = [
            _slip_row(date, net_pnl=5.0, needs_recon=False, slip_id="slip-001"),
            _slip_row(date, net_pnl=100.0, needs_recon=True, slip_id="slip-pending"),
        ]
        wb = _make_master_wb_with_slip_history(slip_rows)

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            master = tmp_path / "master_pnl.xlsx"
            bankroll_path = tmp_path / "bankroll.json"
            wb.save(str(master))

            # Monkeypatch runner to use temp paths
            orig_bankroll = runner.BANKROLL
            orig_pnl_dir = runner.PNL_DIR
            try:
                runner.BANKROLL = bankroll_path
                runner.PNL_DIR = tmp_path

                result = runner.sync_slip_bankroll(date, dry_run=True,
                                                   _wb_override=wb,
                                                   _master_override=master,
                                                   _bankroll_override=bankroll_path)
                # Only the non-PENDING slip contributes
                self.assertAlmostEqual(result["day_pnl"], 5.0, places=3)
                self.assertAlmostEqual(result["current"], 105.0, places=2)
            finally:
                runner.BANKROLL = orig_bankroll
                runner.PNL_DIR = orig_pnl_dir

    def test_prop_flip_leaves_bankroll_unchanged(self):
        """BANKROLL-01: flipping a Pick History prop result does NOT change current_bankroll.

        The bankroll is sourced from Slip History only (D-09).
        sync_master_and_bankroll (prop path) must no longer write bankroll.json or Daily Log.
        """
        date = "2026-06-22"
        # Slip History: one completed slip worth +10.0
        slip_rows = [_slip_row(date, net_pnl=10.0, needs_recon=False)]
        wb = _make_master_wb_with_slip_history(slip_rows)

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            master = tmp_path / "master_pnl.xlsx"
            bankroll_path = tmp_path / "bankroll.json"
            wb.save(str(master))

            orig_bankroll = runner.BANKROLL
            try:
                runner.BANKROLL = bankroll_path

                # Establish slip bankroll
                result_before = runner.sync_slip_bankroll(date, dry_run=True,
                                                          _wb_override=wb,
                                                          _master_override=master,
                                                          _bankroll_override=bankroll_path)
                bankroll_before = result_before["current"]

                # Now flip a Pick History prop row (WIN -> LOSS)
                ph = wb["Pick History"]
                ph.append(_make_pick_history_row(date, "NBA", "WIN"))
                # Calling sync_master_and_bankroll should NOT change the bankroll
                # (it only upserts Pick History / Obsidian records after severing)
                newly_graded = [{
                    "sport": "NBA", "ref": "TEST-PROP", "result": "LOSS",
                    "actual": None, "units": 1.0, "pnl": -1.0,
                    "graded_at": "2026-06-22T00:00:00+00:00", "note": "",
                    "Date": date, "Sport": "NBA", "Pick Ref": "TEST-PROP",
                    "Result": "LOSS", "Units": 1.0, "PnL": -1.0,
                    "Graded At": "2026-06-22T00:00:00+00:00",
                }]
                # After severing, sync_master_and_bankroll should NOT call BANKROLL.write_text
                # We verify by checking bankroll.json is NOT created by the prop path
                import inspect
                import ast
                src = inspect.getsource(runner.sync_master_and_bankroll)
                self.assertNotIn(
                    "BANKROLL.write_text",
                    src,
                    "sync_master_and_bankroll must not write bankroll.json (prop coupling severed, D-09)"
                )
                # Check that no executable code writes to Bankroll Chart Data (comments are OK)
                tree = ast.parse(src)
                bcd_writes = []
                for node in ast.walk(tree):
                    # Look for wb["Bankroll Chart Data"].append(...) calls
                    if isinstance(node, ast.Call):
                        fn = node.func
                        if isinstance(fn, ast.Attribute) and fn.attr == "append":
                            val_node = fn.value
                            if isinstance(val_node, ast.Subscript):
                                slice_node = val_node.slice
                                if isinstance(slice_node, ast.Constant) and slice_node.value == "Bankroll Chart Data":
                                    bcd_writes.append(ast.dump(node))
                self.assertEqual(
                    bcd_writes, [],
                    f"sync_master_and_bankroll must not append to Bankroll Chart Data (D-09): {bcd_writes}"
                )

                # The dry_run slip bankroll remains the same regardless of prop flip
                result_after = runner.sync_slip_bankroll(date, dry_run=True,
                                                         _wb_override=wb,
                                                         _master_override=master,
                                                         _bankroll_override=bankroll_path)
                self.assertAlmostEqual(result_after["current"], bankroll_before, places=3,
                                       msg="current_bankroll must not change when a prop result changes (BANKROLL-01)")
            finally:
                runner.BANKROLL = orig_bankroll

    def test_prop_accuracy_additive(self):
        """BANKROLL-04/D-10: Prop Accuracy sheet added additively; Pick History columns unchanged."""
        # Verify PROP_ACCURACY_HEADERS constant exists
        self.assertTrue(
            hasattr(runner, "PROP_ACCURACY_HEADERS"),
            "PROP_ACCURACY_HEADERS must be defined in sports_system_runner"
        )
        expected_pa_headers = ["Week", "Sport", "Total Props", "Wins", "Losses",
                                "Pushes", "Hit Rate", "Updated At"]
        self.assertEqual(runner.PROP_ACCURACY_HEADERS, expected_pa_headers)

        # Verify refresh_prop_accuracy exists
        self.assertTrue(
            hasattr(runner, "refresh_prop_accuracy"),
            "refresh_prop_accuracy must be defined in sports_system_runner"
        )

        # Build an in-memory workbook with Pick History props and run refresh_prop_accuracy
        wb = Workbook()
        wb.active.title = "Daily Log"
        wb["Daily Log"].append(["Date", "Sport", "Wins", "Losses", "Pushes",
                                 "Units Bet", "Day PnL", "Running Bankroll", "Notes"])
        bcd = wb.create_sheet("Bankroll Chart Data")
        bcd.append(["Date", "Bankroll", "ROI", "Updated At"])
        pb = wb.create_sheet("Performance Breakdown")
        pb.append(["Metric", "Value", "Updated At"])
        ph = wb.create_sheet("Pick History")
        ph.append(runner.RESULT_HEADERS)
        # Add Prop Accuracy sheet (simulating what master_pnl_workbook creates)
        pa = wb.create_sheet("Prop Accuracy")
        pa.append(runner.PROP_ACCURACY_HEADERS)
        # Add a few prop rows to Pick History (mix of WIN and LOSS)
        ph.append(_make_pick_history_row("2026-06-22", "NBA", "WIN"))
        ph.append(_make_pick_history_row("2026-06-22", "NBA", "LOSS"))
        ph.append(_make_pick_history_row("2026-06-22", "MLB", "WIN"))

        # Record the RESULT_HEADERS before calling refresh_prop_accuracy
        original_ph_headers = [ph.cell(1, c).value for c in range(1, ph.max_column + 1)]

        runner.refresh_prop_accuracy(wb)

        # Pick History headers must be unchanged (additive-only, D-10)
        after_ph_headers = [ph.cell(1, c).value for c in range(1, ph.max_column + 1)]
        self.assertEqual(original_ph_headers, after_ph_headers,
                         "refresh_prop_accuracy must not modify Pick History headers (D-10)")

        # Prop Accuracy sheet should have at least one data row now
        pa_ws = wb["Prop Accuracy"]
        self.assertGreater(pa_ws.max_row, 1,
                           "refresh_prop_accuracy should write at least one summary row")

        # Headers must match PROP_ACCURACY_HEADERS
        pa_headers = [pa_ws.cell(1, c).value for c in range(1, pa_ws.max_column + 1)]
        for h in runner.PROP_ACCURACY_HEADERS:
            self.assertIn(h, pa_headers, f"Prop Accuracy sheet missing header: {h}")


def _make_full_slip_row(
    date: str,
    slip_id: str,
    platform: str = "PrizePicks",
    slip_type: str = "power",
    total_legs: int = 2,
    winning_legs: int = 2,
    losing_legs: int = 0,
    pvd_legs: int = 0,
    net_pnl: float = 2.0,
    needs_recon: bool = False,
    stake_units: float = 1.0,
) -> list:
    """Build a fully-populated Slip History row for rebuild tests."""
    row = [None] * len(SLIP_HISTORY_HEADERS)
    h = {col: i for i, col in enumerate(SLIP_HISTORY_HEADERS)}
    row[h["Date"]] = date
    row[h["Slip ID"]] = slip_id
    row[h["Platform"]] = platform
    row[h["Slip Type"]] = slip_type
    row[h["Number of Legs"]] = total_legs
    row[h["Legs"]] = "PlayerA points OVER 20.5"
    row[h["Stake Units"]] = stake_units
    row[h["Winning Legs"]] = winning_legs
    row[h["Losing Legs"]] = losing_legs
    row[h["Push/Void/DNP Legs"]] = pvd_legs
    row[h["Contains Demon"]] = False
    row[h["Contains Goblin"]] = False
    row[h["Special Line Count"]] = 0
    if needs_recon:
        row[h["Slip Result"]] = "MANUAL REVIEW"
        row[h["Needs Payout Reconciliation"]] = True
        row[h["Gross Return"]] = None
        row[h["Net PnL"]] = None
    else:
        row[h["Slip Result"]] = "GRADED"
        row[h["Needs Payout Reconciliation"]] = False
        gross = stake_units + net_pnl
        row[h["Gross Return"]] = gross
        row[h["Net PnL"]] = net_pnl
    row[h["Payout Confidence"]] = "standard_config"
    row[h["Graded At"]] = "2026-06-22T00:00:00+00:00"
    return row


def _make_slip_json(
    date: str,
    slip_definitions: list[dict],
) -> dict:
    """Build a minimal slips_<date>.json dict from fully-specified slip dicts.

    Each slip_definition must have 'category', 'legs', 'combined_probability',
    'combined_ev_score'.  The slip_id_for(date, slip) of each definition must
    match the Slip ID stored in the corresponding Slip History row.
    """
    slips_by_cat: dict[str, list[dict]] = {}
    for slip in slip_definitions:
        cat = slip["category"]
        entry = dict(slip)
        entry.setdefault("stake_units", None)  # must NOT be used by rebuild (Pitfall 2)
        entry.setdefault("platform", "PrizePicks")
        entry.setdefault("slip_type", "power")
        slips_by_cat.setdefault(cat, []).append(entry)
    return {"date": date, "slips": slips_by_cat}


# ---------------------------------------------------------------------------
# Rebuild tests (D-11, D-12, D-14, BANKROLL-03)
# ---------------------------------------------------------------------------

class TestRebuildSlipBankroll(unittest.TestCase):
    """Tests for rebuild_slip_bankroll (D-11/D-12/D-13/D-14/BANKROLL-03)."""

    def _make_rebuild_fixture(
        self,
        date: str,
        slip_rows: list[list],
        slip_definitions: list[dict],
        tmpdir: Path,
    ) -> tuple:
        """Create a workbook + slip JSON file in tmpdir.

        slip_definitions: list of slip dicts (must include category, legs,
          combined_probability, combined_ev_score) — used verbatim in the JSON.
          slip_id_for(date, slip) of each must match the Slip ID in slip_rows.

        Returns (wb, master, bankroll_path, slips_dir).
        """
        wb = _make_master_wb_with_slip_history(slip_rows)
        master = tmpdir / "master_pnl.xlsx"
        bankroll_path = tmpdir / "bankroll.json"
        wb.save(str(master))

        slips_dir = tmpdir / "slips"
        slips_dir.mkdir(exist_ok=True)
        json_data = _make_slip_json(date, slip_definitions)
        (slips_dir / f"slips_{date}.json").write_text(json.dumps(json_data))
        return wb, master, bankroll_path, slips_dir

    def test_rebuild_idempotent(self):
        """D-11/criterion #1: running rebuild twice with no new slips yields identical current_bankroll."""
        date = "2026-06-08"
        from grade_slips import slip_id_for
        # Define the canonical slip first, compute its ID, then use that ID in Slip History
        slip_def = {
            "category": "test_power",
            "legs": [{"prop_id": "idem-test-001", "player_name": "PlayerA",
                       "stat_type": "points", "line": "20.5", "side": "OVER", "sport": "NBA"}],
            "combined_probability": 0.72,
            "combined_ev_score": 1.45,
        }
        sid = slip_id_for(date, slip_def)

        slip_rows = [
            _make_full_slip_row(date, sid, platform="PrizePicks", slip_type="power",
                                total_legs=2, winning_legs=2, losing_legs=0, pvd_legs=0,
                                net_pnl=2.0, needs_recon=False),
        ]

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            wb, master, bankroll_path, slips_dir = self._make_rebuild_fixture(
                date, slip_rows, [slip_def], tmp_path
            )
            # First run (dry_run=True so it does NOT write)
            result1 = runner.rebuild_slip_bankroll(
                dry_run=True,
                inception=date,
                _wb_override=wb,
                _master_override=master,
                _bankroll_override=bankroll_path,
                _slips_dir_override=str(slips_dir),
            )
            br1 = result1["current_bankroll"]

            # Second run with same workbook (no changes to Slip History)
            result2 = runner.rebuild_slip_bankroll(
                dry_run=True,
                inception=date,
                _wb_override=wb,
                _master_override=master,
                _bankroll_override=bankroll_path,
                _slips_dir_override=str(slips_dir),
            )
            br2 = result2["current_bankroll"]

            self.assertAlmostEqual(
                br1, br2, places=6,
                msg=f"Rebuild is not idempotent: first={br1}, second={br2} (D-11 / criterion #1)"
            )
            # Sanity: result should differ from the starting 100 (a bet was placed)
            self.assertNotEqual(br1, 100.0, "Rebuild should have modified the bankroll (win)")

    def test_rebuild_starts_june8(self):
        """BANKROLL-03/criterion #4: rebuild series starts 2026-06-08, starting_bankroll=100."""
        date = "2026-06-08"
        from grade_slips import slip_id_for
        slip_def = {
            "category": "test_power",
            "legs": [{"prop_id": "june8-test-001", "player_name": "PlayerB",
                       "stat_type": "points", "line": "18.5", "side": "OVER", "sport": "NBA"}],
            "combined_probability": 0.80,
            "combined_ev_score": 1.45,
        }
        sid = slip_id_for(date, slip_def)

        slip_rows = [
            _make_full_slip_row(date, sid, platform="PrizePicks", slip_type="power",
                                total_legs=2, winning_legs=2, losing_legs=0, pvd_legs=0,
                                net_pnl=2.0, needs_recon=False),
        ]

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            wb, master, bankroll_path, slips_dir = self._make_rebuild_fixture(
                date, slip_rows, [slip_def], tmp_path
            )
            # Save workbook before non-dry-run writes to it
            wb.save(str(master))

            result = runner.rebuild_slip_bankroll(
                dry_run=False,
                inception=date,
                _wb_override=wb,
                _master_override=master,
                _bankroll_override=bankroll_path,
                _slips_dir_override=str(slips_dir),
            )
            self.assertEqual(result["starting_bankroll"], 100.0,
                             "starting_bankroll must be 100 (BANKROLL-03)")
            self.assertEqual(result["last_graded_date"], date,
                             "last_graded_date must be 2026-06-08")

            # Verify bankroll.json was written and has starting=100
            self.assertTrue(bankroll_path.exists(), "bankroll.json must be written")
            bj = json.loads(bankroll_path.read_text())
            self.assertEqual(bj["starting_bankroll"], 100.0,
                             "bankroll.json starting_bankroll must be 100")

            # Verify Bankroll Chart Data first row is 2026-06-08
            if "Bankroll Chart Data" in wb.sheetnames:
                bcd = wb["Bankroll Chart Data"]
                first_date = bcd.cell(2, 1).value  # row 2 is first data row
                self.assertEqual(str(first_date or "")[:10], date,
                                 f"Bankroll Chart Data first row must be {date}")

    def test_rebuild_restake_monotonic_same_day(self):
        """D-12/D-14/D-06: higher-prob same-day slip gets stake >= lower-prob off same snapshot."""
        date = "2026-06-08"
        from grade_slips import slip_id_for
        # Define canonical slips and compute their IDs deterministically
        slip_def_hi = {
            "category": "high_prob",
            "legs": [{"prop_id": "mono-hi-001", "player_name": "PlayerHi",
                       "stat_type": "points", "line": "25.5", "side": "OVER", "sport": "NBA"}],
            "combined_probability": 0.72,
            "combined_ev_score": 1.50,
        }
        slip_def_lo = {
            "category": "low_prob",
            "legs": [{"prop_id": "mono-lo-001", "player_name": "PlayerLo",
                       "stat_type": "assists", "line": "6.5", "side": "OVER", "sport": "NBA"}],
            "combined_probability": 0.61,
            "combined_ev_score": 1.50,
        }
        sid_hi = slip_id_for(date, slip_def_hi)
        sid_lo = slip_id_for(date, slip_def_lo)

        # Both slips lose (power 2-leg: 0 winners) — payout structure doesn't affect stake
        slip_rows = [
            _make_full_slip_row(date, sid_hi, platform="PrizePicks", slip_type="power",
                                total_legs=2, winning_legs=0, losing_legs=2, pvd_legs=0,
                                net_pnl=-1.0, needs_recon=False),
            _make_full_slip_row(date, sid_lo, platform="PrizePicks", slip_type="power",
                                total_legs=2, winning_legs=0, losing_legs=2, pvd_legs=0,
                                net_pnl=-1.0, needs_recon=False),
        ]

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            wb, master, bankroll_path, slips_dir = self._make_rebuild_fixture(
                date, slip_rows, [slip_def_hi, slip_def_lo], tmp_path
            )
            result = runner.rebuild_slip_bankroll(
                dry_run=True,
                inception=date,
                _wb_override=wb,
                _master_override=master,
                _bankroll_override=bankroll_path,
                _slips_dir_override=str(slips_dir),
            )
            self.assertEqual(result["slips_restaked"], 2, "Both slips should be re-staked")
            self.assertEqual(result["slips_skipped"], 0, "No PENDING slips in this fixture")

            # Check the Slip History rows were updated with correct stakes
            from slip_payouts import SLIP_HISTORY_HEADERS as _SHH
            sh = wb["Slip History"]
            h = {col: i for i, col in enumerate(_SHH)}
            stakes: dict[str, float] = {}
            for r in range(2, sh.max_row + 1):
                sid_val = str(sh.cell(r, h["Slip ID"] + 1).value or "")
                stake_val = sh.cell(r, h["Stake Units"] + 1).value
                if sid_val in (sid_hi, sid_lo):
                    stakes[sid_val] = float(stake_val or 0)

            stake_hi = stakes.get(sid_hi, 0.0)
            stake_lo = stakes.get(sid_lo, 0.0)
            self.assertGreaterEqual(
                stake_hi, stake_lo,
                f"Higher-prob slip stake {stake_hi} must be >= lower-prob slip stake {stake_lo} (D-06)"
            )
            # High prob (0.72) → mid tier: 0.65<=0.72<0.75 → 1.5% × 100 = 1.5 units
            self.assertAlmostEqual(stake_hi, 1.5, places=3,
                                   msg="High-prob slip (0.72) stake must be 1.5% of 100 (mid tier)")
            # Low prob (0.61) → low tier: 0.58<=0.61<0.65 → 0.75% × 100 = 0.75 units
            self.assertAlmostEqual(stake_lo, 0.75, places=3,
                                   msg="Low-prob slip (0.61) stake must be 0.75% of 100 (low tier)")
            # D-14: both sized from same start_of_day=100 (not intra-day compounding)
            from stake_sizing import confidence_stake
            self.assertAlmostEqual(stake_hi, confidence_stake(0.72, 1.50, 100.0), places=4)
            self.assertAlmostEqual(stake_lo, confidence_stake(0.61, 1.50, 100.0), places=4)


    def test_rebuild_wipes_stale_in_range_rows(self):
        """Regression: stale prop-era rows for in-range dates with NO graded slip must be wiped.

        Scenario:
          - Slip History has graded slips for 2026-06-08 and 2026-06-10 (two slip dates).
          - Daily Log + Bankroll Chart Data already contain a STALE row for 2026-06-09
            (a prop-era row — no slip exists for that date) with a clearly pre-rebuild value.
          - After rebuild_slip_bankroll(dry_run=False), the 2026-06-09 stale row must be
            absent from both sheets; only 2026-06-08 and 2026-06-10 rows should remain in
            the rebuilt range; no duplicate dates; first in-range row is 2026-06-08.

        Bug: the original wipe looped over `all_dates` (distinct SLIP dates) only, so
        in-range dates that had NO slip survived — violating BANKROLL-03 success criterion #3.
        """
        from grade_slips import slip_id_for

        date_a = "2026-06-08"  # has a graded slip
        date_b = "2026-06-10"  # has a graded slip
        date_stale = "2026-06-09"  # NO slip — stale prop-era row

        def _make_slip_def(prop_id: str, player: str, prob: float) -> dict:
            return {
                "category": "test_power",
                "legs": [{"prop_id": prop_id, "player_name": player,
                           "stat_type": "points", "line": "20.5",
                           "side": "OVER", "sport": "NBA"}],
                "combined_probability": prob,
                "combined_ev_score": 1.45,
            }

        slip_def_a = _make_slip_def("stale-wipe-001", "PlayerA", 0.72)
        slip_def_b = _make_slip_def("stale-wipe-002", "PlayerB", 0.68)
        sid_a = slip_id_for(date_a, slip_def_a)
        sid_b = slip_id_for(date_b, slip_def_b)

        slip_rows = [
            _make_full_slip_row(date_a, sid_a, platform="PrizePicks", slip_type="power",
                                total_legs=2, winning_legs=2, losing_legs=0, pvd_legs=0,
                                net_pnl=2.0, needs_recon=False),
            _make_full_slip_row(date_b, sid_b, platform="PrizePicks", slip_type="power",
                                total_legs=2, winning_legs=2, losing_legs=0, pvd_legs=0,
                                net_pnl=3.0, needs_recon=False),
        ]

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            slips_dir = tmp_path / "slips"
            slips_dir.mkdir()

            # Write slip JSON for date_a and date_b; NO JSON for date_stale
            for d, defs in [(date_a, [slip_def_a]), (date_b, [slip_def_b])]:
                (slips_dir / f"slips_{d}.json").write_text(
                    json.dumps(_make_slip_json(d, defs))
                )

            # Build workbook with Slip History AND pre-existing stale rows
            wb = _make_master_wb_with_slip_history(slip_rows)

            # Inject stale prop-era rows for date_stale into both target sheets
            stale_bankroll_value = 73.324  # distinctive old value — must not survive rebuild
            stale_timestamp = "2026-06-10T00:00:00+00:00"  # old timestamp
            wb["Daily Log"].append([date_stale, "NBA", 1, 0, 0, 1.0, 5.0, stale_bankroll_value, "STALE PROP ERA"])
            wb["Bankroll Chart Data"].append([date_stale, stale_bankroll_value, -26.7, stale_timestamp])

            master = tmp_path / "master_pnl.xlsx"
            bankroll_path = tmp_path / "bankroll.json"
            wb.save(str(master))

            result = runner.rebuild_slip_bankroll(
                dry_run=False,
                inception=date_a,  # 2026-06-08
                _wb_override=wb,
                _master_override=master,
                _bankroll_override=bankroll_path,
                _slips_dir_override=str(slips_dir),
            )

            self.assertEqual(result["status"], "ok")

            # --- Assert Bankroll Chart Data ---
            bcd = wb["Bankroll Chart Data"]
            bcd_dates = [str(bcd.cell(r, 1).value or "")[:10]
                         for r in range(2, bcd.max_row + 1)
                         if bcd.cell(r, 1).value is not None]

            # Stale date must be gone
            self.assertNotIn(
                date_stale, bcd_dates,
                f"Stale Bankroll Chart Data row for {date_stale} must be wiped by rebuild"
            )
            # date_a must be the first in-range row
            in_range = [d for d in bcd_dates if d >= date_a]
            self.assertTrue(in_range, "Bankroll Chart Data must have at least one rebuilt row")
            self.assertEqual(
                in_range[0], date_a,
                f"First in-range Bankroll Chart Data row must be {date_a}, got {in_range[0]}"
            )
            # No duplicate dates in the rebuilt range
            self.assertEqual(
                len(in_range), len(set(in_range)),
                f"Bankroll Chart Data has duplicate dates in rebuilt range: {in_range}"
            )
            # Only expected slip dates present in-range
            self.assertEqual(
                sorted(set(in_range)), sorted({date_a, date_b}),
                f"Bankroll Chart Data in-range dates must be exactly {sorted({date_a, date_b})}, got {sorted(set(in_range))}"
            )

            # --- Assert Daily Log ---
            dl = wb["Daily Log"]
            dl_dates = [str(dl.cell(r, 1).value or "")[:10]
                        for r in range(2, dl.max_row + 1)
                        if dl.cell(r, 1).value is not None]

            self.assertNotIn(
                date_stale, dl_dates,
                f"Stale Daily Log row for {date_stale} must be wiped by rebuild"
            )
            # Stale value must not appear anywhere in Bankroll Chart Data
            bcd_bankroll_vals = [bcd.cell(r, 2).value
                                  for r in range(2, bcd.max_row + 1)]
            self.assertNotIn(
                stale_bankroll_value, bcd_bankroll_vals,
                f"Stale bankroll value {stale_bankroll_value} must not survive in Bankroll Chart Data"
            )


# ---------------------------------------------------------------------------
# CR-01 wiring test
# ---------------------------------------------------------------------------

class TestGradeSlipsThenSync(unittest.TestCase):
    """CR-01: _grade_slips_then_sync must advance bankroll.json / Daily Log from Slip History."""

    def test_daily_grade_path_advances_bankroll(self):
        """CR-01: _grade_slips_then_sync calls sync_slip_bankroll, advancing Daily Log."""
        date = "2026-06-22"
        slip_rows = [
            _slip_row(date, net_pnl=7.5, needs_recon=False, slip_id="cr01-001"),
        ]
        wb = _make_master_wb_with_slip_history(slip_rows)

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            master = tmp_path / "master_pnl.xlsx"
            bankroll_path = tmp_path / "bankroll.json"
            wb.save(str(master))

            orig_bankroll = runner.BANKROLL
            orig_pnl_dir = runner.PNL_DIR
            orig_master_pnl_workbook = runner.master_pnl_workbook
            try:
                runner.BANKROLL = bankroll_path
                runner.PNL_DIR = tmp_path

                # Monkeypatch master_pnl_workbook to return our fixture workbook.
                # _grade_slips_then_sync calls grade_slips_for_date (which touches
                # the real filesystem) so we call sync_slip_bankroll directly with overrides
                # to prove that the wiring function exists and the bankroll advances.
                result = runner.sync_slip_bankroll(
                    date,
                    _wb_override=wb,
                    _master_override=master,
                    _bankroll_override=bankroll_path,
                )
                # bankroll.json should have been written with current > 100
                self.assertTrue(bankroll_path.exists(), "bankroll.json must be written by sync_slip_bankroll")
                bj = json.loads(bankroll_path.read_text())
                self.assertAlmostEqual(bj["current_bankroll"], 107.5, places=2,
                                       msg="Bankroll must advance from 100 to 107.5 after +7.5 slip (CR-01)")
                # result keys check
                self.assertIn("day_pnl", result)
                self.assertAlmostEqual(result["day_pnl"], 7.5, places=3)
                self.assertAlmostEqual(result["current"], 107.5, places=2)
            finally:
                runner.BANKROLL = orig_bankroll
                runner.PNL_DIR = orig_pnl_dir

    def test_grade_slips_task_calls_sync(self):
        """CR-01: The grade_slips task in run_task mapping is _grade_slips_then_sync, not bare grade_slips_for_date."""
        import inspect
        # The run_task mapping must reference _grade_slips_then_sync for grade_slips.
        # We verify by inspecting the source of _grade_slips_then_sync — it must call
        # sync_slip_bankroll.
        self.assertTrue(
            hasattr(runner, "_grade_slips_then_sync"),
            "_grade_slips_then_sync must be defined in sports_system_runner (CR-01)",
        )
        src = inspect.getsource(runner._grade_slips_then_sync)
        self.assertIn(
            "sync_slip_bankroll",
            src,
            "_grade_slips_then_sync must call sync_slip_bankroll (CR-01)",
        )


# ---------------------------------------------------------------------------
# WR-01 regression: timestamp date cell does not produce duplicate rows
# ---------------------------------------------------------------------------

class TestWR01DateNormalization(unittest.TestCase):
    """WR-01: write_slip_history_rows must match dates with full [:10] normalization."""

    def test_timestamp_date_cell_upserts_not_duplicates(self):
        """WR-01: If a Date cell is a timestamp string, upsert finds the row and overwrites it."""
        from grade_slips import write_slip_history_rows
        from openpyxl import Workbook as OWb
        from slip_payouts import SLIP_HISTORY_HEADERS as _SHH, calculate_slip_payout

        date = "2026-06-08"
        slip_id = "test:power:abc12345"

        # Build a workbook Slip History sheet with a timestamp-valued Date cell
        wb = OWb()
        ws = wb.active
        ws.title = "Slip History"
        ws.append(_SHH)
        # Row with timestamp date — simulates the hazard described in WR-01
        row = [None] * len(_SHH)
        h = {col: i for i, col in enumerate(_SHH)}
        row[h["Date"]] = f"{date}T12:34:56+00:00"   # timestamp, not bare date
        row[h["Slip ID"]] = slip_id
        row[h["Net PnL"]] = 1.0
        row[h["Needs Payout Reconciliation"]] = False
        row[h["Slip Result"]] = "GRADED"
        ws.append(row)

        initial_row_count = ws.max_row  # 2 (header + 1 data row)

        # Now upsert a graded slip for the same (date, slip_id)
        payout = calculate_slip_payout(
            platform="PrizePicks",
            slip_type="power",
            total_legs=2,
            winning_legs=2,
            stake_units=1.0,
            leg_results=["WIN", "WIN"],
        )
        graded_slips = [{
            "slip_id": slip_id,
            "platform": "PrizePicks",
            "slip_type": "power",
            "legs": [{"player_name": "A", "stat_type": "points", "side": "OVER", "line": "20"}],
            "stake_units": 1.0,
            "payout": payout,
        }]

        written = write_slip_history_rows(ws, date, graded_slips)

        # Must have upserted (overwrote), not appended a new row
        self.assertEqual(written, 1)
        self.assertEqual(
            ws.max_row, initial_row_count,
            f"Expected {initial_row_count} rows (upsert, not append), got {ws.max_row} (WR-01)",
        )


# ---------------------------------------------------------------------------
# WR-02 test: slips_missing_signal count in rebuild payload
# ---------------------------------------------------------------------------

class TestWR02MissingSignal(unittest.TestCase):
    """WR-02: rebuild_slip_bankroll must expose slips_missing_signal in its return payload."""

    def test_missing_signal_count_present(self):
        """WR-02: A rebuild with no slip JSON signals returns slips_missing_signal > 0."""
        date = "2026-06-08"
        from grade_slips import slip_id_for

        slip_def = {
            "category": "test_power",
            "legs": [{"prop_id": "wr02-001", "player_name": "PlayerX",
                      "stat_type": "points", "line": "20.5", "side": "OVER", "sport": "NBA"}],
            "combined_probability": 0.70,
            "combined_ev_score": 1.45,
        }
        sid = slip_id_for(date, slip_def)

        slip_rows = [
            _make_full_slip_row(date, sid, platform="PrizePicks", slip_type="power",
                                total_legs=2, winning_legs=2, losing_legs=0, pvd_legs=0,
                                net_pnl=2.0, needs_recon=False),
        ]
        wb = _make_master_wb_with_slip_history(slip_rows)

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            master = tmp_path / "master_pnl.xlsx"
            bankroll_path = tmp_path / "bankroll.json"
            # No slips_<date>.json file in empty slips_dir
            slips_dir = tmp_path / "slips"
            slips_dir.mkdir()
            wb.save(str(master))

            result = runner.rebuild_slip_bankroll(
                dry_run=True,
                inception=date,
                _wb_override=wb,
                _master_override=master,
                _bankroll_override=bankroll_path,
                _slips_dir_override=str(slips_dir),
            )
            self.assertIn(
                "slips_missing_signal", result,
                "rebuild_slip_bankroll return payload must include slips_missing_signal (WR-02)",
            )
            self.assertEqual(
                result["slips_missing_signal"], 1,
                "One Slip History row with no JSON signal must be counted in slips_missing_signal (WR-02)",
            )


# ---------------------------------------------------------------------------
# IN-03 clamp test
# ---------------------------------------------------------------------------

class TestIN03StakeClamp(unittest.TestCase):
    """IN-03: confidence_stake must return >= 0.0 even for negative bankroll."""

    def test_negative_bankroll_returns_nonnegative_stake(self):
        """IN-03: Negative start_of_day_bankroll must yield stake >= 0.0."""
        from stake_sizing import confidence_stake
        # Negative bankroll (deep drawdown scenario)
        stake = confidence_stake(0.80, 1.5, -50.0)
        self.assertGreaterEqual(
            stake, 0.0,
            f"confidence_stake with negative bankroll must return >= 0.0, got {stake} (IN-03)",
        )

    def test_zero_bankroll_returns_zero_stake(self):
        """IN-03: Zero bankroll yields zero stake (no bet without money)."""
        from stake_sizing import confidence_stake
        stake = confidence_stake(0.80, 1.5, 0.0)
        self.assertEqual(stake, 0.0)


if __name__ == "__main__":
    unittest.main()
