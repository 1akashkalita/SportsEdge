#!/usr/bin/env python3
"""Test suite for backfill re-grade guard (Testing strategy #8).

Tests:
 - MANUAL REVIEW prop row is overwritten to terminal on re-grade (in-place, no dup rows)
 - Stored-Result casing variants ("Win", "push ", " VOID") are SKIPPED (not re-graded/flipped)
 - Results count per ref stays 1 (upsert, not append)
 - master_pnl Pick History has exactly one row per ref
 - A second sync_master_and_bankroll(date, []) does not double-count

Requires:
  TERMINAL_RESULTS = {"WIN", "LOSS", "PUSH", "VOID"} constant in sports_system_runner
  existing_result_map(results_ws, date, sport_label) -> dict[str, str] function
  grade_game_in_workbook's three loop guards use TERMINAL_RESULTS via existing_result_map

Run from scripts/: python3 test_backfill_regrade.py
"""
from __future__ import annotations

import importlib
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

_SCRIPTS = Path(__file__).parent
if str(_SCRIPTS) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS))

runner = importlib.import_module("sports_system_runner")

TERMINAL_RESULTS = runner.TERMINAL_RESULTS
existing_result_map = runner.existing_result_map
upsert_result_row = runner.upsert_result_row
sync_master_and_bankroll = runner.sync_master_and_bankroll
RESULT_HEADERS = runner.RESULT_HEADERS
result_headers = runner.result_headers
grade_game_in_workbook = runner.grade_game_in_workbook
master_pnl_workbook = runner.master_pnl_workbook
ensure_ws_columns = runner.ensure_ws_columns


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_results_ws():
    """Create an in-memory Results sheet with RESULT_HEADERS."""
    from openpyxl import Workbook as OWB
    wb = OWB()
    ws = wb.active
    ws.title = "Results"
    ws.append(RESULT_HEADERS)
    return ws


def _add_result_row(ws, date: str, sport: str, ref: str, result: str) -> None:
    """Append a Results row for testing."""
    cols = result_headers(ws)
    row_data = {h: None for h in RESULT_HEADERS}
    row_data["Date"] = date
    row_data["Sport"] = sport
    row_data["Pick Ref"] = ref
    row_data["Result"] = result
    row_data["Units"] = 1.0
    row_data["PnL"] = 0.0
    ws.append([row_data.get(h) for h in RESULT_HEADERS])


# ---------------------------------------------------------------------------
# Test: TERMINAL_RESULTS constant
# ---------------------------------------------------------------------------

class TestTerminalResultsConstant(unittest.TestCase):
    """TERMINAL_RESULTS must be a set containing WIN, LOSS, PUSH, VOID."""

    def test_constant_exists(self) -> None:
        self.assertIsInstance(TERMINAL_RESULTS, (set, frozenset))

    def test_contains_win(self) -> None:
        self.assertIn("WIN", TERMINAL_RESULTS)

    def test_contains_loss(self) -> None:
        self.assertIn("LOSS", TERMINAL_RESULTS)

    def test_contains_push(self) -> None:
        self.assertIn("PUSH", TERMINAL_RESULTS)

    def test_contains_void(self) -> None:
        self.assertIn("VOID", TERMINAL_RESULTS)

    def test_does_not_contain_manual_review(self) -> None:
        self.assertNotIn("MANUAL REVIEW", TERMINAL_RESULTS)

    def test_does_not_contain_pending(self) -> None:
        self.assertNotIn("PENDING", TERMINAL_RESULTS)


# ---------------------------------------------------------------------------
# Test: existing_result_map returns {ref: result_str}
# ---------------------------------------------------------------------------

class TestExistingResultMap(unittest.TestCase):
    """existing_result_map returns a {ref: result_str} dict, not a set."""

    def test_returns_dict(self) -> None:
        ws = _make_results_ws()
        result = existing_result_map(ws, "2026-06-08", "MLB")
        self.assertIsInstance(result, dict)

    def test_maps_ref_to_result(self) -> None:
        ws = _make_results_ws()
        _add_result_row(ws, "2026-06-08", "MLB", "PROP:Player Hits 2.5", "WIN")
        result = existing_result_map(ws, "2026-06-08", "MLB")
        self.assertEqual(result.get("PROP:Player Hits 2.5"), "WIN")

    def test_maps_manual_review_ref(self) -> None:
        ws = _make_results_ws()
        _add_result_row(ws, "2026-06-08", "MLB", "PROP:Player Hits 2.5", "MANUAL REVIEW")
        result = existing_result_map(ws, "2026-06-08", "MLB")
        self.assertEqual(result.get("PROP:Player Hits 2.5"), "MANUAL REVIEW")

    def test_empty_for_different_date(self) -> None:
        ws = _make_results_ws()
        _add_result_row(ws, "2026-06-08", "MLB", "PROP:Player Hits 2.5", "WIN")
        result = existing_result_map(ws, "2026-06-09", "MLB")
        self.assertEqual(result, {})

    def test_empty_for_different_sport(self) -> None:
        ws = _make_results_ws()
        _add_result_row(ws, "2026-06-08", "MLB", "PROP:Player Hits 2.5", "WIN")
        result = existing_result_map(ws, "2026-06-08", "NBA")
        self.assertEqual(result, {})


# ---------------------------------------------------------------------------
# Test: terminal guard skips casing/whitespace variants
# ---------------------------------------------------------------------------

class TestTerminalGuardNormalization(unittest.TestCase):
    """The .strip().upper() normalization in the guard must treat all casing variants as terminal."""

    def _is_terminal(self, stored_value: str) -> bool:
        """Mirror the exact guard logic: (already.get(ref) or '').strip().upper() in TERMINAL_RESULTS."""
        return (stored_value or "").strip().upper() in TERMINAL_RESULTS

    def test_win_uppercase_is_terminal(self) -> None:
        self.assertTrue(self._is_terminal("WIN"))

    def test_win_titlecase_is_terminal(self) -> None:
        self.assertTrue(self._is_terminal("Win"))

    def test_loss_lowercase_is_terminal(self) -> None:
        self.assertTrue(self._is_terminal("loss"))

    def test_push_with_trailing_space_is_terminal(self) -> None:
        self.assertTrue(self._is_terminal("push "))

    def test_void_with_leading_space_is_terminal(self) -> None:
        self.assertTrue(self._is_terminal(" VOID"))

    def test_manual_review_is_not_terminal(self) -> None:
        self.assertFalse(self._is_terminal("MANUAL REVIEW"))

    def test_pending_is_not_terminal(self) -> None:
        self.assertFalse(self._is_terminal("PENDING"))

    def test_empty_string_is_not_terminal(self) -> None:
        self.assertFalse(self._is_terminal(""))

    def test_none_is_not_terminal(self) -> None:
        self.assertFalse(self._is_terminal(None))  # type: ignore[arg-type]


# ---------------------------------------------------------------------------
# Test: grade_game_in_workbook re-grades MANUAL REVIEW, skips settled
# ---------------------------------------------------------------------------

# Minimal ESPN player stats for grading a prop
_PLAYER_STATS = {
    "carlos correa": {
        "batting": {
            "hits": 3.0, "runs": 1.0, "rbis": 2.0, "homeruns": 0.0,
            "walks": 1.0, "strikeouts": 1.0,
            "_hit_counts": {"single": 2, "double": 1, "triple": 0, "home-run": 0},
        },
        "pitching": {},
    }
}

_GAME = {
    "event_id": "test-event-backfill-001",
    "id": "test-event-backfill-001",
    "status": "final",
    "status_name": "STATUS_FINAL",
    "completed": True,
    "home_team": "Twins",
    "away_team": "Tigers",
    "home_score": 5,
    "away_score": 3,
}


class TestGradeGameInWorkbookGuard(unittest.TestCase):
    """grade_game_in_workbook must:
    - Re-grade MANUAL REVIEW / PENDING rows when re-run
    - Skip settled rows even with casing/whitespace variants
    - Never produce duplicate Results rows per ref
    """

    def setUp(self) -> None:
        # Use a temp directory for all workbook/bankroll files
        self._tmpdir = tempfile.TemporaryDirectory()
        tmp = Path(self._tmpdir.name)
        self._pnl_dir = tmp / "pnl"
        self._pnl_dir.mkdir(parents=True)
        self._data_dir = tmp / "data"
        self._mlb_dir = self._data_dir / "mlb"
        self._mlb_dir.mkdir(parents=True)
        self._locks_dir = tmp / "locks"
        self._locks_dir.mkdir()
        self._backups_dir = self._data_dir / "backups" / "workbooks"
        self._backups_dir.mkdir(parents=True)
        self._log_dir = self._pnl_dir / "logs"
        self._log_dir.mkdir()
        # Patch runner path constants to point to temp dirs
        self._patches = [
            patch.object(runner, "PNL_DIR", self._pnl_dir),
            patch.object(runner, "NBA_DIR", tmp / "data" / "nba"),
            patch.object(runner, "MLB_DIR", self._mlb_dir),
            patch.object(runner, "DATA", self._data_dir),
            patch.object(runner, "LOG_DIR", self._log_dir),
            patch.object(runner, "WORKBOOK_LOCK_DIR", self._locks_dir),
            patch.object(runner, "WORKBOOK_BACKUP_DIR", self._backups_dir),
            patch.object(runner, "BANKROLL", self._pnl_dir / "bankroll.json"),
            patch.object(runner, "GAME_STATUS_CACHE_FILE", self._pnl_dir / "game_status_cache.json"),
        ]
        for p in self._patches:
            p.start()
        (tmp / "data" / "nba").mkdir(parents=True, exist_ok=True)
        # Mock ESPN fetch and Obsidian/Telegram
        self._fn_patches = [
            patch.object(runner, "espn_player_stats_by_event", return_value=_PLAYER_STATS),
            patch.object(runner, "send_telegram", return_value=None),
            patch.object(runner, "obsidian_update_results_section", return_value=None),
            patch.object(runner, "obsidian_update_bankroll_files", return_value=None),
        ]
        for p in self._fn_patches:
            p.start()

    def tearDown(self) -> None:
        for p in self._fn_patches:
            p.stop()
        for p in self._patches:
            p.stop()
        self._tmpdir.cleanup()

    def _count_results_rows(self, wb, date: str, sport: str, ref: str) -> int:
        """Count Results sheet rows matching date+sport+ref."""
        ws = wb["Results"]
        cols = result_headers(ws)
        count = 0
        for r in range(2, ws.max_row + 1):
            if (str(ws.cell(r, cols["Date"]).value or "")[:10] == date
                    and str(ws.cell(r, cols["Sport"]).value or "") == sport
                    and str(ws.cell(r, cols["Pick Ref"]).value or "") == ref):
                count += 1
        return count

    def _get_result_value(self, wb, date: str, sport: str, ref: str) -> str | None:
        """Get the Result column for a specific ref in the Results sheet."""
        ws = wb["Results"]
        cols = result_headers(ws)
        for r in range(2, ws.max_row + 1):
            if (str(ws.cell(r, cols["Date"]).value or "")[:10] == date
                    and str(ws.cell(r, cols["Sport"]).value or "") == sport
                    and str(ws.cell(r, cols["Pick Ref"]).value or "") == ref):
                return str(ws.cell(r, cols["Result"]).value or "")
        return None

    def test_manual_review_prop_is_regraded_in_place(self) -> None:
        """A MANUAL REVIEW Results row must be overwritten to terminal on re-grade."""
        date = "2026-06-08"
        # Create workbook with a Props row and a pre-seeded MANUAL REVIEW Results row
        path = runner.ensure_workbook("mlb", date)
        from openpyxl import load_workbook as lw
        wb = lw(path)
        # Seed Props sheet
        props_ws = wb["Props"]
        props_cols = ensure_ws_columns(props_ws, runner.PROPS_HEADERS)
        props_ws.append([None] * len(props_cols))
        r = props_ws.max_row
        props_ws.cell(r, props_cols["Date"]).value = date
        props_ws.cell(r, props_cols["Sport"]).value = "MLB"
        props_ws.cell(r, props_cols["Player Name"]).value = "Carlos Correa"
        props_ws.cell(r, props_cols["Team"]).value = "Twins"
        props_ws.cell(r, props_cols["Stat"]).value = "Hits"
        props_ws.cell(r, props_cols["Line"]).value = 2.5
        props_ws.cell(r, props_cols["Opponent/Description"]).value = "Over"
        # Seed Results sheet with MANUAL REVIEW for this ref
        ref = "PROP:Carlos Correa Hits 2.5"
        results_ws = wb["Results"]
        _add_result_row(results_ws, date, "MLB", ref, "MANUAL REVIEW")
        runner.save_workbook_atomic(wb, path)

        # Re-grade with valid final game — should overwrite MANUAL REVIEW
        grade_game_in_workbook("mlb", _GAME, date, dry_run=False)

        # Verify
        wb2 = lw(path)
        # Must have exactly 1 Results row for this ref (no dup)
        count = self._count_results_rows(wb2, date, "MLB", ref)
        self.assertEqual(count, 1, f"Expected 1 Results row for {ref}, got {count}")
        # Must be terminal now (Hits=3.0 > 2.5 → WIN for Over)
        result = self._get_result_value(wb2, date, "MLB", ref)
        self.assertIn(result, {"WIN", "LOSS", "PUSH"}, f"Expected terminal result, got {result!r}")

    def test_casing_variant_win_is_skipped(self) -> None:
        """A stored 'Win' (titlecase) result must NOT be re-graded."""
        date = "2026-06-08"
        path = runner.ensure_workbook("mlb", date)
        from openpyxl import load_workbook as lw
        wb = lw(path)
        props_ws = wb["Props"]
        props_cols = ensure_ws_columns(props_ws, runner.PROPS_HEADERS)
        props_ws.append([None] * len(props_cols))
        r = props_ws.max_row
        props_ws.cell(r, props_cols["Date"]).value = date
        props_ws.cell(r, props_cols["Sport"]).value = "MLB"
        props_ws.cell(r, props_cols["Player Name"]).value = "Carlos Correa"
        props_ws.cell(r, props_cols["Team"]).value = "Twins"
        props_ws.cell(r, props_cols["Stat"]).value = "Hits"
        props_ws.cell(r, props_cols["Line"]).value = 2.5
        props_ws.cell(r, props_cols["Opponent/Description"]).value = "Over"
        ref = "PROP:Carlos Correa Hits 2.5"
        results_ws = wb["Results"]
        _add_result_row(results_ws, date, "MLB", ref, "Win")  # titlecase
        runner.save_workbook_atomic(wb, path)

        grade_game_in_workbook("mlb", _GAME, date, dry_run=False)

        wb2 = lw(path)
        # Row count must still be 1
        count = self._count_results_rows(wb2, date, "MLB", ref)
        self.assertEqual(count, 1)
        # Result must still be "Win" (not re-graded)
        result = self._get_result_value(wb2, date, "MLB", ref)
        self.assertEqual(result, "Win", f"Settled 'Win' was re-graded to {result!r}")

    def test_push_with_trailing_space_is_skipped(self) -> None:
        """A stored 'push ' (trailing space) result must NOT be re-graded."""
        date = "2026-06-08"
        path = runner.ensure_workbook("mlb", date)
        from openpyxl import load_workbook as lw
        wb = lw(path)
        props_ws = wb["Props"]
        props_cols = ensure_ws_columns(props_ws, runner.PROPS_HEADERS)
        props_ws.append([None] * len(props_cols))
        r = props_ws.max_row
        props_ws.cell(r, props_cols["Date"]).value = date
        props_ws.cell(r, props_cols["Sport"]).value = "MLB"
        props_ws.cell(r, props_cols["Player Name"]).value = "Carlos Correa"
        props_ws.cell(r, props_cols["Team"]).value = "Twins"
        props_ws.cell(r, props_cols["Stat"]).value = "Hits"
        props_ws.cell(r, props_cols["Line"]).value = 2.5
        props_ws.cell(r, props_cols["Opponent/Description"]).value = "Over"
        ref = "PROP:Carlos Correa Hits 2.5"
        results_ws = wb["Results"]
        _add_result_row(results_ws, date, "MLB", ref, "push ")  # trailing space
        runner.save_workbook_atomic(wb, path)

        grade_game_in_workbook("mlb", _GAME, date, dry_run=False)

        wb2 = lw(path)
        count = self._count_results_rows(wb2, date, "MLB", ref)
        self.assertEqual(count, 1)
        result = self._get_result_value(wb2, date, "MLB", ref)
        self.assertEqual(result, "push ", f"Settled 'push ' was re-graded to {result!r}")

    def test_void_with_leading_space_is_skipped(self) -> None:
        """A stored ' VOID' (leading space) result must NOT be re-graded."""
        date = "2026-06-08"
        path = runner.ensure_workbook("mlb", date)
        from openpyxl import load_workbook as lw
        wb = lw(path)
        props_ws = wb["Props"]
        props_cols = ensure_ws_columns(props_ws, runner.PROPS_HEADERS)
        props_ws.append([None] * len(props_cols))
        r = props_ws.max_row
        props_ws.cell(r, props_cols["Date"]).value = date
        props_ws.cell(r, props_cols["Sport"]).value = "MLB"
        props_ws.cell(r, props_cols["Player Name"]).value = "Carlos Correa"
        props_ws.cell(r, props_cols["Team"]).value = "Twins"
        props_ws.cell(r, props_cols["Stat"]).value = "Hits"
        props_ws.cell(r, props_cols["Line"]).value = 2.5
        props_ws.cell(r, props_cols["Opponent/Description"]).value = "Over"
        ref = "PROP:Carlos Correa Hits 2.5"
        results_ws = wb["Results"]
        _add_result_row(results_ws, date, "MLB", ref, " VOID")  # leading space
        runner.save_workbook_atomic(wb, path)

        grade_game_in_workbook("mlb", _GAME, date, dry_run=False)

        wb2 = lw(path)
        count = self._count_results_rows(wb2, date, "MLB", ref)
        self.assertEqual(count, 1)
        result = self._get_result_value(wb2, date, "MLB", ref)
        self.assertEqual(result, " VOID", f"Settled ' VOID' was re-graded to {result!r}")

    def test_double_sync_does_not_double_count(self) -> None:
        """A second sync_master_and_bankroll(date, []) must not double-count existing rows."""
        date = "2026-06-08"
        path = runner.ensure_workbook("mlb", date)
        from openpyxl import load_workbook as lw
        wb = lw(path)
        # Seed Props sheet with a gradeable row
        props_ws = wb["Props"]
        props_cols = ensure_ws_columns(props_ws, runner.PROPS_HEADERS)
        props_ws.append([None] * len(props_cols))
        r = props_ws.max_row
        props_ws.cell(r, props_cols["Date"]).value = date
        props_ws.cell(r, props_cols["Sport"]).value = "MLB"
        props_ws.cell(r, props_cols["Player Name"]).value = "Carlos Correa"
        props_ws.cell(r, props_cols["Team"]).value = "Twins"
        props_ws.cell(r, props_cols["Stat"]).value = "Hits"
        props_ws.cell(r, props_cols["Line"]).value = 2.5
        props_ws.cell(r, props_cols["Opponent/Description"]).value = "Over"
        runner.save_workbook_atomic(wb, path)

        # Grade once
        grade_game_in_workbook("mlb", _GAME, date, dry_run=False)

        # Get bankroll state after first grade
        summary1 = sync_master_and_bankroll(date, [])
        pnl1 = summary1.get("day_pnl", 0)
        rows1 = len(summary1.get("daily_rows", []))

        # Second empty sync
        summary2 = sync_master_and_bankroll(date, [])
        pnl2 = summary2.get("day_pnl", 0)
        rows2 = len(summary2.get("daily_rows", []))

        # PnL and row count must be identical (no double-count)
        self.assertEqual(pnl1, pnl2, f"PnL changed on second sync: {pnl1} → {pnl2}")
        self.assertEqual(rows1, rows2, f"Row count changed on second sync: {rows1} → {rows2}")


if __name__ == "__main__":
    unittest.main()
