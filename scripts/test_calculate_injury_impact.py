#!/usr/bin/env python3
"""Integration tests for calculate_injury_impact.py.

These tests copy the real NBA workbook to a temp directory and never modify the
source workbook.
"""
from __future__ import annotations

import hashlib
import json
import math
import shutil
import subprocess
import tempfile
import unittest
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "scripts" / "calculate_injury_impact.py"
TODAY = datetime.now().strftime("%Y-%m-%d")
NBA_DIR = ROOT / "data" / "nba"


def _schema_workbook() -> Path | None:
    """A real NBA workbook to source the live schema from — today's if the cron has
    already produced it, else the most recent available. The test clears the props and
    seeds its own, so only the SCHEMA (sheets/columns) matters. This removes a flaky
    time-of-day dependence on the daily-picks cron having run yet (NBA is offseason in
    summer, so today's workbook may not exist at early-morning CI time)."""
    today = NBA_DIR / f"nba_{TODAY}.xlsx"
    if today.exists():
        return today
    # Dated workbooks only (nba_YYYY-...) — exclude specials like nba_finals_tracker.xlsx,
    # which do not carry the daily Player Props schema. Date-prefixed names sort latest-last.
    candidates = sorted(NBA_DIR.glob("nba_[0-9]*.xlsx"))
    return candidates[-1] if candidates else None


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def header_map(ws) -> dict[str, int]:
    return {str(ws.cell(1, c).value or "").strip(): c for c in range(1, ws.max_column + 1)}


def append_fake_prop(ws, player: str, stat: str, projection: float, line: float, confidence: str = "C") -> int:
    headers = header_map(ws)
    row = ws.max_row + 1
    values = {
        "Date": TODAY,
        "Sport": "NBA",
        "Projection ID": f"TEST-{player}-{stat}",
        "Player Name": player,
        "Team": "LAL",
        "Opponent/Description": "TEST",
        "Stat": stat,
        "Line": line,
        "Odds Type": "standard",
        "Start Time UTC": f"{TODAY} 23:59:00 UTC",
        "Status": "ACTIVE",
        "Reasoning": "integration test seed",
        "Confidence": confidence,
        "Units": 1,
        "Platform": "PrizePicks",
        "Model Projection": projection,
        "Edge": round(projection - line, 3),
        "Model Over Probability": 0.5,
        "EV": -0.0455,
        "Edge Type Tags": "Value",
        "Opening Line": line,
        "Demon Available": False,
        "Goblin Available": False,
        "Injury Flag": "",
        "Correlation Group": "TEST",
        "Slip ID": "TEST",
    }
    for key, value in values.items():
        if key in headers:
            ws.cell(row, headers[key]).value = value
    return row


def normal_cdf(x: float) -> float:
    return 0.5 * (1 + math.erf(x / math.sqrt(2)))


def expected_over_probability(projection: float, line: float, stat: str) -> float:
    stat_l = stat.lower()
    if "assist" in stat_l:
        sigma = max(1.2, abs(line) * 0.22)
    elif "rebound" in stat_l:
        sigma = max(1.5, abs(line) * 0.20)
    elif "fantasy" in stat_l:
        sigma = max(5.0, abs(line) * 0.16)
    else:
        sigma = max(2.5, abs(line) * 0.18)
    return round(max(0.05, min(0.95, normal_cdf((projection - line) / sigma))), 4)


def expected_ev(prob: float) -> float:
    return round(prob * 0.909 - (1 - prob), 4)


class CalculateInjuryImpactIntegrationTest(unittest.TestCase):
    def test_anthony_davis_out_adjusts_matching_temp_workbook_rows(self) -> None:
        source_workbook = _schema_workbook()
        if source_workbook is None:
            self.skipTest("no NBA workbook available to source the schema from")
        before_hash = sha256(source_workbook)
        with tempfile.TemporaryDirectory(prefix="injury_impact_test_") as td:
            temp_workbook = Path(td) / "nba_injury_impact_test.xlsx"
            shutil.copy2(source_workbook, temp_workbook)
            wb = load_workbook(temp_workbook)
            ws = wb["Player Props"]
            # Clear any existing prop rows so the count of adjusted rows is deterministic
            # regardless of what live props the source workbook happened to contain
            # (e.g. real same-day LAL props would otherwise inflate rows_updated).
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)
            lebron_row = append_fake_prop(ws, "LeBron James", "points", 25.0, 27.5, "C")
            reaves_row = append_fake_prop(ws, "Austin Reaves", "assists", 6.0, 6.5, "C")
            wb.save(temp_workbook)
            wb.close()

            proc = subprocess.run(
                [
                    "python3",
                    str(SCRIPT),
                    "--player",
                    "Anthony Davis",
                    "--team",
                    "LAL",
                    "--status",
                    "OUT",
                    "--date",
                    TODAY,
                    "--workbook",
                    str(temp_workbook),
                ],
                cwd=ROOT,
                capture_output=True,
                text=True,
                timeout=120,
            )
            self.assertEqual(proc.returncode, 0, proc.stderr + proc.stdout)
            payload = json.loads(proc.stdout[proc.stdout.find("{"):])
            self.assertEqual(payload["applied"]["rows_updated"], 2)

            wb = load_workbook(temp_workbook, data_only=True)
            ws = wb["Player Props"]
            headers = header_map(ws)
            checks = [
                {
                    "row": lebron_row,
                    "player": "LeBron James",
                    "stat": "points",
                    "old_projection": 25.0,
                    "line": 27.5,
                    "expected_projection": 27.0,
                    "expected_direction": "positive",
                    "expected_final_tier": "SKIP",
                    "expected_playable": "NO",
                    "reason_contains": ["edge -0.500 <= 0", "EV -0.1222 <= 0", "below breakeven"],
                },
                {
                    "row": reaves_row,
                    "player": "Austin Reaves",
                    "stat": "assists",
                    "old_projection": 6.0,
                    "line": 6.5,
                    "expected_projection": 6.72,
                    "expected_direction": "positive",
                    "expected_final_tier": "B",
                    "expected_playable": "YES",
                    "reason_contains": ["adjusted metrics support playable tier"],
                },
            ]
            for check in checks:
                row = check["row"]
                player = check["player"]
                stat = check["stat"]
                old_projection = check["old_projection"]
                line = check["line"]
                expected_projection = check["expected_projection"]
                self.assertEqual(ws.cell(row, headers["Player Name"]).value, player)
                self.assertAlmostEqual(float(ws.cell(row, headers["Model Projection"]).value), expected_projection, places=3)
                expected_edge = round(expected_projection - line, 3)
                self.assertAlmostEqual(float(ws.cell(row, headers["Edge"]).value), expected_edge, places=3)
                prob = expected_over_probability(expected_projection, line, stat)
                ev = expected_ev(prob)
                self.assertAlmostEqual(float(ws.cell(row, headers["Model Over Probability"]).value), prob, places=4)
                self.assertAlmostEqual(float(ws.cell(row, headers["EV"]).value), ev, places=4)
                self.assertEqual(ws.cell(row, headers["Pre Injury Tier"]).value, "C")
                self.assertEqual(ws.cell(row, headers["Injury Direction"]).value, check["expected_direction"])
                self.assertEqual(ws.cell(row, headers["Injury Suggested Tier Delta"]).value, "+1 tier")
                self.assertEqual(ws.cell(row, headers["Final Tier"]).value, check["expected_final_tier"])
                self.assertEqual(ws.cell(row, headers["Confidence"]).value, check["expected_final_tier"])
                self.assertEqual(ws.cell(row, headers["Injury Adjusted Playable"]).value, check["expected_playable"])
                adjustment_reason = str(ws.cell(row, headers["Injury Adjustment Reason"]).value)
                for expected_text in check["reason_contains"]:
                    self.assertIn(expected_text, adjustment_reason)
                if ev <= 0 or prob < 1 / (1 + 0.909) or expected_edge <= 0:
                    self.assertNotIn(ws.cell(row, headers["Final Tier"]).value, {"A", "B"})
                    self.assertEqual(ws.cell(row, headers["Injury Adjusted Playable"]).value, "NO")
                self.assertEqual(ws.cell(row, headers["Injury Adjusted"]).value, "YES")
                self.assertIn("Anthony Davis OUT", str(ws.cell(row, headers["Injury Flag"]).value))
                reasoning = str(ws.cell(row, headers["Reasoning"]).value)
                self.assertIn(f"projection {old_projection} -> {expected_projection}", reasoning)
                self.assertIn(f"final {check['expected_final_tier']}", reasoning)
            wb.close()

        self.assertEqual(sha256(source_workbook), before_hash, "Source workbook changed during integration test")


if __name__ == "__main__":
    unittest.main(verbosity=2)
