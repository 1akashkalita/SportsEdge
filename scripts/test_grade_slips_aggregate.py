#!/usr/bin/env python3
"""Offline unittest for grade_slips aggregate layer (Wave 2).

Tests:
  (a) 2-leg power, both WIN → GRADED, 3.0x, +2u, reconcile=False
  (b) 2-leg power, 1 WIN + 1 LOSS → GRADED, 0x, -1u, reconcile=False
  (c) slip with absent player leg (abstain) → MANUAL REVIEW, net=None,
      reconcile=True — explicitly NOT WIN / NOT LOSS (money-safety)
  (d) idempotency: write_slip_history_rows called twice for same Slip ID
      → data rows unchanged (no duplicate append)
  (e) separation: slip row lands in 'Slip History' sheet only; 'Results'
      row count is untouched (SLIPS-04)

All tests are fully offline (inject box scores, openpyxl in-memory workbook).
No network calls.

Run from scripts/:
    python3 test_grade_slips_aggregate.py
"""
from __future__ import annotations

import sys
import unittest
from pathlib import Path

# Ensure scripts/ is on sys.path for sibling imports.
_SCRIPTS = Path(__file__).parent
if str(_SCRIPTS) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS))

import openpyxl
from grade_slips import (
    LEG_PENDING,
    grade_slip,
    slip_id_for,
    write_slip_history_rows,
)
from slip_payouts import (
    SLIP_HISTORY_HEADERS,
    ensure_slip_history_sheet,
    load_payout_config,
)


# ---------------------------------------------------------------------------
# Offline fixtures
# ---------------------------------------------------------------------------

# NBA flat row: LeBron James — 30 pts / 10 reb / 5 ast
_NBA_LEBRON: dict = {
    "points": 30.0,
    "rebounds": 10.0,
    "assists": 5.0,
    "steals": 2.0,
    "blocks": 1.0,
    "turnovers": 2.0,
}

# MLB batting row: Freddie Freeman — 3 hits / 2 runs / 1 RBI
_MLB_BATTING_FREEMAN: dict = {
    "hits": 3.0,
    "runs": 2.0,
    "rbis": 1.0,
    "homeruns": 0.0,
    "walks": 1.0,
    "strikeouts": 1.0,
    "atbats": 4.0,
    "_hit_counts": {"single": 3, "double": 0, "triple": 0, "home-run": 0},
}

BOX_SCORES = {
    "NBA": {
        "lebron james": _NBA_LEBRON,
    },
    "MLB": {
        "freddie freeman": {"batting": _MLB_BATTING_FREEMAN, "pitching": {}},
    },
}

_DATE = "2026-06-22"

_CONFIG = load_payout_config()


def _make_slip(
    slip_type: str,
    legs: list[dict],
    category: str = "safest_2_leg",
) -> dict:
    """Build a minimal slip dict for testing."""
    return {
        "date": _DATE,
        "category": category,
        "platform": "PrizePicks",
        "slip_type": slip_type,
        "stake_units": 1.0,
        "leg_count": len(legs),
        "legs": legs,
    }


def _nba_leg(stat: str, line: float, side: str = "OVER", player: str = "LeBron James") -> dict:
    return {
        "player_name": player,
        "stat_type": stat,
        "line": line,
        "side": side,
        "sport": "NBA",
        "prop_id": f"NBA:{player}:{stat}:{line}",
    }


def _mlb_leg(stat: str, line: float, side: str = "OVER", player: str = "Freddie Freeman") -> dict:
    return {
        "player_name": player,
        "stat_type": stat,
        "line": line,
        "side": side,
        "sport": "MLB",
        "prop_id": f"MLB:{player}:{stat}:{line}",
    }


# ---------------------------------------------------------------------------
# Test cases
# ---------------------------------------------------------------------------

class TestGradeSlipPower(unittest.TestCase):
    """(a) + (b): power slip payout paths."""

    def test_power_both_win(self) -> None:
        """2-leg power all-WIN → GRADED, 3.0x, net +2.0, reconcile=False."""
        slip = _make_slip(
            "power",
            [
                _nba_leg("points", 25.5),   # 30 pts > 25.5 → WIN
                _nba_leg("rebounds", 8.5),  # 10 reb > 8.5  → WIN
            ],
        )
        result = grade_slip(slip, BOX_SCORES, config=_CONFIG)
        p = result["payout"]

        self.assertEqual(p["slip_result"], "GRADED", "slip_result must be GRADED")
        self.assertEqual(p["winning_legs"], 2)
        self.assertEqual(p["standard_payout_multiplier"], 3.0,
                         "PrizePicks power 2-leg = 3.0x")
        self.assertAlmostEqual(p["net_pnl"], 2.0, places=6,
                               msg="1u stake × 3x - 1u = +2u net")
        self.assertAlmostEqual(p["gross_return"], 3.0, places=6)
        self.assertFalse(p["needs_payout_reconciliation"],
                         "All legs resolved → no reconciliation needed")

    def test_power_one_loss(self) -> None:
        """2-leg power with one LOSS → gross 0, net -1u, reconcile=False."""
        slip = _make_slip(
            "power",
            [
                _nba_leg("points", 25.5),   # 30 pts > 25.5 → WIN
                _nba_leg("assists", 7.5),   # 5 ast < 7.5  → LOSS
            ],
        )
        result = grade_slip(slip, BOX_SCORES, config=_CONFIG)
        p = result["payout"]

        self.assertEqual(p["slip_result"], "GRADED", "power loss must be GRADED (resolved)")
        self.assertEqual(p["winning_legs"], 1)
        self.assertAlmostEqual(p["gross_return"], 0.0, places=6,
                               msg="power with any LOSS → 0 gross")
        self.assertAlmostEqual(p["net_pnl"], -1.0, places=6,
                               msg="lose the stake: 0 - 1u = -1u")
        self.assertFalse(p["needs_payout_reconciliation"],
                         "All legs resolved → no reconciliation needed")


class TestGradeSlipPending(unittest.TestCase):
    """(c): money-safety — abstain leg must never produce WIN or LOSS."""

    def test_pending_leg_is_manual_review(self) -> None:
        """Slip with absent player leg → MANUAL REVIEW, net=None, reconcile=True."""
        slip = _make_slip(
            "power",
            [
                _nba_leg("points", 25.5),                   # 30 pts → WIN
                _nba_leg("points", 10.0, player="Ghost Player"),  # absent → PENDING
            ],
        )
        result = grade_slip(slip, BOX_SCORES, config=_CONFIG)
        p = result["payout"]

        self.assertEqual(p["slip_result"], "MANUAL REVIEW",
                         "Any abstain leg must yield MANUAL REVIEW")
        self.assertTrue(p["needs_payout_reconciliation"],
                        "needs_payout_reconciliation must be True on abstain leg")
        self.assertIsNone(p["net_pnl"],
                          "net_pnl must be None when reconciliation needed (never fabricated)")
        self.assertIsNone(p["gross_return"],
                          "gross_return must be None when reconciliation needed")
        # Money-safety: explicitly assert NOT fabricated WIN or LOSS.
        self.assertNotEqual(p["slip_result"], "WIN",
                            "MONEY-SAFETY: abstain-leg slip must NOT be WIN")
        self.assertNotEqual(p["slip_result"], "LOSS",
                            "MONEY-SAFETY: abstain-leg slip must NOT be LOSS")

    def test_leg_grades_include_pending_token(self) -> None:
        """grade_slip leg_grades should contain LEG_PENDING for the absent leg."""
        slip = _make_slip(
            "power",
            [
                _nba_leg("points", 25.5),
                _nba_leg("points", 10.0, player="Ghost Player"),
            ],
        )
        result = grade_slip(slip, BOX_SCORES, config=_CONFIG)
        leg_results = [g["result"] for g in result["leg_grades"]]
        self.assertIn(LEG_PENDING, leg_results,
                      "leg_grades must include LEG_PENDING for unresolved leg")


class TestFlexPayout(unittest.TestCase):
    """3-leg flex, 2-of-3 WIN → 1.0x / net 0.0."""

    def test_flex_2_of_3(self) -> None:
        slip = _make_slip(
            "flex",
            [
                _nba_leg("points", 25.5),   # WIN
                _nba_leg("rebounds", 8.5),  # WIN
                _nba_leg("assists", 7.5),   # LOSS (5 ast < 7.5)
            ],
            category="highest_ev",
        )
        result = grade_slip(slip, BOX_SCORES, config=_CONFIG)
        p = result["payout"]

        self.assertEqual(p["slip_result"], "GRADED")
        self.assertEqual(p["winning_legs"], 2)
        self.assertAlmostEqual(p["standard_payout_multiplier"], 1.0, places=6,
                               msg="PrizePicks flex 3-leg 2-win = 1.0x")
        self.assertAlmostEqual(p["gross_return"], 1.0, places=6)
        self.assertAlmostEqual(p["net_pnl"], 0.0, places=6)
        self.assertFalse(p["needs_payout_reconciliation"])


class TestSlipIdFor(unittest.TestCase):
    """Determinism and distinctness of slip_id_for."""

    def test_stable_across_reruns(self) -> None:
        slip = _make_slip("power", [_nba_leg("points", 25.5)])
        id1 = slip_id_for(_DATE, slip)
        id2 = slip_id_for(_DATE, slip)
        self.assertEqual(id1, id2, "Same slip + date must always yield the same ID")

    def test_distinct_for_different_legs(self) -> None:
        slip_a = _make_slip("power", [_nba_leg("points", 25.5)])
        slip_b = _make_slip("power", [_nba_leg("rebounds", 8.5)])
        self.assertNotEqual(
            slip_id_for(_DATE, slip_a),
            slip_id_for(_DATE, slip_b),
            "Different legs must yield different IDs",
        )

    def test_distinct_for_different_dates(self) -> None:
        slip = _make_slip("power", [_nba_leg("points", 25.5)])
        self.assertNotEqual(
            slip_id_for("2026-06-20", slip),
            slip_id_for("2026-06-22", slip),
            "Same legs on different dates must yield different IDs",
        )

    def test_format_contains_date_and_category(self) -> None:
        slip = _make_slip("power", [_nba_leg("points", 25.5)])
        sid = slip_id_for(_DATE, slip)
        self.assertTrue(sid.startswith(f"{_DATE}:"), "ID must start with date")
        self.assertIn("safest_2_leg", sid, "ID must contain category")


class TestWriteSlipHistoryIdempotency(unittest.TestCase):
    """(d): idempotent upsert — second write for same Slip ID adds no row."""

    def _make_graded(self) -> dict:
        slip = _make_slip("power", [_nba_leg("points", 25.5), _nba_leg("rebounds", 8.5)])
        g = grade_slip(slip, BOX_SCORES, config=_CONFIG)
        g["slip_id"] = slip_id_for(_DATE, slip)
        return g

    def test_second_write_no_duplicate(self) -> None:
        wb = openpyxl.Workbook()
        ws = ensure_slip_history_sheet(wb)

        graded = [self._make_graded()]

        # First write — appends one row.
        written1 = write_slip_history_rows(ws, _DATE, graded)
        data_rows_after_1 = ws.max_row - 1  # subtract header row

        # Second write — must overwrite, not append.
        written2 = write_slip_history_rows(ws, _DATE, graded)
        data_rows_after_2 = ws.max_row - 1

        self.assertEqual(written1, 1, "First write should report 1 row written")
        self.assertEqual(written2, 1, "Second write should also report 1 row (upserted)")
        self.assertEqual(data_rows_after_1, 1,
                         "After first write: 1 data row expected")
        self.assertEqual(data_rows_after_2, 1,
                         "After second write: still 1 data row (no duplicate)")

        # Confirm Slip ID appears exactly once.
        slip_id_col = SLIP_HISTORY_HEADERS.index("Slip ID") + 1
        found_ids = [
            ws.cell(r, slip_id_col).value
            for r in range(2, ws.max_row + 1)
        ]
        target_id = graded[0]["slip_id"]
        self.assertEqual(
            found_ids.count(target_id),
            1,
            f"Slip ID {target_id!r} must appear exactly once",
        )


class TestSlipHistorySeparation(unittest.TestCase):
    """(e): SLIPS-04 — slip rows land only in 'Slip History'; Results untouched."""

    def test_slip_rows_not_in_results(self) -> None:
        wb = openpyxl.Workbook()
        # Create a "Slip History" sheet.
        ws_slip = ensure_slip_history_sheet(wb)
        # Create a "Results" sheet (simulate existing prop rows).
        ws_results = wb.create_sheet("Results")
        ws_results.append(["Date", "Player", "Result"])
        ws_results.append([_DATE, "LeBron James", "WIN"])
        results_rows_before = ws_results.max_row

        slip = _make_slip("power", [_nba_leg("points", 25.5), _nba_leg("rebounds", 8.5)])
        graded = [grade_slip(slip, BOX_SCORES, config=_CONFIG)]
        graded[0]["slip_id"] = slip_id_for(_DATE, slip)

        write_slip_history_rows(ws_slip, _DATE, graded)

        # Results sheet must be untouched.
        self.assertEqual(
            ws_results.max_row,
            results_rows_before,
            "write_slip_history_rows must not touch the Results sheet",
        )
        # Slip History sheet must have a data row.
        self.assertGreater(
            ws_slip.max_row,
            1,
            "Slip History sheet must have at least one data row after write",
        )
        # Slip ID must appear in Slip History, NOT in Results.
        slip_id_col = SLIP_HISTORY_HEADERS.index("Slip ID") + 1
        slip_ids_in_slip_history = [
            ws_slip.cell(r, slip_id_col).value for r in range(2, ws_slip.max_row + 1)
        ]
        self.assertIn(
            graded[0]["slip_id"],
            slip_ids_in_slip_history,
            "Slip ID must appear in Slip History",
        )
        # Confirm Results columns don't have a Slip ID column with our value.
        for col in range(1, ws_results.max_column + 1):
            for row in range(1, ws_results.max_row + 1):
                self.assertNotEqual(
                    ws_results.cell(row, col).value,
                    graded[0]["slip_id"],
                    "Slip ID must NOT appear in Results sheet",
                )


class TestStatNormalization(unittest.TestCase):
    """Combo stat types from DFS payload must resolve via normalization."""

    def test_hits_runs_rbis_resolves(self) -> None:
        """'hits runs rbis' (space-sep) must resolve correctly for MLB batter."""
        slip = _make_slip(
            "power",
            [_mlb_leg("hits runs rbis", 5.0)],  # Freeman: 3+2+1=6 > 5 → WIN
        )
        result = grade_slip(slip, BOX_SCORES, config=_CONFIG)
        leg = result["leg_grades"][0]
        self.assertEqual(leg["result"], "WIN",
                         "'hits runs rbis' must resolve to WIN for Freddie Freeman 6 vs 5.0 OVER")
        self.assertAlmostEqual(leg["actual"], 6.0, places=3,
                               msg="hits(3)+runs(2)+rbis(1)=6.0")

    def test_hits_runs_rbis_loss(self) -> None:
        """'hits runs rbis' OVER a high line → LOSS."""
        slip = _make_slip(
            "power",
            [_mlb_leg("hits runs rbis", 7.0)],  # 6 < 7 → LOSS
        )
        result = grade_slip(slip, BOX_SCORES, config=_CONFIG)
        leg = result["leg_grades"][0]
        self.assertEqual(leg["result"], "LOSS",
                         "'hits runs rbis' must resolve to LOSS for 6 vs 7.0 OVER")


if __name__ == "__main__":
    unittest.main()
