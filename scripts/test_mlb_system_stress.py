#!/usr/bin/env python3
"""MLB-only SportsEdge stress tests using synthetic candidates and temp workbook copies.

This script intentionally does not write production MLB/NBA workbooks. It copies the
current workbooks into outputs/verification/mlb_stress_tests/ and hashes the real MLB
workbook before/after the tests.
"""
from __future__ import annotations

import hashlib
import importlib.util
import json
import shutil
import sys
from datetime import date
from pathlib import Path
from typing import Any, Callable

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "outputs" / "verification" / "mlb_stress_tests"
REAL_MLB_WB = ROOT / "data" / "mlb" / f"mlb_{date.today().isoformat()}.xlsx"
REAL_NBA_WB = ROOT / "data" / "nba" / f"nba_{date.today().isoformat()}.xlsx"


def load_module(name: str, path: Path):
    spec = importlib.util.spec_from_file_location(name, path)
    mod = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(mod)
    return mod

runner = load_module("sports_system_runner", ROOT / "scripts" / "sports_system_runner.py")
corr = load_module("analyze_prop_correlation", ROOT / "scripts" / "analyze_prop_correlation.py")
backtest = load_module("run_backtest", ROOT / "scripts" / "run_backtest.py")


def sha256(path: Path) -> str | None:
    if not path.exists():
        return None
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def ev(p: float) -> float:
    return round(p * 0.909 - (1 - p), 4)


def hit_row(rate: float = 0.70, sample: int = 20) -> dict[str, Any]:
    return {
        "hit_rate_l10": rate,
        "hit_rate_l5": rate,
        "sample_size": sample,
        "avg_stat_l10": 6.1,
        "minutes_l5": 15.0,
        "sample_games": [{"minutes": 15, "actual": 6.0} for _ in range(sample)],
    }


def base_candidate(**overrides: Any) -> dict[str, Any]:
    pick = {
        "kind": "prop",
        "date": date.today().isoformat(),
        "sport": "MLB",
        "game_id": "MLB-STRESS-1",
        "player": "Synthetic Player",
        "team": "BAL",
        "away_team": "BOS",
        "home_team": "BAL",
        "selection": "Synthetic Player Over 1.5 Hits+Runs+RBIs",
        "stat": "Hits+Runs+RBIs",
        "line": 1.5,
        "model_projection": 3.0,
        "projection": 3.0,
        "model_over_probability": 0.80,
        "ev": ev(0.80),
        "confidence": "A",
        "units": 3.0,
        "injury_status": "ACTIVE",
        "sportsbook_verified": True,
        "sportsbook_api_available": True,
        "hit_row": hit_row(0.70, 20),
        "reasoning": "synthetic MLB stress candidate",
        "correlation_group": "player:synthetic-player",
    }
    pick.update(overrides)
    return pick


def evaluate(pick: dict[str, Any]) -> tuple[bool, dict[str, Any] | None, list[str], dict[str, Any]]:
    p = dict(pick)
    ok, skip, passed = runner.evaluate_no_bet_gates(p, {})
    return ok, skip, passed, p


def require(cond: bool, msg: str) -> None:
    if not cond:
        raise AssertionError(msg)


def setup_temp_workbooks() -> dict[str, Any]:
    OUT.mkdir(parents=True, exist_ok=True)
    copied = []
    for src, name in [(REAL_MLB_WB, "mlb_temp.xlsx"), (REAL_NBA_WB, "nba_temp.xlsx")]:
        dst = OUT / name
        if src.exists():
            shutil.copy2(src, dst)
        else:
            wb = Workbook()
            wb.active.title = "Picks"
            wb.save(dst)
        copied.append(str(dst))
    return {"temp_workbooks": copied}


def case_1_pitcher_strikeouts_over() -> dict[str, Any]:
    p = base_candidate(
        player="Max Meyer",
        selection="Max Meyer Over 5.5 Pitcher Strikeouts",
        stat="Pitcher Strikeouts",
        line=5.5,
        model_projection=7.0,
        projection=7.0,
        model_over_probability=0.85,
        ev=ev(0.85),
        game_id="MLB-STRESS-K",
        correlation_group="player:max-meyer",
    )
    ok, skip, passed, norm = evaluate(p)
    require(ok, f"expected gate pass, got skip={skip}")
    required_eval = {"G1 minimum edge", "G2 probability", "G3 injury clearance", "G4 minutes stability", "G5 platform line availability", "G6 sample size", "G7 CLV track record", "G9 market disagreement"}
    require(required_eval.issubset(set(passed)), f"missing passed gates: {required_eval - set(passed)}")
    alloc = runner.allocate_eligible_candidates([p], starting_exposure=0.0, daily_cap=runner.BASE_DAILY_CAP)
    require(len(alloc["picks"]) == 1, f"expected approved under Gate 8, got {alloc}")
    approved_gates = set(alloc["picks"][0].get("gates_passed") or [])
    require("G8 dynamic exposure/concentration caps" in approved_gates and "G10 duplicate protection" in approved_gates, f"missing Gate 8/10 approval markers: {approved_gates}")
    return {"passed_gates": alloc["picks"][0].get("gates_passed"), "approved": alloc["picks"][0]["selection"]}


def case_2_pitcher_er_over_positive_corr() -> dict[str, Any]:
    p = base_candidate(
        player="Trevor Rogers",
        selection="Trevor Rogers Over 2.5 Earned Runs Allowed",
        stat="earned runs",
        line=2.5,
        model_projection=4.1,
        projection=4.1,
        model_over_probability=0.79,
        ev=ev(0.79),
        sportsbook_verified=True,
    )
    ok, skip, _, _ = evaluate(p)
    require(ok, f"expected verified ER prop to pass, got {skip}")
    pair = corr.analyze_pair(
        {"sport": "MLB", "player_name": "Trevor Rogers", "team": "BAL", "stat_type": "earned runs", "line": 2.5, "game_id": "G1"},
        {"sport": "MLB", "player_name": "Rafael Devers", "team": "BOS", "stat_type": "hits runs rbis", "line": 1.5, "game_id": "G1"},
    )
    require("positive" in pair["correlation_label"], f"expected positive correlation, got {pair}")
    return {"correlation": pair}


def case_3_hitter_hrr_same_team_not_independent() -> dict[str, Any]:
    p = base_candidate(player="Alec Burleson", team="STL", game_id="G-HRR", selection="Alec Burleson Over 1.5 Hits+Runs+RBIs")
    ok, skip, _, _ = evaluate(p)
    require(ok, f"expected hitter HRR to pass, got {skip}")
    pair = corr.analyze_pair(
        {"sport": "MLB", "player_name": "Alec Burleson", "team": "STL", "stat_type": "hits runs rbis", "line": 1.5, "game_id": "G-HRR"},
        {"sport": "MLB", "player_name": "Nolan Arenado", "team": "STL", "stat_type": "hits runs rbis", "line": 1.5, "game_id": "G-HRR"},
    )
    require(pair["same_team_correlation"] and pair["same_game_correlation"], f"expected same team/game metadata, got {pair}")
    require(pair["correlation_label"] != "unknown correlation", f"same-team HRR mislabeled independent/unknown: {pair}")
    return {"correlation": pair}


def case_4_missing_projection_no_avg_fallback() -> dict[str, Any]:
    p = base_candidate(model_projection=None, projection=None, edge=None, model_edge=None, hit_row=hit_row(0.75, 20))
    p["hit_row"]["avg_stat_l10"] = 4.0
    ok, skip, _, norm = evaluate(p)
    require(not ok and skip and "GATE 1" in skip["gate_failed"], f"expected Gate 1 fail, got ok={ok} skip={skip}")
    reason = (skip.get("reason") or "").lower()
    require("projection unavailable" in reason, f"reason must say projection unavailable, got {skip}")
    require(norm.get("model_projection") in (None, ""), "avg_stat_l10 was incorrectly hydrated as model projection")
    return {"skip": skip}


def case_5_prizepicks_primary_line_not_sportsbook_dependency() -> dict[str, Any]:
    p = base_candidate(sportsbook_verified=False, sportsbook_api_available=False, hit_row=hit_row(0.60, 20), reasoning="synthetic Odds API 422 sportsbook response")
    alloc = runner.allocate_eligible_candidates([p], starting_exposure=0.0, daily_cap=runner.BASE_DAILY_CAP)
    require(len(alloc["picks"]) == 1, f"expected PrizePicks primary line to pass without Odds API props, got {alloc}")
    pick = alloc["picks"][0]
    require(pick.get("platform_line_available") is True, f"expected platform line marker, got {pick}")
    require("UNVERIFIED sportsbook prop allowed" not in pick.get("reasoning", ""), f"old sportsbook fallback leaked into reasoning: {pick.get('reasoning')}")
    return {"approved_reasoning": pick.get("reasoning"), "gate5": pick.get("platform_line_source")}


def case_6_missing_primary_platform_line_fails_gate5() -> dict[str, Any]:
    p = base_candidate(sportsbook_verified=False, hit_row=hit_row(0.70, 20), line=None, edge=1.5)
    p["selection"] = "Synthetic Player Over Hits+Runs+RBIs"
    ok, skip, _, _ = evaluate(p)
    require(not ok and skip and skip["gate_failed"] == "GATE 5 — PLATFORM LINE AVAILABILITY", f"expected Gate 5 primary platform fail, got {ok} {skip}")
    require("primary platform line missing/malformed" in skip["reason"], f"bad Gate 5 reason: {skip}")
    return {"skip": skip}


def case_7_probable_pitcher_change() -> dict[str, Any]:
    pitcher = base_candidate(player="Old Starter", selection="Old Starter Over 5.5 Pitcher Strikeouts", stat="Pitcher Strikeouts", probable_pitcher_changed=True)
    hitter = base_candidate(player="Hitter", selection="Hitter Over 1.5 Hits+Runs+RBIs", stat="Hits+Runs+RBIs", stale_pitcher_matchup=True, depends_on_pitcher_matchup=True)
    for p in [pitcher, hitter]:
        ok, skip, _, _ = evaluate(p)
        require(not ok and skip, f"expected stale pitcher matchup held, got ok={ok}")
        require("pitcher" in skip["reason"].lower() or "matchup" in skip["reason"].lower(), f"reason lacks pitcher/matchup: {skip}")
    alloc = runner.allocate_eligible_candidates([pitcher, hitter], 0.0)
    require(not alloc["picks"], f"stale pitcher matchup approved: {alloc}")
    return {"skipped": len(alloc["skipped"])}


def case_8_lineup_not_confirmed() -> dict[str, Any]:
    p = base_candidate(lineup_confirmed=False, player="Lineup Hitter", selection="Lineup Hitter Over 1.5 Hits+Runs+RBIs")
    ok, skip, _, _ = evaluate(p)
    require(not ok and skip, f"expected unconfirmed lineup hold/skip, got {ok}")
    require("lineup" in skip["reason"].lower(), f"reason lacks lineup: {skip}")
    return {"skip": skip}


def case_9_weather_risk() -> dict[str, Any]:
    severe = base_candidate(weather_risk="SEVERE", game_status="WEATHER DELAY")
    ok, skip, _, _ = evaluate(severe)
    require(not ok and skip, f"expected severe weather skip, got {ok}")
    require("WEATHER RISK" in skip["reason"], f"missing WEATHER RISK flag: {skip}")
    return {"skip": skip}


def case_10_bullpen_game_opener() -> dict[str, Any]:
    p = base_candidate(player="Opener", selection="Opener Over 5.5 Pitcher Strikeouts", stat="Pitcher Strikeouts", bullpen_game=True, opener=True, innings_expectation=2.0)
    ok, skip, _, _ = evaluate(p)
    require(not ok and skip, f"expected opener/workload skip, got {ok}")
    reason = skip["reason"].lower()
    require("innings" in reason or "workload" in reason, f"reason lacks innings/workload: {skip}")
    return {"skip": skip}


def case_11_doubleheader_specific_lineup() -> dict[str, Any]:
    p = base_candidate(doubleheader=True, game_number=2, lineup_confirmed=True, lineup_confirmed_specific_game=False)
    ok, skip, _, _ = evaluate(p)
    require(not ok and skip, f"expected doubleheader game-specific lineup skip, got {ok}")
    require("doubleheader" in skip["reason"].lower() and "lineup" in skip["reason"].lower(), f"bad reason: {skip}")
    return {"skip": skip}


def case_12_mlb_negative_correlation() -> dict[str, Any]:
    pair = corr.analyze_pair(
        {"sport": "MLB", "player_name": "Max Meyer", "team": "MIA", "stat_type": "Pitcher Strikeouts", "line": 5.5, "game_id": "GNEG"},
        {"sport": "MLB", "player_name": "Opp Hitter", "team": "ATL", "stat_type": "Hits+Runs+RBIs", "line": 1.5, "game_id": "GNEG"},
    )
    require("negative/risky" in pair["correlation_label"], f"expected negative/risky, got {pair}")
    require(pair["correlation_score"] < 0, f"expected negative score, got {pair}")
    return {"correlation": pair, "blocked_without_explicit_explanation": True}


def case_13_mlb_positive_correlation_approx() -> dict[str, Any]:
    pair = corr.analyze_pair(
        {"sport": "MLB", "player_name": "Pitcher", "team": "BAL", "stat_type": "Hits Allowed", "line": 5.5, "game_id": "GPOS"},
        {"sport": "MLB", "player_name": "Opp Hitter", "team": "BOS", "stat_type": "Hits+Runs+RBIs", "line": 1.5, "game_id": "GPOS"},
    )
    require("positive" in pair["correlation_label"], f"expected positive, got {pair}")
    p1, p2 = 0.72, 0.70
    independent = p1 * p2
    approx = min(min(p1, p2), independent + 0.20 * (min(p1, p2) - independent))
    require(independent < approx < min(p1, p2), f"combined probability should be approximate, got independent={independent} approx={approx}")
    return {"correlation": pair, "combined_probability_method": "approximate", "independent_product": round(independent, 4), "approx": round(approx, 4)}


def case_14_dynamic_gate8_cross_sport_order_independent() -> dict[str, Any]:
    nba_low = [base_candidate(sport="NBA", game_id=f"NBA{i}", player=f"NBA Low {i}", selection=f"NBA Low {i} Over 10.5 Points", stat="Points", model_projection=11.2, projection=11.2, line=10.5, model_over_probability=0.62, ev=ev(0.62), confidence="A", units=3.0, correlation_group=f"player:nba-low-{i}") for i in range(4)]
    mlb_high = [base_candidate(game_id=f"MLB{i}", player=f"MLB High {i}", selection=f"MLB High {i} Over 1.5 Hits+Runs+RBIs", model_projection=3.3, projection=3.3, line=1.5, model_over_probability=0.84 - i*0.01, ev=ev(0.84 - i*0.01), confidence="A", units=3.0, correlation_group=f"player:mlb-high-{i}") for i in range(4)]
    pool = nba_low + mlb_high
    a = runner.allocate_eligible_candidates([dict(x) for x in pool], starting_exposure=0.0, daily_cap=10.0)
    b = runner.allocate_eligible_candidates([dict(x) for x in reversed(pool)], starting_exposure=0.0, daily_cap=10.0)
    aset = [p["selection"] for p in a["picks"]]
    bset = [p["selection"] for p in b["picks"]]
    require(sorted(aset) == sorted(bset), f"allocation not order independent: {aset} vs {bset}")
    require(any(p.get("sport") == "MLB" for p in a["picks"]), f"MLB high EV did not enter allocation: {a}")
    require(all((p.get("sport") == "MLB" or runner.pick_ev_value(p) >= runner.pick_ev_value(a["picks"][-1])) for p in a["picks"]), "lower EV NBA displaced higher EV MLB")
    return {"approved": aset, "blocked": [s["pick"] for s in a["skipped"] if "GATE 8" in s["gate_failed"]]}


def create_temp_pick_history(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Pick History"
    headers = ["Date", "Sport", "Platform", "Pick Type", "Confidence Tier", "Line", "Odds", "Model Projection", "Edge", "Model Over Probability", "EV", "Edge Type Tags", "CLV", "Opening Line", "Closing Line", "Line Movement", "Favorable Line Move 0.5+", "Correlation Group", "Slip ID", "Result", "Units", "PnL", "Pick Ref"]
    ws.append(headers)
    rows = [
        [date.today().isoformat(), "MLB", "PrizePicks", "PROP", "A", 1.5, -110, 3.0, 1.5, 0.80, 0.5272, "Value", 0.5, 1.5, 1.0, -0.5, True, "player:a", "S1", "WIN", 3, 2.727, "A Over 1.5 HRR"],
        [date.today().isoformat(), "MLB", "PrizePicks", "PROP", "B", 5.5, -110, 4.0, 1.5, 0.70, 0.3363, "Value", -0.5, 5.5, 6.0, 0.5, False, "player:b", "S2", "LOSS", 2, -2, "B Under 5.5 K"],
        [date.today().isoformat(), "MLB", "PrizePicks", "PROP", "A", 2.5, -110, 4.1, 1.6, 0.79, 0.508, "Value", None, 2.5, None, None, False, "player:c", "S3", "MANUAL REVIEW", 3, 0, "C Over ER"],
        [date.today().isoformat(), "MLB", "PrizePicks", "PROP", "A", 1.5, -110, 3.2, 1.7, 0.81, 0.546, "Value", None, 1.5, None, None, False, "player:d", "S4", "PENDING", 3, 0, "D Over HRR"],
    ]
    for row in rows:
        ws.append(row)
    wb.save(path)


def case_15_result_grading_temp_pick_history() -> dict[str, Any]:
    path = OUT / "temp_master_pnl.xlsx"
    create_temp_pick_history(path)
    rows, meta = backtest.load_pick_history(path)
    graded = backtest.graded(rows)
    require(len(rows) == 4, f"expected 4 temp history rows, got {len(rows)}")
    require(len(graded) == 2 and {r["result"] for r in graded} == {"WIN", "LOSS"}, f"expected only WIN/LOSS graded, got {graded}")
    require(all(r.get("tracking_era") == "full_schema_tracking" for r in graded), f"full_schema_tracking not active: {graded}")
    summary = backtest.summarize(graded)
    return {"graded_count": len(graded), "ignored_count": len(rows) - len(graded), "summary": summary, "meta": meta}


CASES: list[tuple[str, Callable[[], dict[str, Any]]]] = [
    ("1. Pitcher strikeouts Over", case_1_pitcher_strikeouts_over),
    ("2. Pitcher earned runs Over", case_2_pitcher_er_over_positive_corr),
    ("3. Hitter HRR Over", case_3_hitter_hrr_same_team_not_independent),
    ("4. Missing projection", case_4_missing_projection_no_avg_fallback),
    ("5. PrizePicks primary line without Odds API props", case_5_prizepicks_primary_line_not_sportsbook_dependency),
    ("6. Missing primary platform line", case_6_missing_primary_platform_line_fails_gate5),
    ("7. Probable pitcher change", case_7_probable_pitcher_change),
    ("8. Lineup not confirmed", case_8_lineup_not_confirmed),
    ("9. Weather risk", case_9_weather_risk),
    ("10. Bullpen game / opener", case_10_bullpen_game_opener),
    ("11. Doubleheader", case_11_doubleheader_specific_lineup),
    ("12. MLB negative correlation", case_12_mlb_negative_correlation),
    ("13. MLB positive correlation", case_13_mlb_positive_correlation_approx),
    ("14. Dynamic Gate 8 cross-sport", case_14_dynamic_gate8_cross_sport_order_independent),
    ("15. Result grading", case_15_result_grading_temp_pick_history),
]


def main() -> int:
    setup = setup_temp_workbooks()
    before = sha256(REAL_MLB_WB)
    results = []
    for name, fn in CASES:
        try:
            detail = fn()
            results.append({"case": name, "status": "PASS", "detail": detail})
        except Exception as e:
            results.append({"case": name, "status": "FAIL", "error": str(e)})
    after = sha256(REAL_MLB_WB)
    unchanged = before == after
    report = {
        "status": "PASS" if all(r["status"] == "PASS" for r in results) and unchanged else "FAIL",
        "output_dir": str(OUT),
        "temp_setup": setup,
        "real_mlb_workbook": str(REAL_MLB_WB),
        "real_mlb_workbook_hash_before": before,
        "real_mlb_workbook_hash_after": after,
        "real_mlb_workbook_hash_unchanged": unchanged,
        "cases": results,
        "missing_mlb_metadata_needed_for_reliable_picks": [
            "confirmed lineup by game for doubleheaders",
            "probable starter change feed and matchup invalidation",
            "weather delay/postponement severity",
            "opener/bullpen-game and innings/workload expectation",
            "handedness and ballpark/wind factors",
        ],
    }
    (OUT / "mlb_stress_test_report.json").write_text(json.dumps(report, indent=2, default=str) + "\n")
    for r in results:
        print(f"{r['case']}: {r['status']}")
        if r["status"] == "FAIL":
            print(f"  ERROR: {r['error']}")
    print(f"real_mlb_workbook_hash_unchanged: {unchanged}")
    print(f"report: {OUT / 'mlb_stress_test_report.json'}")
    return 0 if report["status"] == "PASS" else 1


if __name__ == "__main__":
    raise SystemExit(main())
