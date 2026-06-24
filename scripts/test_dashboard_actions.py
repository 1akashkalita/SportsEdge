#!/usr/bin/env python3
"""test_dashboard_actions.py — ACTION-01..04 tests for Phase 3 safe write actions.

Covers:
- TestRefreshAction: ACTION-01 a/b/c — /action/refresh POST route (async subprocess,
  lock-aware refusal, task whitelist enforcement)
- TestStatusEndpoint: ACTION-01d — /api/status GET endpoint returns locked + last_run
- TestMarkPlaced: ACTION-02 — mark_placed() writes Placed/Placed At additively
- TestAddNote: ACTION-03 — add_note() writes Operator Note; grading Notes untouched
- TestActionFourHardLine: ACTION-04 a/b/c — gate output unchanged, caps unchanged,
  writes only touch Slip History

RED scaffold: all tests fail or error at Wave 1. As later waves (03-02, 03-03) land
the implementations, these tests turn GREEN automatically — no test modifications needed.
"""
from __future__ import annotations

import importlib.util
import os
import sys
import tempfile
import unittest
from pathlib import Path
from typing import Any
from unittest.mock import patch, MagicMock

from openpyxl import Workbook

SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

import dashboard  # noqa: E402
import dashboard_writes  # noqa: E402


# ---------------------------------------------------------------------------
# Slip History headers (from slip_payouts.py:18-24) — 23 columns
# ---------------------------------------------------------------------------

SLIP_HISTORY_HEADERS: list[str] = [
    "Date", "Slip ID", "Platform", "Slip Type", "Number of Legs", "Legs",
    "Stake Units", "Winning Legs", "Losing Legs", "Push/Void/DNP Legs",
    "Contains Demon", "Contains Goblin", "Special Line Count", "Slip Result",
    "Standard Payout Multiplier", "Estimated Payout Multiplier",
    "Actual Payout Multiplier", "Payout Confidence", "Gross Return", "Net PnL",
    "Needs Payout Reconciliation", "Graded At", "Notes",
]


# ---------------------------------------------------------------------------
# Fixture builders
# ---------------------------------------------------------------------------

def _canonical_pick() -> dict[str, Any]:
    """Return a minimal gate-ready pick dict for evaluate_no_bet_gates().

    Field set matches test_dynamic_gate8.py:16-54 (cand() helper) exactly.
    All gate-required fields are present; tests that call evaluate_no_bet_gates
    should use this fixture directly.
    """
    return {
        "kind": "prop",
        "date": "2026-06-24",
        "sport": "MLB",
        "game_id": "g1",
        "projection_id": "proj-1",
        "selection": "Player A Over 0.5 Hits",
        "line": 0.5,
        "odds": "standard",
        "score": 3,
        "confidence": "A",
        "units": 1.0,
        "player": "Player A",
        "player_team": "Player A",
        "team": "T1",
        "stat": "hits",
        "model_projection": 0.7,
        "edge": 1.5,
        "model_over_probability": 0.65,
        "ev": 0.2,
        "edge_type_tags": "projection_edge",
        "injury_status": "ACTIVE",
        "sportsbook_verified": True,
        "hit_row": {"sample_size": 20, "hit_rate_l10": 0.7},
        "reasoning": "test fixture",
        "line_timing": "pregame",
        "line_timing_confidence": "high",
        "line_timing_reason": "test fixture pregame",
        "live_line_flag": False,
        "stale_line_flag": False,
    }


def _make_slip_history_wb(rows: list[dict[str, Any]] | None = None) -> Workbook:
    """Build an in-memory Workbook with a Slip History sheet using the exact 23-column header.

    Args:
        rows: Optional list of row dicts keyed by SLIP_HISTORY_HEADERS column names.

    Returns:
        Workbook with a single "Slip History" sheet populated with headers and rows.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Slip History"
    ws.append(SLIP_HISTORY_HEADERS)
    for row in (rows or []):
        ws.append([row.get(h) for h in SLIP_HISTORY_HEADERS])
    return wb


# ---------------------------------------------------------------------------
# TestRefreshAction — ACTION-01 a/b/c
# ---------------------------------------------------------------------------

class TestRefreshAction(unittest.TestCase):
    """ACTION-01: /action/refresh POST route — async subprocess, lock-awareness, whitelist.

    RED at Wave 1: the route /action/refresh does not exist yet (404/405).
    Turns GREEN in 03-03 (Wave 3) when the route is implemented.
    """

    def setUp(self) -> None:
        self.client = dashboard.app.test_client()

    def test_refresh_triggers_subprocess(self) -> None:
        """ACTION-01a — POST /action/refresh with a valid task returns 302 and spawns async.

        The route must return immediately (never block inline on the runner) and the
        subprocess must have been spawned. A mocked spawn confirms the invocation
        without actually running the runner.

        RED: route /action/refresh does not exist → 404.
        """
        with patch("threading.Thread") as mock_thread_cls:
            mock_thread = MagicMock()
            mock_thread_cls.return_value = mock_thread
            resp = self.client.post(
                "/action/refresh",
                data={"task": "nba_daily_picks"},
            )
        # Expect POST→redirect (302) — never blocks inline
        self.assertEqual(resp.status_code, 302, f"Expected 302 redirect, got {resp.status_code}")
        # The thread must have been started (async, not inline)
        mock_thread.start.assert_called_once()

    def test_refresh_refused_when_locked(self) -> None:
        """ACTION-01b — When the runner lock is held, POST flashes a warning and does NOT spawn.

        The lock probe must detect the held lock and refuse to spawn a concurrent run.

        RED: route /action/refresh does not exist → 404.
        """
        with patch("dashboard._runner_is_locked", return_value=True):
            with patch("threading.Thread") as mock_thread_cls:
                resp = self.client.post(
                    "/action/refresh",
                    data={"task": "nba_daily_picks"},
                )
        # Must redirect (flash + redirect) without spawning
        self.assertEqual(resp.status_code, 302, f"Expected 302 redirect, got {resp.status_code}")
        mock_thread_cls.assert_not_called()

    def test_refresh_invalid_task_rejected(self) -> None:
        """ACTION-01c — A task NOT in ALLOWED_TASKS is rejected with no spawn.

        The whitelist guard must refuse unknown task names before any subprocess is
        attempted.

        RED: route /action/refresh does not exist → 404.
        """
        with patch("threading.Thread") as mock_thread_cls:
            resp = self.client.post(
                "/action/refresh",
                data={"task": "delete_everything"},
            )
        # Must redirect (flash error) without spawning
        self.assertEqual(resp.status_code, 302, f"Expected 302 redirect, got {resp.status_code}")
        mock_thread_cls.assert_not_called()


# ---------------------------------------------------------------------------
# TestStatusEndpoint — ACTION-01d
# ---------------------------------------------------------------------------

class TestStatusEndpoint(unittest.TestCase):
    """ACTION-01d: /api/status returns locked + last_run fields.

    RED at Wave 1: the route /api/status does not exist yet (404).
    Turns GREEN in 03-03 (Wave 3) when the route is implemented.
    """

    def setUp(self) -> None:
        self.client = dashboard.app.test_client()

    def test_status_fields(self) -> None:
        """ACTION-01d — GET /api/status?task=<t> returns JSON with 'locked' and 'last_run'.

        Both fields must be present in the response JSON regardless of their current
        values — the contract is structural, not behavioral.

        RED: route /api/status does not exist → 404.
        """
        resp = self.client.get("/api/status?task=nba_daily_picks")
        self.assertEqual(resp.status_code, 200, f"Expected 200 from /api/status, got {resp.status_code}")
        data = resp.get_json()
        self.assertIsNotNone(data, "Expected JSON response body from /api/status")
        self.assertIn("locked", data, "Response JSON must contain 'locked' field")
        self.assertIn("last_run", data, "Response JSON must contain 'last_run' field")


# ---------------------------------------------------------------------------
# TestMarkPlaced — ACTION-02
# ---------------------------------------------------------------------------

class TestMarkPlaced(unittest.TestCase):
    """ACTION-02: mark_placed() adds Placed/Placed At additively, toggles the correct row.

    RED at Wave 1: mark_placed raises NotImplementedError.
    Turns GREEN in 03-02 (Wave 2) when mark_placed is implemented.
    """

    def test_mark_placed_additive(self) -> None:
        """ACTION-02 — mark_placed adds Placed/Placed At columns to Slip History, toggles row.

        Assertions:
        - After mark_placed, the Placed/Placed At columns exist on the sheet (additive).
        - The matched (Date, Slip ID) row has Placed set correctly.
        - All original 23 SLIP_HISTORY_HEADERS columns are still present and unchanged.

        RED: mark_placed raises NotImplementedError.
        """
        # Build a temporary workbook with one slip row
        wb = _make_slip_history_wb([
            {"Date": "2026-06-24", "Slip ID": "slip-001", "Notes": "grading-note"},
        ])

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name
        self.addCleanup(os.unlink, tmp_path)
        wb.save(tmp_path)

        # Patch PNL_DIR so dashboard_writes resolves master_pnl.xlsx to our temp file
        # by patching the actual path used inside mark_placed
        master_path = Path(tmp_path)

        with patch.object(dashboard_writes, "PNL_DIR", master_path.parent):
            # Rename the temp file to match what mark_placed will look for
            import shutil
            pnl_master = master_path.parent / "master_pnl.xlsx"
            self.addCleanup(lambda p=pnl_master: p.unlink(missing_ok=True))
            shutil.copy(tmp_path, pnl_master)

            dashboard_writes.mark_placed("2026-06-24", "slip-001", True)

            # Reload and verify
            from openpyxl import load_workbook
            wb2 = load_workbook(pnl_master)
            ws2 = wb2["Slip History"]
            headers2 = [ws2.cell(1, c).value for c in range(1, ws2.max_column + 1)]

        # Additive: Placed and Placed At must be present
        self.assertIn("Placed", headers2, "Placed column must be added by mark_placed")
        self.assertIn("Placed At", headers2, "Placed At column must be added by mark_placed")

        # Original 23 headers must still be present
        for h in SLIP_HISTORY_HEADERS:
            self.assertIn(h, headers2, f"Original header {h!r} must not be removed")

        # The matched row must have Placed == True
        placed_col = headers2.index("Placed") + 1
        row2_placed = ws2.cell(2, placed_col).value
        self.assertTrue(row2_placed, "Placed must be True on the matched row after mark_placed(..., True)")

        # The Notes column must be byte-for-byte unchanged
        notes_col = headers2.index("Notes") + 1
        row2_notes = ws2.cell(2, notes_col).value
        self.assertEqual(row2_notes, "grading-note", "Grading-owned Notes column must be unchanged")


# ---------------------------------------------------------------------------
# TestAddNote — ACTION-03
# ---------------------------------------------------------------------------

class TestAddNote(unittest.TestCase):
    """ACTION-03: add_note() sets Operator Note additively; grading Notes column untouched.

    RED at Wave 1: add_note raises NotImplementedError.
    Turns GREEN in 03-02 (Wave 2) when add_note is implemented.
    """

    def test_add_note_additive(self) -> None:
        """ACTION-03 — add_note sets Operator Note on matched row; grading Notes byte-identical.

        Assertions:
        - After add_note, the Operator Note column exists on the sheet (additive).
        - The matched row has Operator Note set to the provided text.
        - The grading-owned Notes column value is byte-for-byte unchanged.

        RED: add_note raises NotImplementedError.
        """
        wb = _make_slip_history_wb([
            {"Date": "2026-06-24", "Slip ID": "slip-002", "Notes": "grading-owned"},
        ])

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name
        self.addCleanup(os.unlink, tmp_path)
        wb.save(tmp_path)

        master_path = Path(tmp_path)
        with patch.object(dashboard_writes, "PNL_DIR", master_path.parent):
            import shutil
            pnl_master = master_path.parent / "master_pnl.xlsx"
            self.addCleanup(lambda p=pnl_master: p.unlink(missing_ok=True))
            shutil.copy(tmp_path, pnl_master)

            dashboard_writes.add_note("2026-06-24", "slip-002", "my operator note")

            from openpyxl import load_workbook
            wb2 = load_workbook(pnl_master)
            ws2 = wb2["Slip History"]
            headers2 = [ws2.cell(1, c).value for c in range(1, ws2.max_column + 1)]

        # Additive: Operator Note must be present
        self.assertIn("Operator Note", headers2, "Operator Note column must be added by add_note")

        # The matched row must have the note set
        note_col = headers2.index("Operator Note") + 1
        row2_note = ws2.cell(2, note_col).value
        self.assertEqual(row2_note, "my operator note", "Operator Note must be set on the matched row")

        # Grading Notes must be byte-for-byte unchanged
        notes_col = headers2.index("Notes") + 1
        row2_notes = ws2.cell(2, notes_col).value
        self.assertEqual(row2_notes, "grading-owned", "Grading-owned Notes column must be unchanged after add_note")


# ---------------------------------------------------------------------------
# TestActionFourHardLine — ACTION-04 a/b/c
# ---------------------------------------------------------------------------

class TestActionFourHardLine(unittest.TestCase):
    """ACTION-04: no dashboard action changes gate logic, grades, EV, or exposure caps.

    Uses the importlib runner-load idiom (test_dynamic_gate8.py:1-13) so the runner
    is loaded once in setUpClass without top-level import side-effects.

    RED at Wave 1 for ACTION-04a and ACTION-04c (need implemented write helpers).
    ACTION-04b passes at Wave 1 (cap constants are static assertions).
    """

    runner: Any = None  # populated by setUpClass

    @classmethod
    def setUpClass(cls) -> None:
        """Load the runner once via importlib (the ONLY supported load idiom in tests)."""
        mod_path = Path(__file__).with_name("sports_system_runner.py")
        spec = importlib.util.spec_from_file_location("sports_system_runner", mod_path)
        assert spec and spec.loader, "Could not create importlib spec for sports_system_runner.py"
        cls.runner = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(cls.runner)
        cls.runner.load_suppressed_edge_types = lambda: {}  # stub external I/O

    def test_mark_placed_does_not_alter_gate_output(self) -> None:
        """ACTION-04a — evaluate_no_bet_gates output is bit-identical before and after mark_placed.

        The gate gauntlet must be completely unaffected by the write action — same (ok,
        skip_record, gates_passed) tuple before and after.

        RED: mark_placed raises NotImplementedError.
        """
        pick = _canonical_pick()

        # Gate output before any write
        before_ok, before_skip, before_passed = self.runner.evaluate_no_bet_gates(pick, {})

        # Perform the write action (on a temp workbook)
        wb = _make_slip_history_wb([
            {"Date": "2026-06-24", "Slip ID": "slip-a04a", "Notes": "grading"},
        ])
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name
        self.addCleanup(os.unlink, tmp_path)
        wb.save(tmp_path)

        master_path = Path(tmp_path)
        with patch.object(dashboard_writes, "PNL_DIR", master_path.parent):
            import shutil
            pnl_master = master_path.parent / "master_pnl.xlsx"
            self.addCleanup(lambda p=pnl_master: p.unlink(missing_ok=True))
            shutil.copy(tmp_path, pnl_master)
            dashboard_writes.mark_placed("2026-06-24", "slip-a04a", True)

        # Gate output after write — must be identical
        after_ok, after_skip, after_passed = self.runner.evaluate_no_bet_gates(pick, {})

        self.assertEqual(before_ok, after_ok, "Gate ok must be unchanged after mark_placed")
        self.assertEqual(before_skip, after_skip, "Gate skip_record must be unchanged after mark_placed")
        self.assertEqual(before_passed, after_passed, "Gate gates_passed must be unchanged after mark_placed")

    def test_exposure_caps_unchanged(self) -> None:
        """ACTION-04b — PER_PLAYER_CAP and PER_GAME_CAP are 6.0 (exposure caps never touched).

        These are static runner constants; the test asserts they have not been removed
        or altered. References PER_PLAYER_CAP and PER_GAME_CAP only (the active v3.0 cap constants).

        PASSES at Wave 1: the cap constants are immutable module-level values.
        """
        self.assertEqual(
            self.runner.PER_PLAYER_CAP,
            6.0,
            "PER_PLAYER_CAP must remain 6.0 — no action may alter exposure caps",
        )
        self.assertEqual(
            self.runner.PER_GAME_CAP,
            6.0,
            "PER_GAME_CAP must remain 6.0 — no action may alter exposure caps",
        )

    def test_write_only_touches_slip_history(self) -> None:
        """ACTION-04c — After a write, Picks/Skipped Picks/CLV Tracker sheets are untouched.

        Build a workbook with multiple sheets, run a write action, and confirm the
        non-Slip-History sheets have neither been added nor modified.

        RED: mark_placed raises NotImplementedError.
        """
        # Build a multi-sheet workbook to simulate a sport workbook
        wb = Workbook()
        ws_slips = wb.active
        ws_slips.title = "Slip History"
        ws_slips.append(SLIP_HISTORY_HEADERS)
        ws_slips.append(["2026-06-24", "slip-a04c"] + [None] * (len(SLIP_HISTORY_HEADERS) - 2))

        ws_picks = wb.create_sheet("Picks")
        ws_picks.append(["Date", "Sport", "Selection"])
        ws_picks.append(["2026-06-24", "MLB", "Player X Over 0.5 Hits"])

        ws_skipped = wb.create_sheet("Skipped Picks")
        ws_skipped.append(["Date", "Sport", "Pick", "Gate Failed"])
        ws_skipped.append(["2026-06-24", "MLB", "Player Y Over 0.5 Hits", "G1"])

        ws_clv = wb.create_sheet("CLV Tracker")
        ws_clv.append(["Date", "Selection", "CLV"])
        ws_clv.append(["2026-06-24", "Player Z Over 0.5 Hits", "1.02"])

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name
        self.addCleanup(os.unlink, tmp_path)
        wb.save(tmp_path)

        master_path = Path(tmp_path)
        with patch.object(dashboard_writes, "PNL_DIR", master_path.parent):
            import shutil
            pnl_master = master_path.parent / "master_pnl.xlsx"
            self.addCleanup(lambda p=pnl_master: p.unlink(missing_ok=True))
            shutil.copy(tmp_path, pnl_master)
            dashboard_writes.mark_placed("2026-06-24", "slip-a04c", True)

            from openpyxl import load_workbook
            wb2 = load_workbook(pnl_master)

        # Picks sheet: row 2 must be byte-identical
        ws_picks2 = wb2["Picks"]
        self.assertEqual(ws_picks2.cell(2, 1).value, "2026-06-24", "Picks sheet Date must be unchanged")
        self.assertEqual(ws_picks2.cell(2, 2).value, "MLB", "Picks sheet Sport must be unchanged")
        self.assertEqual(ws_picks2.cell(2, 3).value, "Player X Over 0.5 Hits", "Picks sheet Selection must be unchanged")

        # Skipped Picks sheet: row 2 must be byte-identical
        ws_skipped2 = wb2["Skipped Picks"]
        self.assertEqual(ws_skipped2.cell(2, 4).value, "G1", "Skipped Picks Gate Failed must be unchanged")

        # CLV Tracker sheet: row 2 must be byte-identical
        ws_clv2 = wb2["CLV Tracker"]
        self.assertEqual(ws_clv2.cell(2, 3).value, "1.02", "CLV Tracker CLV must be unchanged")


# ---------------------------------------------------------------------------
# Test runner
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    unittest.main()
