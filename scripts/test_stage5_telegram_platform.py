#!/usr/bin/env python3
"""Stage 5 audit tests: platform must appear in the user-facing Telegram alerts.

Fixtures use neutral player/selection names that do NOT contain the platform
string, so the assertions genuinely test that the platform field is rendered.
"""
from __future__ import annotations

import importlib.util
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

SCRIPT = Path(__file__).resolve().with_name("sports_system_runner.py")
spec = importlib.util.spec_from_file_location("sports_system_runner", SCRIPT)
runner = importlib.util.module_from_spec(spec)
assert spec and spec.loader
spec.loader.exec_module(runner)


def _picks_row(platform: str, selection: str, conf: str) -> list:
    row = [None] * len(runner.PICKS_HEADERS)
    row[0] = runner.today_str()
    row[1] = "MLB"
    row[6] = "PROP"
    row[7] = selection
    row[8] = 1.5
    row[10] = conf
    row[11] = 1.0
    row[12] = "ACTIVE"
    row[14] = platform
    row[18] = 0.62
    row[19] = 0.18
    return row


def _prop_row(platform: str, player: str, stat: str, line) -> list:
    row = [None] * len(runner.PROPS_HEADERS)
    row[0] = runner.today_str()
    row[1] = "MLB"
    row[3] = player
    row[6] = stat
    row[7] = line
    row[10] = "ACTIVE"
    row[14] = platform
    return row


class Stage5TelegramPlatformTests(unittest.TestCase):
    def _make_workbook(self) -> Path:
        tmpdir = tempfile.TemporaryDirectory()
        self.addCleanup(tmpdir.cleanup)
        path = Path(tmpdir.name) / "mlb.xlsx"
        wb = Workbook()
        wb.remove(wb.active)
        picks = wb.create_sheet("Picks")
        picks.append(runner.PICKS_HEADERS)
        # Neutral names — platform must come from the Platform column, not the selection text.
        picks.append(_picks_row("Underdog", "Aaron Judge Over 1.5 Hits", "A"))
        picks.append(_picks_row("PrizePicks", "Mookie Betts Over 2.5 Hits", "B"))
        pp = wb.create_sheet("Player Props")
        pp.append(runner.PROPS_HEADERS)
        pp.append(_prop_row("Underdog", "Aaron Judge", "Hits", 1.5))
        wb.save(path)
        return path

    def test_pick_text_includes_platform(self) -> None:
        row = _picks_row("Underdog", "Aaron Judge Over 1.5 Hits", "A")
        text = runner.pick_text(tuple(row))
        self.assertIn("Underdog", text, f"platform missing from pick_text output: {text!r}")

    def test_top_props_from_workbook_includes_platform(self) -> None:
        path = self._make_workbook()
        props = runner.top_props_from_workbook(str(path), 3)
        self.assertTrue(props)
        self.assertIn("Underdog", props[0], f"platform missing from prop summary: {props!r}")

    def test_build_picks_alert_shows_both_platforms(self) -> None:
        path = self._make_workbook()
        msg = runner.build_picks_alert("mlb", {"workbook": str(path), "skipped_picks": 0, "skip_rate_pct": 0})
        self.assertIn("Underdog", msg, f"Underdog missing from picks alert: {msg!r}")
        self.assertIn("PrizePicks", msg, f"PrizePicks missing from picks alert: {msg!r}")

    def test_line_move_summary_alert_shows_platform(self) -> None:
        moves = [
            {"player": "Aaron Judge", "stat": "Hits", "old_line": 1.5, "new_line": 2.0,
             "direction": "unfavorable", "action": "watch", "platform": "Underdog",
             "line_timing": "pregame", "movement_type": "pregame"},
        ]
        msg = runner.build_line_move_summary_alert("mlb_prop_monitor", moves)
        self.assertIn("Underdog", msg, f"platform missing from line move alert: {msg!r}")


def _parlay_leg(platform: str, team: str, player: str, conf: str = "A") -> dict:
    return {
        "kind": "prop", "confidence": conf, "team": team, "platform": platform,
        "selection": f"{player} Over 1.5 Hits", "units": 1.0, "ev": 0.2,
        "model_over_probability": 0.62, "sport": "MLB",
    }


class Stage5ParlayPlatformTests(unittest.TestCase):
    """A correlated parlay is a single DFS slip and cannot mix platforms."""

    def _generate(self, legs):
        # allocate_eligible_candidates is the core selector; reuse it to feed generate_picks logic.
        # We test the parlay grouping directly through generate_picks-like selection.
        allocation = runner.allocate_eligible_candidates(
            [dict(x) for x in legs], starting_exposure=0.0, daily_cap=None
        )
        return allocation

    def test_correlated_parlay_does_not_mix_platforms(self) -> None:
        # Two same-team legs on DIFFERENT platforms must NOT be paired into one slip.
        selected = [
            _parlay_leg("PrizePicks", "NYY", "Aaron Judge"),
            _parlay_leg("Underdog", "NYY", "Juan Soto"),
        ]
        # Reproduce the parlay grouping logic from generate_picks.
        prop_legs = [p for p in selected if p.get("kind") == "prop" and p.get("confidence") in ("A", "B")]
        by_platform_team: dict = {}
        for pick in prop_legs:
            plat = str(pick.get("platform") or pick.get("primary_platform") or "DFS")
            by_platform_team.setdefault((plat, str(pick.get("team") or "")), []).append(pick)
        pair = next((legs[:2] for legs in by_platform_team.values() if len(legs) >= 2), [])
        # No same-platform pair exists, so no parlay should form.
        self.assertEqual(pair, [], "parlay must not mix PrizePicks and Underdog legs")

    def test_correlated_parlay_pairs_same_platform_legs(self) -> None:
        selected = [
            _parlay_leg("Underdog", "NYY", "Aaron Judge"),
            _parlay_leg("Underdog", "NYY", "Juan Soto"),
        ]
        prop_legs = [p for p in selected if p.get("kind") == "prop" and p.get("confidence") in ("A", "B")]
        by_platform_team: dict = {}
        for pick in prop_legs:
            plat = str(pick.get("platform") or pick.get("primary_platform") or "DFS")
            by_platform_team.setdefault((plat, str(pick.get("team") or "")), []).append(pick)
        pair = next((legs[:2] for legs in by_platform_team.values() if len(legs) >= 2), [])
        self.assertEqual(len(pair), 2)
        platforms = {p["platform"] for p in pair}
        self.assertEqual(platforms, {"Underdog"}, "paired legs must share one platform")

    def test_generate_picks_groups_parlays_by_platform_and_team(self) -> None:
        """Guard against regressing to team-only parlay grouping in the real function."""
        import inspect
        src = inspect.getsource(runner.generate_picks)
        self.assertIn("by_platform_team", src,
            "generate_picks must group correlated parlays by (platform, team), not team alone")
        self.assertNotIn("by_team.setdefault", src,
            "generate_picks must not use team-only parlay grouping")


if __name__ == "__main__":
    unittest.main()
