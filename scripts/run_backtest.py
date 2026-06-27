#!/usr/bin/env python3
"""Backtest SportsEdge historical pick performance from master_pnl.xlsx."""
from __future__ import annotations

import argparse
import json
import math
import re
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from slip_payouts import summarize_slip_history_rows
from workbook_io import safe_load_workbook, safe_save_workbook

ROOT = Path(__file__).resolve().parents[1]
MASTER_PNL = ROOT / "data" / "pnl" / "master_pnl.xlsx"
JSON_DIR = ROOT / "data" / "research"
REPORT_PATH = ROOT / "data" / "pnl" / "backtest_report.xlsx"
VALID_RESULTS = {"WIN", "LOSS", "PUSH"}
FULL_SCHEMA_TRACKING_START = date(2026, 6, 9)
FULL_SCHEMA_REQUIRED_FIELDS = ["confidence_tier", "ev", "line", "model_over_probability", "platform"]
MARKET_CONTEXT_FIELDS = [
    "game_total",
    "spread",
    "team_implied_total",
    "opponent_implied_total",
    "moneyline_implied_probability",
    "fd_total",
    "dk_total",
    "fd_spread",
    "dk_spread",
    "total_disagreement",
    "spread_disagreement",
    "market_movement_direction",
    "clv_baseline_total",
    "clv_baseline_spread",
    "market_context_available",
    "market_context_source",
]
MARKET_CONTEXT_MIN_RECOMMENDATION_SAMPLE = 50
TRACKING_ERA_NOTE = "Historical rows before full schema tracking can be used for raw ROI only, not model/tier/CLV calibration."
EDGE_TAGS = [
    "Value",
    "Line Movement",
    "Sharp Fade",
    "Injury Spot",
    "Matchup Exploit",
    "Trend",
    "Correlated",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--date", default="today", help="Report date YYYY-MM-DD or today")
    parser.add_argument("--input", default=str(MASTER_PNL), help="master_pnl.xlsx path")
    parser.add_argument("--json-output", default=None, help="Override JSON output path")
    parser.add_argument("--xlsx-output", default=str(REPORT_PATH), help="Override XLSX report path")
    return parser.parse_args()


def resolve_date(value: str | None) -> str:
    if not value or value == "today":
        return date.today().isoformat()
    return datetime.fromisoformat(value).date().isoformat()


def norm_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")


def to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        if math.isnan(float(value)):
            return None
        return float(value)
    text = str(value).strip().replace("%", "")
    text = re.sub(r"[^0-9.\-]+", "", text)
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return None
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(text).date()
    except ValueError:
        return None


def canonical_result(value: Any) -> str:
    text = str(value or "").strip().upper()
    if text in {"W", "WON"}:
        return "WIN"
    if text in {"L", "LOST"}:
        return "LOSS"
    if text in {"P", "VOID"}:
        return "PUSH"
    return text


def pick_type_from(row: dict[str, Any]) -> str:
    explicit = row.get("pick_type") or row.get("type")
    if explicit:
        text = str(explicit).lower()
    else:
        text = str(row.get("pick_ref") or row.get("pick") or "").lower()
    if "parlay" in text or "correlated" in text:
        return "Parlays"
    if "prop" in text or "pts" in text or "reb" in text or "ast" in text or "strikeout" in text or "hits" in text:
        return "Props"
    if "spread" in text or re.search(r"\b[+-]\d+(\.\d+)?\b", text):
        return "Spreads"
    if "total" in text or "over " in text or "under " in text:
        return "Totals"
    return "Unknown"


def confidence_tier_from(row: dict[str, Any]) -> str:
    for key in ("confidence_tier", "tier", "confidence"):
        if row.get(key):
            text = str(row[key]).strip().upper()
            match = re.search(r"\b([ABC])\b", text)
            if match:
                return match.group(1)
    text = " ".join(str(row.get(k) or "") for k in ("pick_ref", "pick", "notes"))
    match = re.search(r"\b(?:TIER|CONFIDENCE)[:=\s-]*([ABC])\b", text, re.I)
    return match.group(1).upper() if match else ""


def platform_from(row: dict[str, Any]) -> str:
    explicit = row.get("platform") or row.get("book") or row.get("sportsbook")
    return str(explicit).strip() if explicit else ""


def edge_tags_from(row: dict[str, Any]) -> list[str]:
    explicit = row.get("edge_type_tags") or row.get("edge_type") or row.get("tags") or row.get("flag_type")
    if not explicit:
        return []
    text = str(explicit).lower()
    tags: list[str] = []
    patterns = {
        "Value": ["value", "+ev", "positive ev"],
        "Line Movement": ["line movement", "line move", "moved", "0.5+", "steam"],
        "Sharp Fade": ["sharp fade", "fade"],
        "Injury Spot": ["injury", "out", "questionable", "minutes risk"],
        "Matchup Exploit": ["matchup", "bad matchup", "exploit"],
        "Trend": ["trend", "trending"],
        "Correlated": ["correlated", "correlation", "stack", "parlay"],
    }
    for tag, needles in patterns.items():
        if any(n in text for n in needles) or tag.lower() in text:
            tags.append(tag)
    return tags


def line_movement_flag(row: dict[str, Any]) -> bool:
    text = " ".join(str(row.get(k) or "") for k in ("line_movement", "notes", "edge_type", "tags", "pick_ref", "pick")).lower()
    return bool(re.search(r"0\.5\+|\+0\.5|favorable move|line movement|line move|moved", text))


def parse_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    text = str(value or "").strip().lower()
    return text in {"true", "yes", "y", "1", "x"}


LINE_TIMING_VALUES = {"pregame", "live", "in_game", "halftime", "stale", "unknown"}
LIVE_DIAGNOSTIC_TIMINGS = {"live", "in_game", "halftime"}


def line_timing_from(row: dict[str, Any]) -> str:
    """Normalize line timing fields from Pick History/workbook rows.

    Historical rows before timing capture are intentionally classified as
    unknown so they are excluded from pregame model calibration rather than
    silently mixed with confirmed pregame picks.
    """
    for key in ("line_timing", "line timing", "timing", "line_timing_status"):
        value = row.get(key)
        text = str(value or "").strip().lower().replace("-", "_").replace(" ", "_")
        if text in LINE_TIMING_VALUES:
            return text
    if parse_bool(row.get("live_line_flag") or row.get("live line flag")):
        return "live"
    if parse_bool(row.get("stale_line_flag") or row.get("stale line flag")):
        return "stale"
    return "unknown"


def pregame_calibration_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Only confirmed pregame rows are eligible for model/tier/CLV calibration."""
    return [r for r in rows if line_timing_from(r) == "pregame"]


def line_timing_counts(rows: list[dict[str, Any]]) -> dict[str, Any]:
    counts = {f"{timing}_count": 0 for timing in ("pregame", "live", "in_game", "halftime", "unknown", "stale")}
    reasons: dict[str, int] = defaultdict(int)
    for r in rows:
        timing = line_timing_from(r)
        counts[f"{timing}_count"] = counts.get(f"{timing}_count", 0) + 1
        if timing != "pregame":
            reason = str(r.get("line_timing_reason") or r.get("line timing reason") or f"{timing} timing excluded from pregame calibration")
            reasons[reason[:240]] += 1
    counts["excluded_from_pregame_calibration_count"] = sum(counts.get(f"{timing}_count", 0) for timing in ("live", "in_game", "halftime", "unknown", "stale"))
    counts["timing_exclusion_reasons"] = dict(sorted(reasons.items(), key=lambda kv: (-kv[1], kv[0])))
    return counts


def signal_flags(row: dict[str, Any]) -> dict[str, bool]:
    text = " ".join(str(row.get(k) or "") for k in ("notes", "pick_ref", "pick", "tags", "flag_type")).lower()
    return {
        "goblin_available": parse_bool(row.get("goblin_available")) or "goblin" in text,
        "demon_available": parse_bool(row.get("demon_available")) or "demon" in text,
        "line_movement_0_5_plus": parse_bool(row.get("favorable_line_move_0_5")) or line_movement_flag(row),
        "standard_over": "over" in text and "under" not in text,
    }


def full_schema_fields_populated(row: dict[str, Any]) -> bool:
    return all(not missing_value(row.get(field)) for field in FULL_SCHEMA_REQUIRED_FIELDS)


def market_context_available(row: dict[str, Any]) -> bool:
    if parse_bool(row.get("market_context_available")):
        return True
    return any(to_float(row.get(field)) is not None for field in MARKET_CONTEXT_FIELDS if field not in {"market_context_available", "market_context_source", "market_movement_direction"})


def market_context_full_schema_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        r for r in rows
        if r.get("tracking_era") == "full_schema_tracking"
        and r.get("result") in VALID_RESULTS
        and market_context_available(r)
    ]


def market_context_group_summary(rows: list[dict[str, Any]], label: str, predicate) -> dict[str, Any]:
    yes = [r for r in rows if predicate(r)]
    no = [r for r in rows if not predicate(r)]
    return {
        f"{label}": summarize(yes),
        f"not_{label}": summarize(no),
        "sample_size_warning": len(yes) < 20 or len(no) < 20,
    }


def market_context_validation(rows: list[dict[str, Any]]) -> dict[str, Any]:
    eligible = [r for r in rows if market_context_available(r) and r.get("result") in VALID_RESULTS]
    by_sport: dict[str, Any] = {}
    for sport in ("NBA", "MLB"):
        sport_rows = [r for r in eligible if str(r.get("sport") or "").upper() == sport]
        sport_full_schema_rows = [r for r in sport_rows if r.get("tracking_era") == "full_schema_tracking"]
        implied_values = [to_float(r.get("team_implied_total")) for r in sport_rows]
        implied_values = [v for v in implied_values if v is not None]
        avg_implied = round(sum(implied_values) / len(implied_values), 3) if implied_values else None
        spread_values = [abs(to_float(r.get("spread")) or 0) for r in sport_rows if to_float(r.get("spread")) is not None]
        avg_abs_spread = round(sum(spread_values) / len(spread_values), 3) if spread_values else None
        rising = [r for r in sport_rows if "total_rising" in str(r.get("market_movement_direction") or "").lower()]
        falling = [r for r in sport_rows if "total_falling" in str(r.get("market_movement_direction") or "").lower()]
        disagreement = [r for r in sport_rows if (to_float(r.get("total_disagreement")) or 0) > 0 or (to_float(r.get("spread_disagreement")) or 0) > 0]
        no_disagreement = [r for r in sport_rows if r not in disagreement]
        by_sport[sport] = {
            "market_context_graded_rows": len(sport_rows),
            "full_schema_market_context_graded_rows": len(sport_full_schema_rows),
            "average_team_implied_total": avg_implied,
            "average_absolute_spread": avg_abs_spread,
            "team_implied_total_above_average": summarize([r for r in sport_rows if avg_implied is not None and (to_float(r.get("team_implied_total")) or -999999) > avg_implied]),
            "team_implied_total_below_or_equal_average": summarize([r for r in sport_rows if avg_implied is not None and to_float(r.get("team_implied_total")) is not None and (to_float(r.get("team_implied_total")) or 0) <= avg_implied]),
            "games_with_rising_totals": summarize(rising),
            "games_with_falling_totals": summarize(falling),
            "high_spread_games": summarize([r for r in sport_rows if avg_abs_spread is not None and to_float(r.get("spread")) is not None and abs(to_float(r.get("spread")) or 0) > avg_abs_spread]),
            "close_spread_games": summarize([r for r in sport_rows if avg_abs_spread is not None and to_float(r.get("spread")) is not None and abs(to_float(r.get("spread")) or 0) <= avg_abs_spread]),
            "fd_dk_disagreement": summarize(disagreement),
            "no_fd_dk_disagreement": summarize(no_disagreement),
            "movement_data_available_rows": len(rising) + len(falling),
            "recommendation_allowed": len(sport_full_schema_rows) >= MARKET_CONTEXT_MIN_RECOMMENDATION_SAMPLE,
            "recommendation_policy": "Do not adjust projections, confidence tiers, or gates until at least 50 graded full-schema picks with market context exist.",
        }
    total_full_schema_context = len(market_context_full_schema_rows(rows))
    return {
        "policy": "Research-only validation: market context is stored next to PrizePicks props for backtesting and must not affect projections, confidence tiers, approved picks, or gates yet.",
        "minimum_graded_full_schema_sample_for_recommendation": MARKET_CONTEXT_MIN_RECOMMENDATION_SAMPLE,
        "total_market_context_graded_rows": len(eligible),
        "total_full_schema_market_context_rows": total_full_schema_context,
        "recommendation_allowed": total_full_schema_context >= MARKET_CONTEXT_MIN_RECOMMENDATION_SAMPLE,
        "recommendation": "HOLD — collect at least 50 graded full-schema picks with market context before recommending model/gate/tier adjustments." if total_full_schema_context < MARKET_CONTEXT_MIN_RECOMMENDATION_SAMPLE else "Sample threshold met; review sport-level correlations before any implementation change.",
        "by_sport": by_sport,
    }


def tracking_era_from(row: dict[str, Any]) -> str:
    row_date = row.get("date_obj")
    if isinstance(row_date, date) and row_date >= FULL_SCHEMA_TRACKING_START:
        return "full_schema_tracking"
    if full_schema_fields_populated(row):
        return "full_schema_tracking"
    return "legacy_history"


def full_schema_tracking_active(rows: list[dict[str, Any]], meta: dict[str, Any]) -> bool:
    headers = set(meta.get("headers") or [])
    return all(field in headers for field in FULL_SCHEMA_REQUIRED_FIELDS)


def clv_from(row: dict[str, Any]) -> float | None:
    for key in ("clv", "closing_line_value", "clv_units"):
        if key in row:
            val = to_float(row.get(key))
            if val is not None:
                return val
    notes = str(row.get("notes") or "")
    match = re.search(r"\bCLV\b[:=\s]+([+\-]?\d+(?:\.\d+)?)", notes, re.I)
    return float(match.group(1)) if match else None


def load_pick_history(path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    if not path.exists():
        raise FileNotFoundError(f"Missing workbook: {path}")
    wb = safe_load_workbook(path, data_only=True)
    if "Pick History" not in wb.sheetnames:
        raise ValueError(f"Workbook has no Pick History sheet: {path}")
    ws = wb["Pick History"]
    headers = [norm_header(c.value) for c in ws[1]]
    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None and v != "" for v in values):
            continue
        raw = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers)) if headers[i]}
        result = canonical_result(raw.get("result"))
        units = to_float(raw.get("units") or raw.get("units_bet") or raw.get("stake")) or 0.0
        pnl = to_float(raw.get("pnl") or raw.get("profit_loss") or raw.get("profit/loss")) or 0.0
        row_date = parse_date(raw.get("date"))
        row = dict(raw)
        row.update({
            "date_obj": row_date,
            "date": row_date.isoformat() if row_date else raw.get("date"),
            "sport": str(raw.get("sport") or "Unknown").upper(),
            "result": result,
            "units": units,
            "pnl": pnl,
            "pick_type": pick_type_from(raw),
            "confidence_tier": confidence_tier_from(raw),
            "platform": platform_from(raw),
            "edge_tags": edge_tags_from(raw),
            "clv": clv_from(raw),
            "line_timing": line_timing_from(raw),
            **signal_flags(raw),
        })
        row["tracking_era"] = tracking_era_from(row)
        rows.append(row)
    meta = {"sheet_rows": ws.max_row - 1, "headers": headers, "workbook": str(path)}
    return rows, meta


def load_optional_sheet(path: Path, sheet_name: str) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    if not path.exists():
        return [], {"exists": False, "headers": [], "rows": 0}
    wb = safe_load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return [], {"exists": False, "headers": [], "rows": 0}
    ws = wb[sheet_name]
    headers_raw = [c.value for c in ws[1]]
    headers = [str(h or "").strip() for h in headers_raw]
    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None and v != "" for v in values):
            continue
        rows.append({headers[i]: values[i] if i < len(values) else None for i in range(len(headers)) if headers[i]})
    return rows, {"exists": True, "headers": headers, "rows": len(rows)}


def slip_summary(rows: list[dict[str, Any]]) -> dict[str, Any]:
    base = summarize_slip_history_rows(rows)
    def group(key: str) -> dict[str, Any]:
        out: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for r in rows:
            out[str(r.get(key) or "Unknown")].append(r)
        return {k: summarize_slip_history_rows(v) for k, v in sorted(out.items())}
    base.update({
        "by_slip_type": group("Slip Type"),
        "by_number_of_legs": group("Number of Legs"),
        "by_platform": group("Platform"),
        "unreconciled_special_line_count": sum(1 for r in rows if r.get("Needs Payout Reconciliation") and (r.get("Contains Demon") or r.get("Contains Goblin"))),
        "standard_exact_slips": sum(1 for r in rows if not r.get("Contains Demon") and not r.get("Contains Goblin") and not r.get("Needs Payout Reconciliation")),
        "exact_manual_reconciled_special_slips": sum(1 for r in rows if (r.get("Contains Demon") or r.get("Contains Goblin")) and str(r.get("Payout Confidence") or "") == "exact_manual"),
    })
    return base


def graded(rows: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    return [r for r in rows if r.get("result") in VALID_RESULTS]


def summarize(rows: list[dict[str, Any]]) -> dict[str, Any]:
    total = len(rows)
    wins = sum(1 for r in rows if r["result"] == "WIN")
    losses = sum(1 for r in rows if r["result"] == "LOSS")
    pushes = sum(1 for r in rows if r["result"] == "PUSH")
    units = sum(float(r.get("units") or 0) for r in rows)
    pnl = sum(float(r.get("pnl") or 0) for r in rows)
    clvs = [float(r["clv"]) for r in rows if r.get("clv") is not None]
    decisions = wins + losses
    win_rate = wins / decisions * 100 if decisions else None
    roi = pnl / units * 100 if units else None
    return {
        "total_picks": total,
        "wins": wins,
        "losses": losses,
        "pushes": pushes,
        "win_rate_pct": round(win_rate, 2) if win_rate is not None else None,
        "total_units_wagered": round(units, 3),
        "total_profit_loss": round(pnl, 3),
        "roi_pct": round(roi, 2) if roi is not None else None,
        "average_clv": round(sum(clvs) / len(clvs), 3) if clvs else None,
        "profitable": pnl > 0,
        "sample_size_warning": total < 20,
    }


def group_summary(rows: list[dict[str, Any]], key: str, expected: list[str] | None = None) -> dict[str, dict[str, Any]]:
    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for r in rows:
        value = str(r.get(key) or "").strip()
        if not value or value.lower() == "unknown":
            continue
        groups[value].append(r)
    if expected:
        for item in expected:
            groups.setdefault(item, [])
    return {name: summarize(items) for name, items in sorted(groups.items())}


def edge_type_summary(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for r in rows:
        for tag in r.get("edge_tags") or []:
            groups[tag].append(r)
    for tag in EDGE_TAGS:
        groups.setdefault(tag, [])
    return {name: summarize(items) for name, items in sorted(groups.items())}


def missing_value(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def data_quality_metrics(rows: list[dict[str, Any]]) -> dict[str, Any]:
    total = len(rows)
    fields = {
        "tier": lambda r: missing_value(r.get("confidence_tier")),
        "clv": lambda r: r.get("clv") is None,
        "platform": lambda r: missing_value(r.get("platform")),
        "edge_tags": lambda r: not bool(r.get("edge_tags")),
        "line": lambda r: to_float(r.get("line")) is None,
        "model_over_probability": lambda r: to_float(r.get("model_over_probability")) is None,
        "ev": lambda r: to_float(r.get("ev")) is None,
        "odds": lambda r: missing_value(r.get("odds")),
    }
    out: dict[str, Any] = {"total_graded_picks": total}
    for name, predicate in fields.items():
        missing = sum(1 for r in rows if predicate(r))
        out[f"missing_{name}_count"] = missing
        out[f"missing_{name}_pct"] = round((missing / total * 100), 2) if total else None
    unknown_timing = sum(1 for r in rows if line_timing_from(r) == "unknown")
    stale_timing = sum(1 for r in rows if line_timing_from(r) == "stale")
    out["unknown_timing_count"] = unknown_timing
    out["unknown_timing_pct"] = round((unknown_timing / total * 100), 2) if total else None
    out["stale_timing_count"] = stale_timing
    out["stale_timing_pct"] = round((stale_timing / total * 100), 2) if total else None
    missing_market_context = sum(1 for r in rows if not market_context_available(r))
    out["missing_market_context_count"] = missing_market_context
    out["missing_market_context_pct"] = round((missing_market_context / total * 100), 2) if total else None
    return out


def rolling_rows(rows: list[dict[str, Any]], report_date: date) -> list[dict[str, Any]]:
    start = report_date - timedelta(days=30)
    return [r for r in rows if r.get("date_obj") and start <= r["date_obj"] <= report_date]


def phi_correlation_binary(xs: list[int], ys: list[int]) -> float | None:
    if len(xs) < 2 or len(xs) != len(ys):
        return None
    n = len(xs)
    mx = sum(xs) / n
    my = sum(ys) / n
    cov = sum((x - mx) * (y - my) for x, y in zip(xs, ys))
    vx = sum((x - mx) ** 2 for x in xs)
    vy = sum((y - my) ** 2 for y in ys)
    if vx == 0 or vy == 0:
        return None
    return cov / math.sqrt(vx * vy)


def signal_validation(rows: list[dict[str, Any]]) -> dict[str, Any]:
    movement = [r for r in rows if r.get("line_movement_0_5_plus")]
    no_movement = [r for r in rows if not r.get("line_movement_0_5_plus")]
    goblin = [r for r in rows if r.get("goblin_available")]
    goblin_standard_over = [r for r in goblin if r.get("standard_over")]
    demon = [r for r in rows if r.get("demon_available")]
    demon_clvs = [float(r["clv"]) for r in demon if r.get("clv") is not None]
    return {
        "line_movement_0_5_plus": summarize(movement),
        "without_line_movement": summarize(no_movement),
        "line_movement_predictive": (
            (summarize(movement).get("win_rate_pct") or 0) > (summarize(no_movement).get("win_rate_pct") or 0)
            if movement and no_movement else None
        ),
        "goblin_available_all": summarize(goblin),
        "goblin_available_standard_over_only": summarize(goblin_standard_over),
        "goblin_standard_over_won_more_often": (
            (summarize(goblin_standard_over).get("win_rate_pct") or 0) > 50 if goblin_standard_over else None
        ),
        "demon_available_all": summarize(demon),
        "demon_average_clv": round(sum(demon_clvs) / len(demon_clvs), 3) if demon_clvs else None,
        "demon_correlates_with_line_value": (sum(1 for x in demon_clvs if x > 0) / len(demon_clvs) > 0.5) if demon_clvs else None,
    }


def ev_analysis(rows: list[dict[str, Any]]) -> dict[str, Any]:
    with_ev = [r for r in rows if to_float(r.get("ev")) is not None]
    positive = [r for r in with_ev if (to_float(r.get("ev")) or 0) > 0]
    nonpositive = [r for r in with_ev if (to_float(r.get("ev")) or 0) <= 0]
    return {
        "with_ev_count": len(with_ev),
        "positive_ev": summarize(positive),
        "nonpositive_ev": summarize(nonpositive),
    }


def clv_calibration(rows: list[dict[str, Any]]) -> dict[str, Any]:
    with_clv = [r for r in rows if r.get("clv") is not None and r.get("result") in {"WIN", "LOSS"}]
    pos = [r for r in with_clv if float(r["clv"]) > 0]
    nonpos = [r for r in with_clv if float(r["clv"]) <= 0]
    xs = [1 if float(r["clv"]) > 0 else 0 for r in with_clv]
    ys = [1 if r["result"] == "WIN" else 0 for r in with_clv]
    pos_summary = summarize(pos)
    return {
        "with_clv_count": len(with_clv),
        "positive_clv": pos_summary,
        "nonpositive_clv": summarize(nonpos),
        "positive_clv_win_correlation_phi": round(phi_correlation_binary(xs, ys), 4) if phi_correlation_binary(xs, ys) is not None else None,
        "calibration_issue": bool(pos and (pos_summary.get("win_rate_pct") or 0) < 52),
    }


def recommendation_items(payload: dict[str, Any]) -> list[dict[str, Any]]:
    recs: list[dict[str, Any]] = []
    recommendation_base = payload.get("backtest_era", {}).get("full_schema_tracking", payload["all_time"])
    recommendation_sections = [
        ("tier", recommendation_base["by_confidence_tier"]),
        ("pick_type", recommendation_base["by_pick_type"]),
        ("edge_type", {k: v for k, v in recommendation_base["by_edge_type"].items() if k != "Untagged"}),
    ]
    for section_name, section in recommendation_sections:
        for name, metrics in section.items():
            n = metrics.get("total_picks", 0)
            roi = metrics.get("roi_pct")
            if roi is None or n < 20:
                continue
            if roi < -5:
                recs.append({"severity": "DROP_REVIEW", "recommendation": f"CONSIDER DROPPING: {section_name} {name}", "basis": metrics})
            if roi > 10:
                recs.append({"severity": "HIGH_PERFORMER", "recommendation": f"HIGH PERFORMING: {section_name} {name} — increase sizing", "basis": metrics})
    untagged_missing = payload.get("data_quality", {}).get("full_schema_tracking", {}).get("missing_edge_tags_count", 0)
    if untagged_missing >= 20:
        recs.append({"severity": "DATA_QUALITY", "recommendation": "DATA QUALITY ISSUE: edge type tags missing in full_schema_tracking rows; cannot fully evaluate edge performance.", "basis": payload.get("data_quality", {}).get("full_schema_tracking", {})})
    dq = payload.get("data_quality", {}).get("full_schema_tracking", {})
    if dq.get("missing_tier_count", 0):
        recs.append({"severity": "DATA_QUALITY", "recommendation": "DATA QUALITY ISSUE: confidence tier missing in full_schema_tracking rows; cannot evaluate tier performance for those rows.", "basis": {"missing_tier_count": dq.get("missing_tier_count"), "missing_tier_pct": dq.get("missing_tier_pct")}})
    if dq.get("missing_platform_count", 0):
        recs.append({"severity": "DATA_QUALITY", "recommendation": "DATA QUALITY ISSUE: platform missing in full_schema_tracking rows; cannot fully evaluate platform performance.", "basis": {"missing_platform_count": dq.get("missing_platform_count"), "missing_platform_pct": dq.get("missing_platform_pct")}})
    if dq.get("missing_clv_count", 0):
        recs.append({"severity": "DATA_QUALITY", "recommendation": "DATA QUALITY ISSUE: CLV missing in full_schema_tracking rows; cannot fully evaluate CLV calibration.", "basis": {"missing_clv_count": dq.get("missing_clv_count"), "missing_clv_pct": dq.get("missing_clv_pct")}})
    clv = recommendation_base["clv_calibration"]
    pos = clv.get("positive_clv", {})
    if pos.get("total_picks", 0) and (pos.get("win_rate_pct") or 0) < 52:
        recs.append({"severity": "CALIBRATION", "recommendation": "CLV CALIBRATION ISSUE — model may need adjustment", "basis": clv})
    injury = recommendation_base["by_edge_type"].get("Injury Spot", {})
    if injury.get("total_picks", 0) and (injury.get("win_rate_pct") or 0) < 45:
        recs.append({"severity": "EDGE_REVIEW", "recommendation": "INJURY SPOTS NOT WORKING — review logic", "basis": injury})
    if not recs:
        recs.append({"severity": "INFO", "recommendation": "No keep/drop recommendation met the configured sample-size and ROI thresholds.", "basis": {}})
    return recs


def build_payload(rows: list[dict[str, Any]], meta: dict[str, Any], report_date_text: str, slip_rows: list[dict[str, Any]] | None = None, slip_meta: dict[str, Any] | None = None, conditional_rows: list[dict[str, Any]] | None = None, conditional_meta: dict[str, Any] | None = None) -> dict[str, Any]:
    report_dt = datetime.fromisoformat(report_date_text).date()
    graded_rows = graded(rows)
    slip_rows = slip_rows or []
    conditional_rows = conditional_rows or []
    rolling = rolling_rows(graded_rows, report_dt)
    legacy_rows = [r for r in graded_rows if r.get("tracking_era") == "legacy_history"]
    full_schema_rows = [r for r in graded_rows if r.get("tracking_era") == "full_schema_tracking"]
    calibration_all = pregame_calibration_rows(graded_rows)
    calibration_full_schema = pregame_calibration_rows(full_schema_rows)
    calibration_rolling = pregame_calibration_rows(rolling)

    def bundle(items: list[dict[str, Any]]) -> dict[str, Any]:
        return {
            "summary": summarize(items),
            "by_confidence_tier": group_summary(items, "confidence_tier", ["A", "B", "C"]),
            "by_pick_type": group_summary(items, "pick_type", ["Props", "Spreads", "Totals", "Parlays"]),
            "by_edge_type": edge_type_summary(items),
            "by_sport": group_summary(items, "sport", ["NBA", "MLB"]),
            "by_platform": group_summary(items, "platform", ["PrizePicks", "Dabble", "Underdog"]),
            "signal_validation": signal_validation(items),
            "ev_analysis": ev_analysis(items),
            "clv_calibration": clv_calibration(items),
        }

    payload = {
        "report_date": report_date_text,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source": meta,
        "row_counts": {
            "pick_history_rows": len(rows),
            "graded_rows": len(graded_rows),
            "pending_or_ungraded_rows": len(rows) - len(graded_rows),
            "rolling_30_day_graded_rows": len(rolling),
            "legacy_history_graded_rows": len(legacy_rows),
            "full_schema_tracking_graded_rows": len(full_schema_rows),
        },
        "tracking_era_note": TRACKING_ERA_NOTE,
        "sample_size_limitations": [],
        "data_quality": {
            "all_time": data_quality_metrics(graded_rows),
            "legacy_history": data_quality_metrics(legacy_rows),
            "full_schema_tracking": data_quality_metrics(full_schema_rows),
            "rolling_30_day": data_quality_metrics(rolling),
        },
        "all_time_raw_results": summarize(graded_rows),
        "legacy_history_raw_results": summarize(legacy_rows),
        "full_schema_tracking_raw_results": summarize(full_schema_rows),
        "all_time": bundle(calibration_all),
        "full_schema_tracking": bundle(calibration_full_schema),
        "rolling_30_day": bundle(calibration_rolling),
        "line_timing_calibration": {
            "policy": "Only line_timing=pregame rows are included in normal model/tier/EV/CLV calibration. Live/in_game/halftime rows are diagnostic-only; unknown/stale rows are data-quality exclusions.",
            "all_time": line_timing_counts(graded_rows),
            "full_schema_tracking": line_timing_counts(full_schema_rows),
            "rolling_30_day": line_timing_counts(rolling),
            "pregame_summary": summarize(calibration_all),
            "live_diagnostic_summary": summarize([r for r in graded_rows if line_timing_from(r) in LIVE_DIAGNOSTIC_TIMINGS]),
            "unknown_timing_summary": summarize([r for r in graded_rows if line_timing_from(r) == "unknown"]),
            "stale_timing_summary": summarize([r for r in graded_rows if line_timing_from(r) == "stale"]),
        },
        "backtest_era": {
            "note": TRACKING_ERA_NOTE,
            "definitions": {
                "legacy_history": "Rows before full Pick History schema tracking began.",
                "full_schema_tracking": "Rows on/after 2026-06-09, or rows where Confidence Tier, EV, Line, Model Over Probability, and Platform are populated.",
            },
            "full_schema_tracking_active_for_future_picks": full_schema_tracking_active(rows, meta),
            "pick_history_is_leg_level": True,
            "bankroll_should_use_slip_history_when_slip_ids_exist": True,
            "all_time_raw_results": summarize(graded_rows),
            "legacy_history_raw_results": summarize(legacy_rows),
            "full_schema_tracking": bundle(calibration_full_schema),
        },
    }
    payload["slip_level_analysis"] = slip_summary(slip_rows)
    payload["market_context_validation"] = market_context_validation(graded_rows)
    payload["slip_history_status"] = slip_meta or {"exists": False, "rows": 0}
    payload["conditional_specials"] = {"status": conditional_meta or {"exists": False, "rows": 0}, "rows": conditional_rows[:200]}
    payload["sample_size_limitations"].append(TRACKING_ERA_NOTE)
    if len(graded_rows) < 20:
        payload["sample_size_limitations"].append("Fewer than 20 graded all-time picks; recommendations are directional only.")
    if not any(r.get("clv") is not None for r in graded_rows):
        payload["sample_size_limitations"].append("No CLV values found in Pick History; CLV calibration and demon line-value analysis are unavailable.")
    if not any(r.get("goblin_available") or r.get("demon_available") for r in graded_rows):
        payload["sample_size_limitations"].append("No demon/goblin tags found in graded Pick History rows; signal analysis is unavailable until those fields are logged.")
    if not any(r.get("line_movement_0_5_plus") for r in graded_rows):
        payload["sample_size_limitations"].append("No 0.5+ favorable line-movement tags found in graded Pick History rows.")
    market_ctx = payload.get("market_context_validation", {})
    if not market_ctx.get("recommendation_allowed"):
        payload["sample_size_limitations"].append("Market context is research-only: fewer than 50 graded full-schema picks with market-context fields; do not adjust projections, confidence tiers, approved picks, or gates yet.")
    payload["recommendations"] = recommendation_items(payload)
    return payload


def flatten_section(section_name: str, data: dict[str, dict[str, Any]], window: str) -> list[list[Any]]:
    rows = []
    for name, metrics in data.items():
        rows.append([window, section_name, name, metrics.get("total_picks"), metrics.get("wins"), metrics.get("losses"), metrics.get("pushes"), metrics.get("win_rate_pct"), metrics.get("total_units_wagered"), metrics.get("total_profit_loss"), metrics.get("roi_pct"), metrics.get("average_clv"), metrics.get("profitable"), metrics.get("sample_size_warning")])
    return rows


def write_sheet(ws, headers: list[str], rows: list[list[Any]]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for row in rows:
        ws.append(row)
    for col in range(1, len(headers) + 1):
        width = min(50, max(12, max((len(str(ws.cell(r, col).value or "")) for r in range(1, ws.max_row + 1)), default=12) + 2))
        ws.column_dimensions[get_column_letter(col)].width = width


def write_report(payload: dict[str, Any], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    summary_rows = []
    for window in ("all_time", "full_schema_tracking", "rolling_30_day"):
        b = payload[window]
        summary_rows.append([window, "overall", "ALL", *[b["summary"].get(k) for k in ("total_picks", "wins", "losses", "pushes", "win_rate_pct", "total_units_wagered", "total_profit_loss", "roi_pct", "average_clv", "profitable", "sample_size_warning")]])
        summary_rows += flatten_section("confidence_tier", b["by_confidence_tier"], window)
        summary_rows += flatten_section("pick_type", b["by_pick_type"], window)
        summary_rows += flatten_section("sport", b["by_sport"], window)
        summary_rows += flatten_section("platform", b["by_platform"], window)
    headers = ["Window", "Group", "Name", "Total Picks", "Wins", "Losses", "Pushes", "Win Rate %", "Units Wagered", "Profit/Loss", "ROI %", "Avg CLV", "Profitable", "Sample Warning"]
    write_sheet(ws, headers, summary_rows)

    ws = wb.create_sheet("Backtest Era")
    era = payload.get("backtest_era", {})
    era_rows = [
        ["note", "", era.get("note") or payload.get("tracking_era_note"), None, None, None, None, None, None, None, None, None],
        ["full_schema_tracking_active_for_future_picks", "", era.get("full_schema_tracking_active_for_future_picks"), None, None, None, None, None, None, None, None, None],
    ]
    for name, metrics in (
        ("all_time_raw_results", payload.get("all_time_raw_results", {})),
        ("legacy_history_raw_results", payload.get("legacy_history_raw_results", {})),
        ("full_schema_tracking_raw_results", payload.get("full_schema_tracking_raw_results", {})),
    ):
        era_rows.append([name, "raw_results", "ALL", *[metrics.get(k) for k in ("total_picks", "wins", "losses", "pushes", "win_rate_pct", "total_units_wagered", "total_profit_loss", "roi_pct", "average_clv")]])
    full_schema = era.get("full_schema_tracking", {})
    if full_schema:
        era_rows += flatten_section("confidence_tier", full_schema.get("by_confidence_tier", {}), "full_schema_tracking")
        era_rows += flatten_section("platform", full_schema.get("by_platform", {}), "full_schema_tracking")
        era_rows += flatten_section("edge_type", full_schema.get("by_edge_type", {}), "full_schema_tracking")
    write_sheet(ws, ["Era", "Group", "Name/Value", "Total Picks", "Wins", "Losses", "Pushes", "Win Rate %", "Units Wagered", "Profit/Loss", "ROI %", "Avg CLV", "Profitable", "Sample Warning"], era_rows)

    ws = wb.create_sheet("Edge Type Analysis")
    rows = []
    for window in ("all_time", "full_schema_tracking", "rolling_30_day"):
        rows += flatten_section("edge_type", payload[window]["by_edge_type"], window)
    write_sheet(ws, headers, rows)

    ws = wb.create_sheet("Signal Validation")
    signal_rows = []
    for window in ("all_time", "full_schema_tracking", "rolling_30_day"):
        sig = payload[window]["signal_validation"]
        for name, metrics in sig.items():
            if isinstance(metrics, dict):
                signal_rows.append([window, name, metrics.get("total_picks"), metrics.get("wins"), metrics.get("losses"), metrics.get("pushes"), metrics.get("win_rate_pct"), metrics.get("roi_pct"), metrics.get("average_clv")])
            else:
                signal_rows.append([window, name, metrics, None, None, None, None, None, None])
    write_sheet(ws, ["Window", "Signal", "Total/Value", "Wins", "Losses", "Pushes", "Win Rate %", "ROI %", "Avg CLV"], signal_rows)

    ws = wb.create_sheet("EV Analysis")
    ev_rows = []
    for window in ("all_time", "full_schema_tracking", "rolling_30_day"):
        ev = payload[window]["ev_analysis"]
        ev_rows.append([window, "with_ev_count", ev.get("with_ev_count"), None, None, None, None])
        for name in ("positive_ev", "nonpositive_ev"):
            m = ev.get(name, {})
            ev_rows.append([window, name, m.get("total_picks"), m.get("wins"), m.get("losses"), m.get("win_rate_pct"), m.get("roi_pct")])
    write_sheet(ws, ["Window", "Metric", "Value/Total", "Wins", "Losses", "Win Rate %", "ROI %"], ev_rows)

    ws = wb.create_sheet("CLV Calibration")
    clv_rows = []
    for window in ("all_time", "full_schema_tracking", "rolling_30_day"):
        clv = payload[window]["clv_calibration"]
        clv_rows.append([window, "with_clv_count", clv.get("with_clv_count"), None, None, None, None])
        clv_rows.append([window, "positive_clv_win_correlation_phi", clv.get("positive_clv_win_correlation_phi"), None, None, None, None])
        clv_rows.append([window, "calibration_issue", clv.get("calibration_issue"), None, None, None, None])
        for name in ("positive_clv", "nonpositive_clv"):
            m = clv.get(name, {})
            clv_rows.append([window, name, m.get("total_picks"), m.get("wins"), m.get("losses"), m.get("win_rate_pct"), m.get("roi_pct")])
    write_sheet(ws, ["Window", "Metric", "Value/Total", "Wins", "Losses", "Win Rate %", "ROI %"], clv_rows)

    ws = wb.create_sheet("Market Context")
    market_payload = payload.get("market_context_validation", {})
    market_rows = [["ALL", "policy", market_payload.get("policy"), None, None, None, None, None, None]]
    market_rows.append(["ALL", "minimum_sample_for_recommendation", market_payload.get("minimum_graded_full_schema_sample_for_recommendation"), None, None, None, None, None, None])
    market_rows.append(["ALL", "total_full_schema_market_context_rows", market_payload.get("total_full_schema_market_context_rows"), None, None, None, None, None, None])
    market_rows.append(["ALL", "recommendation", market_payload.get("recommendation"), None, None, None, None, None, None])
    for sport, sport_payload in (market_payload.get("by_sport") or {}).items():
        market_rows.append([sport, "market_context_graded_rows", sport_payload.get("market_context_graded_rows"), None, None, None, None, None, None])
        market_rows.append([sport, "full_schema_market_context_graded_rows", sport_payload.get("full_schema_market_context_graded_rows"), None, None, None, None, None, None])
        market_rows.append([sport, "average_team_implied_total", sport_payload.get("average_team_implied_total"), None, None, None, None, None, None])
        market_rows.append([sport, "average_absolute_spread", sport_payload.get("average_absolute_spread"), None, None, None, None, None, None])
        for name in ("team_implied_total_above_average", "team_implied_total_below_or_equal_average", "games_with_rising_totals", "games_with_falling_totals", "high_spread_games", "close_spread_games", "fd_dk_disagreement", "no_fd_dk_disagreement"):
            m = sport_payload.get(name, {})
            market_rows.append([sport, name, m.get("total_picks"), m.get("wins"), m.get("losses"), m.get("pushes"), m.get("win_rate_pct"), m.get("roi_pct"), m.get("average_clv")])
    write_sheet(ws, ["Sport", "Segment", "Total/Value", "Wins", "Losses", "Pushes", "Win Rate %", "ROI %", "Avg CLV"], market_rows)

    ws = wb.create_sheet("Line Timing Calibration")
    timing_payload = payload.get("line_timing_calibration", {})
    timing_rows = [["policy", "", timing_payload.get("policy"), None, None, None, None, None, None, None]]
    for window in ("all_time", "full_schema_tracking", "rolling_30_day"):
        counts = timing_payload.get(window, {})
        timing_rows.append([window, "counts", "pregame", counts.get("pregame_count"), None, None, None, None, None, None])
        timing_rows.append([window, "counts", "live", counts.get("live_count"), None, None, None, None, None, None])
        timing_rows.append([window, "counts", "in_game", counts.get("in_game_count"), None, None, None, None, None, None])
        timing_rows.append([window, "counts", "halftime", counts.get("halftime_count"), None, None, None, None, None, None])
        timing_rows.append([window, "counts", "unknown", counts.get("unknown_count"), None, None, None, None, None, None])
        timing_rows.append([window, "counts", "stale", counts.get("stale_count"), None, None, None, None, None, None])
        timing_rows.append([window, "excluded_from_pregame_calibration", "ALL_NON_PREGAME", counts.get("excluded_from_pregame_calibration_count"), None, None, None, None, None, None])
        for reason, count in (counts.get("timing_exclusion_reasons") or {}).items():
            timing_rows.append([window, "timing_exclusion_reason", reason, count, None, None, None, None, None, None])
    for name in ("pregame_summary", "live_diagnostic_summary", "unknown_timing_summary", "stale_timing_summary"):
        m = timing_payload.get(name, {})
        timing_rows.append(["all_time", name, "summary", m.get("total_picks"), m.get("wins"), m.get("losses"), m.get("pushes"), m.get("win_rate_pct"), m.get("total_profit_loss"), m.get("roi_pct")])
    write_sheet(ws, ["Window", "Section", "Name/Reason", "Count/Total", "Wins", "Losses", "Pushes", "Win Rate %", "Profit/Loss", "ROI %"], timing_rows)

    ws = wb.create_sheet("Slip Level")
    slip = payload.get("slip_level_analysis", {})
    slip_rows = [["overall", "ALL", slip.get("slips"), slip.get("exact_slips"), slip.get("unreconciled_slips"), slip.get("exact_units"), slip.get("exact_net_pnl"), slip.get("exact_roi_pct")]]
    for group_name in ("by_slip_type", "by_number_of_legs", "by_platform"):
        for name, metrics in (slip.get(group_name) or {}).items():
            slip_rows.append([group_name, name, metrics.get("slips"), metrics.get("exact_slips"), metrics.get("unreconciled_slips"), metrics.get("exact_units"), metrics.get("exact_net_pnl"), metrics.get("exact_roi_pct")])
    write_sheet(ws, ["Group", "Name", "Slips", "Exact Slips", "Unreconciled Slips", "Exact Units", "Exact Net PnL", "Exact ROI %"], slip_rows)

    ws = wb.create_sheet("Conditional Specials")
    cond_rows = []
    for r in payload.get("conditional_specials", {}).get("rows", []):
        cond_rows.append([r.get("Date"), r.get("Sport"), r.get("Emoji"), r.get("Line Type"), r.get("Player"), r.get("Stat"), r.get("Special Line"), r.get("Probability"), r.get("Break-even Multiplier"), r.get("Required Use Multiplier"), r.get("Conditional Instruction"), r.get("Status"), r.get("Actual Multiplier"), r.get("Multiplier Known")])
    write_sheet(ws, ["Date", "Sport", "Emoji", "Line Type", "Player", "Stat", "Line", "Probability", "Break-even", "Required", "Instruction", "Status", "Actual Multiplier", "Known"], cond_rows)

    ws = wb.create_sheet("Data Quality")
    dq_rows = []
    for window in ("all_time", "full_schema_tracking", "rolling_30_day"):
        dq = payload.get("data_quality", {}).get(window, {})
        total = dq.get("total_graded_picks")
        for field in ("tier", "clv", "platform", "edge_tags", "line", "model_over_probability", "ev", "odds"):
            dq_rows.append([window, field, dq.get(f"missing_{field}_count"), dq.get(f"missing_{field}_pct"), total])
        dq_rows.append([window, "unknown_timing", dq.get("unknown_timing_count"), dq.get("unknown_timing_pct"), total])
        dq_rows.append([window, "stale_timing", dq.get("stale_timing_count"), dq.get("stale_timing_pct"), total])
        dq_rows.append([window, "market_context", dq.get("missing_market_context_count"), dq.get("missing_market_context_pct"), total])
    write_sheet(ws, ["Window", "Field", "Missing Count", "Missing %", "Total Graded Picks"], dq_rows)

    ws = wb.create_sheet("Recommendations")
    rec_rows = [[r.get("severity"), r.get("recommendation"), json.dumps(r.get("basis", {}), sort_keys=True)] for r in payload.get("recommendations", [])]
    for warning in payload.get("sample_size_limitations", []):
        rec_rows.append(["SAMPLE_LIMIT", warning, ""])
    write_sheet(ws, ["Severity", "Recommendation", "Basis"], rec_rows)
    safe_save_workbook(wb, path)


def main() -> int:
    args = parse_args()
    report_date_text = resolve_date(args.date)
    input_path = Path(args.input)
    rows, meta = load_pick_history(input_path)
    slip_rows, slip_meta = load_optional_sheet(input_path, "Slip History")
    conditional_rows, conditional_meta = load_optional_sheet(input_path, "Conditional Specials")
    payload = build_payload(rows, meta, report_date_text, slip_rows, slip_meta, conditional_rows, conditional_meta)
    json_path = Path(args.json_output) if args.json_output else JSON_DIR / f"backtest_{report_date_text}.json"
    json_path.parent.mkdir(parents=True, exist_ok=True)
    json_path.write_text(json.dumps(payload, indent=2, sort_keys=True, default=str) + "\n")
    market_context_path = JSON_DIR / f"market_context_backtest_{report_date_text}.json"
    market_context_path.write_text(json.dumps(payload.get("market_context_validation", {}), indent=2, sort_keys=True, default=str) + "\n")
    xlsx_path = Path(args.xlsx_output)
    write_report(payload, xlsx_path)
    print(json.dumps({
        "status": "ok",
        "date": report_date_text,
        "json": str(json_path),
        "market_context_json": str(market_context_path),
        "xlsx": str(xlsx_path),
        "row_counts": payload["row_counts"],
        "slip_history_status": payload.get("slip_history_status"),
        "slip_level_analysis": payload.get("slip_level_analysis"),
        "line_timing_calibration": payload.get("line_timing_calibration", {}),
        "recommendation_count": len(payload.get("recommendations", [])),
        "sample_size_limitations": payload.get("sample_size_limitations", []),
    }, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
