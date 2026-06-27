#!/usr/bin/env python3
"""Build DFS slips from SportsEdge projections and correlations.

Slips are restricted to the gauntlet-vetted universe (APPROVED picks plus Gate-8
exposure/concentration-cap-held picks from the day's workbook), partitioned by
platform so every slip carries its real platform and never mixes Underdog with
PrizePicks legs.
"""
from __future__ import annotations

import argparse
import json
import math
import os
import re
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any

import analyze_prop_correlation as corr
from slip_payouts import load_payout_config, payout_multiplier
from stake_sizing import apply_confidence_stakes

ROOT = Path(__file__).resolve().parents[1]
SLIP_DIR = ROOT / "data" / "research" / "slips"
DATA_DIR = ROOT / "data"
CALIBRATION_PATH = DATA_DIR / "research" / "calibration.json"
HERMES_ENV = Path.home() / ".hermes" / ".env"

# --- EV-based slip-type reasoning engine (all behavior flag-gated, default OFF) ---
# Eligibility floor and shrink anchor: the model's per-leg over_probability must
# already exceed 0.5238 to be eligible (is_eligible), so it is the natural anchor
# the calibration shrink pulls overconfident probabilities back toward.
EV_SHRINK_ANCHOR = 0.5238
# Default conservative cushion: power is only chosen when P'_all clears the
# break-even by this fraction, so model overconfidence cannot flip a true-negative
# slip into an apparent-positive one. Operator-tunable via EV_SLIP_MARGIN.
EV_MARGIN_DEFAULT = 0.15
# Calibration ratio used when a sport is uncalibrated (n_with_mop < N_GATE). It is
# intentionally pessimistic (heavy shrink) AND marks the leg untrustworthy.
EV_UNCALIBRATED_RATIO = 1.30
# Minimum graded sample before a sport's empirical calibration ratio is trusted.
EV_CALIBRATION_N_GATE = 30
# A leg with fewer than this many historical samples is untrustworthy for power.
EV_MIN_LEG_SAMPLE = 8


def _env_value(key: str) -> str | None:
    """Read a config value from process env, then ~/.hermes/.env (KEY=VALUE).

    Self-contained mirror of sports_system_runner.env_value; the runner is NOT
    imported here because importing it triggers heavy module-load side effects.
    """
    value = os.environ.get(key)
    if value:
        return value.strip().strip('"').strip("'")
    try:
        if not HERMES_ENV.exists():
            return None
        for line in HERMES_ENV.read_text().splitlines():
            stripped = line.strip()
            if not stripped or stripped.startswith("#") or "=" not in stripped:
                continue
            k, v = stripped.split("=", 1)
            if k.strip() == key:
                return v.strip().strip('"').strip("'") or None
    except Exception:
        return None
    return None


def _env_bool(key: str, default: bool) -> bool:
    raw = _env_value(key)
    if raw is None:
        return default
    return str(raw).strip().lower() in {"1", "true", "yes", "on", "enabled"}


def _ev_slip_type_enabled() -> bool:
    """Master flag gating the EV chooser AND leg-count expansion (default OFF)."""
    return _env_bool("ENABLE_EV_SLIP_TYPE", False)


def _calibrated_probabilities_enabled() -> bool:
    """Source-level probability calibration stopgap (default OFF).

    When ON, generate_projections already shrank each leg's over_probability at the
    projection source, so calibration_ratio must NOT shrink a calibrated sport's legs
    a second time (no double-shrink). Flag OFF => slip math is byte-identical to today."""
    return _env_bool("USE_CALIBRATED_PROBABILITIES", False)


def _ev_margin() -> float:
    raw = _env_value("EV_SLIP_MARGIN")
    if raw is None:
        return EV_MARGIN_DEFAULT
    try:
        val = float(raw)
        return val if val >= 0 else EV_MARGIN_DEFAULT
    except (TypeError, ValueError):
        return EV_MARGIN_DEFAULT


def _load_calibration() -> dict[str, Any]:
    """Load calibration.json; return an empty dict on any failure (then every
    sport reads as uncalibrated -> untrustworthy -> conservative flex)."""
    try:
        if CALIBRATION_PATH.exists():
            return json.loads(CALIBRATION_PATH.read_text())
    except Exception:
        pass
    return {}


def calibration_ratio(sport: Any, calibration: dict[str, Any]) -> tuple[float, bool]:
    """Return (ratio, trusted) for a sport.

    ratio = the latest "computed" audit raw_ratio for that sport when the sport's
    fingerprint n_with_mop >= EV_CALIBRATION_N_GATE; otherwise EV_UNCALIBRATED_RATIO
    and trusted=False. raw_ratio is model_implied/empirical (>1 == overconfident).
    """
    sport_key = str(sport or "").upper()
    # Coordination with the source-level probability calibration stopgap: when
    # USE_CALIBRATED_PROBABILITIES is on AND this sport's over_probabilities were
    # already shrunk at projection time (shrink factor < 1.0), the legs are
    # pre-calibrated — return an identity ratio (trusted) so shrink_probability is a
    # no-op and we never double-shrink. Sports NOT calibrated at source (e.g.
    # gate-not-met NBA) fall through to the existing uncalibrated logic below. Lazy,
    # guarded import: any failure degrades to today's behavior.
    if _calibrated_probabilities_enabled():
        try:
            from calibration import probability_shrink_factor_from_cfg
            if probability_shrink_factor_from_cfg(calibration or {}, sport_key) < 1.0:
                return 1.0, True
        except Exception:
            pass
    fingerprints = (calibration or {}).get("fingerprints", {}) or {}
    n_with_mop = 0
    fp = fingerprints.get(sport_key)
    if isinstance(fp, dict):
        try:
            n_with_mop = int(fp.get("n_with_mop") or 0)
        except (TypeError, ValueError):
            n_with_mop = 0
    if n_with_mop < EV_CALIBRATION_N_GATE:
        return EV_UNCALIBRATED_RATIO, False
    ratio: float | None = None
    for entry in (calibration or {}).get("audit", []) or []:
        if not isinstance(entry, dict):
            continue
        if str(entry.get("sport") or "").upper() != sport_key:
            continue
        if entry.get("reason") == "computed" and entry.get("raw_ratio") is not None:
            try:
                ratio = float(entry.get("raw_ratio"))
            except (TypeError, ValueError):
                continue
    if ratio is None or ratio <= 0:
        return EV_UNCALIBRATED_RATIO, False
    return ratio, True


def shrink_probability(p: float, ratio: float) -> float:
    """Pull an overconfident probability back toward the eligibility anchor.

    p' = anchor + (p - anchor) / ratio. ratio>1 shrinks toward the anchor;
    ratio==1 is the identity. Result is clamped to [0, 1].
    """
    if ratio is None or ratio <= 0:
        ratio = 1.0
    shrunk = EV_SHRINK_ANCHOR + (float(p) - EV_SHRINK_ANCHOR) / ratio
    return max(0.0, min(1.0, shrunk))


def poisson_binomial(probs: list[float]) -> list[float]:
    """Distribution of the number of successes over independent Bernoulli legs.

    Returns a list of length len(probs)+1 where index k is P(exactly k hits).
    O(n^2) DP; n<=6 here.
    """
    dist = [1.0]
    for p in probs:
        p = max(0.0, min(1.0, float(p)))
        nxt = [0.0] * (len(dist) + 1)
        for k, prob in enumerate(dist):
            nxt[k] += prob * (1.0 - p)
            nxt[k + 1] += prob * p
        dist = nxt
    return dist


def ev_power(probs: list[float], full_win_multiplier: float) -> float:
    """Expected gross-return multiple of a power slip: P(all hit) * mult."""
    p_all = 1.0
    for p in probs:
        p_all *= max(0.0, min(1.0, float(p)))
    return p_all * float(full_win_multiplier)


def ev_flex(probs: list[float], win_table: dict[int, float]) -> float:
    """Expected gross-return multiple of a flex slip via Poisson-binomial.

    win_table maps winning-leg-count -> gross multiplier. EV = sum_k P(k) * mult[k].
    """
    dist = poisson_binomial(probs)
    ev = 0.0
    for k, mult in win_table.items():
        if 0 <= int(k) < len(dist):
            ev += dist[int(k)] * float(mult)
    return ev

SAFE_FLAGS = {"BAD MATCHUP"}
KAT_NAMES = {"karl-anthony towns", "karl anthony towns"}

# Sports whose workbooks may hold vetted picks.
VETTED_SPORTS = ("nba", "mlb")
# Gate-8 cap-held picks are vetted: they passed every quality gate and were held
# back only by the bankroll exposure/concentration cap. MISSING-EV / MISSING-PROB
# Gate-8 rows are quality hard-stops and are NOT in this set.
GATE8_VETTED_MARKERS = (
    "GATE 8 — DYNAMIC EXPOSURE CAP",
    "GATE 8 — CONCENTRATION CAP",
)
# Tokens dropped when comparing a projection's stat_type to a vetted pick's text.
STAT_STOPWORDS = {
    "over", "under", "inn", "1st", "2nd", "3rd", "pitcher", "hitter",
    "total", "allowed", "plus", "home",
}


def _strip_accents(text: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFKD", text) if not unicodedata.combining(c))


def canonical_name(name: Any) -> str:
    """Casefold, strip accents, drop punctuation, collapse whitespace."""
    cleaned = _strip_accents(str(name or "")).casefold()
    cleaned = re.sub(r"[^a-z0-9 ]", " ", cleaned)
    return " ".join(cleaned.split())


def stat_token_set(text: Any) -> set[str]:
    """Normalize a stat label or pick fragment to comparable tokens."""
    normalized = _strip_accents(str(text or "")).lower()
    normalized = normalized.replace("+", " ").replace("_", " ").replace("&", " ")
    return {w for w in re.findall(r"[a-z0-9]+", normalized) if w not in STAT_STOPWORDS}


def pick_stat_tokens(pick_text: Any, player: Any) -> set[str]:
    """Tokens describing the stat in a vetted pick, minus player name and numbers."""
    player_words = set(canonical_name(player).split())
    tokens = stat_token_set(pick_text)
    return {t for t in tokens if t not in player_words and not re.fullmatch(r"[0-9]+", t)}


def _to_float(value: Any) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def workbook_path(sport: str, date: str) -> Path:
    return DATA_DIR / sport.lower() / f"{sport.lower()}_{date}.xlsx"


def load_vetted_keys(date: str) -> dict[str, list[dict[str, Any]]] | None:
    """Collect vetted picks (APPROVED + Gate-8 cap-held) from each sport workbook.

    Returns a mapping sport-upper -> list of {player, line, platform, pick_text}.
    Returns None when NO workbook exists for the date (caller falls back to the
    historical is_eligible behavior for backfill).
    """
    try:
        import openpyxl
    except ImportError:
        return None

    found_any = False
    vetted: dict[str, list[dict[str, Any]]] = {}
    for sport in VETTED_SPORTS:
        path = workbook_path(sport, date)
        if not path.exists():
            continue
        found_any = True
        sport_key = sport.upper()
        entries = vetted.setdefault(sport_key, [])
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception:
            continue
        try:
            _collect_approved(wb, sport_key, entries)
            _collect_gate8(wb, sport_key, entries)
        finally:
            wb.close()

    if not found_any:
        return None
    return vetted


def _header_index(row: tuple[Any, ...]) -> dict[str, int]:
    return {str(h): i for i, h in enumerate(row) if h is not None}


def _collect_approved(wb: Any, sport_key: str, entries: list[dict[str, Any]]) -> None:
    if "Picks" not in wb.sheetnames:
        return
    rows = list(wb["Picks"].iter_rows(values_only=True))
    if not rows:
        return
    idx = _header_index(rows[0])
    for col in ("Status", "Selection", "Player/Team", "Line", "Platform"):
        if col not in idx:
            return
    for row in rows[1:]:
        if str(row[idx["Status"]] or "").strip().upper() != "APPROVED":
            continue
        line = _to_float(row[idx["Line"]])
        if line is None:
            continue
        entries.append({
            "player": row[idx["Player/Team"]],
            "line": line,
            "platform": str(row[idx["Platform"]] or "").strip(),
            "pick_text": str(row[idx["Selection"]] or ""),
        })


def _collect_gate8(wb: Any, sport_key: str, entries: list[dict[str, Any]]) -> None:
    if "Skipped Picks" not in wb.sheetnames:
        return
    rows = list(wb["Skipped Picks"].iter_rows(values_only=True))
    if not rows:
        return
    idx = _header_index(rows[0])
    for col in ("Gate Failed", "Pick", "Player/Team", "Line", "Platform"):
        if col not in idx:
            return
    for row in rows[1:]:
        gate = str(row[idx["Gate Failed"]] or "")
        if not any(marker in gate for marker in GATE8_VETTED_MARKERS):
            continue
        line = _to_float(row[idx["Line"]])
        if line is None:
            continue
        entries.append({
            "player": row[idx["Player/Team"]],
            "line": line,
            "platform": str(row[idx["Platform"]] or "").strip(),
            "pick_text": str(row[idx["Pick"]] or ""),
        })


def filter_to_vetted(projections: list[dict[str, Any]], vetted: dict[str, list[dict[str, Any]]]) -> list[dict[str, Any]]:
    """Keep only projections that positively match a vetted pick (fail-safe).

    A projection is kept iff some vetted pick of the same sport shares its
    (canonical player, line, platform) AND the projection's stat tokens are a
    subset of the pick's stat tokens. When several projections at the same key
    match a single pick (e.g. bare "hits" vs "hits runs rbis"), only the
    projection whose token set is the largest subset of the pick is kept — so a
    bare component prop is NOT vetted by a combo pick. Anything not positively
    matched is excluded.
    """
    # Index vetted picks by (sport, canonical player, line, platform).
    pick_index: dict[tuple[str, str, float, str], list[set[str]]] = {}
    for sport_key, picks in vetted.items():
        for pick in picks:
            line = _to_float(pick.get("line"))
            if line is None:
                continue
            key = (
                sport_key.upper(),
                canonical_name(pick.get("player")),
                round(line, 4),
                str(pick.get("platform") or "").strip(),
            )
            tokens = pick_stat_tokens(pick.get("pick_text"), pick.get("player"))
            pick_index.setdefault(key, []).append(tokens)

    # Group projections by their key so we can resolve the best subset match.
    by_key: dict[tuple[str, str, float, str], list[dict[str, Any]]] = {}
    for proj in projections:
        line = _to_float(proj.get("pp_line"))
        if line is None:
            continue
        key = (
            str(proj.get("sport") or "").upper(),
            canonical_name(proj.get("player_name")),
            round(line, 4),
            str(proj.get("platform") or "").strip(),
        )
        by_key.setdefault(key, []).append(proj)

    kept: list[dict[str, Any]] = []
    for key, projs in by_key.items():
        pick_token_sets = pick_index.get(key)
        if not pick_token_sets:
            continue
        for pick_tokens in pick_token_sets:
            # Projections whose stat tokens are a (non-empty) subset of this pick.
            subset_matches = [
                (proj, stat_token_set(proj.get("stat_type")))
                for proj in projs
            ]
            subset_matches = [
                (proj, toks) for proj, toks in subset_matches
                if toks and toks <= pick_tokens
            ]
            if not subset_matches:
                continue
            best_len = max(len(toks) for _, toks in subset_matches)
            for proj, toks in subset_matches:
                if len(toks) == best_len and proj not in kept:
                    kept.append(proj)
    return kept


def resolve_date(value: str | None) -> str:
    return corr.resolve_date(value)


def load_correlations(date: str) -> dict[str, Any]:
    path = SLIP_DIR / f"prop_correlations_{date}.json"
    if not path.exists():
        projections = corr.load_all_projections(date)
        payload = corr.analyze(projections, date)
        corr.write_output(payload, date)
        return payload
    return json.loads(path.read_text())


def projection_key(row: dict[str, Any]) -> str:
    enriched = dict(row)
    enriched.setdefault("sport", str(row.get("sport") or "").upper())
    return row.get("prop_id") or corr.prop_id(enriched)


def is_eligible(prop: dict[str, Any], allow_skip: bool = False) -> bool:
    if str(prop.get("line_timing") or "pregame").lower() != "pregame":
        return False
    if prop.get("diagnostic_only"):
        return False
    if not allow_skip and str(prop.get("confidence_tier")) == "SKIP":
        return False
    if float(prop.get("edge") or 0) <= 0:
        return False
    if float(prop.get("expected_value") or 0) <= 0:
        return False
    if float(prop.get("over_probability") or 0) <= 0.5238:
        return False
    if int(prop.get("sample_size") or 0) < 2:
        return False
    flags = {str(f) for f in prop.get("flags") or []}
    if flags - SAFE_FLAGS:
        return False
    return True


def score_safety(prop: dict[str, Any]) -> float:
    p = float(prop.get("over_probability") or 0)
    ev = float(prop.get("expected_value") or 0)
    edge = float(prop.get("edge") or 0)
    sample = min(int(prop.get("sample_size") or 0), 20) / 20
    tier_bonus = {"A": 0.08, "B": 0.04, "C": 0.01}.get(str(prop.get("confidence_tier")), 0)
    return p * 2.0 + ev * 0.8 + min(edge / 10.0, 1.0) * 0.4 + sample * 0.2 + tier_bonus


def score_ev(prop: dict[str, Any]) -> float:
    return float(prop.get("expected_value") or 0) * 2 + float(prop.get("edge") or 0) / 10 + float(prop.get("over_probability") or 0)


def correlation_lookup(correlation_payload: dict[str, Any]) -> dict[frozenset[str], dict[str, Any]]:
    out: dict[frozenset[str], dict[str, Any]] = {}
    for pair in correlation_payload.get("pairs", []):
        out[frozenset([pair["prop_a"], pair["prop_b"]])] = pair
    return out


def get_pair(pair_map: dict[frozenset[str], dict[str, Any]], a: dict[str, Any], b: dict[str, Any]) -> dict[str, Any] | None:
    return pair_map.get(frozenset([projection_key(a), projection_key(b)]))


def pair_label(pair_map: dict[frozenset[str], dict[str, Any]], a: dict[str, Any], b: dict[str, Any]) -> str:
    pair = get_pair(pair_map, a, b)
    return str(pair.get("correlation_label")) if pair else "unknown correlation"


def has_bad_pair(legs: list[dict[str, Any]], pair_map: dict[frozenset[str], dict[str, Any]], conservative: bool = False) -> bool:
    seen = set()
    for leg in legs:
        key = projection_key(leg)
        if key in seen:
            return True
        seen.add(key)
    for i, a in enumerate(legs):
        for b in legs[i + 1:]:
            label = pair_label(pair_map, a, b)
            if label == "negative/risky correlation":
                return True
            if conservative and (label in {"strong positive correlation", "moderate positive correlation"} or a.get("player_name") == b.get("player_name")):
                return True
    return False


def combined_probability_details(legs: list[dict[str, Any]], pair_map: dict[frozenset[str], dict[str, Any]], correlated: bool = False) -> dict[str, Any]:
    """Return a documented combined probability estimate.

    Independent slips use the exact independence assumption: product(p_i).
    Correlated slips do not have a real joint distribution model here, so they are
    explicitly marked approximate. Positive correlation moves the product toward,
    but never above, the weakest individual leg probability. Negative/risky
    correlation reduces the product.
    """
    probs = [float(leg.get("over_probability") or 0) for leg in legs]
    product = 1.0
    for prob in probs:
        product *= prob
    weakest = min(probs) if probs else 0.0
    labels: list[str] = []
    strong_pos = 0
    moderate_pos = 0
    negative = 0
    weak_or_unknown = 0
    for i, a in enumerate(legs):
        for b in legs[i + 1:]:
            label = pair_label(pair_map, a, b)
            labels.append(label)
            if label == "strong positive correlation":
                strong_pos += 1
            elif label == "moderate positive correlation":
                moderate_pos += 1
            elif label == "negative/risky correlation":
                negative += 1
            else:
                weak_or_unknown += 1

    is_approximate = correlated or strong_pos > 0 or moderate_pos > 0 or negative > 0
    adjusted = product
    formula = "product(p_i)"
    method = "exact_independent_product"
    note = "Exact only under the independent-leg assumption."

    if strong_pos or moderate_pos:
        # Heuristic: move product toward the weakest leg by a bounded correlation
        # factor. This keeps P(A and B) <= min(P(A), P(B)) for positive correlation.
        rho = min(strong_pos * 0.35 + moderate_pos * 0.20, 0.75)
        adjusted = product + rho * max(0.0, weakest - product)
        formula = "product(p_i) + rho * (min(p_i) - product(p_i)); rho=min(0.35*strong_pairs + 0.20*moderate_pairs, 0.75)"
        method = "approximate_positive_correlation_adjustment"
        note = "Approximate correlation estimate, not a real joint probability model; capped at weakest individual leg."
    if negative:
        adjusted *= max(0.40, 1 - 0.25 * negative)
        formula += "; then * max(0.40, 1 - 0.25*negative_pairs)"
        method = "approximate_negative_correlation_adjustment" if not (strong_pos or moderate_pos) else method
        note = "Approximate correlation estimate with negative/risky-pair penalty, not a real joint probability model."

    adjusted = round(max(0.0, min(adjusted, weakest if (strong_pos or moderate_pos) else 0.99)), 4)
    return {
        "combined_probability": adjusted,
        "independent_probability_product": round(product, 4),
        "combined_probability_method": method,
        "combined_probability_formula": formula,
        "combined_probability_is_exact": not is_approximate,
        "combined_probability_is_approximate": is_approximate,
        "combined_probability_note": note,
        "correlation_pair_labels": labels,
    }


def combined_probability(legs: list[dict[str, Any]], pair_map: dict[frozenset[str], dict[str, Any]], correlated: bool = False) -> float:
    return float(combined_probability_details(legs, pair_map, correlated)["combined_probability"])


def _flex_win_table(platform: str, leg_count: int) -> dict[int, float] | None:
    """Gross multipliers per winning-leg-count for a platform's flex slip.

    Returns None when the platform has no flex table for that leg count (e.g.
    PrizePicks n==2, or any Underdog leg count — Underdog tables are empty).
    """
    table: dict[int, float] = {}
    for k in range(leg_count, 0, -1):
        mult = payout_multiplier(platform, "flex", leg_count, k)
        if mult is not None:
            table[k] = float(mult)
    return table or None


def _leg_trustworthy(leg: dict[str, Any], calibration: dict[str, Any]) -> tuple[bool, float]:
    """A leg is trustworthy for a power recommendation only when its sport is
    calibrated (>= N_GATE graded), its sample size >= EV_MIN_LEG_SAMPLE, and it
    carries a probability. Returns (trustworthy, calibration_ratio_used).
    """
    ratio, sport_trusted = calibration_ratio(leg.get("sport"), calibration)
    prob = _to_float(leg.get("over_probability"))
    try:
        sample = int(leg.get("sample_size") or 0)
    except (TypeError, ValueError):
        sample = 0
    trustworthy = bool(sport_trusted and prob is not None and sample >= EV_MIN_LEG_SAMPLE)
    return trustworthy, ratio


def choose_slip_type(
    legs: list[dict[str, Any]],
    platform: Any,
    leg_count: int,
    combined_probability_details_result: dict[str, Any],
    calibration: dict[str, Any] | None = None,
) -> str:
    """Choose 'power' vs 'flex' by conservative, calibration-shrunk expected value.

    Flag OFF (default): byte-identical to today's mechanical rule —
    "power" if leg_count == 2 else "flex". The name-override semantics in
    make_slip are preserved by make_slip itself, not here.

    Flag ON: pick power ONLY when every guardrail holds (see DECISION RULE).
    Any impossibility or untrusted input falls back to the mechanical result.
    """
    mechanical = "power" if leg_count == 2 else "flex"
    if not _ev_slip_type_enabled():
        return mechanical

    platform_name = platform if platform is not None else "PrizePicks"
    cal = calibration if calibration is not None else _load_calibration()

    full_win_mult = payout_multiplier(platform_name, "power", leg_count, leg_count)
    flex_table = _flex_win_table(platform_name, leg_count)
    # No power table (or no flex alternative) -> EV comparison impossible.
    # PrizePicks n==2 has a power table but NO flex table -> always power
    # (preserves current behavior). Underdog has neither -> mechanical fallback.
    if full_win_mult is None:
        return mechanical
    if flex_table is None:
        # Power exists but flex does not (PrizePicks 2-leg): power is the only option.
        return "power"

    # Per-leg shrunk probabilities + trustworthiness.
    shrunk: list[float] = []
    all_trustworthy = True
    for leg in legs:
        prob = _to_float(leg.get("over_probability"))
        trustworthy, ratio = _leg_trustworthy(leg, cal)
        if not trustworthy:
            all_trustworthy = False
        if prob is None:
            # Missing probability: cannot compute EV safely -> conservative flex.
            return "flex"
        shrunk.append(shrink_probability(prob, ratio))

    ev_power_val = ev_power(shrunk, full_win_mult)
    ev_flex_val = ev_flex(shrunk, flex_table)
    p_all = 1.0
    for p in shrunk:
        p_all *= p
    break_even = 1.0 / float(full_win_mult)
    margin = _ev_margin()

    correlated = bool(combined_probability_details_result.get("combined_probability_is_approximate"))

    pick_power = (
        ev_power_val >= ev_flex_val
        and p_all >= break_even * (1.0 + margin)
        and all_trustworthy
        and not correlated
    )
    return "power" if pick_power else "flex"


def _best_shrunk_ev(
    legs: list[dict[str, Any]],
    platform: Any,
    calibration: dict[str, Any],
) -> float | None:
    """Best calibration-shrunk EV (max of power-EV and flex-EV) for a leg set.

    Returns None when the platform lacks a payout table for that leg count or a
    leg is missing a probability — i.e. the EV is unknown and the set must not be
    used to justify expansion. Power-EV is only counted when the chooser would
    actually allow power for the set (so an untrusted/correlated set cannot use
    the higher power multiplier to look good); flex-EV always counts when a flex
    table exists.
    """
    platform_name = platform if platform is not None else "PrizePicks"
    leg_count = len(legs)
    full_win_mult = payout_multiplier(platform_name, "power", leg_count, leg_count)
    flex_table = _flex_win_table(platform_name, leg_count)
    if full_win_mult is None and flex_table is None:
        return None

    shrunk: list[float] = []
    for leg in legs:
        prob = _to_float(leg.get("over_probability"))
        if prob is None:
            return None
        _trustworthy, ratio = _leg_trustworthy(leg, calibration)
        shrunk.append(shrink_probability(prob, ratio))

    details = combined_probability_details(legs, {}, False)
    recommended = choose_slip_type(legs, platform_name, leg_count, details, calibration=calibration)
    candidates: list[float] = []
    if flex_table is not None:
        candidates.append(ev_flex(shrunk, flex_table))
    if full_win_mult is not None and recommended == "power":
        candidates.append(ev_power(shrunk, full_win_mult))
    if not candidates:
        return None
    return max(candidates)


def _ev_annotations(
    legs: list[dict[str, Any]],
    platform: Any,
    leg_count: int,
    combined_probability_details_result: dict[str, Any],
    calibration: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Compute the additive EV annotation block attached to every slip.

    Always computed (regardless of the flag) for operator visibility. Uses the
    same calibration-shrunk inputs the chooser would use. Never raises.
    """
    platform_name = platform if platform is not None else "PrizePicks"
    cal = calibration if calibration is not None else _load_calibration()
    margin = _ev_margin()

    full_win_mult = payout_multiplier(platform_name, "power", leg_count, leg_count)
    flex_table = _flex_win_table(platform_name, leg_count)

    shrunk: list[float] = []
    all_trustworthy = True
    ratios: dict[str, float] = {}
    for leg in legs:
        prob = _to_float(leg.get("over_probability"))
        trustworthy, ratio = _leg_trustworthy(leg, cal)
        ratios[str(leg.get("sport") or "").upper()] = ratio
        if not trustworthy:
            all_trustworthy = False
        if prob is not None:
            shrunk.append(shrink_probability(prob, ratio))

    ev_power_val: float | None = None
    ev_flex_val: float | None = None
    if shrunk and len(shrunk) == leg_count:
        if full_win_mult is not None:
            ev_power_val = round(ev_power(shrunk, full_win_mult), 6)
        if flex_table is not None:
            ev_flex_val = round(ev_flex(shrunk, flex_table), 6)

    recommended = choose_slip_type(
        legs, platform_name, leg_count, combined_probability_details_result, calibration=cal
    )
    return {
        "ev_power": ev_power_val,
        "ev_flex": ev_flex_val,
        "ev_recommended_type": recommended,
        "ev_margin_used": margin,
        "ev_calibration_ratio": ratios,
        "ev_all_legs_trustworthy": all_trustworthy,
    }


def leg_summary(prop: dict[str, Any]) -> dict[str, Any]:
    return {
        "prop_id": projection_key(prop),
        "sport": prop.get("sport"),
        "player_name": prop.get("player_name"),
        "team": prop.get("team"),
        "platform": prop.get("platform"),
        "stat_type": prop.get("stat_type"),
        "side": "OVER",
        "line": prop.get("pp_line"),
        "projection": prop.get("projection"),
        "edge": prop.get("edge"),
        "over_probability": prop.get("over_probability"),
        "expected_value": prop.get("expected_value"),
        "confidence_tier": prop.get("confidence_tier"),
        "flags": prop.get("flags", []),
    }


def slip_platform(legs: list[dict[str, Any]]) -> str | None:
    """Real platform shared by all legs, or None for unlabeled fixtures."""
    platforms = {leg.get("platform") for leg in legs}
    if len(platforms) == 1:
        return next(iter(platforms))
    return None


def make_slip(category: str, name: str, legs: list[dict[str, Any]], pair_map: dict[frozenset[str], dict[str, Any]], correlated: bool = False, explanation: str = "") -> dict[str, Any]:
    # Defensive dedup: never emit a slip with the same exact prop twice. Distinct
    # stats for one player (KAT/correlated categories) are preserved; only an
    # exact (player, stat, line) repeat is dropped.
    deduped: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    for leg in legs:
        key = projection_key(leg)
        if key in seen_ids:
            continue
        seen_ids.add(key)
        deduped.append(leg)
    legs = deduped
    probability = combined_probability_details(legs, pair_map, correlated)
    leg_count = len(legs)
    # Real platform shared by all legs. Unlabeled fixtures (platform=None) are
    # treated as a single legacy group whose payout table is PrizePicks.
    platform = slip_platform(legs)
    payout_platform = platform if platform is not None else "PrizePicks"
    # Slip type: when ENABLE_EV_SLIP_TYPE is OFF (default), choose_slip_type returns
    # the byte-identical mechanical result ("power" if leg_count==2 else "flex").
    # When ON, it returns the conservative EV-based recommendation.
    slip_type = choose_slip_type(legs, payout_platform, leg_count, probability)
    # Name-override is preserved unchanged: an explicit "power" in the slip name
    # forces a power slip in either flag state (operator-named power categories).
    if "power" in name.lower():
        slip_type = "power"
    # Additive EV annotations (always present, for operator visibility). These are
    # NEW keys only; they never rename/remove existing keys and never change the
    # slip_type when the flag is OFF.
    ev_annotations = _ev_annotations(legs, payout_platform, leg_count, probability)
    # Underdog payout tables are intentionally empty; payout_multiplier returns
    # None there. Keep None (rendered as "n/a"); never substitute another
    # platform's numbers — a wrong payout mis-states real money.
    standard_payout_multiplier = payout_multiplier(payout_platform, slip_type, leg_count, leg_count)
    return {
        "category": category,
        "name": name,
        "platform": platform if platform is not None else "PrizePicks",
        "slip_type": slip_type,
        "stake_units": 1.0,
        "is_correlated": correlated,
        "legs": [leg_summary(x) for x in legs],
        "leg_count": leg_count,
        "standard_payout_multiplier_if_perfect": standard_payout_multiplier,
        **probability,
        **ev_annotations,
        "combined_ev_score": round(sum(float(x.get("expected_value") or 0) for x in legs), 4),
        "explanation": explanation,
    }


def first_valid_combo(candidates: list[dict[str, Any]], n: int, pair_map: dict[frozenset[str], dict[str, Any]], conservative: bool = False) -> list[dict[str, Any]]:
    import itertools
    for combo in itertools.combinations(candidates, n):
        legs = list(combo)
        if not has_bad_pair(legs, pair_map, conservative=conservative):
            return legs
    return []


SLIP_CATEGORIES = ["safest_2_leg", "safest_3_leg", "highest_ev", "correlated_upside", "diversified", "kat_based", "ev_extended"]

# Higher-leg expansion (behind ENABLE_EV_SLIP_TYPE) considers only PrizePicks and
# caps the candidate pool before itertools.combinations to bound cost under the
# 120s build_slips subprocess timeout.
EV_EXTENDED_PLATFORMS = {"PrizePicks"}
EV_EXTENDED_POOL_CAP = 12


def platform_groups(eligible: list[dict[str, Any]]) -> list[tuple[Any, list[dict[str, Any]]]]:
    """Partition eligible props by their real platform.

    Each group's legs all share one platform, so any slip built within a group is
    single-platform by construction. Unlabeled fixtures (platform=None) collapse
    into one legacy group, preserving historical behavior for the test suite.
    """
    groups: dict[Any, list[dict[str, Any]]] = {}
    order: list[Any] = []
    for prop in eligible:
        key = prop.get("platform")
        if key not in groups:
            groups[key] = []
            order.append(key)
    for prop in eligible:
        groups[prop.get("platform")].append(prop)
    return [(key, groups[key]) for key in order]


def _build_ev_extended_slips(
    eligible: list[dict[str, Any]],
    pair_map: dict[frozenset[str], dict[str, Any]],
) -> list[dict[str, Any]]:
    """EV-justified higher-leg (n=4..6) PrizePicks slips, behind ENABLE_EV_SLIP_TYPE.

    Emits a higher-leg slip only when its best calibration-shrunk EV exceeds the
    best 2/3-leg shrunk EV by EV_MARGIN. PrizePicks only; hard-capped at the max
    leg count present in the PrizePicks payout table (6). The candidate pool is
    capped by score before itertools.combinations to bound cost.
    """
    if not _ev_slip_type_enabled():
        return []
    # Single-platform group by construction; expansion is PrizePicks only.
    platform = slip_platform(eligible)
    payout_platform = platform if platform is not None else "PrizePicks"
    if payout_platform not in EV_EXTENDED_PLATFORMS:
        return []

    # Hard cap = max leg count present in the platform's payout table.
    power_cfg = (load_payout_config().get(payout_platform, {}) or {}).get("power", {}) or {}
    flex_cfg = (load_payout_config().get(payout_platform, {}) or {}).get("flex", {}) or {}
    leg_keys = [int(k) for k in list(power_cfg.keys()) + list(flex_cfg.keys()) if str(k).isdigit()]
    if not leg_keys:
        return []
    max_legs = min(6, max(leg_keys))
    if max_legs < 4:
        return []

    calibration = _load_calibration()

    # Cap the candidate pool by EV-score to bound the combinatorial cost.
    pool = sorted(eligible, key=score_ev, reverse=True)[:EV_EXTENDED_POOL_CAP]

    # Baseline: best valid 2- and 3-leg shrunk EV from the same pool.
    baseline_best: float | None = None
    for n in (2, 3):
        legs = first_valid_combo(pool, n, pair_map, conservative=True)
        if not legs:
            continue
        ev = _best_shrunk_ev(legs, payout_platform, calibration)
        if ev is not None and (baseline_best is None or ev > baseline_best):
            baseline_best = ev
    if baseline_best is None:
        return []

    import itertools

    margin = _ev_margin()
    threshold = baseline_best * (1.0 + margin)
    slips: list[dict[str, Any]] = []
    for n in range(4, max_legs + 1):
        if len(pool) < n:
            break
        best_legs: list[dict[str, Any]] | None = None
        best_ev: float | None = None
        for combo in itertools.combinations(pool, n):
            legs = list(combo)
            if has_bad_pair(legs, pair_map, conservative=True):
                continue
            ev = _best_shrunk_ev(legs, payout_platform, calibration)
            if ev is None:
                continue
            if best_ev is None or ev > best_ev:
                best_ev = ev
                best_legs = legs
        if best_legs is not None and best_ev is not None and best_ev > threshold:
            slips.append(make_slip(
                "ev_extended",
                f"EV-justified {n}-leg",
                best_legs,
                pair_map,
                False,
                f"EV-justified {n}-leg: calibration-shrunk EV {best_ev:.3f} exceeds best 2/3-leg EV {baseline_best:.3f} by margin {margin:.2f}.",
            ))
    return slips


def _build_category_slips(eligible: list[dict[str, Any]], correlation_payload: dict[str, Any], pair_map: dict[frozenset[str], dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    """Run the per-category combo logic over one (already single-platform) pool."""
    by_safety = sorted(eligible, key=score_safety, reverse=True)
    by_ev = sorted(eligible, key=score_ev, reverse=True)
    slips: dict[str, list[dict[str, Any]]] = {k: [] for k in SLIP_CATEGORIES}

    legs = first_valid_combo(by_safety, 2, pair_map, conservative=True)
    if legs:
        slips["safest_2_leg"].append(make_slip("safest_2_leg", "Safest 2-leg", legs, pair_map, False, "Highest model probability legs with positive EV and no strong overlap."))
    legs = first_valid_combo(by_safety, 3, pair_map, conservative=True)
    if legs:
        slips["safest_3_leg"].append(make_slip("safest_3_leg", "Safest 3-leg", legs, pair_map, False, "Three high-probability independent legs."))
    legs = first_valid_combo(by_ev, 2, pair_map, conservative=True)
    if legs:
        slips["highest_ev"].append(make_slip("highest_ev", "Highest EV 2-leg", legs, pair_map, False, "Highest EV independent legs that pass audit constraints."))
    legs = first_valid_combo(by_ev, 3, pair_map, conservative=True)
    if legs:
        slips["highest_ev"].append(make_slip("highest_ev", "Highest EV 3-leg", legs, pair_map, False, "Top EV independent 3-leg combination without strong overlap."))

    # Diversified: require unique players and avoid strong/moderate correlation.
    diversified = first_valid_combo(by_safety, 3, pair_map, conservative=True)
    if diversified:
        slips["diversified"].append(make_slip("diversified", "Diversified 3-leg", diversified, pair_map, False, "Avoids same-player overlap and strong positive correlation."))

    # Correlated upside: prefer explicitly positive pairs (within this platform).
    # NOTE: prop_id excludes platform, so the same player/stat/line on two
    # platforms shares a prop_id. The correlation file pairs those as a "self
    # pair" (prop_a == prop_b). Skipping equal ids — and requiring two distinct
    # objects — prevents a slip with two identical legs.
    correlated_pairs: list[tuple[float, dict[str, Any], dict[str, Any], dict[str, Any]]] = []
    by_id = {projection_key(p): p for p in eligible}
    for pair in correlation_payload.get("pairs", []):
        if pair.get("prop_a") == pair.get("prop_b"):
            continue
        if pair.get("correlation_label") in {"strong positive correlation", "moderate positive correlation"}:
            a, b = by_id.get(pair.get("prop_a")), by_id.get(pair.get("prop_b"))
            if a and b and a is not b:
                correlated_pairs.append((score_ev(a) + score_ev(b), pair, a, b))
    for _, pair, a, b in sorted(correlated_pairs, key=lambda x: x[0], reverse=True)[:3]:
        slips["correlated_upside"].append(make_slip("correlated_upside", "Correlated upside pair", [a, b], pair_map, True, f"Labeled correlated: {pair.get('explanation')}"))

    kats = [p for p in eligible if str(p.get("player_name", "")).lower() in KAT_NAMES]
    kats_by_ev = sorted(kats, key=score_ev, reverse=True)
    if len(kats_by_ev) >= 2:
        slips["kat_based"].append(make_slip("kat_based", "Best KAT-only stack", kats_by_ev[:2], pair_map, True, "Special KAT handling: overlapping KAT props intentionally allowed in KAT category."))
    if kats_by_ev:
        anchor = kats_by_ev[0]
        independent = [p for p in by_safety if p is not anchor and p.get("player_name") != anchor.get("player_name") and pair_label(pair_map, anchor, p) not in {"strong positive correlation", "moderate positive correlation", "negative/risky correlation"}]
        if independent:
            slips["kat_based"].append(make_slip("kat_based", "KAT anchor + safest independent leg", [anchor, independent[0]], pair_map, False, "KAT anchor paired with safest independent non-overlap leg."))
        pos = [p for p in by_ev if p is not anchor and pair_label(pair_map, anchor, p) in {"strong positive correlation", "moderate positive correlation"}]
        if pos:
            slips["kat_based"].append(make_slip("kat_based", "KAT anchor + positively correlated Knicks/game leg", [anchor, pos[0]], pair_map, True, "KAT anchor paired with positive correlation."))
        noncorr_ev = [p for p in by_ev if p is not anchor and p.get("player_name") != anchor.get("player_name") and pair_label(pair_map, anchor, p) not in {"strong positive correlation", "moderate positive correlation", "negative/risky correlation"}]
        if noncorr_ev:
            slips["kat_based"].append(make_slip("kat_based", "KAT anchor + highest-EV non-correlated leg", [anchor, noncorr_ev[0]], pair_map, False, "KAT anchor paired with highest-EV non-correlated leg."))

    # EV-justified higher-leg (n=4..6) PrizePicks slips (behind ENABLE_EV_SLIP_TYPE;
    # produces nothing when the flag is OFF -> identical to today).
    slips["ev_extended"].extend(_build_ev_extended_slips(eligible, pair_map))

    return slips


def build_slips(projections: list[dict[str, Any]], correlation_payload: dict[str, Any], date: str, vetted_source: str = "fallback_is_eligible") -> dict[str, Any]:
    for row in projections:
        row["prop_id"] = projection_key(row)
    pair_map = correlation_lookup(correlation_payload)
    eligible = [p for p in projections if is_eligible(p)]

    # Partition the vetted/eligible pool by platform and run the per-category combo
    # logic WITHIN each platform. Every emitted slip is therefore single-platform,
    # carries its real platform, and has dedup'd legs (prop_id is unique per
    # player/stat/line within a platform). A platform that cannot form a >=2-leg
    # combo contributes no slips.
    slips: dict[str, list[dict[str, Any]]] = {k: [] for k in SLIP_CATEGORIES}
    for _platform, group in platform_groups(eligible):
        if len(group) < 2:
            continue
        group_slips = _build_category_slips(group, correlation_payload, pair_map)
        for category, group_list in group_slips.items():
            # Drop any slip that defensive dedup collapsed below 2 legs.
            slips[category].extend(s for s in group_list if s.get("leg_count", 0) >= 2)

    platform_breakdown: dict[str, int] = {}
    for prop in eligible:
        label = str(prop.get("platform") or "unlabeled")
        platform_breakdown[label] = platform_breakdown.get(label, 0) + 1

    avoid_pairing = [p for p in correlation_payload.get("pairs", []) if p.get("correlation_label") == "negative/risky correlation"][:25]
    return {
        "date": date,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "projection_count": len(projections),
        "eligible_count": len(eligible),
        "vetted_source": vetted_source,
        "platform_breakdown": platform_breakdown,
        "slips": slips,
        "avoid_pairing": avoid_pairing,
        "warnings": [] if eligible else ["No eligible positive-EV props available."],
    }


def _payout_display(value: Any) -> str:
    """Underdog tables are empty -> payout None. Render gracefully as n/a."""
    if isinstance(value, (int, float)):
        return f"{value}x"
    return "n/a"


def render_markdown(payload: dict[str, Any]) -> str:
    lines = [f"# SportsEdge Slips {payload['date']}", ""]
    for category, slips in payload["slips"].items():
        lines += [f"## {category}", ""]
        if not slips:
            lines += ["No slip generated.", ""]
            continue
        for slip in slips:
            lines.append(f"### {slip['name']}")
            prob_label = "Approx combined probability" if slip.get("combined_probability_is_approximate") else "Combined probability"
            lines.append(f"Platform: {slip.get('platform')} | {prob_label}: {slip['combined_probability']:.1%} | EV score: {slip['combined_ev_score']:+.2f} | Type: {slip.get('slip_type')} | Perfect payout: {_payout_display(slip.get('standard_payout_multiplier_if_perfect'))} | Correlated: {slip['is_correlated']}")
            if slip.get("combined_probability_is_approximate"):
                lines.append(f"Probability note: {slip.get('combined_probability_note')} Formula: {slip.get('combined_probability_formula')}")
            lines.append(slip.get("explanation") or "")
            for leg in slip["legs"]:
                lines.append(f"- {leg['sport']} {leg['player_name']} {leg['stat_type']} OVER {leg['line']} | Proj {leg['projection']} | P {float(leg['over_probability']):.1%} | EV {float(leg['expected_value']):+.2f} | Tier {leg['confidence_tier']}")
            lines.append("")
    lines += ["## avoid_pairing", ""]
    for pair in payload.get("avoid_pairing", [])[:20]:
        lines.append(f"- {pair.get('player_a')} + {pair.get('player_b')}: {pair.get('explanation')}")
    return "\n".join(lines).rstrip() + "\n"


def write_outputs(payload: dict[str, Any], date: str) -> tuple[Path, Path]:
    SLIP_DIR.mkdir(parents=True, exist_ok=True)
    json_path = SLIP_DIR / f"slips_{date}.json"
    md_path = SLIP_DIR / f"slips_{date}.md"
    json_path.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n")
    md_path.write_text(render_markdown(payload))
    return json_path, md_path


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--date", default="today")
    args = parser.parse_args()
    date = resolve_date(args.date)
    projections = corr.load_all_projections(date)

    # Restrict to the gauntlet-vetted universe when a workbook exists for the date.
    # When no workbook exists (historical backfill), fall back to is_eligible.
    vetted = load_vetted_keys(date)
    if vetted is not None:
        projections = filter_to_vetted(projections, vetted)
        vetted_source = "workbook"
    else:
        vetted_source = "fallback_is_eligible"

    correlation_payload = load_correlations(date)
    payload = build_slips(projections, correlation_payload, date, vetted_source=vetted_source)

    # D-04: read start-of-day bankroll; fall back to stake_units=1.0 on any failure
    # (never crash, never stake 0 when bankroll is unknown — D-04 literal fallback).
    bankroll_path = ROOT / "data" / "pnl" / "bankroll.json"
    start_of_day_bankroll: float | None = None
    try:
        if bankroll_path.exists():
            bk = json.loads(bankroll_path.read_text(encoding="utf-8"))
            raw = float(bk.get("current_bankroll") or 0)
            if raw > 0:
                start_of_day_bankroll = raw
            else:
                print("[build_slips] WARNING: bankroll.json current_bankroll <= 0; using fallback stake_units=1.0")
        else:
            print("[build_slips] WARNING: bankroll.json not found; using fallback stake_units=1.0")
    except Exception as exc:
        print(f"[build_slips] WARNING: bankroll.json read failed ({exc}); using fallback stake_units=1.0")

    if start_of_day_bankroll is not None:
        # D-01: apply stakes per category — payload["slips"] is dict[str, list], NOT a flat list
        staked_slips: dict[str, list[dict[str, Any]]] = {}
        for category, slip_list in payload["slips"].items():
            staked_slips[category] = apply_confidence_stakes(slip_list, start_of_day_bankroll)
        payload["slips"] = staked_slips
    # else: leave stake_units=1.0 as set by make_slip (D-04 literal fallback —
    # do NOT call apply_confidence_stakes(slips, 1.0), which would yield 0.025)

    json_path, md_path = write_outputs(payload, date)
    print(json.dumps({
        "status": "ok",
        "date": date,
        "json": str(json_path),
        "markdown": str(md_path),
        "vetted_source": vetted_source,
        "eligible_count": payload["eligible_count"],
        "platform_breakdown": payload.get("platform_breakdown", {}),
        "slip_counts": {k: len(v) for k, v in payload["slips"].items()},
        "avoid_pairing_count": len(payload.get("avoid_pairing", [])),
    }, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
