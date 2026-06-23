#!/usr/bin/env python3
"""Tests for D-06: Prop Accuracy refresh on the daily sync_slip_bankroll path.

TestPropAccuracyRefresh covers:
  - non-dry-run populates the Prop Accuracy sheet (mirrors rebuild_slip_bankroll:5681)
  - dry-run does NOT populate / mutate the Prop Accuracy sheet (Pitfall 3 guard)
"""
from __future__ import annotations

import importlib.util
import json
import sys
import tempfile
import unittest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

# Load runner via importlib to avoid import side effects
MOD_PATH = Path(__file__).with_name("sports_system_runner.py")
spec = importlib.util.spec_from_file_location("sports_system_runner", MOD_PATH)
runner = importlib.util.module_from_spec(spec)  # type: ignore[arg-type]
assert spec and spec.loader
spec.loader.exec_module(runner)  # type: ignore[union-attr]

from openpyxl import Workbook  # noqa: E402
from slip_payouts import SLIP_HISTORY_HEADERS as _SHH  # noqa: E402


# ---------------------------------------------------------------------------
# Fixture helpers
# ---------------------------------------------------------------------------

def _make_sync_wb(
    pick_history_rows: list[list] | None = None,
    slip_history_rows: list[list] | None = None,
) -> Workbook:
    """Build an in-memory master_pnl-style workbook with the sheets sync_slip_bankroll touches.

    Creates: Daily Log, Slip History, Pick History, Bankroll Chart Data, Performance Breakdown.
    Column positions follow the canonical header lists from the runner / slip_payouts.
    """
    wb = Workbook()
    wb.active.title = "Daily Log"

    # Minimal Daily Log headers (sync_slip_bankroll uses positional columns 0..8)
    wb["Daily Log"].append(["Date", "Source", "C3", "C4", "C5", "Units Bet", "Net PnL", "Running Bankroll", "Notes"])

    # Slip History sheet
    sh = wb.create_sheet("Slip History")
    sh.append(_SHH)
    if slip_history_rows:
        for row in slip_history_rows:
            sh.append(row)

    # Pick History sheet — required by refresh_prop_accuracy
    ph = wb.create_sheet("Pick History")
    ph.append(runner.RESULT_HEADERS)
    if pick_history_rows:
        for row in pick_history_rows:
            ph.append(row)

    # Bankroll Chart Data (optional; sync checks `if ... in wb.sheetnames`)
    wb.create_sheet("Bankroll Chart Data")
    wb["Bankroll Chart Data"].append(["Date", "Bankroll", "ROI", "Updated At"])

    # Performance Breakdown (optional)
    wb.create_sheet("Performance Breakdown")

    return wb


def _make_ph_row(date: str, sport: str, result: str) -> list:
    """Build a Pick History row with the minimum fields for refresh_prop_accuracy."""
    row: list = [None] * len(runner.RESULT_HEADERS)
    h = {col: idx for idx, col in enumerate(runner.RESULT_HEADERS)}
    row[h["Date"]] = date
    row[h["Sport"]] = sport
    row[h["Result"]] = result
    return row


# ---------------------------------------------------------------------------
# TestPropAccuracyRefresh — D-06 daily path
# ---------------------------------------------------------------------------

class TestPropAccuracyRefresh(unittest.TestCase):
    """D-06: sync_slip_bankroll refreshes Prop Accuracy on the non-dry-run path."""

    _DATE = "2026-06-20"

    def _bankroll_json(self, tmp: Path) -> Path:
        """Write a minimal bankroll.json and return its path."""
        p = tmp / "bankroll.json"
        p.write_text(json.dumps({"starting_bankroll": 100.0, "current_bankroll": 100.0}))
        return p

    def _make_wb_with_picks(self) -> Workbook:
        """Build a workbook with Pick History rows so refresh_prop_accuracy produces rows."""
        pick_rows = [
            _make_ph_row(self._DATE, "NBA", "WIN"),
            _make_ph_row(self._DATE, "NBA", "LOSS"),
            _make_ph_row(self._DATE, "MLB", "WIN"),
        ]
        return _make_sync_wb(pick_history_rows=pick_rows)

    def test_non_dry_run_populates_prop_accuracy(self) -> None:
        """After a non-dry-run sync, Prop Accuracy sheet is created and has data rows."""
        wb = self._make_wb_with_picks()

        # Prop Accuracy sheet should not exist yet (or be header-only if present)
        if "Prop Accuracy" in wb.sheetnames:
            # Clear it so we can verify refresh_prop_accuracy populated it fresh
            del wb["Prop Accuracy"]

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            bankroll_path = self._bankroll_json(tmp_path)
            # Use _master_override=None so save_workbook_atomic is skipped
            runner.sync_slip_bankroll(
                self._DATE,
                dry_run=False,
                _wb_override=wb,
                _master_override=None,
                _bankroll_override=bankroll_path,
            )

        self.assertIn(
            "Prop Accuracy",
            wb.sheetnames,
            "Prop Accuracy sheet must be created by sync_slip_bankroll non-dry-run",
        )
        pa = wb["Prop Accuracy"]
        # Should have header row + at least 1 data row
        self.assertGreater(
            pa.max_row,
            1,
            "Prop Accuracy must have at least one data row after non-dry-run sync with pick history",
        )

    def test_dry_run_does_not_populate_prop_accuracy(self) -> None:
        """dry_run=True must NOT create or mutate the Prop Accuracy sheet (Pitfall 3)."""
        wb = self._make_wb_with_picks()

        # Ensure Prop Accuracy is absent before the call
        if "Prop Accuracy" in wb.sheetnames:
            del wb["Prop Accuracy"]

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            bankroll_path = self._bankroll_json(tmp_path)
            runner.sync_slip_bankroll(
                self._DATE,
                dry_run=True,
                _wb_override=wb,
                _master_override=None,
                _bankroll_override=bankroll_path,
            )

        # On dry_run, Prop Accuracy must NOT have been populated (may or may not exist)
        if "Prop Accuracy" in wb.sheetnames:
            pa = wb["Prop Accuracy"]
            # If it exists it must have only the header row (no data rows written)
            self.assertEqual(
                pa.max_row,
                1,
                "dry_run must NOT write data rows to Prop Accuracy (Pitfall 3 guard)",
            )
        # else: sheet absent — also acceptable (proves no mutation)


if __name__ == "__main__":
    unittest.main()
