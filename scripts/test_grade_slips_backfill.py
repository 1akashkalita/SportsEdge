#!/usr/bin/env python3
"""Offline unittest for grade_slips Wave 3 backfill layer.

Tests (all fully offline, no network, no real data/ workbooks):

  (a) ensure_slip_defs invokes build_slips.py subprocess when the slip file is
      MISSING, and is a no-op (does NOT call subprocess) when the file already
      EXISTS — verifies the idempotent build gate.

  (b) Idempotent multi-date backfill — running backfill_range (or
      grade_slips_for_date) TWICE over a small date set with injected box
      scores and temp workbooks produces the SAME Slip History row count both
      times (no duplicate appends); each Slip ID appears exactly once.

  (c) Slip vs prop separation — SLIPS-04: Slip History rows land only in the
      "Slip History" sheet; a "Results" sheet present in the same workbook is
      never touched.

Run from scripts/:
    python3 test_grade_slips_backfill.py
"""
from __future__ import annotations

import json
import sys
import tempfile
import unittest
import unittest.mock
from pathlib import Path
from typing import Any

# Ensure scripts/ is on sys.path for sibling imports.
_SCRIPTS = Path(__file__).parent
if str(_SCRIPTS) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS))

import openpyxl

from grade_slips import (
    LEG_PENDING,
    backfill_range,
    ensure_slip_defs,
    grade_slips_for_date,
    write_slip_history_rows,
    grade_slip,
    slip_id_for,
    _SLIPS_DIR,
)
from slip_payouts import (
    SLIP_HISTORY_HEADERS,
    ensure_slip_history_sheet,
    load_payout_config,
)


# ---------------------------------------------------------------------------
# Offline fixtures
# ---------------------------------------------------------------------------

_DATE_A = "2026-01-15"
_DATE_B = "2026-01-16"

# NBA flat row: LeBron James — 30 pts / 10 reb / 5 ast
_NBA_LEBRON: dict = {
    "points": 30.0,
    "rebounds": 10.0,
    "assists": 5.0,
}

# MLB batting row: Freddie Freeman — 3 hits / 2 runs / 1 RBI
_MLB_BATTING_FREEMAN: dict = {
    "hits": 3.0,
    "runs": 2.0,
    "rbis": 1.0,
    "homeruns": 0.0,
}

# Injected box scores (offline — no ESPN network call).
_BOX_SCORES: dict[str, dict] = {
    "NBA": {"lebron james": _NBA_LEBRON},
    "MLB": {"freddie freeman": {"batting": _MLB_BATTING_FREEMAN, "pitching": {}}},
}

_CONFIG = load_payout_config()


def _make_slip_json(date: str, category: str = "safest_2_leg") -> dict[str, Any]:
    """Build a minimal slip definition dict (mimics slips_<date>.json)."""
    leg_a = {
        "player_name": "LeBron James",
        "stat_type": "points",
        "line": 25.5,
        "side": "OVER",
        "sport": "NBA",
        "prop_id": f"NBA:LeBron James:points:25.5:{date}",
        "confidence_tier": "A",
        "edge": 4.5,
        "over_probability": 0.75,
    }
    leg_b = {
        "player_name": "LeBron James",
        "stat_type": "rebounds",
        "line": 8.5,
        "side": "OVER",
        "sport": "NBA",
        "prop_id": f"NBA:LeBron James:rebounds:8.5:{date}",
        "confidence_tier": "B",
        "edge": 1.5,
        "over_probability": 0.65,
    }
    return {
        "date": date,
        "slips": {
            category: [
                {
                    "date": date,
                    "category": category,
                    "platform": "PrizePicks",
                    "slip_type": "power",
                    "stake_units": 1.0,
                    "leg_count": 2,
                    "legs": [leg_a, leg_b],
                }
            ]
        },
    }


def _make_slip_file(slip_dir: Path, date: str, category: str = "safest_2_leg") -> Path:
    """Write a slips_<date>.json fixture file into slip_dir and return its path."""
    path = slip_dir / f"slips_{date}.json"
    path.write_text(json.dumps(_make_slip_json(date, category)))
    return path


def _data_rows(ws: Any) -> list[tuple]:
    """Return all data rows (row 2+) from a worksheet as tuples."""
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Include only rows with at least one non-None cell.
        if any(v is not None for v in row):
            rows.append(tuple(row))
    return rows


# ---------------------------------------------------------------------------
# (a) ensure_slip_defs — subprocess gate
# ---------------------------------------------------------------------------

class TestEnsureSlipDefs(unittest.TestCase):
    """ensure_slip_defs invokes the builder for missing files and is a no-op when present."""

    def test_calls_subprocess_when_file_missing(self) -> None:
        """When the slip file is absent, ensure_slip_defs should call subprocess.run."""
        fake_date = "2099-01-01"  # Cannot exist on disk.

        with unittest.mock.patch("grade_slips.subprocess.run") as mock_run:
            # Simulate build_slips.py returning success but NOT creating the file
            # (we only test that subprocess.run is called, not the file creation).
            mock_run.return_value = unittest.mock.Mock(returncode=0, stderr="")

            result = ensure_slip_defs(fake_date)

        mock_run.assert_called_once()
        call_args = mock_run.call_args
        cmd = call_args[0][0]  # The command list.
        self.assertTrue(
            any("build_slips" in str(arg) for arg in cmd),
            f"Expected build_slips.py in command, got {cmd}",
        )
        self.assertIn(fake_date, cmd, "Date must be passed to build_slips.py")

    def test_no_subprocess_when_file_exists(self) -> None:
        """When the slip file already exists, ensure_slip_defs must NOT call subprocess."""
        # Use a date that definitely has a slip file on disk.
        # June 8 was confirmed present in the slip_dir inventory.
        real_date = "2026-06-08"
        real_slip = _SLIPS_DIR / f"slips_{real_date}.json"

        if not real_slip.exists():
            self.skipTest(f"Fixture file {real_slip} not present — cannot test no-op path")

        with unittest.mock.patch("grade_slips.subprocess.run") as mock_run:
            result = ensure_slip_defs(real_date)

        mock_run.assert_not_called()
        self.assertTrue(result, "Should return True (file exists)")

    def test_subprocess_called_once_not_twice_for_same_missing_date(self) -> None:
        """Each call invokes at most one subprocess (no retry loop)."""
        fake_date = "2099-12-31"
        with unittest.mock.patch("grade_slips.subprocess.run") as mock_run:
            mock_run.return_value = unittest.mock.Mock(returncode=1, stderr="no projections")
            ensure_slip_defs(fake_date)
        self.assertEqual(mock_run.call_count, 1, "Exactly one subprocess call per missing date")

    def test_subprocess_timeout_is_handled_gracefully(self) -> None:
        """A TimeoutExpired from subprocess.run must NOT propagate — returns False gracefully."""
        import subprocess as _sp
        fake_date = "2099-02-28"
        with unittest.mock.patch("grade_slips.subprocess.run") as mock_run:
            mock_run.side_effect = _sp.TimeoutExpired(cmd=["build_slips.py"], timeout=120)
            result = ensure_slip_defs(fake_date)
        self.assertFalse(result, "Should return False when subprocess times out")


# ---------------------------------------------------------------------------
# (b) Idempotent multi-date backfill
# ---------------------------------------------------------------------------

class TestIdempotentBackfill(unittest.TestCase):
    """Running grade_slips_for_date or backfill_range twice never adds duplicate rows."""

    def _run_grade_twice_and_compare(
        self,
        slip_dir: Path,
        master_wb_path: Path,
        date: str,
    ) -> tuple[int, int]:
        """
        Grade *date* twice using grade_slips_for_date (with temp workbook paths)
        and return (row_count_after_first, row_count_after_second).

        Patches grade_slips internals to redirect workbook writes to the temp dir.
        """
        import grade_slips as _gs

        # Override the slip dir so grade_slips_for_date reads from our temp fixture.
        original_slips_dir = _gs._SLIPS_DIR
        _gs._SLIPS_DIR = slip_dir
        try:
            # We need to redirect workbook writes.  grade_slips_for_date uses
            # ensure_workbook + safe_load_workbook + save_workbook_atomic from the runner.
            # We intercept master_pnl_workbook and ensure_workbook to use temp paths.
            master_wb = openpyxl.Workbook()
            from slip_payouts import ensure_slip_history_sheet
            ensure_slip_history_sheet(master_wb)
            master_wb.save(str(master_wb_path))

            sport_paths: dict[str, Path] = {}

            def _fake_ensure_workbook(sport: str, dt: str) -> Path:
                key = f"{sport}_{dt}"
                if key not in sport_paths:
                    p = master_wb_path.parent / f"{sport}_{dt}_test.xlsx"
                    wb = openpyxl.Workbook()
                    ensure_slip_history_sheet(wb)
                    wb.save(str(p))
                    sport_paths[key] = p
                return sport_paths[key]

            def _fake_safe_load_workbook(path: Path):  # type: ignore[return]
                return openpyxl.load_workbook(str(path))

            def _fake_save_atomic(wb: Any, path: Path) -> None:
                wb.save(str(path))

            def _fake_master_pnl() -> tuple[Any, Path]:
                wb = openpyxl.load_workbook(str(master_wb_path))
                return wb, master_wb_path

            with (
                unittest.mock.patch.object(_gs, "ensure_workbook", _fake_ensure_workbook),
                unittest.mock.patch.object(_gs, "safe_load_workbook", _fake_safe_load_workbook),
                unittest.mock.patch.object(_gs, "save_workbook_atomic", _fake_save_atomic),
                unittest.mock.patch.object(_gs, "master_pnl_workbook", _fake_master_pnl),
                unittest.mock.patch.object(_gs, "build_date_box_scores", return_value=_BOX_SCORES),
            ):
                # First pass.
                _gs.grade_slips_for_date(date, dry_run=False)
                count_after_first = _count_master_rows(master_wb_path)

                # Second pass (idempotency check).
                _gs.grade_slips_for_date(date, dry_run=False)
                count_after_second = _count_master_rows(master_wb_path)

        finally:
            _gs._SLIPS_DIR = original_slips_dir

        return count_after_first, count_after_second

    def test_second_pass_adds_no_rows(self) -> None:
        """Row count in master Slip History is identical after first and second run."""
        with tempfile.TemporaryDirectory() as tmp:
            slip_dir = Path(tmp) / "slips"
            slip_dir.mkdir()
            _make_slip_file(slip_dir, _DATE_A)

            master_wb_path = Path(tmp) / "master_pnl.xlsx"
            first, second = self._run_grade_twice_and_compare(slip_dir, master_wb_path, _DATE_A)

        self.assertGreater(first, 0, "At least one slip row written on first pass")
        self.assertEqual(first, second, "Second pass must not add duplicate rows")

    def test_each_slip_id_appears_once_after_two_runs(self) -> None:
        """Each unique (Date, Slip ID) pair appears exactly once in Slip History."""
        with tempfile.TemporaryDirectory() as tmp:
            slip_dir = Path(tmp) / "slips"
            slip_dir.mkdir()
            _make_slip_file(slip_dir, _DATE_A, category="safest_2_leg")
            master_wb_path = Path(tmp) / "master_pnl.xlsx"

            self._run_grade_twice_and_compare(slip_dir, master_wb_path, _DATE_A)

            wb = openpyxl.load_workbook(str(master_wb_path))
            ws = wb["Slip History"]
            # Collect (Date, Slip ID) pairs.
            date_col = SLIP_HISTORY_HEADERS.index("Date") + 1
            slip_id_col = SLIP_HISTORY_HEADERS.index("Slip ID") + 1
            seen: list[tuple] = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    continue
                key = (row[date_col - 1], row[slip_id_col - 1])
                seen.append(key)

        slip_id_counts: dict[tuple, int] = {}
        for key in seen:
            slip_id_counts[key] = slip_id_counts.get(key, 0) + 1

        for key, count in slip_id_counts.items():
            self.assertEqual(count, 1, f"Slip ID {key} appears {count} times (expected 1)")

    def test_multi_date_backfill_idempotent(self) -> None:
        """backfill_range over two dates produces no duplicates on a second call."""
        import grade_slips as _gs

        with tempfile.TemporaryDirectory() as tmp:
            slip_dir = Path(tmp) / "slips"
            slip_dir.mkdir()
            _make_slip_file(slip_dir, _DATE_A)
            _make_slip_file(slip_dir, _DATE_B)

            master_wb_path = Path(tmp) / "master_pnl.xlsx"
            master_wb = openpyxl.Workbook()
            ensure_slip_history_sheet(master_wb)
            master_wb.save(str(master_wb_path))

            sport_paths: dict[str, Path] = {}

            def _fake_ensure_workbook(sport: str, dt: str) -> Path:
                key = f"{sport}_{dt}"
                if key not in sport_paths:
                    p = master_wb_path.parent / f"{sport}_{dt}_test.xlsx"
                    wb = openpyxl.Workbook()
                    ensure_slip_history_sheet(wb)
                    wb.save(str(p))
                    sport_paths[key] = p
                return sport_paths[key]

            def _fake_safe_load(path: Path):  # type: ignore
                return openpyxl.load_workbook(str(path))

            def _fake_save(wb: Any, path: Path) -> None:
                wb.save(str(path))

            def _fake_master() -> tuple[Any, Path]:
                return openpyxl.load_workbook(str(master_wb_path)), master_wb_path

            original_slips_dir = _gs._SLIPS_DIR
            _gs._SLIPS_DIR = slip_dir
            try:
                with (
                    unittest.mock.patch.object(_gs, "ensure_workbook", _fake_ensure_workbook),
                    unittest.mock.patch.object(_gs, "safe_load_workbook", _fake_safe_load),
                    unittest.mock.patch.object(_gs, "save_workbook_atomic", _fake_save),
                    unittest.mock.patch.object(_gs, "master_pnl_workbook", _fake_master),
                    unittest.mock.patch.object(_gs, "build_date_box_scores", return_value=_BOX_SCORES),
                    unittest.mock.patch.object(_gs, "ensure_slip_defs", side_effect=lambda d: True),
                ):
                    backfill_range(_DATE_A, _DATE_B, dry_run=False)
                    count_first = _count_master_rows(master_wb_path)

                    backfill_range(_DATE_A, _DATE_B, dry_run=False)
                    count_second = _count_master_rows(master_wb_path)
            finally:
                _gs._SLIPS_DIR = original_slips_dir

        self.assertGreater(count_first, 0, "At least one row after first backfill")
        self.assertEqual(count_first, count_second, "Idempotent: second backfill must not grow row count")


def _count_master_rows(master_wb_path: Path) -> int:
    """Count data rows (row 2+) in the Slip History sheet of master_wb_path."""
    wb = openpyxl.load_workbook(str(master_wb_path))
    if "Slip History" not in wb.sheetnames:
        return 0
    ws = wb["Slip History"]
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            count += 1
    return count


# ---------------------------------------------------------------------------
# (c) Slip vs prop separation — SLIPS-04
# ---------------------------------------------------------------------------

class TestSlipVsPropSeparation(unittest.TestCase):
    """Slip History writes must not affect Results / Pick History sheets (SLIPS-04)."""

    def test_results_sheet_untouched(self) -> None:
        """After writing slip rows, a co-located Results sheet has the same row count."""
        wb = openpyxl.Workbook()
        ws_slip = wb.create_sheet("Slip History")
        # Add the header row to Slip History.
        ws_slip.append(SLIP_HISTORY_HEADERS)

        # Pre-populate a Results sheet with 3 data rows.
        ws_results = wb.create_sheet("Results")
        ws_results.append(["Date", "Player", "Stat", "Result"])
        ws_results.append(["2026-01-15", "LeBron James", "points", "WIN"])
        ws_results.append(["2026-01-15", "Freddie Freeman", "hits", "LOSS"])
        ws_results.append(["2026-01-15", "Shane Bieber", "strikeouts", "WIN"])
        results_row_count_before = sum(
            1 for row in ws_results.iter_rows(min_row=2, values_only=True)
            if any(v is not None for v in row)
        )

        # Build two graded slips and write to Slip History.
        slip_a = {
            "date": _DATE_A,
            "category": "safest_2_leg",
            "platform": "PrizePicks",
            "slip_type": "power",
            "stake_units": 1.0,
            "leg_count": 2,
            "legs": [
                {
                    "player_name": "LeBron James",
                    "stat_type": "points",
                    "line": 25.5,
                    "side": "OVER",
                    "sport": "NBA",
                    "prop_id": "NBA:LeBron James:points:25.5",
                },
                {
                    "player_name": "LeBron James",
                    "stat_type": "rebounds",
                    "line": 8.5,
                    "side": "OVER",
                    "sport": "NBA",
                    "prop_id": "NBA:LeBron James:rebounds:8.5",
                },
            ],
        }
        slip_b = dict(slip_a)
        slip_b["category"] = "highest_ev"

        graded_a = grade_slip(slip_a, _BOX_SCORES, config=_CONFIG)
        graded_b = grade_slip(slip_b, _BOX_SCORES, config=_CONFIG)
        # Fix slip IDs to include date (grade_slip uses slip["date"]).
        graded_a["slip_id"] = slip_id_for(_DATE_A, slip_a)
        graded_b["slip_id"] = slip_id_for(_DATE_A, slip_b)

        rows_written = write_slip_history_rows(ws_slip, _DATE_A, [graded_a, graded_b])

        # Verify: Slip History got new rows.
        self.assertGreater(rows_written, 0, "Should write at least one slip row")
        slip_data_rows = sum(
            1 for row in ws_slip.iter_rows(min_row=2, values_only=True)
            if any(v is not None for v in row)
        )
        self.assertGreater(slip_data_rows, 0, "Slip History must have data rows")

        # Verify: Results sheet is completely unchanged.
        results_row_count_after = sum(
            1 for row in ws_results.iter_rows(min_row=2, values_only=True)
            if any(v is not None for v in row)
        )
        self.assertEqual(
            results_row_count_before,
            results_row_count_after,
            "Results sheet row count must be unchanged after slip write (SLIPS-04)",
        )

    def test_pick_history_sheet_untouched(self) -> None:
        """After slip writes, a Pick History sheet row count is unchanged."""
        wb = openpyxl.Workbook()
        ws_slip = wb.create_sheet("Slip History")
        ws_slip.append(SLIP_HISTORY_HEADERS)

        ws_ph = wb.create_sheet("Pick History")
        ws_ph.append(["Date", "Player", "Pick Ref", "Grade"])
        ws_ph.append(["2026-01-15", "Freddie Freeman", "MLB:Freeman:hits:2.5", "WIN"])
        ws_ph.append(["2026-01-15", "LeBron James", "NBA:James:pts:28.5", "LOSS"])
        ph_count_before = sum(
            1 for row in ws_ph.iter_rows(min_row=2, values_only=True)
            if any(v is not None for v in row)
        )

        slip = {
            "date": _DATE_A,
            "category": "correlated_upside",
            "platform": "PrizePicks",
            "slip_type": "power",
            "stake_units": 1.0,
            "leg_count": 2,
            "legs": [
                {
                    "player_name": "LeBron James",
                    "stat_type": "points",
                    "line": 25.5,
                    "side": "OVER",
                    "sport": "NBA",
                    "prop_id": "NBA:LeBron James:points:25.5:ph_test",
                },
                {
                    "player_name": "LeBron James",
                    "stat_type": "rebounds",
                    "line": 8.5,
                    "side": "OVER",
                    "sport": "NBA",
                    "prop_id": "NBA:LeBron James:rebounds:8.5:ph_test",
                },
            ],
        }
        graded = grade_slip(slip, _BOX_SCORES, config=_CONFIG)
        graded["slip_id"] = slip_id_for(_DATE_A, slip)
        write_slip_history_rows(ws_slip, _DATE_A, [graded])

        ph_count_after = sum(
            1 for row in ws_ph.iter_rows(min_row=2, values_only=True)
            if any(v is not None for v in row)
        )
        self.assertEqual(ph_count_before, ph_count_after, "Pick History must be untouched (SLIPS-04)")

    def test_slip_history_rows_not_in_results_content(self) -> None:
        """Slip History row content (Slip ID, payout, category) is NOT in Results sheet."""
        wb = openpyxl.Workbook()
        ws_slip = wb.create_sheet("Slip History")
        ws_slip.append(SLIP_HISTORY_HEADERS)
        ws_results = wb.create_sheet("Results")
        ws_results.append(["Date", "Player", "Stat", "Result"])

        slip = {
            "date": _DATE_A,
            "category": "safest_2_leg",
            "platform": "PrizePicks",
            "slip_type": "power",
            "stake_units": 1.0,
            "leg_count": 2,
            "legs": [
                {
                    "player_name": "LeBron James",
                    "stat_type": "points",
                    "line": 25.5,
                    "side": "OVER",
                    "sport": "NBA",
                    "prop_id": "NBA:LeBron James:points:25.5:content_test",
                },
                {
                    "player_name": "LeBron James",
                    "stat_type": "rebounds",
                    "line": 8.5,
                    "side": "OVER",
                    "sport": "NBA",
                    "prop_id": "NBA:LeBron James:rebounds:8.5:content_test",
                },
            ],
        }
        graded = grade_slip(slip, _BOX_SCORES, config=_CONFIG)
        sid = slip_id_for(_DATE_A, slip)
        graded["slip_id"] = sid
        write_slip_history_rows(ws_slip, _DATE_A, [graded])

        # Collect all values in Results sheet.
        results_values = set()
        for row in ws_results.iter_rows(values_only=True):
            for v in row:
                if v is not None:
                    results_values.add(str(v))

        self.assertNotIn(sid, results_values, "Slip ID must NOT appear in Results sheet")
        self.assertNotIn("PrizePicks", results_values, "platform string must not bleed into Results")


# ---------------------------------------------------------------------------
# Write_slip_history_rows idempotency unit test (in-memory)
# ---------------------------------------------------------------------------

class TestWriteSlipHistoryRowsIdempotency(unittest.TestCase):
    """Direct test of write_slip_history_rows upsert: second write never appends a duplicate."""

    def _make_graded_slip(self, prop_id_suffix: str = "") -> dict[str, Any]:
        slip = {
            "date": _DATE_A,
            "category": "safest_2_leg",
            "platform": "PrizePicks",
            "slip_type": "power",
            "stake_units": 1.0,
            "leg_count": 2,
            "legs": [
                {
                    "player_name": "LeBron James",
                    "stat_type": "points",
                    "line": 25.5,
                    "side": "OVER",
                    "sport": "NBA",
                    "prop_id": f"NBA:LBJ:pts:25.5{prop_id_suffix}",
                },
                {
                    "player_name": "LeBron James",
                    "stat_type": "rebounds",
                    "line": 8.5,
                    "side": "OVER",
                    "sport": "NBA",
                    "prop_id": f"NBA:LBJ:reb:8.5{prop_id_suffix}",
                },
            ],
        }
        g = grade_slip(slip, _BOX_SCORES, config=_CONFIG)
        g["slip_id"] = slip_id_for(_DATE_A, slip)
        return g

    def test_no_duplicate_on_second_write(self) -> None:
        """Calling write_slip_history_rows twice for the same slip yields 1 data row."""
        wb = openpyxl.Workbook()
        ws = wb.create_sheet("Slip History")
        ws.append(SLIP_HISTORY_HEADERS)

        graded = [self._make_graded_slip(":nodup")]

        write_slip_history_rows(ws, _DATE_A, graded)
        count_after_first = _count_ws_rows(ws)

        write_slip_history_rows(ws, _DATE_A, graded)
        count_after_second = _count_ws_rows(ws)

        self.assertEqual(count_after_first, count_after_second,
                         "Second write must not append duplicate rows")
        self.assertEqual(count_after_first, 1, "Exactly one row for one slip")

    def test_different_slip_ids_append(self) -> None:
        """Two slips with different prop IDs (different Slip IDs) both persist."""
        wb = openpyxl.Workbook()
        ws = wb.create_sheet("Slip History")
        ws.append(SLIP_HISTORY_HEADERS)

        graded_a = [self._make_graded_slip(":a")]
        graded_b = [self._make_graded_slip(":b")]

        write_slip_history_rows(ws, _DATE_A, graded_a)
        write_slip_history_rows(ws, _DATE_A, graded_b)
        count = _count_ws_rows(ws)

        self.assertEqual(count, 2, "Two distinct slips must produce two rows")


def _count_ws_rows(ws: Any) -> int:
    """Count data rows (row 2+) in an in-memory worksheet."""
    return sum(
        1 for row in ws.iter_rows(min_row=2, values_only=True)
        if any(v is not None for v in row)
    )


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    unittest.main(verbosity=2)
