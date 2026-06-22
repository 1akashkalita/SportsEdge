#!/usr/bin/env python3
"""Stage 1 platform-output regression tests for first-class DFS sources."""
from __future__ import annotations

import importlib.util
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

SCRIPT = Path(__file__).resolve().with_name("sports_system_runner.py")
spec = importlib.util.spec_from_file_location("sports_system_runner", SCRIPT)
runner = importlib.util.module_from_spec(spec)
assert spec and spec.loader
spec.loader.exec_module(runner)


def underdog_prop(line=1.5) -> dict:
    return {
        "projection_id": "ud-1",
        "player_name": "Underdog Player",
        "team": "TST",
        "description": "OPP",
        "stat_name": "Hits",
        "stat_type": "hits",
        "line_score": line,
        "odds_type": "standard",
        "status": "active",
        "start_time": "2026-06-10T20:00:00Z",
        "line_timing": "pregame",
        "line_timing_reason": "test fixture pregame",
        "platform": "Underdog",
        "primary_platform": "Underdog",
    }


class Stage1PlatformOutputTests(unittest.TestCase):
    def test_skipped_pick_schema_has_platform_column(self) -> None:
        self.assertIn("Platform", runner.SKIPPED_PICK_HEADERS)
        # Platform is appended at the end so positional row.append() stays
        # column-aligned for legacy workbooks that get migrated in place.
        self.assertEqual(runner.SKIPPED_PICK_HEADERS[-1], "Platform")

    def test_obsidian_line_moves_use_actual_platform(self) -> None:
        captured = []
        with patch.object(runner, "obsidian_sync", side_effect=lambda payload: captured.append(payload) or {"ok": True}):
            runner.obsidian_append_line_moves([
                {
                    "player": "Underdog Player",
                    "stat": "Hits",
                    "old_line": 1.5,
                    "new_line": 2.0,
                    "direction": "unfavorable",
                    "platform": "Underdog",
                    "line_timing": "pregame",
                    "movement_type": "pregame",
                }
            ], sport="MLB", date="2026-06-10")
        markdown = captured[0]["data"]["line_moves_markdown"]
        self.assertIn("* Platform: Underdog", markdown)
        self.assertNotIn("* Platform: PrizePicks", markdown)

    def test_prop_monitor_refresh_writes_underdog_platform_and_line_move_platform(self) -> None:
        tmpdir = tempfile.TemporaryDirectory()
        self.addCleanup(tmpdir.cleanup)
        wb_path = Path(tmpdir.name) / "mlb_test.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Player Props"
        ws.append(runner.PROPS_HEADERS)
        ws.append([
            runner.today_str(), "MLB", "ud-1", "Underdog Player", "TST", "OPP", "Hits", 1.5,
            "standard", "2026-06-10T20:00:00Z", "ACTIVE", "Existing active row",
            "", "", "Underdog", "", "", "", "", "", "", "", "", "", "", "",
        ] + ["" for _ in runner.LINE_TIMING_FIELDS])
        wb.save(wb_path)
        details_dir = Path(tmpdir.name) / "details"

        with patch.object(runner, "ensure_dirs", return_value=None), \
             patch.object(runner, "run_fetch_dfs_props", return_value=None), \
             patch.object(runner, "first_class_dfs_props_latest", return_value=[underdog_prop(line=2.0)]), \
             patch.object(runner, "fetch_sportsbook_prop_lines", return_value=({}, {"credits_remaining": "DISABLED"})), \
             patch.object(runner, "ensure_workbook", return_value=wb_path), \
             patch.object(runner, "obsidian_append_line_moves", return_value=None), \
             patch.object(runner, "PROP_MONITOR_DIR", details_dir):
            result = runner.prop_monitor("mlb")

        self.assertEqual(result["line_moves"][0]["platform"], "Underdog")
        wb2 = load_workbook(wb_path, read_only=True, data_only=True)
        try:
            ws2 = wb2["Player Props"]
            headers = [c.value for c in next(ws2.iter_rows(min_row=1, max_row=1))]
            platform_col = headers.index("Platform") + 1
            reason_col = headers.index("Reasoning") + 1
            self.assertEqual(ws2.cell(2, platform_col).value, "Underdog")
            self.assertIn("Underdog", ws2.cell(2, reason_col).value)
            self.assertNotIn("PrizePicks", ws2.cell(2, reason_col).value)
        finally:
            wb2.close()

    def test_approved_pick_summary_includes_platform(self) -> None:
        pick = {
            "selection": "Underdog Player Over 1.5 Hits",
            "sport": "MLB",
            "kind": "prop",
            "platform": "Underdog",
            "confidence": "A",
            "units": 1.0,
            "model_over_probability": 0.62,
            "ev": 0.18,
        }
        summary = runner.approved_pick_summary(pick)
        self.assertEqual(summary["platform"], "Underdog")

    def test_clv_preview_includes_platform(self) -> None:
        row = {"Pick Ref": "Underdog Player Over 1.5 Hits", "Platform": "Underdog"}
        preview = runner.clv_preview_row(row, our=1.5, closing=2.0, val=-0.5, status="NEGATIVE CLV")
        self.assertEqual(preview["platform"], "Underdog")

    def test_skipped_picks_legacy_workbook_migration_appends_platform_at_end(self) -> None:
        tmpdir = tempfile.TemporaryDirectory()
        self.addCleanup(tmpdir.cleanup)
        path = Path(tmpdir.name) / "mlb_test.xlsx"
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("Skipped Picks")
        legacy = ["Date", "Sport", "Pick", "Gate Failed", "Reason", "What Edge Would Have Been", "Result", "Pick Type", "Player/Team", "Line", "Probability", "EV", "Units", "Logged At"] + runner.LINE_TIMING_FIELDS
        ws.append(legacy)
        ws.append(["2026-06-09", "MLB", "Old Pick", "GATE 1", "old reason", 0.05, "", "PROP", "Old Player", 1.5, 0.55, 0.1, 1.0, "2026-06-09T10:00:00Z"] + ["" for _ in runner.LINE_TIMING_FIELDS])
        wb.save(path)

        original_workbook_path = runner.workbook_path
        original_ensure_dirs = runner.ensure_dirs
        runner.workbook_path = lambda sport, date=None: path
        runner.ensure_dirs = lambda: None
        try:
            runner.ensure_workbook("mlb")
        finally:
            runner.workbook_path = original_workbook_path
            runner.ensure_dirs = original_ensure_dirs

        wb2 = load_workbook(path, read_only=True)
        try:
            ws2 = wb2["Skipped Picks"]
            headers = [c.value for c in next(ws2.iter_rows(min_row=1, max_row=1))]
            self.assertEqual(headers[-1], "Platform")
            # Legacy data preserved.
            self.assertEqual(ws2.cell(2, 3).value, "Old Pick")
        finally:
            wb2.close()


if __name__ == "__main__":
    unittest.main()
