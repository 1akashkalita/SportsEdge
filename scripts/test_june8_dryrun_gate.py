#!/usr/bin/env python3
"""RESULTS-07 hard gate: ≥80% of non-Fantasy-Score MANUAL REVIEW prop rows for
the June 8 MLB backlog resolve to WIN/LOSS/PUSH after Layer-1 hardening.

Dry-run only — does NOT write to any workbook.

Idempotency design (GAP 3 fix):
  The denominator is a FIXED pre-backfill snapshot loaded from
  testdata/june8_manual_review_snapshot.json — not the live workbook's current
  MANUAL REVIEW rows.  The backfill has already resolved those rows; re-reading
  the live workbook would return 0 rows and produce a spurious 0/0 failure.
  The snapshot pins the original 37-row non-Fantasy set exactly as it existed
  before the June 23 backfill ran, making this test stable across all future runs.

Strategy (Testing strategy #11 from the trustworthy-results spec):
  1. Load the FIXED denominator from testdata/june8_manual_review_snapshot.json.
     Rows with dnp_void=true are excluded (DNP→VOID, never Layer-1-recoverable).
  2. Exclude Fantasy Score composites (Hitter/Pitcher Fantasy Score, Fantasy Points) —
     the snapshot already omits them, but the filter is kept as defense-in-depth.
  3. For each denominator row:
       a. Parse player / stat / line from the PROP: Pick Ref using parse_prop_ref.
       b. Match the Game column to an ESPN game event_id (name-based matching).
       c. Load player stats via espn_player_stats_by_event (Layer-1 hardened).
       d. Call stat_value_for_prop — if it returns a non-None value, the stat is
          resolvable.  Grade using the default side ("Over" — the Props sheet rows
          store opponent abbreviation, not "Over"/"Under", so the original grading
          always defaulted to "Over").
  4. Assert resolved / denominator >= 0.80.
     On failure, print the measured numerator/denominator breakdown so the operator
     sees the real ceiling rather than a bare assertion failure.
  5. If ESPN is unreachable (network error), skip with a clear message rather than
     fail — unreachable ESPN is a network condition, not a gate regression.

Run from scripts/:
    cd scripts && python3 test_june8_dryrun_gate.py
"""

from __future__ import annotations

import importlib
import json
import sys
import unittest
from pathlib import Path

# Ensure scripts/ is on sys.path for sibling imports (per CLAUDE.md)
SCRIPTS = Path(__file__).parent
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))

runner = importlib.import_module("sports_system_runner")

DATE = "2026-06-08"
WORKBOOK = SCRIPTS.parent / "data" / "mlb" / f"mlb_{DATE}.xlsx"
SNAPSHOT_PATH = SCRIPTS / "testdata" / "june8_manual_review_snapshot.json"

# Fantasy Score stat keywords (exact stat strings that are Layer-2 residue only).
_FANTASY_STAT_KEYWORDS = frozenset(
    {
        "fantasy score",
        "fantasy points",
        "hitter fantasy score",
        "pitcher fantasy score",
        "nba fantasy score",
        "nba fantasy points",
    }
)


def _is_fantasy_stat(ref: str) -> bool:
    """Return True if the PROP: ref refers to a Fantasy Score composite stat."""
    ref_lower = ref.lower()
    return any(kw in ref_lower for kw in _FANTASY_STAT_KEYWORDS)


def _game_label_to_event_id(game_label: str, espn_games: list[dict]) -> str | None:
    """Match a Result-sheet Game label (e.g. 'Houston Astros @ Los Angeles Angels')
    to an ESPN game dict and return the event_id, or None if no match.

    Uses team_aliases for fuzzy matching so "Athletics" == "Oakland Athletics" etc.
    """
    if not game_label:
        return None
    parts = [p.strip() for p in str(game_label).split("@")]
    if len(parts) != 2:
        return None
    away_label, home_label = parts[0], parts[1]
    away_aliases = runner.team_aliases(away_label)
    home_aliases = runner.team_aliases(home_label)
    for g in espn_games:
        g_home_aliases = runner.team_aliases(g.get("home_team") or "")
        g_away_aliases = runner.team_aliases(g.get("away_team") or "")
        if away_aliases & g_away_aliases and home_aliases & g_home_aliases:
            return str(g.get("event_id") or "")
    return None


def _load_snapshot_denominator() -> list[dict]:
    """Load the FIXED pre-backfill denominator from the snapshot fixture.

    Returns only the non-DNP rows (dnp_void == false) and non-Fantasy rows.
    This is the stable denominator: independent of the current workbook state.
    """
    with open(SNAPSHOT_PATH, encoding="utf-8") as f:
        data = json.load(f)
    rows = data.get("rows", [])
    # Exclude DNP→VOID rows (never recoverable to WIN/LOSS/PUSH via Layer-1)
    rows = [r for r in rows if not r.get("dnp_void")]
    # Defense-in-depth: exclude any Fantasy Score rows that might slip in
    rows = [r for r in rows if not _is_fantasy_stat(r.get("pick_ref", ""))]
    return rows


# _read_june8_manual_review_rows is retained (unused) for reference only.
# The gate no longer uses the live workbook as the denominator source.
def _read_june8_manual_review_rows() -> list[dict]:
    """Read all MANUAL REVIEW prop rows from the June 8 Results sheet.

    NOTE: This function is NO LONGER USED as the denominator source.
    After the June 23 backfill resolved all non-Fantasy rows, this function
    returns 0 rows, causing the gate to report a spurious 0/0 failure.
    The gate now uses _load_snapshot_denominator() instead (idempotent).
    Retained here for reference only.
    """
    from openpyxl import load_workbook  # type: ignore[import]

    wb = load_workbook(str(WORKBOOK), read_only=True, data_only=True)
    ws = wb["Results"]
    headers = [c.value for c in ws[1]]
    rows = []
    for vals in ws.iter_rows(min_row=2, values_only=True):
        if not vals or vals[0] is None:
            continue
        rd = dict(zip(headers, vals))
        if (
            str(rd.get("Date") or "")[:10] == DATE
            and str(rd.get("Sport") or "").upper() == "MLB"
            and str(rd.get("Result") or "").upper() == "MANUAL REVIEW"
        ):
            ref = str(rd.get("Pick Ref") or "")
            if ref.startswith("PROP:"):
                rows.append(rd)
    wb.close()
    return rows


class TestJune8DryRunGate(unittest.TestCase):
    """RESULTS-07 gate: ≥80% non-Fantasy MANUAL REVIEW prop rows resolve after Layer-1.

    Idempotent: uses a fixed pre-backfill snapshot fixture as the denominator,
    not the live workbook's current MANUAL REVIEW state.
    """

    def test_snapshot_fixture_integrity(self) -> None:
        """Fixture integrity: snapshot exists, has a non-zero fixed denominator,
        and every non-DNP row has a parseable PROP: ref (stat AND line resolvable).
        """
        self.assertTrue(
            SNAPSHOT_PATH.exists(),
            f"Snapshot fixture not found: {SNAPSHOT_PATH}",
        )

        with open(SNAPSHOT_PATH, encoding="utf-8") as f:
            data = json.load(f)

        all_rows = data.get("rows", [])
        self.assertGreater(len(all_rows), 0, "Snapshot has no rows")

        dnp_rows = [r for r in all_rows if r.get("dnp_void")]
        non_dnp_rows = [r for r in all_rows if not r.get("dnp_void")]

        # The fixture must have exactly the documented set
        self.assertGreaterEqual(
            len(all_rows),
            37,
            f"Expected ≥37 snapshot rows (37 documented), got {len(all_rows)}",
        )
        self.assertEqual(
            len(dnp_rows),
            2,
            f"Expected exactly 2 DNP→VOID rows (Nick Martinez + Masataka Yoshida), "
            f"got {len(dnp_rows)}: {[r['pick_ref'] for r in dnp_rows]}",
        )
        self.assertEqual(
            len(non_dnp_rows),
            35,
            f"Expected exactly 35 denominator rows (non-DNP), got {len(non_dnp_rows)}",
        )

        # Every non-DNP row must have a parseable PROP: ref (stat AND line non-None)
        unparseable = []
        for row in non_dnp_rows:
            ref = row.get("pick_ref", "")
            _player, stat, line = runner.parse_prop_ref(ref)
            if stat is None or line is None:
                unparseable.append(ref)

        self.assertEqual(
            len(unparseable),
            0,
            f"Some snapshot rows have unparseable PROP: refs (stat or line is None):\n"
            + "\n".join(f"  {r}" for r in unparseable),
        )

    def test_gate_eighty_percent_non_fantasy_resolve(self) -> None:
        """RESULTS-07 gate: ≥80% of the FIXED pre-backfill non-Fantasy MANUAL REVIEW
        prop rows resolve via Layer-1.  Idempotent — uses snapshot denominator.
        """
        # Step 1 — load the FIXED denominator from the snapshot fixture
        self.assertTrue(
            SNAPSHOT_PATH.exists(),
            f"Snapshot fixture not found: {SNAPSHOT_PATH}",
        )

        denominator_rows = _load_snapshot_denominator()
        denominator = len(denominator_rows)

        self.assertGreater(
            denominator,
            0,
            "Snapshot denominator is 0 — fixture may be empty or all rows are DNP",
        )

        print(f"\nJune 8 pre-backfill snapshot: {denominator} denominator rows (non-DNP, non-Fantasy)")

        # Step 2 — fetch June 8 ESPN games (deterministic for a historical date)
        try:
            espn_games = runner.espn_scoreboard_games_for_date("mlb", DATE)
        except Exception as exc:
            self.skipTest(
                f"ESPN API unreachable — skipping gate (network condition, not a regression): {exc}"
            )

        if not espn_games:
            self.skipTest(
                "ESPN returned no June 8 MLB games — skipping gate (network condition, not a regression)"
            )

        # Build event_id → player_stats cache (one ESPN call per game)
        event_stats_cache: dict[str, dict] = {}

        # Step 3 — grade each denominator row via Layer-1
        resolved: list[dict] = []         # WIN/LOSS/PUSH
        still_manual: list[dict] = []     # stat_value returned None → unresolvable

        for row in denominator_rows:
            ref = row.get("pick_ref", "")
            game_label = row.get("game", "")

            # Parse player / stat / line from PROP: ref using parse_prop_ref
            player, stat, line = runner.parse_prop_ref(ref)

            if stat is None or line is None:
                # Ref unparseable — counts as still-manual (should not occur per fixture integrity test)
                still_manual.append({
                    "ref": ref,
                    "reason": "parse_prop_ref returned None stat or line",
                })
                continue

            # Find the ESPN event_id for this game
            event_id = _game_label_to_event_id(game_label, espn_games)
            if not event_id:
                still_manual.append({
                    "ref": ref,
                    "reason": f"game label not matched to ESPN event: '{game_label}'",
                })
                continue

            # Load player stats (cached per event_id)
            if event_id not in event_stats_cache:
                try:
                    event_stats_cache[event_id] = runner.espn_player_stats_by_event(
                        "mlb", event_id
                    )
                except Exception as exc:
                    self.skipTest(
                        f"ESPN API unreachable during stat load — skipping gate "
                        f"(network condition, not a regression): {exc}"
                    )
            player_stats = event_stats_cache[event_id]

            # Call Layer-1 stat resolution
            # The original Props rows had Opponent/Description = opponent abbreviation (not
            # "Over"/"Under"), so grade_prop defaulted to "Over" for all of these rows.
            # We replicate that default here.
            actual_val, src, conf = runner.stat_value_for_prop(player_stats, player, stat)

            if actual_val is not None:
                # Stat resolved — grade with default "Over" side (matching original grading).
                side = "Over"
                diff = actual_val - line
                if diff == 0:
                    result = "PUSH"
                elif side == "Over":
                    result = "WIN" if diff > 0 else "LOSS"
                else:
                    result = "WIN" if diff < 0 else "LOSS"

                resolved.append({
                    "ref": ref,
                    "result": result,
                    "actual": actual_val,
                    "line": line,
                    "src": src,
                    "conf": conf,
                })
            else:
                still_manual.append({
                    "ref": ref,
                    "reason": f"stat_value_for_prop returned None for {player!r} {stat!r}",
                })

        numerator = len(resolved)
        rate = numerator / denominator if denominator else 0.0

        # Print detailed breakdown always so the operator can see it
        print(f"\n=== June 8 Layer-1 Dry-Run Resolution ===")
        print(f"Denominator (pre-backfill non-Fantasy non-DNP rows): {denominator}")
        print(f"Resolved (WIN/LOSS/PUSH): {numerator}")
        print(f"Still unresolvable: {len(still_manual)}")
        print(f"Resolution rate: {numerator}/{denominator} = {rate:.1%}")
        print()

        if resolved:
            print("Resolved rows (first 10):")
            for r in resolved[:10]:
                print(f"  {r['result']:6s}  {r['ref']}  (actual={r['actual']} vs Over {r['line']})")

        if still_manual:
            print(f"\nStill unresolvable ({len(still_manual)}):")
            for r in still_manual[:20]:
                print(f"  {r['ref']}  — {r['reason']}")

        self.assertGreaterEqual(
            rate,
            0.80,
            f"RESULTS-07 gate FAILED: Layer-1 resolved {numerator}/{denominator} = {rate:.1%} "
            f"of the pre-backfill non-Fantasy-Score MANUAL REVIEW prop rows for June 8 "
            f"(required ≥ 80%).\n"
            f"Still unresolvable breakdown:\n"
            + "\n".join(f"  {r['ref']} — {r['reason']}" for r in still_manual),
        )

        print(
            f"\nRESULTS-07 gate PASSED: {numerator}/{denominator} = {rate:.1%} "
            f"non-Fantasy MANUAL REVIEW rows resolve after Layer-1 hardening."
        )


if __name__ == "__main__":
    unittest.main(verbosity=2)
