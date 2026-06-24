#!/usr/bin/env python3
"""test_dashboard_data.py — DASH-04 + D-01/D-02 test suite for the dashboard data layer.

Covers:
- test_read_only_untouched (DASH-04): read_only load leaves source mtime + sha256 unchanged.
- test_lock_tolerant (DASH-04/D-01): WorkbookAccessError → accessor returns empty/None without raising.
- test_missing_is_empty (DASH-04): missing workbook/JSON → empty state, no exception.
- test_today_matches_runner (D-02): dashboard_data's "today" == naive datetime.now() (no ZoneInfo).
- test_write_in_progress (D-01): live lock pid → True; dead/stale pid → False.

All tests except test_today_matches_runner are RED until plan 02 builds dashboard_data.py.
test_today_matches_runner is also RED because it imports dashboard_data.
"""
from __future__ import annotations

import hashlib
import json
import os
import sys
import tempfile
import time
import unittest
from datetime import datetime
from pathlib import Path
from typing import Any
from unittest.mock import patch

import openpyxl
from openpyxl import Workbook

SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

# RED until plan 02 creates scripts/dashboard_data.py — that is the intended state.
import dashboard_data  # noqa: E402

from workbook_io import WorkbookAccessError  # noqa: E402


# ---------------------------------------------------------------------------
# Helpers to build synthetic workbooks for fixtures
# ---------------------------------------------------------------------------

_SAMPLE_HEADERS = ["Platform", "Sport", "Pick Type", "Selection", "Edge", "Model Over Probability", "EV", "Confidence"]


def _make_picks_wb(rows: list[dict[str, Any]]) -> Workbook:
    """Build an in-memory workbook with a Picks sheet populated from sample rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Picks"
    ws.append(_SAMPLE_HEADERS)
    for row in rows:
        ws.append([row.get(h) for h in _SAMPLE_HEADERS])
    return wb


# ---------------------------------------------------------------------------
# DASH-04: read_only load must not modify the source file
# ---------------------------------------------------------------------------

class TestReadOnlyUntouched(unittest.TestCase):
    """DASH-04 — read_only workbook load must leave source mtime + sha256 byte-identical."""

    def test_read_only_untouched(self) -> None:
        """read_only=True load via dashboard_data reader must not alter the source file.

        Protocol:
        1. Write a synthetic .xlsx to a TemporaryDirectory.
        2. Capture st_mtime + sha256 of the file bytes before the read.
        3. Call dashboard_data's sheet-reader with read_only mode.
        4. Assert st_mtime AND sha256 are byte-identical after the read.

        This is the DASH-04 write-prevention contract — mirrors the live-verified
        pattern from 01-RESEARCH.md "Read-only proves source untouched."
        """
        sample_rows = [
            {"Platform": "PrizePicks", "Sport": "MLB", "Pick Type": "prop",
             "Selection": "over", "Edge": 1.2, "Model Over Probability": 0.62,
             "EV": 0.08, "Confidence": "HIGH"},
        ]
        wb = _make_picks_wb(sample_rows)

        with tempfile.TemporaryDirectory() as tmpdir:
            xlsx_path = Path(tmpdir) / "nba_2026-06-24.xlsx"
            wb.save(xlsx_path)

            # Capture baseline
            before_mtime = xlsx_path.stat().st_mtime
            before_hash = hashlib.sha256(xlsx_path.read_bytes()).hexdigest()

            # Override module path constants so the reader targets our temp file
            orig_nba = dashboard_data.NBA_DIR
            orig_mlb = dashboard_data.MLB_DIR
            try:
                dashboard_data.NBA_DIR = Path(tmpdir)
                dashboard_data.MLB_DIR = Path(tmpdir)
                dashboard_data.read_sheet_rows(xlsx_path, "Picks")
            finally:
                dashboard_data.NBA_DIR = orig_nba
                dashboard_data.MLB_DIR = orig_mlb

            # Assertions: source file must be byte-identical
            after_mtime = xlsx_path.stat().st_mtime
            after_hash = hashlib.sha256(xlsx_path.read_bytes()).hexdigest()

            self.assertEqual(
                before_mtime,
                after_mtime,
                "read_only load must not change st_mtime of the source workbook",
            )
            self.assertEqual(
                before_hash,
                after_hash,
                "read_only load must not change the sha256 of the source workbook bytes",
            )


# ---------------------------------------------------------------------------
# DASH-04 / D-01: WorkbookAccessError must not propagate — return empty state
# ---------------------------------------------------------------------------

class TestLockTolerant(unittest.TestCase):
    """DASH-04/D-01 — WorkbookAccessError from safe_load_workbook must NOT propagate.

    The accessor must return last-known-good / None / [] without raising.
    """

    def test_lock_tolerant(self) -> None:
        """WorkbookAccessError on load → accessor returns None/empty, never raises.

        Protocol:
        1. Monkeypatch dashboard_data.safe_load_workbook to raise WorkbookAccessError.
        2. Call dashboard_data.read_sheet_rows — it must NOT raise.
        3. Assert the return value is None or [] (empty last-known-good state).
        """
        with tempfile.TemporaryDirectory() as tmpdir:
            xlsx_path = Path(tmpdir) / "locked.xlsx"
            # Create a real (but unreadable-because-locked) placeholder
            Workbook().save(xlsx_path)

            with patch.object(
                dashboard_data,
                "safe_load_workbook",
                side_effect=WorkbookAccessError("simulated lock contention"),
            ):
                # Must NOT raise — D-01 contract: last-known-good on locked workbook
                result = dashboard_data.read_sheet_rows(xlsx_path, "Picks")

            # Return must be the empty/None fallback state (not an exception)
            self.assertIn(
                result,
                (None, []),
                f"read_sheet_rows must return None or [] on WorkbookAccessError, got {result!r}",
            )


# ---------------------------------------------------------------------------
# DASH-04: missing file → empty state, no exception
# ---------------------------------------------------------------------------

class TestMissingIsEmpty(unittest.TestCase):
    """DASH-04 — missing workbook or JSON path must return empty state, never raise."""

    def test_missing_is_empty(self) -> None:
        """Missing .xlsx and missing JSON both return empty/None without raising.

        Protocol:
        1. Construct a non-existent .xlsx path.
        2. Call dashboard_data.read_sheet_rows — must return None/[].
        3. Construct a non-existent JSON path.
        4. Call dashboard_data.read_json — must return None.
        """
        non_existent_xlsx = Path("/tmp/__nonexistent_dashboard_test__.xlsx")
        non_existent_json = Path("/tmp/__nonexistent_dashboard_test__.json")

        # Ensure they really don't exist
        non_existent_xlsx.unlink(missing_ok=True)
        non_existent_json.unlink(missing_ok=True)

        # Workbook read on missing file: must not raise
        xlsx_result = dashboard_data.read_sheet_rows(non_existent_xlsx, "Picks")
        self.assertIn(
            xlsx_result,
            (None, []),
            f"read_sheet_rows on missing file must return None or [], got {xlsx_result!r}",
        )

        # JSON read on missing file: must not raise
        json_result = dashboard_data.read_json(non_existent_json)
        self.assertIsNone(
            json_result,
            f"read_json on missing file must return None, got {json_result!r}",
        )


# ---------------------------------------------------------------------------
# D-02: "today" must match the pipeline's naive-local today_str()
# ---------------------------------------------------------------------------

class TestTodayMatchesRunner(unittest.TestCase):
    """D-02 — dashboard_data's "today" must match the runner's naive-local today_str().

    The runner uses datetime.now().strftime("%Y-%m-%d") (no ZoneInfo). The dashboard
    must compute the same value or the Today view will point at the wrong workbook
    near midnight (Pitfall 2 in 01-RESEARCH.md).
    """

    def test_today_matches_runner(self) -> None:
        """dashboard_data.today_str() must equal datetime.now().strftime('%Y-%m-%d').

        Also asserts that dashboard_data's source text does NOT import zoneinfo or
        ZoneInfo — preventing a midnight mismatch with the pipeline.
        """
        # The contract value: naive local time, matching the runner's today_str()
        expected = datetime.now().strftime("%Y-%m-%d")
        actual = dashboard_data.today_str()

        self.assertEqual(
            actual,
            expected,
            f"dashboard_data.today_str() = {actual!r} must equal datetime.now().strftime('%Y-%m-%d') = {expected!r}",
        )

        # Belt-and-suspenders: assert the module source has NO zoneinfo import
        source_path = Path(dashboard_data.__file__)
        source_text = source_path.read_text()
        self.assertNotIn(
            "zoneinfo",
            source_text,
            "dashboard_data.py must NOT import zoneinfo — it would mismatch the pipeline's naive-local today",
        )
        self.assertNotIn(
            "ZoneInfo",
            source_text,
            "dashboard_data.py must NOT use ZoneInfo — it would mismatch the pipeline's naive-local today",
        )


# ---------------------------------------------------------------------------
# D-01: write_in_progress() — live pid → True; dead/stale pid → False
# ---------------------------------------------------------------------------

class TestWriteInProgress(unittest.TestCase):
    """D-01 — write_in_progress() badge: live lock pid → True; dead/stale pid → False.

    The badge signal uses the cooperative-lock files in locks/*.xlsx.lock.
    Lock JSON shape (from workbook_io.py:90): {"pid": <int>, "path": <str>, "acquired_at": <iso>}
    Staleness threshold: 600 seconds (matches workbook_io stale_seconds=600).
    """

    _STALE_SECONDS = 600

    def _write_lock(self, lock_dir: Path, name: str, pid: int, age_seconds: float = 0.0) -> None:
        """Write a lock file with the given pid and mtime offset from now."""
        lock_path = lock_dir / f"{name}.xlsx.lock"
        lock_data = {
            "pid": pid,
            "path": str(lock_dir / f"{name}.xlsx"),
            "acquired_at": datetime.now().isoformat(),
        }
        lock_path.write_text(json.dumps(lock_data))
        if age_seconds > 0:
            # Back-date mtime by age_seconds so the staleness check triggers
            old_time = time.time() - age_seconds
            os.utime(lock_path, (old_time, old_time))

    def test_write_in_progress(self) -> None:
        """live pid + fresh lock → True; dead/bogus pid + stale lock → False.

        Covers two sub-cases:
        a) A lock file holding os.getpid() (the live test process) with fresh mtime → True.
        b) A lock file holding a bogus/dead pid OR a mtime older than 600s → False.
        """
        with tempfile.TemporaryDirectory() as tmpdir:
            lock_dir = Path(tmpdir) / "locks"
            lock_dir.mkdir()

            orig_lock_dir = dashboard_data.LOCK_DIR
            try:
                dashboard_data.LOCK_DIR = lock_dir

                # --- Sub-case a: live pid, fresh mtime → True ---
                self._write_lock(lock_dir, "nba_2026-06-24", pid=os.getpid(), age_seconds=0.0)
                result_live = dashboard_data.write_in_progress()
                self.assertTrue(
                    result_live,
                    "write_in_progress() must return True for a fresh lock held by the live process",
                )

                # Clear the lock for the next sub-case
                for lf in lock_dir.glob("*.xlsx.lock"):
                    lf.unlink()

                # --- Sub-case b: bogus (dead) pid, fresh mtime → False ---
                # PID 2 is typically the kernel thread on macOS and never process-killable.
                # Alternatively use a very large pid that cannot exist (> max pid).
                dead_pid = 99999999  # guaranteed not to be a live user process
                self._write_lock(lock_dir, "mlb_2026-06-24", pid=dead_pid, age_seconds=0.0)
                result_dead = dashboard_data.write_in_progress()
                self.assertFalse(
                    result_dead,
                    f"write_in_progress() must return False for a lock held by dead pid {dead_pid}",
                )

                # Clear for the next sub-case
                for lf in lock_dir.glob("*.xlsx.lock"):
                    lf.unlink()

                # --- Sub-case c: live pid but stale mtime (> 600s) → False ---
                self._write_lock(
                    lock_dir,
                    "nba_stale",
                    pid=os.getpid(),
                    age_seconds=self._STALE_SECONDS + 60,
                )
                result_stale = dashboard_data.write_in_progress()
                self.assertFalse(
                    result_stale,
                    "write_in_progress() must return False for a lock older than stale_seconds (600s)",
                )

            finally:
                dashboard_data.LOCK_DIR = orig_lock_dir


if __name__ == "__main__":
    unittest.main()
