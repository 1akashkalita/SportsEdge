#!/usr/bin/env python3
"""Test suite for prop/single-pick PnL = 0 (BANKROLL-01 / GAP 4).

Tests (RED phase — these MUST FAIL before the GREEN fix is applied):
 - A WIN prop row has Result=="WIN" and PnL==0 (not +0.909)
 - A LOSS prop row has Result=="LOSS" and PnL==0 (not -1.0)
 - A SPREAD single-pick row has Result preserved and PnL==0
 - A VOID/PENDING/MANUAL REVIEW prop row has PnL==0
 - A PARLAY row that grades WIN keeps a NON-ZERO PnL (parlays are staked; must NOT be zeroed)

Requires:
  grade_game_in_workbook in sports_system_runner
  odds_profit / pnl_for_result remain UNCHANGED (used by slip-grading path)

Run from scripts/: cd scripts && python3 -m pytest test_prop_pnl_slip_terms.py -x -q
"""
from __future__ import annotations

import importlib
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

_SCRIPTS = Path(__file__).parent
if str(_SCRIPTS) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS))

runner = importlib.import_module("sports_system_runner")

grade_game_in_workbook = runner.grade_game_in_workbook
odds_profit = runner.odds_profit
ensure_ws_columns = runner.ensure_ws_columns
save_workbook_atomic = runner.save_workbook_atomic
result_headers = runner.result_headers
RESULT_HEADERS = runner.RESULT_HEADERS
PICKS_HEADERS = runner.PICKS_HEADERS
PROPS_HEADERS = runner.PROPS_HEADERS
PARLAY_HEADERS = runner.PARLAY_HEADERS


# ---------------------------------------------------------------------------
# Shared game / player-stats fixture
# ---------------------------------------------------------------------------

_GAME = {
    "event_id": "test-event-pnl-001",
    "id": "test-event-pnl-001",
    "status": "final",
    "status_name": "STATUS_FINAL",
    "completed": True,
    "home_team": "Yankees",
    "away_team": "Red Sox",
    "home_score": 5,
    "away_score": 3,
    # Odds-API scores list required by final_scores() for spread/total grading
    "scores": [
        {"name": "Yankees", "score": "5"},
        {"name": "Red Sox", "score": "3"},
    ],
}

# Mookie Betts: Hits=3.0 → Over 2.5 = WIN; Walks=0.0 → Over 1.5 = LOSS
_PLAYER_STATS = {
    "mookie betts": {
        "batting": {
            "hits": 3.0, "runs": 1.0, "rbis": 1.0, "homeruns": 0.0,
            "walks": 0.0, "strikeouts": 1.0,
            "_hit_counts": {"single": 3, "double": 0, "triple": 0, "home-run": 0},
        },
        "pitching": {},
    }
}

_DATE = "2026-06-23"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _get_pnl_and_result(wb, date: str, sport: str, ref: str) -> tuple[str | None, float | None]:
    """Return (result_str, pnl_float) for the first matching Results row."""
    ws = wb["Results"]
    cols = result_headers(ws)
    for r in range(2, ws.max_row + 1):
        if (str(ws.cell(r, cols["Date"]).value or "")[:10] == date
                and str(ws.cell(r, cols["Sport"]).value or "") == sport
                and str(ws.cell(r, cols["Pick Ref"]).value or "") == ref):
            result_val = str(ws.cell(r, cols["Result"]).value or "")
            raw_pnl = ws.cell(r, cols["PnL"]).value
            pnl_val = float(raw_pnl) if raw_pnl is not None else None
            return result_val, pnl_val
    return None, None


def _seed_prop_row(props_ws, props_cols: dict, date: str, player: str, stat: str,
                   line: float, side: str = "Over") -> None:
    """Append one row to the Props sheet."""
    props_ws.append([None] * len(props_cols))
    r = props_ws.max_row
    for col_name, value in [
        ("Date", date), ("Sport", "MLB"), ("Player Name", player),
        ("Team", "Yankees"), ("Stat", stat), ("Line", line),
        ("Opponent/Description", side), ("Platform", "PrizePicks"),
    ]:
        if col_name in props_cols:
            props_ws.cell(r, props_cols[col_name]).value = value


def _seed_picks_row(picks_ws, picks_cols: dict, date: str, pick_type: str,
                    selection: str, home_team: str, away_team: str,
                    line: float = 0.0, odds: float | None = None) -> None:
    """Append one row to the Picks sheet (for SPREAD/TOTAL)."""
    picks_ws.append([None] * len(picks_cols))
    r = picks_ws.max_row
    for col_name, value in [
        ("Date", date), ("Sport", "MLB"), ("Pick Type", pick_type),
        ("Selection", selection), ("Home Team", home_team), ("Away Team", away_team),
        ("Line", line), ("Odds", odds), ("Units", 1.0),
        ("Platform", "FanDuel"),
    ]:
        if col_name in picks_cols:
            picks_ws.cell(r, picks_cols[col_name]).value = value


def _seed_parlay_row(parlays_ws, parlay_cols: dict, date: str, name: str,
                     legs: str, units: float = 0.5) -> None:
    """Append one row to the Correlated Parlays sheet."""
    parlays_ws.append([None] * len(parlay_cols))
    r = parlays_ws.max_row
    for col_name, value in [
        ("Date", date), ("Sport", "MLB"), ("Parlay Name", name),
        ("Legs", legs), ("Units", units),
    ]:
        if col_name in parlay_cols:
            parlays_ws.cell(r, parlay_cols[col_name]).value = value


# ---------------------------------------------------------------------------
# Main test class
# ---------------------------------------------------------------------------

class TestPropPnLSlipTerms(unittest.TestCase):
    """BANKROLL-01: PROP and single-pick SPREAD/TOTAL rows must write PnL=0.

    The WIN/LOSS Result carries the accuracy signal; real money PnL is computed
    only by grade_slips at the Slip History level.
    """

    def setUp(self) -> None:
        self._tmpdir = tempfile.TemporaryDirectory()
        tmp = Path(self._tmpdir.name)
        self._pnl_dir = tmp / "pnl"
        self._pnl_dir.mkdir(parents=True)
        self._data_dir = tmp / "data"
        self._mlb_dir = self._data_dir / "mlb"
        self._mlb_dir.mkdir(parents=True)
        self._locks_dir = tmp / "locks"
        self._locks_dir.mkdir()
        self._backups_dir = self._data_dir / "backups" / "workbooks"
        self._backups_dir.mkdir(parents=True)
        self._log_dir = self._pnl_dir / "logs"
        self._log_dir.mkdir()

        self._patches = [
            patch.object(runner, "PNL_DIR", self._pnl_dir),
            patch.object(runner, "NBA_DIR", tmp / "data" / "nba"),
            patch.object(runner, "MLB_DIR", self._mlb_dir),
            patch.object(runner, "DATA", self._data_dir),
            patch.object(runner, "LOG_DIR", self._log_dir),
            patch.object(runner, "WORKBOOK_LOCK_DIR", self._locks_dir),
            patch.object(runner, "WORKBOOK_BACKUP_DIR", self._backups_dir),
            patch.object(runner, "BANKROLL", self._pnl_dir / "bankroll.json"),
            patch.object(runner, "GAME_STATUS_CACHE_FILE", self._pnl_dir / "game_status_cache.json"),
        ]
        for p in self._patches:
            p.start()
        (tmp / "data" / "nba").mkdir(parents=True, exist_ok=True)

        self._fn_patches = [
            patch.object(runner, "espn_player_stats_by_event", return_value=_PLAYER_STATS),
            patch.object(runner, "send_telegram", return_value=None),
            patch.object(runner, "obsidian_update_results_section", return_value=None),
            patch.object(runner, "obsidian_update_bankroll_files", return_value=None),
        ]
        for p in self._fn_patches:
            p.start()

    def tearDown(self) -> None:
        for p in self._fn_patches:
            p.stop()
        for p in self._patches:
            p.stop()
        self._tmpdir.cleanup()

    # -----------------------------------------------------------------------
    # Test 1: WIN prop row PnL must be 0 (not +0.909 or similar positive value)
    # -----------------------------------------------------------------------

    def test_win_prop_pnl_is_zero(self) -> None:
        """BANKROLL-01: a WIN prop row must carry PnL=0, not a money figure.

        Fixture: Mookie Betts Hits Over 2.5 → actual=3.0 → WIN.
        Current code (before GREEN fix) calls odds_profit(result, units, None)
        which returns 0.909 for a WIN. This test will FAIL until the fix is applied.
        """
        date = _DATE
        path = runner.ensure_workbook("mlb", date)
        from openpyxl import load_workbook as lw
        wb = lw(path)

        props_ws = wb["Props"]
        props_cols = ensure_ws_columns(props_ws, PROPS_HEADERS)
        _seed_prop_row(props_ws, props_cols, date, "Mookie Betts", "Hits", 2.5, "Over")
        save_workbook_atomic(wb, path)

        grade_game_in_workbook("mlb", _GAME, date, dry_run=False)

        from openpyxl import load_workbook as lw2
        wb2 = lw2(path)
        ref = "PROP:Mookie Betts Hits 2.5"
        result_val, pnl_val = _get_pnl_and_result(wb2, date, "MLB", ref)

        # Result must be WIN (accuracy signal preserved)
        self.assertEqual(result_val, "WIN", f"Expected Result=WIN, got {result_val!r}")
        # PnL must be exactly 0.0 (BANKROLL-01: no standalone money PnL on prop rows)
        self.assertEqual(pnl_val, 0.0, f"Expected PnL=0.0 for WIN prop row, got {pnl_val}")

    # -----------------------------------------------------------------------
    # Test 2: LOSS prop row PnL must be 0 (not -1.0)
    # -----------------------------------------------------------------------

    def test_loss_prop_pnl_is_zero(self) -> None:
        """BANKROLL-01: a LOSS prop row must carry PnL=0, not -1.0.

        Fixture: Mookie Betts Walks Over 1.5 → actual=0.0 → LOSS.
        """
        date = _DATE
        path = runner.ensure_workbook("mlb", date)
        from openpyxl import load_workbook as lw
        wb = lw(path)

        props_ws = wb["Props"]
        props_cols = ensure_ws_columns(props_ws, PROPS_HEADERS)
        _seed_prop_row(props_ws, props_cols, date, "Mookie Betts", "Walks", 1.5, "Over")
        save_workbook_atomic(wb, path)

        grade_game_in_workbook("mlb", _GAME, date, dry_run=False)

        from openpyxl import load_workbook as lw2
        wb2 = lw2(path)
        ref = "PROP:Mookie Betts Walks 1.5"
        result_val, pnl_val = _get_pnl_and_result(wb2, date, "MLB", ref)

        # Result must be LOSS (accuracy signal preserved)
        self.assertEqual(result_val, "LOSS", f"Expected Result=LOSS, got {result_val!r}")
        # PnL must be 0.0 (BANKROLL-01)
        self.assertEqual(pnl_val, 0.0, f"Expected PnL=0.0 for LOSS prop row, got {pnl_val}")

    # -----------------------------------------------------------------------
    # Test 3: SPREAD single-pick row PnL must be 0
    # -----------------------------------------------------------------------

    def test_spread_single_pick_pnl_is_zero(self) -> None:
        """BANKROLL-01: a single SPREAD row must carry PnL=0.

        Fixture: Yankees -1.5 at home → home_score=5, away_score=3 → margin=2 > 1.5 → WIN.
        """
        date = _DATE
        path = runner.ensure_workbook("mlb", date)
        from openpyxl import load_workbook as lw
        wb = lw(path)

        picks_ws = wb["Picks"]
        picks_cols = ensure_ws_columns(picks_ws, PICKS_HEADERS)
        # Yankees -1.5 home team, home covers if home_score - away_score > 1.5 → 5-3=2 > 1.5 ✓
        _seed_picks_row(picks_ws, picks_cols, date, "SPREAD", "Yankees -1.5",
                        "Yankees", "Red Sox", line=-1.5, odds=-110)
        save_workbook_atomic(wb, path)

        grade_game_in_workbook("mlb", _GAME, date, dry_run=False)

        from openpyxl import load_workbook as lw2
        wb2 = lw2(path)
        ref = "SPREAD:Yankees -1.5"
        result_val, pnl_val = _get_pnl_and_result(wb2, date, "MLB", ref)

        # Result must be WIN or a valid terminal (accuracy preserved)
        self.assertIn(result_val, {"WIN", "LOSS", "PUSH"}, f"Expected terminal result, got {result_val!r}")
        # PnL must be 0.0 (BANKROLL-01: no standalone bankroll PnL for single-pick spreads)
        self.assertEqual(pnl_val, 0.0, f"Expected PnL=0.0 for SPREAD single-pick row, got {pnl_val}")

    # -----------------------------------------------------------------------
    # Test 4: VOID prop row PnL stays 0 (existing behavior preserved)
    # -----------------------------------------------------------------------

    def test_void_prop_pnl_is_zero(self) -> None:
        """A VOID prop row already writes PnL=0; must stay 0 after fix."""
        date = _DATE
        void_game = dict(_GAME)
        void_game["status"] = "void"
        void_game["status_name"] = "STATUS_POSTPONED"
        void_game["completed"] = False

        path = runner.ensure_workbook("mlb", date)
        from openpyxl import load_workbook as lw
        wb = lw(path)

        props_ws = wb["Props"]
        props_cols = ensure_ws_columns(props_ws, PROPS_HEADERS)
        _seed_prop_row(props_ws, props_cols, date, "Mookie Betts", "Hits", 2.5, "Over")
        save_workbook_atomic(wb, path)

        grade_game_in_workbook("mlb", void_game, date, dry_run=False)

        from openpyxl import load_workbook as lw2
        wb2 = lw2(path)
        ref = "PROP:Mookie Betts Hits 2.5"
        result_val, pnl_val = _get_pnl_and_result(wb2, date, "MLB", ref)

        self.assertEqual(result_val, "VOID", f"Expected Result=VOID for postponed game, got {result_val!r}")
        self.assertEqual(pnl_val, 0.0, f"Expected PnL=0.0 for VOID prop, got {pnl_val}")

    # -----------------------------------------------------------------------
    # Test 5: PARLAY row PnL is NOT zeroed (parlays are staked — keep money PnL)
    # -----------------------------------------------------------------------

    def test_parlay_win_pnl_is_nonzero(self) -> None:
        """Money-safety guard: a parlay WIN row MUST keep non-zero PnL.

        Parlays are staked bets (grade_slips / Slip History). They must NOT be
        zeroed by the prop PnL fix. This test verifies the fix doesn't regress
        the parlay money path.
        """
        date = _DATE
        path = runner.ensure_workbook("mlb", date)
        from openpyxl import load_workbook as lw
        wb = lw(path)

        # Seed a prop row to create leg results in graded[]
        props_ws = wb["Props"]
        props_cols = ensure_ws_columns(props_ws, PROPS_HEADERS)
        _seed_prop_row(props_ws, props_cols, date, "Mookie Betts", "Hits", 2.5, "Over")

        # Seed a Picks row for another leg
        picks_ws = wb["Picks"]
        picks_cols = ensure_ws_columns(picks_ws, PICKS_HEADERS)
        _seed_picks_row(picks_ws, picks_cols, date, "SPREAD", "Yankees -1.5",
                        "Yankees", "Red Sox", line=-1.5, odds=-110)

        # Seed a parlay whose legs are the two above (using declared leg refs)
        parlays_ws = wb["Correlated Parlays"]
        parlay_cols = ensure_ws_columns(parlays_ws, PARLAY_HEADERS)
        leg1 = "PROP:Mookie Betts Hits 2.5"
        leg2 = "SPREAD:Yankees -1.5"
        _seed_parlay_row(parlays_ws, parlay_cols, date, "TestParlay",
                         f"{leg1}|{leg2}", units=1.0)
        save_workbook_atomic(wb, path)

        grade_game_in_workbook("mlb", _GAME, date, dry_run=False)

        from openpyxl import load_workbook as lw2
        wb2 = lw2(path)
        parlay_ref = "PARLAY:TestParlay"
        result_val, pnl_val = _get_pnl_and_result(wb2, date, "MLB", parlay_ref)

        # Parlay must have graded
        self.assertIsNotNone(result_val, f"Parlay row was not graded; result is None")
        # If parlay graded (any terminal result), PnL must be nonzero
        # (odds_profit for WIN with 1 unit = ~0.909; for LOSS = -1.0; for PUSH = 0)
        # We accept any result but assert: PnL != 0 for WIN and PnL == -1 for LOSS,
        # and specifically that zero-ing logic was NOT applied to parlay rows.
        # The simplest safe assertion: if result is WIN, pnl > 0; if LOSS, pnl < 0.
        if result_val == "WIN":
            self.assertGreater(pnl_val, 0, f"Parlay WIN must have positive PnL; got {pnl_val}")
        elif result_val == "LOSS":
            self.assertLess(pnl_val, 0, f"Parlay LOSS must have negative PnL; got {pnl_val}")
        elif result_val == "PUSH":
            # PUSH → 0 is correct; test still passes (zero is expected for push)
            pass
        else:
            # Parlay abstained (incomplete legs) — acceptable; skip pnl assertion
            pass

    # -----------------------------------------------------------------------
    # Test 6: odds_profit helper is UNCHANGED (used by slip path)
    # -----------------------------------------------------------------------

    def test_odds_profit_win_returns_positive(self) -> None:
        """odds_profit must return a positive value for WIN (slip-grading path still uses it)."""
        result = odds_profit("WIN", 1.0, None)
        self.assertGreater(result, 0, f"odds_profit('WIN', 1.0, None) must be > 0, got {result}")

    def test_odds_profit_loss_returns_negative(self) -> None:
        """odds_profit must return a negative value for LOSS."""
        result = odds_profit("LOSS", 1.0, None)
        self.assertLess(result, 0, f"odds_profit('LOSS', 1.0, None) must be < 0, got {result}")


if __name__ == "__main__":
    unittest.main()
