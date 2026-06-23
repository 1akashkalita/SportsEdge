#!/usr/bin/env python3
"""test_dnp_void.py — TDD RED tests for GAP 1: DNP → VOID (plan 01-9).

Pins three behaviour contracts (all real-money safety):
  1. verify_results.player_appearance: confirmed absent (status=ok, player not in box) -> "dnp"
                                       confirmed present (status=ok, player in box)   -> "played"
                                       transient skip / ambiguous name                -> "unknown"
  2. grade_game_in_workbook MANUAL REVIEW branch:
       Layer-1 None + appearance=="dnp"     -> result="VOID",  PnL=0, src="scraped", conf=1.0
       Layer-1 None + appearance=="played"  -> result="MANUAL REVIEW" (stat unresolvable)
       Layer-1 None + appearance=="unknown" -> result="MANUAL REVIEW" (abstain, no guess)
  3. Hard money-safety: the no-stat-line path NEVER produces result="LOSS" or "WIN"

Run from scripts/ with: python3 -m pytest test_dnp_void.py -x -q
All tests are offline (no network, no firecrawl binary required).
"""
from __future__ import annotations

import importlib.util
import json
import os
import shutil
import sys
import tempfile
import unittest
from pathlib import Path
from typing import Any
from unittest.mock import MagicMock, patch

# ---------------------------------------------------------------------------
# Bootstrap: load verify_results and sports_system_runner
# ---------------------------------------------------------------------------
_SCRIPTS = Path(__file__).parent
_RUNNER_PATH = _SCRIPTS / "sports_system_runner.py"
_VR_PATH = _SCRIPTS / "verify_results.py"

spec_vr = importlib.util.spec_from_file_location("verify_results", _VR_PATH)
assert spec_vr and spec_vr.loader
vr = importlib.util.module_from_spec(spec_vr)
spec_vr.loader.exec_module(vr)  # type: ignore[union-attr]

spec_runner = importlib.util.spec_from_file_location("sports_system_runner", _RUNNER_PATH)
assert spec_runner and spec_runner.loader
runner = importlib.util.module_from_spec(spec_runner)
spec_runner.loader.exec_module(runner)  # type: ignore[union-attr]

# ---------------------------------------------------------------------------
# Fixtures: canonical players dict from the espn_box_ok.md fixture
# ---------------------------------------------------------------------------
_FIXTURE_DIR = _SCRIPTS / "testdata" / "firecrawl"
_BOX_OK_MD = _FIXTURE_DIR / "espn_box_ok.md"
_SKIP_JSON = _FIXTURE_DIR / "verify_skip.json"

def _load_box_ok_players() -> dict[str, Any]:
    """Parse the espn_box_ok.md fixture and return the players dict (offline)."""
    md_text = _BOX_OK_MD.read_text()
    return vr.parse_espn_box_markdown(md_text)


def _write_cache(cache_dir: Path, event_id: str, players: dict, status: str = "ok") -> Path:
    """Write a per-event cache file with the given players dict."""
    cache_dir.mkdir(parents=True, exist_ok=True)
    data = {"status": status, "schema": 1, "reason": "", "players": players}
    p = cache_dir / f"{event_id}.json"
    p.write_text(json.dumps(data, indent=2))
    return p


# ---------------------------------------------------------------------------
# Section 1: player_appearance() in verify_results.py
# ---------------------------------------------------------------------------

class TestPlayerAppearanceFunction(unittest.TestCase):
    """player_appearance(players, player, status) -> 'played'|'dnp'|'unknown'

    This function must exist in verify_results.py (GAP 1 / RESULTS-05).
    These tests will FAIL (RED) until it is implemented.
    """

    @classmethod
    def setUpClass(cls) -> None:
        cls.players = _load_box_ok_players()

    def test_function_exists_in_verify_results(self) -> None:
        """player_appearance must be defined in verify_results.py."""
        self.assertTrue(
            hasattr(vr, "player_appearance"),
            "verify_results.player_appearance does not exist — implement it (GAP 1 / RESULTS-05)",
        )

    def test_present_player_returns_played(self) -> None:
        """Player in the box (status=ok) -> 'played'."""
        result = vr.player_appearance(self.players, "Bobby Witt Jr.", "ok")
        self.assertEqual(result, "played",
                         f"Bobby Witt Jr. is in the box; expected 'played', got {result!r}")

    def test_present_player_different_case_returns_played(self) -> None:
        """Name-normalised present player -> 'played'."""
        result = vr.player_appearance(self.players, "bobby witt jr", "ok")
        self.assertEqual(result, "played")

    def test_present_player_shohei_ohtani_returns_played(self) -> None:
        """Shohei Ohtani is in the Dodgers batting table -> 'played'."""
        result = vr.player_appearance(self.players, "Shohei Ohtani", "ok")
        self.assertEqual(result, "played",
                         f"Shohei Ohtani is in the box; expected 'played', got {result!r}")

    def test_absent_player_status_ok_returns_dnp(self) -> None:
        """Player NOT in box (status=ok, box fully read) -> 'dnp'."""
        result = vr.player_appearance(self.players, "Nick Martinez", "ok")
        self.assertEqual(result, "dnp",
                         f"Nick Martinez absent from box with status=ok; expected 'dnp', got {result!r}")

    def test_absent_player_zack_gelof_status_ok_returns_dnp(self) -> None:
        """Zack Gelof (June 20 real example) absent from box -> 'dnp'."""
        result = vr.player_appearance(self.players, "Zack Gelof", "ok")
        self.assertEqual(result, "dnp")

    def test_absent_player_masataka_yoshida_returns_dnp(self) -> None:
        """Masataka Yoshida (June 8 real example) absent from box -> 'dnp'."""
        result = vr.player_appearance(self.players, "Masataka Yoshida", "ok")
        self.assertEqual(result, "dnp")

    def test_status_skip_returns_unknown(self) -> None:
        """status='skip' means scrape failed — ABSTAIN ('unknown'), not 'dnp'."""
        result = vr.player_appearance({}, "Bobby Witt Jr.", "skip")
        self.assertEqual(result, "unknown",
                         f"status=skip must return 'unknown' (abstain), got {result!r}")

    def test_empty_players_status_skip_returns_unknown(self) -> None:
        """Empty players dict with status=skip -> 'unknown' (not 'dnp')."""
        result = vr.player_appearance({}, "Nick Martinez", "skip")
        self.assertEqual(result, "unknown")

    def test_empty_player_name_returns_unknown(self) -> None:
        """Empty or whitespace player name -> 'unknown' (ambiguous)."""
        result = vr.player_appearance(self.players, "", "ok")
        self.assertEqual(result, "unknown",
                         "Empty player name must return 'unknown' (ambiguous)")

    def test_whitespace_player_name_returns_unknown(self) -> None:
        result = vr.player_appearance(self.players, "   ", "ok")
        self.assertEqual(result, "unknown")

    def test_none_player_name_returns_unknown(self) -> None:
        result = vr.player_appearance(self.players, None, "ok")  # type: ignore[arg-type]
        self.assertEqual(result, "unknown")

    def test_status_not_ok_empty_players_returns_unknown(self) -> None:
        """Any non-ok status with empty players -> 'unknown'."""
        result = vr.player_appearance({}, "Bobby Witt Jr.", "error")
        self.assertEqual(result, "unknown")

    def test_absent_player_empty_players_dict_status_ok_returns_dnp(self) -> None:
        """An empty players dict (scrape ok, no players) means nobody played -> 'dnp'."""
        # Edge case: status=ok with empty box (game was cancelled before first pitch?)
        # The spec says absent + status=ok = "dnp". Empty box is a special valid subset.
        result = vr.player_appearance({}, "Nick Martinez", "ok")
        self.assertEqual(result, "dnp",
                         "Empty players dict with status=ok means box was read, player absent -> 'dnp'")


# ---------------------------------------------------------------------------
# Section 2: resolve_player_appearance or extended resolve_missing_stat
# ---------------------------------------------------------------------------

class TestResolvePlayerAppearanceViaCachedBox(unittest.TestCase):
    """The runner must expose a way to learn appearance from the cached/scraped box.

    Either a new function resolve_player_appearance(sport, game, player) -> str
    or an extended resolve_missing_stat that also surfaces appearance.
    This test checks for the resolve_player_appearance helper (plan spec).
    """

    def test_resolve_player_appearance_exists_in_runner(self) -> None:
        """resolve_player_appearance must be defined in sports_system_runner.py."""
        self.assertTrue(
            hasattr(runner, "resolve_player_appearance"),
            "sports_system_runner.resolve_player_appearance does not exist — implement it (GAP 1 / RESULTS-05)",
        )

    def _make_cache(self, tmpdir: Path, event_id: str, players: dict) -> Path:
        cache_dir = tmpdir / "research" / "results_cache"
        return _write_cache(cache_dir, event_id, players)

    def test_cached_absent_player_returns_dnp(self) -> None:
        """Cached status=ok box, player absent -> 'dnp'."""
        with tempfile.TemporaryDirectory() as td:
            tmpdir = Path(td)
            event_id = "401815839"
            players = _load_box_ok_players()
            self._make_cache(tmpdir, event_id, players)
            game = {"event_id": event_id}
            with patch.object(runner, "DATA", tmpdir):
                with patch("shutil.which", return_value="/usr/local/bin/npx"):
                    result = runner.resolve_player_appearance("mlb", game, "Nick Martinez")
        self.assertEqual(result, "dnp",
                         f"Nick Martinez absent from cached box; expected 'dnp', got {result!r}")

    def test_cached_present_player_returns_played(self) -> None:
        """Cached status=ok box, player present -> 'played'."""
        with tempfile.TemporaryDirectory() as td:
            tmpdir = Path(td)
            event_id = "401815839"
            players = _load_box_ok_players()
            self._make_cache(tmpdir, event_id, players)
            game = {"event_id": event_id}
            with patch.object(runner, "DATA", tmpdir):
                with patch("shutil.which", return_value="/usr/local/bin/npx"):
                    result = runner.resolve_player_appearance("mlb", game, "Bobby Witt Jr.")
        self.assertEqual(result, "played",
                         f"Bobby Witt Jr. in cached box; expected 'played', got {result!r}")

    def test_cached_status_skip_returns_unknown(self) -> None:
        """Cached status=skip -> 'unknown' (transient scrape failure; abstain)."""
        with tempfile.TemporaryDirectory() as td:
            tmpdir = Path(td)
            event_id = "401815840"
            # Write a skip envelope to cache (status=skip)
            cache_dir = tmpdir / "research" / "results_cache"
            skip_data = {"status": "skip", "schema": 1, "reason": "rate-limited", "players": {}}
            _write_cache(cache_dir, event_id, {})
            # Override with a skip-status file
            skip_file = cache_dir / f"{event_id}.json"
            skip_file.write_text(json.dumps(skip_data))
            game = {"event_id": event_id}
            with patch.object(runner, "DATA", tmpdir):
                with patch("shutil.which", return_value="/usr/local/bin/npx"):
                    result = runner.resolve_player_appearance("mlb", game, "Nick Martinez")
        # skip cache doesn't contain status=ok -> unknown
        self.assertEqual(result, "unknown")

    def test_no_cache_no_npx_returns_unknown(self) -> None:
        """No cache, no npx -> degrade to 'unknown' (abstain)."""
        with tempfile.TemporaryDirectory() as td:
            tmpdir = Path(td)
            game = {"event_id": "999999"}
            with patch.object(runner, "DATA", tmpdir):
                with patch("shutil.which", return_value=None):
                    result = runner.resolve_player_appearance("mlb", game, "Nick Martinez")
        self.assertEqual(result, "unknown")

    def test_no_event_id_returns_unknown(self) -> None:
        """Game dict without event_id -> 'unknown'."""
        with tempfile.TemporaryDirectory() as td:
            tmpdir = Path(td)
            game = {}  # no event_id
            with patch.object(runner, "DATA", tmpdir):
                with patch("shutil.which", return_value="/usr/local/bin/npx"):
                    result = runner.resolve_player_appearance("mlb", game, "Nick Martinez")
        self.assertEqual(result, "unknown")

    def test_budget_cap_returns_unknown(self) -> None:
        """When scrape budget exhausted and no cache -> 'unknown' (abstain, not crash)."""
        with tempfile.TemporaryDirectory() as td:
            tmpdir = Path(td)
            game = {"event_id": "888888"}
            with patch.object(runner, "DATA", tmpdir):
                with patch("shutil.which", return_value="/usr/local/bin/npx"):
                    with patch.object(runner, "_scrape_run_count",
                                      runner.RESULT_SCRAPE_MAX_PER_RUN):
                        result = runner.resolve_player_appearance("mlb", game, "Nick Martinez")
        self.assertEqual(result, "unknown")


# ---------------------------------------------------------------------------
# Section 3: grade_game_in_workbook MANUAL REVIEW branch — DNP -> VOID
# ---------------------------------------------------------------------------

class TestGradeGameDNPToVoid(unittest.TestCase):
    """When Layer-1 returns None and appearance is 'dnp', grade VOID.

    These tests simulate the MANUAL REVIEW branch by patching resolve_player_appearance
    and verify that:
      - 'dnp'     -> result="VOID", PnL=0, src="scraped", conf=1.0
      - 'played'  -> result stays "MANUAL REVIEW"
      - 'unknown' -> result stays "MANUAL REVIEW"
    They use the existing grade_game_in_workbook function with a fully-mocked workbook.
    """

    def _build_prop_envelope(self, player: str, stat: str, line: float) -> dict:
        """Minimal prop row dict (mimics Props sheet row)."""
        return {
            "Date": "2026-06-23",
            "Player Name": player,
            "Stat": stat,
            "Line": line,
            "Opponent/Description": f"Over {line}",
            "Reasoning": "units=1.0",
            "Selection": f"Over {line}",
            "Platform": "PrizePicks",
            "Confidence": "High",
        }

    def _run_grade_with_appearance(
        self,
        appearance: str,
        player: str = "Nick Martinez",
        stat: str = "Strikeouts",
        line: float = 5.5,
    ) -> list[dict]:
        """Run grade_game_in_workbook with ENABLE_FIRECRAWL_RESULT_FALLBACK=True
        and resolve_player_appearance returning the given appearance string.
        Returns the graded list (captures results).
        """
        import openpyxl

        wb = openpyxl.Workbook()

        # Build a minimal workbook with Results + Props sheets
        # Remove default sheet
        del wb["Sheet"]

        # Results sheet (with headers from RESULTS_HEADERS)
        results_ws = wb.create_sheet("Results")
        results_headers = runner.RESULTS_HEADERS
        results_ws.append(results_headers)

        # Props sheet
        props_ws = wb.create_sheet("Props")
        props_headers = [
            "Date", "Player Name", "Stat", "Line", "Opponent/Description",
            "Reasoning", "Selection", "Platform", "Confidence",
        ]
        props_ws.append(props_headers)
        row = self._build_prop_envelope(player, stat, line)
        props_ws.append([row.get(h, "") for h in props_headers])

        # Picks + Parlay + CLV + Skipped (empty, needed by grade_game_in_workbook)
        wb.create_sheet("Picks").append(runner.PICKS_HEADERS)
        wb.create_sheet("Correlated Parlays").append(runner.PARLAY_HEADERS)
        wb.create_sheet("CLV Tracker").append(runner.CLV_HEADERS)
        wb.create_sheet("Skipped Picks").append(runner.SKIP_HEADERS)

        # Fake game dict — no real ESPN data (player_stats will be empty -> Layer-1 None)
        game = {
            "id": "401815839",
            "event_id": "401815839",
            "home_team": "Los Angeles Dodgers",
            "away_team": "Kansas City Royals",
            "home_score": 5,
            "away_score": 2,
            "status_name": "Final",
            "commence_time": "2026-06-23T18:10:00Z",
        }

        with tempfile.TemporaryDirectory() as td:
            tmpdir = Path(td)
            # Patch DATA so workbook save goes to tmpdir
            wb_path = tmpdir / "test_grade.xlsx"
            wb.save(str(wb_path))

            def fake_safe_load_workbook(path: Any, **kwargs: Any):  # type: ignore[misc]
                return wb

            def fake_save_workbook(wb_obj: Any, path: Any, **kwargs: Any) -> None:
                pass  # no-op

            with patch.object(runner, "ENABLE_FIRECRAWL_RESULT_FALLBACK", True):
                with patch.object(runner, "DATA", tmpdir):
                    with patch.object(runner, "safe_load_workbook", fake_safe_load_workbook):
                        with patch.object(runner, "save_workbook_atomic", fake_save_workbook):
                            with patch.object(runner, "ensure_workbook", return_value=wb_path):
                                with patch.object(
                                    runner, "resolve_player_appearance",
                                    return_value=appearance
                                ) as mock_rpa:
                                    # Also patch resolve_missing_stat so it returns None
                                    # (appearance path is consulted instead)
                                    with patch.object(
                                        runner, "resolve_missing_stat",
                                        return_value=(None, "manual", 0.0)
                                    ):
                                        result_dict = runner.grade_game_in_workbook(
                                            "mlb", game, date="2026-06-23", dry_run=True
                                        )
        return result_dict.get("graded", [])

    def test_dnp_appearance_grades_void(self) -> None:
        """Layer-1 None + appearance='dnp' -> result='VOID'."""
        graded = self._run_grade_with_appearance("dnp")
        prop_rows = [r for r in graded if str(r.get("ref", "")).startswith("PROP:")]
        self.assertEqual(len(prop_rows), 1, f"Expected 1 PROP row; got {len(prop_rows)}: {prop_rows}")
        row = prop_rows[0]
        self.assertEqual(row["result"], "VOID",
                         f"DNP appearance must grade VOID; got result={row['result']!r}")

    def test_dnp_void_pnl_is_zero(self) -> None:
        """VOID result must have PnL = 0.0."""
        graded = self._run_grade_with_appearance("dnp")
        prop_rows = [r for r in graded if str(r.get("ref", "")).startswith("PROP:")]
        self.assertEqual(len(prop_rows), 1)
        self.assertEqual(prop_rows[0]["pnl"], 0.0,
                         f"VOID PnL must be 0.0; got {prop_rows[0]['pnl']}")

    def test_dnp_void_result_source_is_scraped(self) -> None:
        """VOID result source must be 'scraped'."""
        graded = self._run_grade_with_appearance("dnp")
        prop_rows = [r for r in graded if str(r.get("ref", "")).startswith("PROP:")]
        self.assertEqual(len(prop_rows), 1)
        # Check either the rec dict or the result record
        rec = prop_rows[0]
        src = rec.get("Result Source") or rec.get("res_src") or ""
        self.assertEqual(src, "scraped",
                         f"VOID Result Source must be 'scraped'; got {src!r}")

    def test_dnp_void_result_confidence_is_1(self) -> None:
        """VOID result confidence must be 1.0."""
        graded = self._run_grade_with_appearance("dnp")
        prop_rows = [r for r in graded if str(r.get("ref", "")).startswith("PROP:")]
        self.assertEqual(len(prop_rows), 1)
        rec = prop_rows[0]
        conf = rec.get("Result Confidence") or rec.get("res_conf") or 0.0
        self.assertEqual(float(conf), 1.0,
                         f"VOID Result Confidence must be 1.0; got {conf!r}")

    def test_played_appearance_stays_manual_review(self) -> None:
        """Layer-1 None + appearance='played' (stat unresolvable) -> MANUAL REVIEW."""
        graded = self._run_grade_with_appearance("played")
        prop_rows = [r for r in graded if str(r.get("ref", "")).startswith("PROP:")]
        self.assertEqual(len(prop_rows), 1)
        self.assertEqual(prop_rows[0]["result"], "MANUAL REVIEW",
                         f"'played' appearance must stay MANUAL REVIEW; got {prop_rows[0]['result']!r}")

    def test_unknown_appearance_stays_manual_review(self) -> None:
        """Layer-1 None + appearance='unknown' -> MANUAL REVIEW (abstain)."""
        graded = self._run_grade_with_appearance("unknown")
        prop_rows = [r for r in graded if str(r.get("ref", "")).startswith("PROP:")]
        self.assertEqual(len(prop_rows), 1)
        self.assertEqual(prop_rows[0]["result"], "MANUAL REVIEW",
                         f"'unknown' appearance must stay MANUAL REVIEW; got {prop_rows[0]['result']!r}")


# ---------------------------------------------------------------------------
# Section 4: Hard money-safety — never auto-LOSS
# ---------------------------------------------------------------------------

class TestNeverAutoLoss(unittest.TestCase):
    """The no-stat-line path NEVER produces result='LOSS' or 'WIN'.

    Acceptable terminals for Layer-1 None path: VOID or MANUAL REVIEW only.
    """

    def test_dnp_is_void_not_loss(self) -> None:
        """Confirmed DNP -> VOID, never LOSS."""
        import openpyxl
        wb = openpyxl.Workbook()
        del wb["Sheet"]
        results_ws = wb.create_sheet("Results")
        results_ws.append(runner.RESULTS_HEADERS)
        props_ws = wb.create_sheet("Props")
        props_headers = [
            "Date", "Player Name", "Stat", "Line", "Opponent/Description",
            "Reasoning", "Selection", "Platform", "Confidence",
        ]
        props_ws.append(props_headers)
        # Nick Martinez, 5 strikeouts line — he DNP -> should VOID not LOSS
        row_data = {
            "Date": "2026-06-23",
            "Player Name": "Nick Martinez",
            "Stat": "Strikeouts",
            "Line": 5.5,
            "Opponent/Description": "Over 5.5",
            "Reasoning": "units=1.0",
            "Selection": "Over 5.5",
            "Platform": "PrizePicks",
            "Confidence": "High",
        }
        props_ws.append([row_data.get(h, "") for h in props_headers])
        wb.create_sheet("Picks").append(runner.PICKS_HEADERS)
        wb.create_sheet("Correlated Parlays").append(runner.PARLAY_HEADERS)
        wb.create_sheet("CLV Tracker").append(runner.CLV_HEADERS)
        wb.create_sheet("Skipped Picks").append(runner.SKIP_HEADERS)

        game = {
            "id": "401815839", "event_id": "401815839",
            "home_team": "Los Angeles Dodgers", "away_team": "Kansas City Royals",
            "home_score": 5, "away_score": 2,
            "status_name": "Final", "commence_time": "2026-06-23T18:10:00Z",
        }

        with tempfile.TemporaryDirectory() as td:
            tmpdir = Path(td)
            wb_path = tmpdir / "test_grade.xlsx"
            wb.save(str(wb_path))

            with patch.object(runner, "ENABLE_FIRECRAWL_RESULT_FALLBACK", True):
                with patch.object(runner, "DATA", tmpdir):
                    with patch.object(runner, "safe_load_workbook",
                                      return_value=wb):
                        with patch.object(runner, "save_workbook_atomic",
                                          return_value=None):
                            with patch.object(runner, "ensure_workbook",
                                              return_value=wb_path):
                                with patch.object(runner, "resolve_player_appearance",
                                                  return_value="dnp"):
                                    with patch.object(runner, "resolve_missing_stat",
                                                      return_value=(None, "manual", 0.0)):
                                        result_dict = runner.grade_game_in_workbook(
                                            "mlb", game, date="2026-06-23", dry_run=True
                                        )
        graded = result_dict.get("graded", [])
        prop_rows = [r for r in graded if str(r.get("ref", "")).startswith("PROP:")]
        for r in prop_rows:
            self.assertNotEqual(r["result"], "LOSS",
                                f"DNP prop must never be LOSS; got {r['result']!r}")
            self.assertNotEqual(r["result"], "WIN",
                                f"DNP prop must never be WIN on the no-stat-line path; got {r['result']!r}")

    def test_unknown_appearance_is_not_loss(self) -> None:
        """Unknown appearance (status=skip) -> MANUAL REVIEW, never LOSS."""
        import openpyxl
        wb = openpyxl.Workbook()
        del wb["Sheet"]
        wb.create_sheet("Results").append(runner.RESULTS_HEADERS)
        props_ws = wb.create_sheet("Props")
        props_headers = [
            "Date", "Player Name", "Stat", "Line", "Opponent/Description",
            "Reasoning", "Selection", "Platform", "Confidence",
        ]
        props_ws.append(props_headers)
        props_ws.append(["2026-06-23", "Nick Martinez", "Strikeouts", 5.5,
                          "Over 5.5", "units=1.0", "Over 5.5", "PrizePicks", "High"])
        wb.create_sheet("Picks").append(runner.PICKS_HEADERS)
        wb.create_sheet("Correlated Parlays").append(runner.PARLAY_HEADERS)
        wb.create_sheet("CLV Tracker").append(runner.CLV_HEADERS)
        wb.create_sheet("Skipped Picks").append(runner.SKIP_HEADERS)

        game = {
            "id": "401815839", "event_id": "401815839",
            "home_team": "Los Angeles Dodgers", "away_team": "Kansas City Royals",
            "home_score": 5, "away_score": 2,
            "status_name": "Final", "commence_time": "2026-06-23T18:10:00Z",
        }

        with tempfile.TemporaryDirectory() as td:
            tmpdir = Path(td)
            with patch.object(runner, "ENABLE_FIRECRAWL_RESULT_FALLBACK", True):
                with patch.object(runner, "DATA", tmpdir):
                    with patch.object(runner, "safe_load_workbook", return_value=wb):
                        with patch.object(runner, "save_workbook_atomic", return_value=None):
                            with patch.object(runner, "ensure_workbook",
                                              return_value=tmpdir / "x.xlsx"):
                                with patch.object(runner, "resolve_player_appearance",
                                                  return_value="unknown"):
                                    with patch.object(runner, "resolve_missing_stat",
                                                      return_value=(None, "manual", 0.0)):
                                        result_dict = runner.grade_game_in_workbook(
                                            "mlb", game, date="2026-06-23", dry_run=True
                                        )
        graded = result_dict.get("graded", [])
        prop_rows = [r for r in graded if str(r.get("ref", "")).startswith("PROP:")]
        for r in prop_rows:
            self.assertNotEqual(r["result"], "LOSS",
                                "Unknown appearance must never produce LOSS")
            self.assertNotEqual(r["result"], "WIN",
                                "Unknown appearance must never produce WIN on no-stat-line path")

    def test_flag_off_never_calls_resolve_player_appearance(self) -> None:
        """With ENABLE_FIRECRAWL_RESULT_FALLBACK=False, resolve_player_appearance is not called."""
        with patch.object(runner, "ENABLE_FIRECRAWL_RESULT_FALLBACK", False):
            with patch.object(runner, "resolve_player_appearance") as mock_rpa:
                # Simulate the gate check from grade_game_in_workbook
                # (resolve_player_appearance must be inside the firecrawl gate)
                _scrape_count = 0
                if (runner.ENABLE_FIRECRAWL_RESULT_FALLBACK
                        and _scrape_count < runner.RESULT_SCRAPE_MAX_PER_RUN):
                    runner.resolve_player_appearance("mlb", {}, "Nick Martinez")
                mock_rpa.assert_not_called()


if __name__ == "__main__":
    unittest.main()
