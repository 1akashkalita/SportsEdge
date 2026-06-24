#!/usr/bin/env python3
"""Reproduce the wrong-team prop-binding bug in grade_game_in_workbook.

Symptom (from production): a player prop whose `Team` is an Underdog UUID
(unresolvable by team_aliases) gets bound to the WRONG same-night game by the
5-minute start-time-window fallback in game_matches_row. The player is absent
from that game's box score, so grading emits a spurious
"⚠️ MANUAL REVIEW: ... not found in ESPN box score" Telegram alert.

Real example: Pete Crow-Armstrong (Chicago Cub) was graded against
"Milwaukee Brewers @ Cincinnati Reds".

These tests are fully offline: espn_player_stats_by_event is monkeypatched to
return a per-event box score, and all workbook I/O is stubbed.

Run from scripts/ with: python3 -m pytest test_prop_wrong_team_binding.py -x -q
"""
from __future__ import annotations

import importlib.util
import tempfile
import unittest
from pathlib import Path
from typing import Any
from unittest.mock import patch

import openpyxl

_SCRIPTS = Path(__file__).parent
_RUNNER_PATH = _SCRIPTS / "sports_system_runner.py"
spec_runner = importlib.util.spec_from_file_location("sports_system_runner", _RUNNER_PATH)
assert spec_runner and spec_runner.loader
runner = importlib.util.module_from_spec(spec_runner)
spec_runner.loader.exec_module(runner)  # type: ignore[union-attr]

DATE = "2026-06-23"
# Pete Crow-Armstrong's Underdog team UUID — team_aliases cannot resolve it.
PCA_UUID = "0a2c9da4-7227-4055-bf48-bb4bf5c8f410"
PROP_START = "2026-06-23T23:40:00Z"

# Wrong game: shares the prop's start time, so the start-time-window fallback
# would (incorrectly) bind the prop here. PCA is NOT in this box.
GAME_WRONG = {
    "id": "EVENT_WRONG", "event_id": "EVENT_WRONG",
    "home_team": "Cincinnati Reds", "away_team": "Milwaukee Brewers",
    "home_score": 4, "away_score": 3, "status_name": "Final",
    "status": "final", "completed": True,
    "start_time": PROP_START, "commence_time": PROP_START,
}
# Correct game: different start time; PCA IS in this box with 2 hits.
GAME_RIGHT = {
    "id": "EVENT_RIGHT", "event_id": "EVENT_RIGHT",
    "home_team": "St. Louis Cardinals", "away_team": "Chicago Cubs",
    "home_score": 6, "away_score": 5, "status_name": "Final",
    "status": "final", "completed": True,
    "start_time": "2026-06-23T20:10:00Z", "commence_time": "2026-06-23T20:10:00Z",
}


def _fake_box(sport: str, event_id: str) -> dict[str, Any]:
    eid = str(event_id)
    if eid == "EVENT_RIGHT":
        return {"pete crow-armstrong": {"batting": {"hits": 2.0, "h": 2.0}}}
    if eid == "EVENT_WRONG":
        return {"william contreras": {"batting": {"hits": 1.0, "h": 1.0}}}
    return {}


def _build_workbook() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    del wb["Sheet"]
    wb.create_sheet("Results").append(runner.RESULT_HEADERS)
    props_ws = wb.create_sheet("Props")
    props_headers = [
        "Date", "Sport", "Player Name", "Team", "Stat", "Line",
        "Opponent/Description", "Reasoning", "Selection", "Platform",
        "Confidence", "Start Time UTC",
    ]
    props_ws.append(props_headers)
    # Pete Crow-Armstrong Over 0.5 Hits — Underdog UUID team, NO Game ID.
    props_ws.append([
        DATE, "MLB", "Pete Crow-Armstrong", PCA_UUID, "Hits", 0.5,
        "Pete Crow-Armstrong Over 0.5 Hits", "units=1.0", "Over 0.5",
        "Underdog", "High", PROP_START,
    ])
    wb.create_sheet("Picks").append(runner.PICKS_HEADERS)
    wb.create_sheet("Correlated Parlays").append(runner.PARLAY_HEADERS)
    wb.create_sheet("CLV Tracker").append(runner.CLV_HEADERS)
    wb.create_sheet("Skipped Picks").append(runner.SKIPPED_PICK_HEADERS)
    return wb


def _grade(game: dict[str, Any]) -> list[dict[str, Any]]:
    """Grade one game against a fresh single-prop workbook; return graded rows."""
    wb = _build_workbook()
    with tempfile.TemporaryDirectory() as td:
        tmpdir = Path(td)
        wb_path = tmpdir / "wb.xlsx"
        wb.save(str(wb_path))
        with patch.object(runner, "ENABLE_FIRECRAWL_RESULT_FALLBACK", False), \
             patch.object(runner, "espn_player_stats_by_event", _fake_box), \
             patch.object(runner, "safe_load_workbook", return_value=wb), \
             patch.object(runner, "save_workbook_atomic", return_value=None), \
             patch.object(runner, "ensure_workbook", return_value=wb_path), \
             patch.object(runner, "sync_master_and_bankroll",
                          return_value={"bankroll": {}, "daily_rows": [], "day_pnl": 0, "current": 100}):
            result = runner.grade_game_in_workbook("mlb", game, date=DATE)
    return [r for r in result.get("graded", []) if str(r.get("ref", "")).startswith("PROP:")]


class TestPropWrongTeamBinding(unittest.TestCase):
    def test_prop_not_bound_to_wrong_same_time_game(self) -> None:
        """Grading the Brewers@Reds game must NOT grade Pete Crow-Armstrong's prop.

        PCA is not in that game's box score and the prop is not reliably bound
        to it; binding only via the start-time window is the bug. The prop must
        be skipped here (no spurious MANUAL REVIEW), not graded against the wrong
        team.
        """
        prop_rows = _grade(GAME_WRONG)
        self.assertEqual(
            prop_rows, [],
            f"Prop was mis-bound to the wrong game and graded: {prop_rows}",
        )

    def test_prop_graded_against_correct_game(self) -> None:
        """Grading the Cubs@Cardinals game must grade the prop WIN (2 hits > 0.5)."""
        prop_rows = _grade(GAME_RIGHT)
        self.assertEqual(len(prop_rows), 1, f"Expected exactly one graded prop; got {prop_rows}")
        self.assertEqual(prop_rows[0]["result"], "WIN", f"Expected WIN; got {prop_rows[0]}")


class TestPropBelongsToGameContract(unittest.TestCase):
    """Pin the reliable-identity binding contract directly."""

    def test_game_id_match_binds(self) -> None:
        game = {"event_id": "E1", "home_team": "Chicago Cubs", "away_team": "St. Louis Cardinals"}
        row = {"Game ID": "E1", "Team": PCA_UUID, "Player Name": "Whoever"}
        self.assertTrue(runner.prop_belongs_to_game(game, row, {}))

    def test_nickname_team_binds(self) -> None:
        """Bare nickname 'Yankees' must bind to a game whose team is 'Yankees'."""
        game = {"event_id": "E1", "home_team": "Yankees", "away_team": "Red Sox"}
        row = {"Team": "Yankees", "Player Name": "Aaron Judge"}
        self.assertTrue(runner.prop_belongs_to_game(game, row, {}))

    def test_abbreviation_team_binds(self) -> None:
        game = {"event_id": "E1", "home_team": "Chicago Cubs", "away_team": "St. Louis Cardinals"}
        row = {"Team": "CHC", "Player Name": "Whoever"}
        self.assertTrue(runner.prop_belongs_to_game(game, row, {}))

    def test_uuid_team_does_not_bind_unrelated_game(self) -> None:
        """A UUID team with the player absent from the box must NOT bind."""
        game = {"event_id": "E1", "home_team": "Milwaukee Brewers", "away_team": "Cincinnati Reds"}
        row = {"Team": PCA_UUID, "Player Name": "Pete Crow-Armstrong",
               "Start Time UTC": PROP_START}
        box = {"william contreras": {"batting": {"hits": 1.0}}}
        self.assertFalse(runner.prop_belongs_to_game(game, row, box))

    def test_uuid_team_binds_via_box_membership(self) -> None:
        game = {"event_id": "E2", "home_team": "St. Louis Cardinals", "away_team": "Chicago Cubs"}
        row = {"Team": PCA_UUID, "Player Name": "Pete Crow-Armstrong"}
        box = {"pete crow-armstrong": {"batting": {"hits": 2.0}}}
        self.assertTrue(runner.prop_belongs_to_game(game, row, box))

    def test_strict_membership_rejects_same_surname(self) -> None:
        """Don't bind a prop to a game merely because another player shares the surname."""
        box = {"will smith": {"batting": {"hits": 1.0}}}
        self.assertFalse(runner._player_in_boxscore_strict("Dwight Smith", box))

    def test_strict_membership_exact_and_canonical(self) -> None:
        box = {"pete crow-armstrong": {"batting": {"hits": 2.0}}}
        self.assertTrue(runner._player_in_boxscore_strict("Pete Crow-Armstrong", box))
        self.assertTrue(runner._player_in_boxscore_strict("pete crow-armstrong", box))


if __name__ == "__main__":
    unittest.main()
