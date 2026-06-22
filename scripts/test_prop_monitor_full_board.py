#!/usr/bin/env python3
"""Regression tests for prop monitor full-board availability logic."""
from __future__ import annotations

import importlib.util
import json
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))
SCRIPT = SCRIPT_DIR / "sports_system_runner.py"
spec = importlib.util.spec_from_file_location("sports_system_runner", SCRIPT)
runner = importlib.util.module_from_spec(spec)
assert spec and spec.loader
spec.loader.exec_module(runner)


def make_prop(idx: int, odds_type: str = "demon", status: str = "active", line=1.5) -> dict:
    return {
        "projection_id": f"prop-{idx}",
        "player_name": f"Player {idx}",
        "team": "TST",
        "description": "OPP",
        "stat_name": "Hits",
        "stat_type": "hits",
        "line_score": line,
        "odds_type": odds_type,
        "status": status,
        "start_time": "2026-06-10T20:00:00Z",
        "line_timing": "pregame",
        "line_timing_reason": "test",
    }


def build_workbook(path: Path, existing_ids: list[str]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Player Props"
    ws.append(runner.PROPS_HEADERS)
    for pid in existing_ids:
        idx = int(pid.split("-")[1]) if "-" in pid else 999
        ws.append([
            runner.today_str(), "MLB", pid, f"Player {idx}", "TST", "OPP", "Hits", 1.5,
            "standard", "2026-06-10T20:00:00Z", "ACTIVE", "Existing active row",
            "", "", "PrizePicks", "", "", "", "", "", "", "", "", "", "", "",
        ] + ["" for _ in runner.LINE_TIMING_FIELDS])
    wb.save(path)


class PropMonitorFullBoardTests(unittest.TestCase):
    def run_monitor(self, props: list[dict], existing_ids: list[str]):
        tmpdir = tempfile.TemporaryDirectory()
        self.addCleanup(tmpdir.cleanup)
        wb_path = Path(tmpdir.name) / "mlb_test.xlsx"
        build_workbook(wb_path, existing_ids)
        details_dir = Path(tmpdir.name) / "details"

        with patch.object(runner, "ensure_dirs", return_value=None), \
             patch.object(runner, "run_fetch_dfs_props", return_value=None), \
             patch.object(runner, "first_class_dfs_props_latest", return_value=props), \
             patch.object(runner, "fetch_sportsbook_prop_lines", return_value=({}, {"credits_remaining": "DISABLED"})), \
             patch.object(runner, "ensure_workbook", return_value=wb_path), \
             patch.object(runner, "obsidian_append_line_moves", return_value=None), \
             patch.object(runner, "PROP_MONITOR_DIR", details_dir):
            result = runner.prop_monitor("mlb")

        wb = load_workbook(wb_path, read_only=True, data_only=True)
        try:
            ws = wb["Player Props"]
            rows = {ws.cell(r, 3).value: {
                "status": ws.cell(r, 11).value,
                "notes": ws.cell(r, 12).value,
            } for r in range(2, ws.max_row + 1)}
        finally:
            wb.close()
        return result, rows

    def test_existing_prop_outside_first_300_but_on_full_board_stays_active(self) -> None:
        props = [make_prop(i, "demon") for i in range(1, 351)]
        result, rows = self.run_monitor(props, ["prop-350"])
        self.assertEqual(result["full_active_ids_count"], 350)
        self.assertEqual(rows["prop-350"]["status"], "ACTIVE")
        self.assertNotIn("injury", str(rows["prop-350"]["notes"]).lower())
        self.assertEqual(result["board_availability_count"], 0)
        self.assertEqual(result["confirmed_injury_news_count"], 0)

    def test_existing_prop_absent_from_full_board_becomes_board_unavailable_not_injury_watch(self) -> None:
        props = [make_prop(i, "standard") for i in range(1, 6)]
        result, rows = self.run_monitor(props, ["missing-999"])
        self.assertIn(rows["missing-999"]["status"], {"BOARD_UNAVAILABLE", "WATCH_BOARD"})
        self.assertIn("not confirmed injury/news", rows["missing-999"]["notes"])
        self.assertEqual(result["board_availability_count"], 1)
        self.assertEqual(result["confirmed_injury_news_count"], 0)
        self.assertEqual(result.get("injury_watch_alerts", []), [])

    def test_refresh_includes_all_standard_rows_and_not_first_300_mixed_strategy(self) -> None:
        props = [make_prop(i, "demon") for i in range(1, 301)] + [make_prop(i, "standard") for i in range(301, 321)]
        result, rows = self.run_monitor(props, [])
        self.assertEqual(result["active_props_total"], 320)
        self.assertEqual(result["full_active_ids_count"], 320)
        self.assertEqual(result["standard_rows_refreshed"], 20)
        self.assertGreater(result["standard_rows_refreshed"], 0)
        # If first 300 mixed rows were used, no standard rows would be refreshed in this fixture.
        for i in range(301, 321):
            self.assertIn(f"prop-{i}", rows)
            self.assertEqual(rows[f"prop-{i}"]["status"], "ACTIVE")

    def test_workbook_refresh_cap_does_not_affect_disappearance_logic_and_output_is_capped(self) -> None:
        props = [make_prop(i, "standard") for i in range(1, 11)]
        missing = [f"missing-{i}" for i in range(1, 25)]
        result, _rows = self.run_monitor(props, missing)
        self.assertEqual(result["board_availability_count"], 24)
        self.assertLessEqual(len(result["board_availability_examples"]), 10)
        self.assertIn("full_details_path", result)
        self.assertTrue(Path(result["full_details_path"]).exists())
        details = json.loads(Path(result["full_details_path"]).read_text())
        self.assertEqual(len(details["board_availability_alerts"]), 24)

    def test_no_odds_api_player_prop_call_is_introduced_when_disabled(self) -> None:
        props = [make_prop(i, "standard") for i in range(1, 3)]
        with patch.object(runner, "ENABLE_ODDS_API_PLAYER_PROPS", False):
            result, _rows = self.run_monitor(props, [])
        self.assertFalse(result["prop_comparison_sources"]["odds_api_player_props"])
        self.assertEqual(result["credits_remaining"], "DISABLED")


if __name__ == "__main__":
    unittest.main()
