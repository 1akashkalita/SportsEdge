"""Regression tests for skipped_picks_summary_for_date (quick 260623-lzi).

The function previously read the Skipped Picks sheet with per-row ws.cell() calls
on a read-only openpyxl worksheet, which re-parses the whole sheet XML on every
call (O(n^2)) and made build_recap_alert hang at ~1.5k rows. It now does a single
ws.iter_rows() streaming pass. These tests lock the output contract AND guard
against the O(n^2) regression via a large fixture sheet that would hang the old code.
"""
import importlib.util
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

_SCRIPT = Path(__file__).parent / "sports_system_runner.py"
_spec = importlib.util.spec_from_file_location("ssr_skp_perf", _SCRIPT)
runner = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(runner)

HEADER = ["Date", "Pick Ref", "Player", "Stat", "Line", "Gate", "Result"]


def _write_skipped(path: Path, rows: list[list]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Skipped Picks"
    ws.append(HEADER)
    for r in rows:
        ws.append(r)
    wb.save(path)


class SkippedPicksSummaryTests(unittest.TestCase):
    def setUp(self) -> None:
        self._orig_path = runner.workbook_path
        self._orig_valid = runner.workbook_is_valid

    def tearDown(self) -> None:
        runner.workbook_path = self._orig_path
        runner.workbook_is_valid = self._orig_valid

    def _patch(self, tmp: Path, mlb_file: Path) -> None:
        runner.workbook_is_valid = lambda p: True
        runner.workbook_path = lambda sport, date: mlb_file if sport == "mlb" else tmp / f"{sport}_absent.xlsx"

    def test_counts_and_record(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            mlb = tmp / "mlb_2026-06-23.xlsx"
            _write_skipped(mlb, [
                ["2026-06-23", "r1", "A", "Pts", "5", "G1", "WIN"],
                ["2026-06-23", "r2", "B", "Pts", "5", "G1", "LOSS"],
                ["2026-06-23", "r3", "C", "Pts", "5", "G1", ""],     # counts in total, not W/L
                ["2026-06-22", "r4", "D", "Pts", "5", "G1", "WIN"],  # other date — ignored
            ])
            self._patch(tmp, mlb)
            total, record = runner.skipped_picks_summary_for_date("2026-06-23")
        self.assertEqual(total, 3)
        self.assertEqual(record, "1-1")

    def test_empty_and_missing(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            mlb = tmp / "mlb_2026-06-23.xlsx"
            _write_skipped(mlb, [])  # header only
            self._patch(tmp, mlb)
            self.assertEqual(runner.skipped_picks_summary_for_date("2026-06-23"), (0, "0-0"))

    def test_large_sheet_does_not_hang(self) -> None:
        """2000 rows — the old per-cell O(n^2) read would take minutes/hang here."""
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            mlb = tmp / "mlb_2026-06-23.xlsx"
            rows = []
            for i in range(2000):
                res = "WIN" if i % 2 == 0 else "LOSS"
                rows.append(["2026-06-23", f"r{i}", "P", "Pts", "5", "G1", res])
            _write_skipped(mlb, rows)
            self._patch(tmp, mlb)
            total, record = runner.skipped_picks_summary_for_date("2026-06-23")
        self.assertEqual(total, 2000)
        self.assertEqual(record, "1000-1000")


if __name__ == "__main__":
    unittest.main()
