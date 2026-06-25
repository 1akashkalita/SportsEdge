#!/usr/bin/env python3
"""make_demo_data.py — synthetic workbooks for the dashboard demo / screenshots.

The real `data/` tree holds a live bankroll and betting ledger and is never
published. This builds a small, entirely FICTIONAL dataset under `demo_data/`
(git-ignored) so the localhost dashboard can be run and screenshotted with zero
exposure of real money. All players/lines/results/bankroll below are invented.

    cd scripts
    python3 make_demo_data.py                 # write demo_data/ only
    python3 make_demo_data.py --serve --port 8799   # build + serve the dashboard

Reuses the production read-layer's column contract (dashboard_data) so the demo
renders through the real templates unchanged.
"""
from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook

import dashboard_data

ROOT = Path(__file__).resolve().parent.parent
DEMO_ROOT = ROOT / "demo_data"


def _sheet(wb: Workbook, title: str, headers: list[str], rows: list[dict], first: bool):
    ws = wb.active if first else wb.create_sheet()
    ws.title = title
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
    return ws


# --- Fictional picks (today) -------------------------------------------------

def _picks_workbook(sport: str, date: str, approved: list[dict], skipped: list[dict]) -> Workbook:
    wb = Workbook()
    pick_headers = ["Date", "Sport", "Status", "Pick", "Selection", "Platform",
                    "Edge", "Confidence", "EV", "Probability",
                    "Model Over Probability", "What Edge Would Have Been"]
    _sheet(wb, "Picks", pick_headers,
           [{**a, "Date": date, "Sport": sport, "Status": "APPROVED"} for a in approved], True)
    skip_headers = pick_headers + ["Gate Failed"]
    _sheet(wb, "Skipped Picks", skip_headers,
           [{**s, "Date": date, "Sport": sport} for s in skipped], False)
    return wb


def build(demo_root: Path = DEMO_ROOT, date: str | None = None) -> Path:
    date = date or dashboard_data.today_str()
    nba_dir = demo_root / "nba"
    mlb_dir = demo_root / "mlb"
    pnl_dir = demo_root / "pnl"
    for d in (nba_dir, mlb_dir, pnl_dir, pnl_dir / "logs", demo_root / "locks"):
        d.mkdir(parents=True, exist_ok=True)

    # --- NBA today board (all fictional) ---
    nba_appr = [
        {"Pick": "Points", "Selection": "Jalen Brunson OVER 24.5 Points", "Platform": "PrizePicks",
         "Edge": 2.1, "Confidence": "A", "EV": 0.14, "Probability": 0.63, "Model Over Probability": 0.65},
        {"Pick": "Pts+Reb+Ast", "Selection": "Anthony Edwards OVER 38.5 PRA", "Platform": "Underdog",
         "Edge": 1.6, "Confidence": "B", "EV": 0.09, "Probability": 0.59, "Model Over Probability": 0.60},
        {"Pick": "Assists", "Selection": "Tyrese Haliburton OVER 9.5 Assists", "Platform": "PrizePicks",
         "Edge": 1.4, "Confidence": "B", "EV": 0.07, "Probability": 0.58, "Model Over Probability": 0.58},
    ]
    nba_skip = [
        {"Pick": "Rebounds", "Selection": "Victor Wembanyama OVER 11.5 Rebounds", "Platform": "PrizePicks",
         "Edge": 0.3, "Confidence": "C", "EV": -0.02, "Probability": 0.51,
         "What Edge Would Have Been": 0.3, "Gate Failed": "G2 — MINIMUM PROBABILITY"},
        {"Pick": "Threes", "Selection": "Stephen Curry OVER 4.5 Threes", "Platform": "Underdog",
         "Edge": 0.6, "Confidence": "C", "EV": 0.01, "Probability": 0.53,
         "What Edge Would Have Been": 0.6, "Gate Failed": "G5 — PLATFORM LINE AVAILABILITY"},
    ]
    _picks_workbook("NBA", date, nba_appr, nba_skip).save(nba_dir / f"nba_{date}.xlsx")

    # --- MLB today board (all fictional) ---
    mlb_appr = [
        {"Pick": "Strikeouts", "Selection": "Tarik Skubal OVER 7.5 Strikeouts", "Platform": "PrizePicks",
         "Edge": 1.9, "Confidence": "A", "EV": 0.12, "Probability": 0.62, "Model Over Probability": 0.64},
        {"Pick": "Total Bases", "Selection": "Aaron Judge OVER 1.5 Total Bases", "Platform": "Underdog",
         "Edge": 1.2, "Confidence": "B", "EV": 0.06, "Probability": 0.57, "Model Over Probability": 0.57},
    ]
    mlb_skip = [
        {"Pick": "Hits Allowed", "Selection": "Zack Wheeler UNDER 5.5 Hits Allowed", "Platform": "PrizePicks",
         "Edge": 0.4, "Confidence": "C", "EV": -0.01, "Probability": 0.52,
         "What Edge Would Have Been": 0.4, "Gate Failed": "G9 — MARKET DISAGREEMENT"},
    ]
    _picks_workbook("MLB", date, mlb_appr, mlb_skip).save(mlb_dir / f"mlb_{date}.xlsx")

    # --- master_pnl.xlsx: Slip History, Pick History, Bankroll Chart Data ---
    _master_workbook().save(pnl_dir / "master_pnl.xlsx")
    return demo_root


def _master_workbook() -> Workbook:
    wb = Workbook()

    slip_headers = ["Date", "Slip ID", "Slip Type", "Number of Legs", "Legs",
                    "Standard Payout Multiplier", "Slip Result", "Net PnL",
                    "Placed", "Placed At", "Operator Note"]
    slips = [
        {"Date": "2026-06-23", "Slip ID": "2026-06-23:power:7f3a1c", "Slip Type": "Power",
         "Number of Legs": 2, "Legs": "Jalen Brunson OVER 24.5 Pts; Tarik Skubal OVER 7.5 K",
         "Standard Payout Multiplier": 3.0, "Slip Result": "WIN", "Net PnL": 20.0,
         "Placed": "Yes", "Placed At": "2026-06-23 16:42", "Operator Note": "Both A-tier, uncorrelated"},
        {"Date": "2026-06-23", "Slip ID": "2026-06-23:flex:2b9e44", "Slip Type": "Flex",
         "Number of Legs": 3, "Legs": "Edwards OVER 38.5 PRA; Judge OVER 1.5 TB; Haliburton OVER 9.5 Ast",
         "Standard Payout Multiplier": 2.25, "Slip Result": "LOSS", "Net PnL": -10.0,
         "Placed": "Yes", "Placed At": "2026-06-23 16:50", "Operator Note": ""},
        {"Date": "2026-06-22", "Slip ID": "2026-06-22:power:a1d8f0", "Slip Type": "Power",
         "Number of Legs": 2, "Legs": "Skubal OVER 7.5 K; Wheeler UNDER 5.5 HA",
         "Standard Payout Multiplier": 3.0, "Slip Result": "WIN", "Net PnL": 20.0,
         "Placed": "Yes", "Placed At": "2026-06-22 17:05", "Operator Note": "Correlated pitcher game"},
    ]
    ws = wb.active
    ws.title = "Slip History"
    ws.append(slip_headers)
    for s in slips:
        ws.append([s.get(h, "") for h in slip_headers])

    # Pick History — fictional W/L ledger for the history page aggregates
    ph_headers = ["Date", "Sport", "Confidence Tier", "Result", "Units", "PnL"]
    ph_rows = []
    pattern = [  # (sport, tier, result)
        ("NBA", "A", "WIN"), ("NBA", "A", "WIN"), ("NBA", "B", "LOSS"), ("NBA", "B", "WIN"),
        ("NBA", "C", "LOSS"), ("NBA", "A", "WIN"), ("NBA", "B", "PUSH"), ("NBA", "C", "WIN"),
        ("MLB", "A", "WIN"), ("MLB", "A", "LOSS"), ("MLB", "B", "WIN"), ("MLB", "B", "WIN"),
        ("MLB", "C", "LOSS"), ("MLB", "A", "WIN"), ("MLB", "C", "WIN"), ("MLB", "B", "LOSS"),
    ]
    for i, (sport, tier, res) in enumerate(pattern):
        pnl = 1.0 if res == "WIN" else (-1.0 if res == "LOSS" else 0.0)
        ph_rows.append({"Date": f"2026-06-{8 + i // 2:02d}", "Sport": sport,
                        "Confidence Tier": tier, "Result": res, "Units": 1.0, "PnL": pnl})
    ws2 = wb.create_sheet("Pick History")
    ws2.append(ph_headers)
    for r in ph_rows:
        ws2.append([r.get(h, "") for h in ph_headers])

    # Bankroll Chart Data — fictional curve starting at 100u
    ws3 = wb.create_sheet("Bankroll Chart Data")
    ws3.append(["Date", "Bankroll", "ROI"])
    curve = [100, 101, 99, 103, 106, 104, 108, 107, 111, 109,
             113, 112, 116, 115, 118, 121, 119, 123, 122, 126]
    for i, bank in enumerate(curve):
        d = f"2026-06-{6 + i:02d}"
        roi = round((bank - 100) / 100, 4)
        ws3.append([d, bank, roi])
    return wb


def serve(demo_root: Path = DEMO_ROOT, port: int = 8799) -> int:
    """Point the read-layer at demo_root and run the dashboard (demo only)."""
    dashboard_data.DATA = demo_root
    dashboard_data.NBA_DIR = demo_root / "nba"
    dashboard_data.MLB_DIR = demo_root / "mlb"
    dashboard_data.PNL_DIR = demo_root / "pnl"
    dashboard_data.LOCK_DIR = demo_root / "locks"
    dashboard_data.RUN_LOG_JSONL = demo_root / "pnl" / "logs" / "run_log.jsonl"
    import dashboard
    print(f"DEMO dashboard on http://127.0.0.1:{port}/  (synthetic data, no real money)")
    dashboard.app.run(host="127.0.0.1", port=port, debug=False, use_reloader=False)
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description="Build synthetic dashboard demo data")
    ap.add_argument("--serve", action="store_true", help="serve the dashboard after building")
    ap.add_argument("--port", type=int, default=8799)
    ap.add_argument("--date", default=None, help="date stamp for the today board (default: today)")
    args = ap.parse_args()
    build(DEMO_ROOT, args.date)
    print(f"demo data written to {DEMO_ROOT}")
    if args.serve:
        return serve(DEMO_ROOT, args.port)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
